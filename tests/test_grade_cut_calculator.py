import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from app.grade_cut_calculator import (
    grade5_cut_summary,
    load_grade5_cut_reports,
)


def _save_report_workbook(path: Path, scores: list[float]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "sheet1"
    ws["A1"] = "고사 : 1차 정기시험 교과목 : 테스트과목 만점 : 100.00"
    ws["A2"] = "반번호"
    ws["B2"] = 1
    ws["C2"] = 2
    for idx, row in enumerate(range(3, 13)):
        ws.cell(row, 1).value = idx + 1
        ws.cell(row, 2).value = scores[idx * 2]
        ws.cell(row, 3).value = scores[idx * 2 + 1]
    end_row = len(scores) + 3
    ws.cell(end_row, 1).value = "응시생수"
    wb.save(path)


class GradeCutCalculatorTests(unittest.TestCase):
    def test_grade_cut_uses_selected_report_only(self):
        with tempfile.TemporaryDirectory() as tmp:
            report_path = Path(tmp) / "지필평가 교과목별 일람표.xlsx"
            _save_report_workbook(report_path, list(range(80, 100)))

            report_only = load_grade5_cut_reports(str(report_path))
            report_summary = grade5_cut_summary(report_only[0]["scores"])
            self.assertEqual(report_summary["cut_rows"][0]["score"], 98.0)
            self.assertNotIn("참조 파일", report_only[0]["source_note"])

    def test_non_score_status_cells_are_excluded_and_reported(self):
        with tempfile.TemporaryDirectory() as tmp:
            report_path = Path(tmp) / "지필평가 교과목별 일람표.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "sheet1"
            ws["A1"] = "고사 : 1차 정기시험 교과목 : 테스트과목 만점 : 100.00"
            ws["A2"] = "반번호"
            ws["B2"] = 1
            ws["C2"] = 2
            rows = [
                (1, 100, 90),
                (2, "인정결", 80),
                (3, 70, "질병결"),
                (4, "자퇴", 60),
                (5, 50, "전출"),
            ]
            for row_idx, row in enumerate(rows, start=3):
                for col_idx, value in enumerate(row, start=1):
                    ws.cell(row_idx, col_idx).value = value
            ws["A8"] = "응시생수"
            wb.save(report_path)

            report = load_grade5_cut_reports(str(report_path))[0]
            self.assertEqual(len(report["scores"]), 6)
            self.assertEqual(len(report["excluded_entries"]), 4)
            self.assertIn("비점수 제외 4건", report["source_note"])

    def test_boundary_note_reports_tie_at_cut_score(self):
        scores = [100.0, 90.0, 90.0, 90.0] + [float(v) for v in range(80, 64, -1)]
        summary = grade5_cut_summary(scores)
        first_note = summary["boundary_notes"][0]["messages"]
        self.assertTrue(any("컷 점수 동점" in message for message in first_note))
        self.assertTrue(any("중간석차" in message for message in first_note))


if __name__ == "__main__":
    unittest.main()
