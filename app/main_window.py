"""
PySide6 기반 메인 윈도우 (KICE Shiny 웹앱 2.1.2 화면 구조에 맞춰 재구성).

화면 구조:
- 좌측 사이드바: 입력 파일, 분할점수, 옵션, 분석 실행
- 우측 탭: Data | 전체 성취도 | 문항 분석 | 답지반응분포 | 성취기준 분석 | 도움말

데이터는 모두 로컬에서만 처리되며, 외부로 전송되지 않는다.
"""
from __future__ import annotations

import csv
import html
import hashlib
import json
import math
import os
import queue
import re
import shutil
import socket
import subprocess
import sys
import tempfile
import threading
import traceback
import zipfile
from datetime import datetime
from pathlib import Path
import xml.etree.ElementTree as ET

from PySide6.QtCore import Qt, QSettings, QUrl, QSize, QTimer, QObject, Slot
from PySide6.QtGui import QAction, QActionGroup, QColor, QBrush, QFont, QKeySequence, QShortcut, QIcon, QPixmap, QDesktopServices
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QHBoxLayout, QVBoxLayout, QLabel,
    QPushButton, QLineEdit, QFileDialog, QFrame, QTabWidget, QGroupBox,
    QFormLayout, QDoubleSpinBox, QTableWidget, QTableWidgetItem, QHeaderView,
    QMessageBox, QSplitter, QComboBox, QSizePolicy, QCheckBox, QToolButton,
    QScrollArea, QAbstractItemView, QGridLayout, QTextBrowser, QPlainTextEdit, QInputDialog,
    QDialog, QLayout,
)

from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas

from .data_loader import load_exam, apply_perform, ExamData
from .analysis import (
    analyze_overall, analyze_items, build_score_matrix, analyze_serdap,
    grade_level, reliability_label, LEVELS,
)
from . import charts
from .theme import ThemeManager
from .cuts_loader import load_cut_scores
from .perform_loader import load_perform
from . import fonts as font_pack
from .grade_cut_calculator import (
    GRADE5_CUMULATIVE,
    format_score as _format_score,
    grade5_cut_summary as _grade5_cut_summary,
    load_grade5_cut_reports as _load_grade5_cut_reports,
    relative_grade_cut_points as _official_relative_grade_cut_points,
    relative_grade_labels as _official_relative_grade_labels,
)
from .ai_client import (
    AIProviderConfig, default_endpoint, default_model, list_ollama_models,
    get_ollama_version, list_openai_compatible_models, normalize_endpoint, parse_review_rows,
    check_codex_cli_oauth, find_codex_cli, probe_openai_compatible_chat, run_completion, scrub_personal_data,
)
from .ai_review_logic import (
    ai_review_expected_values_from_rates,
    ai_review_normalize_counts,
    ai_review_rate_summary,
    ai_review_target_threshold,
)
from .widgets import (
    StepperSpinBox, ItemBarDelegate, attach_wheel_zoom, NaturalItem,
    install_frozen_columns, refresh_frozen_columns,
)
from . import __version__

try:
    from PySide6.QtWebEngineWidgets import QWebEngineView
except Exception:
    QWebEngineView = None

try:
    from PySide6.QtWebEngineCore import QWebEnginePage
except Exception:
    QWebEnginePage = None

try:
    from PySide6.QtWebChannel import QWebChannel
except Exception:
    QWebChannel = None


APP_TITLE = "성취평가 결과 분석 (Goedu-Split)"
APP_AUTHOR = "이준서"
APP_VERSION = __version__
APP_COPYRIGHT = "© 2026 이준서. All rights reserved."
SIDEBAR_MIN_BASE_WIDTH = 300
SIDEBAR_DEFAULT_BASE_WIDTH = 360
SIDEBAR_FOLDER_BUTTON_TEXT = "폴더 불러오기"
SIDEBAR_REVERT_BUTTON_TEXT = "복원"
SIDEBAR_CSV_EXPORT_BUTTON_TEXT = "CSV 저장"
SIDEBAR_SPLITER_EXPORT_BUTTON_TEXT = "근거 엑셀 저장"
DEFAULT_CUTS = {"A": 90.0, "B": 80.0, "C": 70.0, "D": 60.0, "E": 40.0}
PERFORM_DEFAULT_RATES = {"A": 0.90, "B": 0.80, "C": 0.70, "D": 0.60, "E": 0.40}
LEVELS_AE = ["A", "B", "C", "D", "E"]
TARGET_RATE_PRESET_TITLE = "목표 성취수준 정답률 설정"
TARGET_RATE_MIN_PERCENT = 67
DEFAULT_TARGET_RATE_PRESETS = {
    "A": {"A": 70, "B": 45, "C": 25, "D": 10, "E": 0},
    "B": {"A": 85, "B": 70, "C": 45, "D": 25, "E": 10},
    "C": {"A": 90, "B": 80, "C": 70, "D": 45, "E": 25},
    "D": {"A": 95, "B": 90, "C": 80, "D": 70, "E": 45},
    "E": {"A": 95, "B": 95, "C": 90, "D": 80, "E": 70},
}
GRADE9_CUMULATIVE = [("1", 4.0), ("2", 11.0), ("3", 23.0), ("4", 40.0), ("5", 60.0), ("6", 77.0), ("7", 89.0), ("8", 96.0), ("9", 100.0)]
AI_EXPECTED_HEADERS = [f"{lv} 예상" for lv in LEVELS_AE]
AI_REVIEW_HEADERS = [
    "구분", "번호/요소", "성취기준 후보", "평가유형", "목표수준 후보", "난이도 후보",
    *AI_EXPECTED_HEADERS, "근거", "다음 확인",
]
AI_SOURCE_TEXT_LIMIT = 120_000
AI_REFERENCE_TEXT_LIMIT = 300_000
AI_SOURCE_CACHE_NAME = "_combined_ai_source_text.txt"
AI_REFERENCE_CACHE_NAME = "_combined_ai_reference_text.txt"
AI_REVIEW_LOCAL_CHUNK_SIZE = 3
AI_REVIEW_CLOUD_CHUNK_SIZE = 5
AI_REVIEW_CHUNK_SOURCE_LIMIT = 9_000
AI_REVIEW_CHUNK_REFERENCE_LIMIT = 8_000
AI_STANDARD_CODE_RE = re.compile(r"\[(?=[^\]\n]*\d)[^\]\n]{4,40}\]")
PRIVACY_REAL_TEXT_ROLE = Qt.UserRole + 601
PORTFOLIO_STUDENT_KEY_ROLE = Qt.UserRole + 602
NEIS_RATES_ROLE = Qt.UserRole + 701
NEIS_SOURCE_KEY_ROLE = Qt.UserRole + 702


def _app_icon_path() -> Path | None:
    here = Path(__file__).resolve()
    bundle_base = Path(getattr(sys, "_MEIPASS", here.parents[1]))
    candidates = [
        bundle_base / "assets" / "app_icon" / "goedusplit.png",
        here.parents[1] / "assets" / "app_icon" / "goedusplit.png",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def _class_no_sort_value(text: str) -> tuple[int, int, str]:
    raw = (text or "").strip()
    if "/" in raw:
        try:
            a, b = raw.split("/", 1)
            return (int(a), int(b), raw)
        except Exception:
            pass
    return (10**9, 10**9, raw)


def _is_fixed_cut_scores(cuts: dict[str, float], tolerance: float = 0.005) -> bool:
    return all(abs(float(cuts.get(lv, -999)) - DEFAULT_CUTS[lv]) <= tolerance for lv in DEFAULT_CUTS)


def _normal_cdf(x: float, mean: float, std: float) -> float:
    if std <= 0:
        return 1.0 if x >= mean else 0.0
    return 0.5 * (1.0 + math.erf((x - mean) / (std * math.sqrt(2.0))))


def _relative_grade_labels(scores: list[float], cumulative: list[tuple[str, float]]) -> list[str]:
    return _official_relative_grade_labels(scores, cumulative)


def _relative_grade_cut_points(scores: list[float], cumulative: list[tuple[str, float]], prefix: str) -> list[dict]:
    return _official_relative_grade_cut_points(scores, cumulative, prefix)


def build_data_empty_state_html() -> str:
    steps = [
        ("입력 데이터", "왼쪽에서 정오표와 문항정보표를 넣습니다. 폴더 일괄 불러오기를 쓰면 관련 파일을 한 번에 채울 수 있습니다."),
        ("분할점수 확인", "추정분할점수 파일이 있으면 불러오고, 없으면 왼쪽 분할점수를 직접 확인합니다."),
        ("분석 실행", "왼쪽 아래의 분석 실행을 누르면 성취도, 문항 분석, 성취기준 결과, 모니터링 자료가 만들어집니다."),
        ("근거 확인", "Data 탭의 문항 칸, 문항 분석, 성취수준별 답지반응을 보며 어느 문항이 쉬웠고 어려웠는지 확인합니다."),
        ("예상정답률 조정", "예상정답률 입력 탭에서 O/X 판단과 근거 추천을 보며 분할점수 산정 근거를 정리합니다."),
        ("모니터링·상담·내보내기", "모니터링 탭과 상담 모드, 포트폴리오, CSV/근거 엑셀 내보내기로 협의와 상담 자료를 마무리합니다."),
    ]
    items = "".join(
        "<li>"
        f"<b>{idx}. {html.escape(title)}</b>"
        f"<span>{html.escape(description)}</span>"
        "</li>"
        for idx, (title, description) in enumerate(steps, start=1)
    )
    return (
        "<h3>처음 시작하기</h3>"
        "<p>아직 분석 자료가 없습니다. 아래 순서대로 진행하면 됩니다.</p>"
        f"<ol>{items}</ol>"
        "<p class='muted'>AI 문항 검토는 선택 기능입니다. 처음에는 AI 설정을 건드리지 않아도 되며, "
        "AI 연결 없이도 기본 분석과 예상정답률 계산, 상담 모드, 내보내기를 사용할 수 있습니다.</p>"
    )

LUCIDE_PATHS = {
    "menu": '<path d="M4 6h16"/><path d="M4 12h16"/><path d="M4 18h16"/>',
    "minus": '<path d="M5 12h14"/>',
    "plus": '<path d="M12 5v14"/><path d="M5 12h14"/>',
    "rotate": '<path d="M3 12a9 9 0 1 0 9-9 9.75 9.75 0 0 0-6.74 2.74L3 8"/><path d="M3 3v5h5"/>',
    "moon": '<path d="M12 3a6 6 0 0 0 9 9 9 9 0 1 1-9-9Z"/>',
    "sun": '<circle cx="12" cy="12" r="4"/><path d="M12 2v2"/><path d="M12 20v2"/><path d="m4.93 4.93 1.41 1.41"/><path d="m17.66 17.66 1.41 1.41"/><path d="M2 12h2"/><path d="M20 12h2"/><path d="m6.34 17.66-1.41 1.41"/><path d="m19.07 4.93-1.41 1.41"/>',
    "shield": '<path d="M20 13c0 5-3.5 7.5-7.66 8.95a1 1 0 0 1-.67-.01C7.5 20.5 4 18 4 13V6a1 1 0 0 1 1-1c2 0 4.5-1.2 6.24-2.72a1.2 1.2 0 0 1 1.52 0C14.5 3.8 17 5 19 5a1 1 0 0 1 1 1z"/>',
    "chevron-down": '<path d="m6 9 6 6 6-6"/>',
    "chevron-right": '<path d="m9 18 6-6-6-6"/>',
}


def _lucide_icon(name: str, color: str, size: int = 18) -> QIcon:
    path = LUCIDE_PATHS.get(name, LUCIDE_PATHS["menu"])
    svg = (
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{size}" height="{size}" '
        f'viewBox="0 0 24 24" fill="none" stroke="{color}" stroke-width="2" '
        f'stroke-linecap="round" stroke-linejoin="round">{path}</svg>'
    )
    pix = QPixmap()
    pix.loadFromData(svg.encode("utf-8"), "SVG")
    return QIcon(pix)


# ---------------------------------------------------------------------------
# 헬퍼 위젯
# ---------------------------------------------------------------------------

class FileSelector(QWidget):
    """좁은 사이드바에서도 잘리지 않는 파일 선택 위젯."""
    BUTTON_BASE_WIDTH = 56

    def __init__(self, label: str, file_filter: str = "Excel (*.xlsx)", parent=None):
        super().__init__(parent)
        layout = QVBoxLayout(self); layout.setContentsMargins(0, 0, 0, 0); layout.setSpacing(4)
        self.label = QLabel(label)
        self.label.setWordWrap(True)
        self.label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.path_edit = QLineEdit(); self.path_edit.setReadOnly(True)
        self.path_edit.setPlaceholderText("xlsx")
        self.path_edit.setMinimumWidth(0)
        self.path_edit.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.btn = QPushButton("찾기")
        self.btn.setToolTip(f"{label} 파일 찾아보기")
        self.btn.setFixedWidth(self.BUTTON_BASE_WIDTH)
        self.btn.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.btn.clicked.connect(self._pick)
        row = QHBoxLayout(); row.setContentsMargins(0, 0, 0, 0); row.setSpacing(6)
        row.addWidget(self.path_edit, 1)
        row.addWidget(self.btn)
        layout.addWidget(self.label)
        layout.addLayout(row)
        self.file_filter = file_filter

    def _pick(self):
        path, _ = QFileDialog.getOpenFileName(self, "파일 선택", "", self.file_filter)
        if path:
            self.path_edit.setText(path)

    def path(self) -> str:
        return self.path_edit.text().strip()


class CanvasHolder(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._layout = QVBoxLayout(self); self._layout.setContentsMargins(0, 0, 0, 0)
        self._canvas: FigureCanvas | None = None

    def set_figure(self, fig):
        if self._canvas is not None:
            self._layout.removeWidget(self._canvas); self._canvas.setParent(None); self._canvas.deleteLater()
        self._canvas = FigureCanvas(fig)
        self._canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        # 마우스 휠로 줌, 더블클릭으로 원상 복귀
        attach_wheel_zoom(self._canvas)
        self._canvas.setToolTip("마우스 휠 = 확대/축소 · 더블클릭 = 원상복귀")
        self._layout.addWidget(self._canvas)


def _set_item(table: QTableWidget, r: int, c: int, value, *, align_right=False, bg=None,
              bold=False, align_left=False, tooltip: str | None = None):
    text = "" if value is None else str(value)
    item = QTableWidgetItem(text)
    if align_left:
        item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
    elif align_right:
        item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
    else:
        item.setTextAlignment(Qt.AlignCenter)
    if bg is not None:
        item.setBackground(QBrush(bg))
    if bold:
        f = item.font(); f.setBold(True); item.setFont(f)
    # 길게 잘릴 가능성 있는 셀은 무조건 툴팁(전체 텍스트). 외부에서 별도 지정하면 그것 우선.
    if tooltip:
        item.setToolTip(tooltip)
    elif len(text) > 8:
        item.setToolTip(text)
    table.setItem(r, c, item)
    return item


def _setup_table(t: QTableWidget, *, stretch_first: bool = False, word_wrap: bool = False,
                 horizontal_scroll: bool = False, row_height: int = 28):
    """공통 표 스타일: 한 줄 표시(말줄임) + 호버 툴팁 + 알맞은 행 높이."""
    t.verticalHeader().setVisible(False)
    t.setAlternatingRowColors(True)
    t.setSelectionBehavior(QTableWidget.SelectRows)
    t.setEditTriggers(QTableWidget.NoEditTriggers)
    t.setWordWrap(word_wrap)
    t.setTextElideMode(Qt.ElideRight)
    # 호버만으로 툴팁이 빨리 뜨도록 mouseTracking 활성화 + 툴팁 노출 시간 길게
    t.setMouseTracking(True)
    t.viewport().setMouseTracking(True)
    if not horizontal_scroll:
        t.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        t.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
    else:
        t.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        t.horizontalHeader().setStretchLastSection(False)
    if stretch_first:
        t.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
    # 행 높이 고정 → 줄바꿈으로 한 행이 너무 커지는 문제 방지
    t.setProperty("baseRowHeight", row_height)
    t.verticalHeader().setSectionResizeMode(QHeaderView.Fixed)
    t.verticalHeader().setDefaultSectionSize(row_height)


def _contrast_text_for_fill(hex_color: str) -> QColor:
    """셀 배경색 위에서 더 잘 보이는 흰색/진한색을 고른다."""
    try:
        h = hex_color.lstrip("#")
        if len(h) == 3:
            h = "".join(ch * 2 for ch in h)
        r, g, b = (int(h[i:i+2], 16) / 255 for i in (0, 2, 4))

        def lin(v):
            return v / 12.92 if v <= 0.03928 else ((v + 0.055) / 1.055) ** 2.4

        lum = 0.2126 * lin(r) + 0.7152 * lin(g) + 0.0722 * lin(b)
        return QColor("#ffffff" if lum < 0.179 else "#111827")
    except Exception:
        return QColor("#111827")


# ---------------------------------------------------------------------------
# 메인 윈도우
# ---------------------------------------------------------------------------

class SpliterWebBridge(QObject):
    """Qt WebEngine 내부 계산기에서 클립보드/파일 저장을 쓰기 위한 브리지."""

    def __init__(self, window):
        super().__init__(window)
        self.window = window

    @Slot(str, str, result=bool)
    def copyText(self, content: str, label: str = "") -> bool:
        try:
            QApplication.clipboard().setText(content or "")
            name = label.strip() if label else "NEIS 표"
            self.window.statusBar().showMessage(f"{name}을(를) 클립보드에 복사했습니다.", 6000)
            return True
        except Exception as exc:
            QMessageBox.critical(self.window, "NEIS 입력표", f"복사하지 못했습니다.\n{exc}")
            return False

    @Slot(str, str, str, result=bool)
    def saveTextFile(self, suggested_name: str, content: str, mime_type: str = "") -> bool:
        safe_name = re.sub(r'[\\/:*?"<>|]+', "_", (suggested_name or "goedu-neis.xls").strip())
        if not safe_name:
            safe_name = "goedu-neis.xls"
        suffix = Path(safe_name).suffix.lower()
        if suffix == ".xls":
            caption = "NEIS 입력표 엑셀 저장"
            file_filter = "Excel에서 열 수 있는 표 (*.xls);;HTML (*.html);;모든 파일 (*.*)"
        elif suffix == ".md":
            caption = "NEIS 입력표 마크다운 저장"
            file_filter = "Markdown (*.md);;텍스트 (*.txt);;모든 파일 (*.*)"
        elif suffix in {".tsv", ".txt"}:
            caption = "NEIS 입력표 텍스트 저장"
            file_filter = "탭 구분 텍스트 (*.tsv *.txt);;모든 파일 (*.*)"
        else:
            caption = "NEIS 입력표 저장"
            file_filter = "모든 파일 (*.*)"
        default_path = Path.home() / "Desktop" / safe_name
        path, _ = QFileDialog.getSaveFileName(self.window, caption, str(default_path), file_filter)
        if not path:
            return False
        target = Path(path)
        if suffix and not target.suffix:
            target = target.with_suffix(suffix)
        try:
            data = content or ""
            if data.startswith("\ufeff"):
                data = data[1:]
            target.write_text(data, encoding="utf-8-sig")
            self.window.statusBar().showMessage(f"NEIS 입력표를 저장했습니다: {target}", 6000)
            return True
        except Exception as exc:
            QMessageBox.critical(self.window, "NEIS 입력표", f"저장하지 못했습니다.\n{exc}")
            return False

    @Slot(result=str)
    def targetRatePresetsJson(self) -> str:
        try:
            return json.dumps(self.window._load_target_rate_presets(), ensure_ascii=False)
        except Exception:
            return json.dumps(DEFAULT_TARGET_RATE_PRESETS, ensure_ascii=False)

    @Slot(str, result=bool)
    def saveTargetRatePresetsJson(self, raw: str) -> bool:
        try:
            presets = json.loads(raw or "{}")
            self.window._save_target_rate_presets(presets)
            self.window._send_spliter_teacher_presets(apply_current=False)
            self.window.statusBar().showMessage(f"{TARGET_RATE_PRESET_TITLE}을 저장했습니다.", 6000)
            return True
        except Exception as exc:
            QMessageBox.critical(self.window, TARGET_RATE_PRESET_TITLE, f"저장하지 못했습니다.\n{exc}")
            return False

    @Slot(result=bool)
    def openTargetRatePresets(self) -> bool:
        try:
            self.window._open_target_rate_presets_dialog()
            return True
        except Exception as exc:
            QMessageBox.critical(self.window, TARGET_RATE_PRESET_TITLE, f"설정창을 열지 못했습니다.\n{exc}")
            return False


class MainWindow(QMainWindow):
    def __init__(self, theme_manager: ThemeManager | None = None):
        super().__init__()
        self.setWindowTitle(APP_TITLE)
        icon_path = _app_icon_path()
        if icon_path is not None:
            self.setWindowIcon(QIcon(str(icon_path)))
        self.resize(1320, 860)
        self.setMinimumSize(1080, 720)

        self.exam: ExamData | None = None
        self.overall = None
        self.item_stats: list = []
        self.perform_data = None
        self.spliter_view = None
        self._spliter_loaded = False
        self._spliter_load_requested = False
        self._spliter_load_signal_connected = False
        self._spliter_pending_payload = None
        self._spliter_pending_project_payload = None
        self._perform_recalc_dirty = False
        self.settings = QSettings("GoeduSplit", "Goedu-Split")
        self.theme = theme_manager
        # 화면 배율(%). 100이 기본이고 50~160 사이에서 조절한다.
        self._zoom = 100
        self._base_font_pt = 13

        # Qt에 번들 폰트 등록 → 시스템에 Gowun Dodum/NanumGothic이 없어도 동작
        try: font_pack.register_fonts()
        except Exception: pass

        self._build_ui()
        self._build_menu()
        self._build_statusbar()
        self._install_shortcuts()
        if self.theme is not None:
            self.theme.changed.connect(self._on_theme_changed)
            charts.set_theme(self.theme.colors)
        self.statusBar().showMessage("좌측에서 정오표·문항정보표를 지정하고 '분석 실행'을 누르세요.")

    def _icon_color(self) -> str:
        if self.theme is not None:
            return self.theme.colors.get("accent", "#2a7770")
        return "#2a7770"

    def _apply_tool_icon(self, button: QToolButton, name: str):
        button.setText("")
        button.setIcon(_lucide_icon(name, self._icon_color()))
        button.setIconSize(QSize(18, 18))

    # ------------------------------------------------------------------ UI
    def _build_ui(self):
        central = QWidget()
        outer = QVBoxLayout(central); outer.setContentsMargins(12, 8, 12, 8); outer.setSpacing(8)
        outer.addWidget(self._build_header_bar())

        self.splitter = QSplitter(Qt.Horizontal, self)
        self.input_panel = self._build_input_panel()
        self.input_panel.setMinimumWidth(self._sidebar_min_width())
        self.splitter.addWidget(self.input_panel)
        self.splitter.addWidget(self._build_tabs())
        self.splitter.setStretchFactor(0, 0); self.splitter.setStretchFactor(1, 1)
        side_w = self._sidebar_default_width()
        self.splitter.setSizes([side_w, max(720, 1320 - side_w)])
        # 좁아질 때 자동 접힘 가능
        self.splitter.setCollapsible(0, True)
        self.splitter.setCollapsible(1, False)
        outer.addWidget(self.splitter, 1)
        self.setCentralWidget(central)

    def _build_header_bar(self) -> QWidget:
        bar = QFrame(); bar.setProperty("role", "headerbar")
        h = QHBoxLayout(bar); h.setContentsMargins(8, 4, 8, 4); h.setSpacing(8)

        # 사이드바 토글 (좁은 창에서 입력 패널 보이기/숨기기)
        self.btn_sidebar = QToolButton()
        self.btn_sidebar.setProperty("role", "iconbtn")
        self.btn_sidebar.setToolTip("입력 패널 보이기/숨기기")
        self.btn_sidebar.clicked.connect(self._toggle_sidebar)
        self._apply_tool_icon(self.btn_sidebar, "menu")
        h.addWidget(self.btn_sidebar)

        title = QLabel("성취평가 결과 분석")
        title.setProperty("role", "appname")
        h.addWidget(title)
        ver = QLabel(f"v{APP_VERSION}")
        ver.setProperty("role", "muted")
        h.addWidget(ver)
        h.addStretch(1)

        self.btn_counsel_mode = QToolButton()
        self.btn_counsel_mode.setCheckable(True)
        self.btn_counsel_mode.setProperty("role", "iconbtn")
        self.btn_counsel_mode.setToolTip(
            "상담 모드 (Ctrl+Shift+H): 검색한 학생 외 이름과 반/번호를 화면에서 가립니다."
        )
        self.btn_counsel_mode.toggled.connect(self._on_counseling_mode_changed)
        self._apply_tool_icon(self.btn_counsel_mode, "shield")
        h.addWidget(self.btn_counsel_mode)

        # 줌 컨트롤
        self.btn_zoom_out = QToolButton()
        self.btn_zoom_out.setToolTip("화면 축소 (Ctrl+−, 최소 50%)")
        self.btn_zoom_out.setProperty("role", "iconbtn")
        self.btn_zoom_out.clicked.connect(lambda: self._set_zoom(self._zoom - 10))
        self._apply_tool_icon(self.btn_zoom_out, "minus")
        self.lbl_zoom = QLabel("100%")
        self.lbl_zoom.setProperty("role", "muted")
        self.lbl_zoom.setMinimumWidth(self._px(46))
        self.lbl_zoom.setAlignment(Qt.AlignCenter)
        self.btn_zoom_in = QToolButton()
        self.btn_zoom_in.setToolTip("화면 확대 (Ctrl++, 최대 160%)")
        self.btn_zoom_in.setProperty("role", "iconbtn")
        self.btn_zoom_in.clicked.connect(lambda: self._set_zoom(self._zoom + 10))
        self._apply_tool_icon(self.btn_zoom_in, "plus")
        self.btn_zoom_reset = QToolButton()
        self.btn_zoom_reset.setToolTip("기본 크기로 (Ctrl+0)")
        self.btn_zoom_reset.setProperty("role", "iconbtn")
        self.btn_zoom_reset.clicked.connect(lambda: self._set_zoom(100))
        self._apply_tool_icon(self.btn_zoom_reset, "rotate")
        for b in (self.btn_zoom_out, self.lbl_zoom, self.btn_zoom_in, self.btn_zoom_reset):
            h.addWidget(b)

        spacer = QLabel("│"); spacer.setProperty("role", "muted"); h.addWidget(spacer)

        # 라이트/다크 토글
        self.btn_theme = QToolButton()
        self.btn_theme.setProperty("role", "iconbtn")
        self.btn_theme.setToolTip("라이트/다크 전환  (Ctrl+1: 라이트 / Ctrl+2: 다크 / Ctrl+3: 자동)")
        self.btn_theme.clicked.connect(self._toggle_theme_button)
        h.addWidget(self.btn_theme)
        self._update_theme_button_icon()
        return bar

    def _install_shortcuts(self):
        QShortcut(QKeySequence("Ctrl+="), self, activated=lambda: self._set_zoom(self._zoom + 10))
        QShortcut(QKeySequence("Ctrl++"), self, activated=lambda: self._set_zoom(self._zoom + 10))
        QShortcut(QKeySequence("Ctrl+-"), self, activated=lambda: self._set_zoom(self._zoom - 10))
        QShortcut(QKeySequence("Ctrl+0"), self, activated=lambda: self._set_zoom(100))
        QShortcut(
            QKeySequence("Ctrl+Shift+H"),
            self,
            activated=lambda: self.btn_counsel_mode.setChecked(not self.btn_counsel_mode.isChecked())
            if hasattr(self, "btn_counsel_mode") else None,
        )

    def _zoom_factor(self) -> float:
        return max(0.5, min(1.6, float(self._zoom) / 100.0))

    def _zoom_percent(self) -> int:
        return round(self._zoom_factor() * 100)

    def _px(self, value: int | float) -> int:
        return max(1, int(round(float(value) * self._zoom_factor())))

    def _sidebar_default_width(self) -> int:
        return min(max(330, self._px(SIDEBAR_DEFAULT_BASE_WIDTH)), 460)

    def _sidebar_min_width(self) -> int:
        return min(max(280, self._px(SIDEBAR_MIN_BASE_WIDTH)), 420)

    def _set_scaled_column_width(self, table: QTableWidget, col: int, base_width: int):
        widths = dict(table.property("baseColumnWidths") or {})
        widths[int(col)] = int(base_width)
        table.setProperty("baseColumnWidths", widths)
        table.setColumnWidth(col, self._px(base_width))

    def _apply_zoom_to_surfaces(self):
        """QSS만으로는 안 바뀌는 고정 픽셀 표면까지 현재 배율에 맞춘다."""
        for table in self.findChildren(QTableWidget):
            base_row = table.property("baseRowHeight") or 28
            row_h = self._px(base_row)
            table.verticalHeader().setDefaultSectionSize(row_h)
            for r in range(table.rowCount()):
                table.setRowHeight(r, row_h)
            widths = table.property("baseColumnWidths") or {}
            for col, base_width in dict(widths).items():
                try:
                    table.setColumnWidth(int(col), self._px(float(base_width)))
                except Exception:
                    pass
            refresh_frozen_columns(table)
            table.viewport().update()
        for stepper in self.findChildren(StepperSpinBox):
            compact = bool(stepper.property("sidebarCompact"))
            button_width = (
                StepperSpinBox.SIDEBAR_BUTTON_BASE_WIDTH
                if compact else StepperSpinBox.BUTTON_BASE_WIDTH
            )
            button_height = (
                StepperSpinBox.SIDEBAR_BUTTON_BASE_HEIGHT
                if compact else StepperSpinBox.BUTTON_BASE_HEIGHT
            )
            value_width = (
                StepperSpinBox.SIDEBAR_VALUE_MIN_BASE_WIDTH
                if compact else StepperSpinBox.VALUE_MIN_BASE_WIDTH
            )
            base_height = (
                StepperSpinBox.SIDEBAR_BASE_HEIGHT
                if compact else StepperSpinBox.BASE_HEIGHT
            )
            stepper.btn_minus.setFixedWidth(self._px(button_width))
            stepper.btn_plus.setFixedWidth(self._px(button_width))
            stepper.spin.setMinimumWidth(self._px(value_width))
            control_h = max(
                self._px(button_height),
                stepper.spin.fontMetrics().height() + self._px(12 if compact else 18),
            )
            stepper_h = max(self._px(base_height), control_h + self._px(4))
            stepper.setFixedHeight(stepper_h)
            stepper.spin.setFixedHeight(control_h)
            stepper.btn_minus.setFixedHeight(control_h)
            stepper.btn_plus.setFixedHeight(control_h)
        for row in self.findChildren(QFrame):
            if row.property("role") == "inputrow":
                base_height = row.property("baseHeight") or 96
                label = next((child for child in row.findChildren(QLabel) if child.parent() is row), None)
                stepper = row.findChild(StepperSpinBox)
                compact = bool(stepper.property("sidebarCompact")) if stepper is not None else False
                row_layout = row.layout()
                if row_layout is not None:
                    row_layout.setSpacing(self._px(4 if compact else 6))
                    row_layout.setContentsMargins(0, 0, 0, self._px(4 if compact else 6))
                dynamic_height = self._px(base_height)
                if label is not None:
                    label_h = max(self._px(24 if compact else 26), label.sizeHint().height() + self._px(2))
                    label.setMinimumHeight(label_h)
                    dynamic_height = max(dynamic_height, label_h + self._px(12))
                if stepper is not None:
                    dynamic_height = max(
                        dynamic_height,
                        (label.minimumHeight() if label is not None else 0)
                        + self._px(4 if compact else 6)
                        + stepper.height()
                        + self._px(12),
                    )
                row.setFixedHeight(dynamic_height)
                row.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        for panel in self.findChildren(QFrame):
            if hasattr(panel, "body_widget"):
                self._refresh_collapsible_panel_size(panel)
        for selector in self.findChildren(FileSelector):
            selector.label.setMinimumWidth(0)
            selector.path_edit.setMinimumWidth(0)
            selector.btn.setFixedWidth(self._px(FileSelector.BUTTON_BASE_WIDTH))
        for button in self.findChildren(QToolButton):
            button.setIconSize(QSize(self._px(18), self._px(18)))
        if hasattr(self, "btn_clear_search"):
            self.btn_clear_search.setFixedWidth(self._px(32))
        if hasattr(self, "lbl_zoom"):
            self.lbl_zoom.setMinimumWidth(max(42, self._px(46)))
        if hasattr(self, "input_panel"):
            self.input_panel.setMinimumWidth(self._sidebar_min_width())
        if hasattr(self, "score_chart_tabs"):
            self.score_chart_tabs.setMinimumHeight(max(190, self._px(220)))
        if hasattr(self, "table_serdap"):
            self.table_serdap.setMaximumHeight(max(70, self._px(110)))
        self._send_spliter_zoom()

    def _set_zoom(self, percent: int):
        percent = max(50, min(160, int(round(percent))))
        self._zoom = percent
        pt = max(7, int(round(self._base_font_pt * self._zoom_factor())))
        if self.theme is not None:
            # 1) QSS 재빌드 + matplotlib rcParams 동기화 + 시그널 → _on_theme_changed → 모든 차트 재렌더
            self.theme.set_base_font_pt(pt)
        else:
            f = self.font(); f.setPointSize(pt); self.setFont(f)
            app = QApplication.instance()
            if app is not None:
                af = app.font(); af.setPointSize(pt); app.setFont(af)
            # ThemeManager가 없을 때도 차트는 다시 그려야 폰트 반영
            self._refresh_all_charts()
        self._apply_zoom_to_surfaces()
        # % 라벨 갱신 (기본 13pt = 100%)
        pct = self._zoom_percent()
        if hasattr(self, "lbl_zoom"):
            self.lbl_zoom.setText(f"{pct}%")
        self.statusBar().showMessage(f"화면 배율 {pct}% · 글자 {pt}pt", 2500)

    def _toggle_sidebar(self):
        sizes = self.splitter.sizes()
        if sizes[0] <= 4:
            # 펼치기
            side_w = self._sidebar_default_width()
            self.splitter.setSizes([side_w, max(600, sum(sizes) - side_w)])
        else:
            # 접기
            self.splitter.setSizes([0, sum(sizes)])

    def _show_sidebar(self):
        if not hasattr(self, "splitter"):
            return
        sizes = self.splitter.sizes()
        total = sum(sizes) if sizes else 1320
        if not sizes or sizes[0] <= 4:
            side_w = self._sidebar_default_width()
            self.splitter.setSizes([side_w, max(600, total - side_w)])
        self.statusBar().showMessage("입력 패널을 열었습니다. 정오표와 문항정보표를 지정한 뒤 분석 실행을 누르세요.", 4000)

    def resizeEvent(self, ev):
        super().resizeEvent(ev)
        try:
            self._apply_responsive_tab_labels()
            w = self.width()
            sizes = self.splitter.sizes()
            if w < 1180 and sizes[0] > 0:
                self.splitter.setSizes([0, sum(sizes)])
            elif w >= 1280 and sizes[0] == 0:
                side_w = self._sidebar_default_width()
                self.splitter.setSizes([side_w, max(600, sum(sizes) - side_w)])
            # KPI 카드 자동 줄바꿈
            if hasattr(self, "_kpis") and self._kpis:
                content_w = max(300, w - sizes[0] - 60)
                cols = 3 if content_w >= 720 else (2 if content_w >= 480 else 1)
                # 현재 cols와 다르면 재배치
                cur_cols = max(1, self.kpi_grid.columnCount() or 1)
                if cur_cols != cols:
                    self._relayout_kpis(cols)
        except Exception:
            pass

    def _toggle_theme_button(self):
        if self.theme is None:
            return
        # auto이면 dark→light로 토글, dark이면 light로, light이면 dark로
        cur = self.theme.effective
        nxt = "light" if cur == "dark" else "dark"
        self.theme.set_mode(nxt)

    def _update_theme_button_icon(self):
        if self.theme is None:
            self._apply_tool_icon(self.btn_theme, "moon")
            return
        eff = self.theme.effective
        # 다크일 때는 해(라이트로 가는 버튼), 라이트일 때는 달(다크로 가는 버튼)
        self._apply_tool_icon(self.btn_theme, "sun" if eff == "dark" else "moon")

    def _refresh_header_icons(self):
        for button, icon_name in (
            (getattr(self, "btn_sidebar", None), "menu"),
            (getattr(self, "btn_counsel_mode", None), "shield"),
            (getattr(self, "btn_zoom_out", None), "minus"),
            (getattr(self, "btn_zoom_in", None), "plus"),
            (getattr(self, "btn_zoom_reset", None), "rotate"),
        ):
            if button is not None:
                self._apply_tool_icon(button, icon_name)
        if hasattr(self, "btn_spliter_summary_toggle") and hasattr(self, "spliter_summary_body"):
            self._apply_tool_icon(
                self.btn_spliter_summary_toggle,
                "chevron-down" if self.spliter_summary_body.isVisible() else "chevron-right",
            )
        if hasattr(self, "btn_theme"):
            self._update_theme_button_icon()

    def _add_sidebar_stepper_row(self, layout: QVBoxLayout, label: str, stepper: StepperSpinBox):
        row = QFrame()
        row.setProperty("role", "inputrow")
        row.setProperty("baseHeight", StepperSpinBox.SIDEBAR_ROW_BASE_HEIGHT)
        row.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        row_layout = QVBoxLayout(row)
        row_layout.setContentsMargins(0, 0, 0, self._px(4))
        row_layout.setSpacing(self._px(4))
        row_layout.setAlignment(Qt.AlignTop)
        lab = QLabel(label)
        lab.setWordWrap(False)
        lab.setProperty("role", "field-label")
        lab.setMinimumHeight(max(self._px(24), lab.sizeHint().height() + self._px(2)))
        lab.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        stepper.setProperty("sidebarCompact", True)
        control_h = max(
            self._px(StepperSpinBox.SIDEBAR_BUTTON_BASE_HEIGHT),
            stepper.spin.fontMetrics().height() + self._px(12),
        )
        stepper_h = max(self._px(StepperSpinBox.SIDEBAR_BASE_HEIGHT), control_h + self._px(4))
        stepper.spin.setMinimumWidth(self._px(StepperSpinBox.SIDEBAR_VALUE_MIN_BASE_WIDTH))
        stepper.setFixedHeight(stepper_h)
        stepper.spin.setFixedHeight(control_h)
        stepper.btn_minus.setFixedWidth(self._px(StepperSpinBox.SIDEBAR_BUTTON_BASE_WIDTH))
        stepper.btn_plus.setFixedWidth(self._px(StepperSpinBox.SIDEBAR_BUTTON_BASE_WIDTH))
        stepper.btn_minus.setFixedHeight(control_h)
        stepper.btn_plus.setFixedHeight(control_h)
        stepper.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        row_layout.addWidget(lab)
        row_layout.addWidget(stepper)
        required_height = max(
            self._px(StepperSpinBox.SIDEBAR_ROW_BASE_HEIGHT),
            lab.minimumHeight() + row_layout.spacing() + stepper.height() + self._px(12),
        )
        row.setFixedHeight(required_height)
        layout.addWidget(row)

    def _build_input_panel(self) -> QWidget:
        outer = QScrollArea(); outer.setWidgetResizable(True)
        outer.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        w = QFrame(); outer.setWidget(w)
        v = QVBoxLayout(w); v.setContentsMargins(12, 12, 12, 12); v.setSpacing(10)

        title = QLabel("입력 데이터")
        title.setProperty("role", "title")
        v.addWidget(title)

        # 입력 파일 그룹
        files_body = QWidget()
        ff = QVBoxLayout(files_body); ff.setContentsMargins(0, 0, 0, 0); ff.setSpacing(8)

        # 폴더 일괄 업로드 버튼 (가장 위에 강조)
        folder_btn = QPushButton(SIDEBAR_FOLDER_BUTTON_TEXT)
        folder_btn.setProperty("role", "primary-soft")
        folder_btn.setMinimumHeight(34)
        folder_btn.clicked.connect(self._pick_folder_batch)
        folder_btn.setToolTip(
            "폴더 안의 정오표/문항정보표/추정분할점수/수행평가 파일을\n"
            "파일 이름 패턴으로 자동 분류해 한 번에 채워 넣습니다."
        )
        ff.addWidget(folder_btn)

        self.fs_response = FileSelector("학생답 정오표 data")
        self.fs_iteminfo = FileSelector("문항정보표")
        self.fs_cuts = FileSelector("예상추정분할점수 조회")
        self.fs_cuts.path_edit.textChanged.connect(self._on_cuts_path_changed)
        ff.addWidget(self.fs_response); ff.addWidget(self.fs_iteminfo); ff.addWidget(self.fs_cuts)
        cuts_help = QLabel("※ 추정분할점수 파일을 지정하면 아래 분할점수가 자동 입력됩니다.")
        cuts_help.setProperty("role", "muted"); cuts_help.setWordWrap(True)
        ff.addWidget(cuts_help)
        v.addWidget(self._make_collapsible_panel("입력 파일", files_body))

        # 2022 개정 5등급 컷 계산
        grade5_body = QWidget()
        gf = QVBoxLayout(grade5_body); gf.setContentsMargins(0, 0, 0, 0); gf.setSpacing(8)
        self.fs_grade5_report = FileSelector("교과목별 일람표")
        gf.addWidget(self.fs_grade5_report)
        self.lbl_grade5_cut_info = QLabel(
            "지필평가 교과목별 일람표의 실제 숫자 점수만 기준으로 계산합니다. 인정결·질병결·자퇴·전출 등 비점수 값은 제외합니다."
        )
        self.lbl_grade5_cut_info.setProperty("role", "muted")
        self.lbl_grade5_cut_info.setWordWrap(True)
        gf.addWidget(self.lbl_grade5_cut_info)
        btn_grade5_cut = QPushButton("1-5등급 컷 계산")
        btn_grade5_cut.clicked.connect(self.calculate_grade5_cuts_from_report)
        gf.addWidget(btn_grade5_cut)
        v.addWidget(self._make_collapsible_panel("1-5등급 컷 계산", grade5_body))

        # 수행평가 (선택)
        perf_body = QWidget()
        pf = QVBoxLayout(perf_body); pf.setContentsMargins(0, 0, 0, 0); pf.setSpacing(8)
        self.chk_perform = QCheckBox("수행평가도 포함하여 분석합니다.")
        self.chk_perform.toggled.connect(self._toggle_perform)
        self.fs_perform = FileSelector("수행평가 결과 (.xlsx)")
        self.fs_perform.setEnabled(False)
        self.fs_perform.path_edit.textChanged.connect(self._on_perform_path_changed)
        self.lbl_perform_info = QLabel("(미설정)")
        self.lbl_perform_info.setProperty("role", "muted"); self.lbl_perform_info.setWordWrap(True)
        pf.addWidget(self.chk_perform); pf.addWidget(self.fs_perform); pf.addWidget(self.lbl_perform_info)
        v.addWidget(self._make_collapsible_panel("수행평가 (선택)", perf_body))

        # 점수 반영비율
        weight_body = QWidget()
        wf = QVBoxLayout(weight_body); wf.setContentsMargins(0, 0, 0, 0); wf.setSpacing(6)
        self.spin_pencil_ratio = StepperSpinBox(value=100, minimum=0, maximum=100, step=5, decimals=0, suffix=" %")
        self.spin_perform_ratio = StepperSpinBox(value=0, minimum=0, maximum=100, step=5, decimals=0, suffix=" %")
        self.spin_perform_ratio.setEnabled(False)
        self.spin_pencil_ratio.valueChanged.connect(self._sync_pencil_to_perform)
        self.spin_perform_ratio.valueChanged.connect(self._sync_perform_to_pencil)
        self._add_sidebar_stepper_row(wf, "지필평가 반영비율", self.spin_pencil_ratio)
        self._add_sidebar_stepper_row(wf, "수행평가 반영비율", self.spin_perform_ratio)
        v.addWidget(self._make_collapsible_panel("반영비율 (지필 + 수행 = 100)", weight_body))

        # 분할점수
        cuts_body = QWidget()
        form = QVBoxLayout(cuts_body); form.setContentsMargins(0, 0, 0, 0); form.setSpacing(6)
        self.spin_cuts = {}
        labels = {"A": "A/B 분할점수", "B": "B/C 분할점수",
                  "C": "C/D 분할점수", "D": "D/E 분할점수",
                  "E": "E/미도달 분할점수"}
        for lv in ["A", "B", "C", "D", "E"]:
            sp = StepperSpinBox(value=DEFAULT_CUTS[lv], minimum=0, maximum=100,
                                step=1.0, decimals=2)
            sp.valueChanged.connect(self._update_cut_box_title)
            sp.valueChanged.connect(lambda *_: self._send_spliter_manual_cuts())
            self.spin_cuts[lv] = sp
            self._add_sidebar_stepper_row(form, labels[lv], sp)
        self.cuts_box = self._make_collapsible_panel("분할점수 (고정 분할 방식)", cuts_body)
        v.addWidget(self.cuts_box)

        # 액션 버튼
        # 액션 버튼들
        run_btn = QPushButton("분석 실행")
        run_btn.setProperty("role", "primary")
        run_btn.setMinimumHeight(40)
        run_btn.clicked.connect(self.run_analysis)
        v.addWidget(run_btn)

        action_row = QHBoxLayout(); action_row.setSpacing(6)
        self.btn_revert = QPushButton(SIDEBAR_REVERT_BUTTON_TEXT)
        self.btn_revert.setToolTip("마지막으로 분석 실행했을 때의 입력값으로 되돌립니다.\n실수로 분할점수나 파일 경로를 건드렸을 때 사용하세요.")
        self.btn_revert.clicked.connect(self._revert_inputs)
        self.btn_revert.setEnabled(False)
        action_row.addWidget(self.btn_revert)
        export_btn = QPushButton(SIDEBAR_CSV_EXPORT_BUTTON_TEXT)
        export_btn.setToolTip("현재 분석 결과를 CSV 파일로 저장합니다.")
        export_btn.clicked.connect(self.export_csv)
        action_row.addWidget(export_btn)
        v.addLayout(action_row)

        spliter_btn = QPushButton(SIDEBAR_SPLITER_EXPORT_BUTTON_TEXT)
        spliter_btn.setToolTip("분석된 학생 성취수준과 문항 정오 데이터를 예상정답률 계산기 추천 근거 엑셀로 저장합니다.")
        spliter_btn.clicked.connect(self.export_spliter_evidence)
        v.addWidget(spliter_btn)

        v.addStretch(1)
        self.exam_info_lbl = QLabel("(분석 전)")
        self.exam_info_lbl.setWordWrap(True)
        self.exam_info_lbl.setProperty("role", "muted")
        v.addWidget(self.exam_info_lbl)

        credit = QLabel(APP_COPYRIGHT)
        credit.setProperty("role", "credit")
        credit.setAlignment(Qt.AlignRight)
        v.addWidget(credit)
        self._update_cut_box_title()
        return outer

    def _current_cut_scores_from_inputs(self) -> dict[str, float]:
        if not hasattr(self, "spin_cuts"):
            return dict(DEFAULT_CUTS)
        return {lv: float(self.spin_cuts[lv].value()) for lv in ["A", "B", "C", "D", "E"]}

    def _update_cut_box_title(self, *_):
        if not hasattr(self, "cuts_box"):
            return
        mode = "고정 분할 방식" if _is_fixed_cut_scores(self._current_cut_scores_from_inputs()) else "추정 분할 방식"
        title = f"분할점수 ({mode})"
        if hasattr(self.cuts_box, "toggle_button"):
            self.cuts_box.toggle_button.setText(title)
        elif hasattr(self.cuts_box, "setTitle"):
            self.cuts_box.setTitle(title)

    def _current_cut_scores_label(self) -> str:
        return "기본" if _is_fixed_cut_scores(self._current_cut_scores_from_inputs()) else "직접"

    def _snapshot_inputs(self):
        """현재 사이드바 입력값 전체를 dict로 보존 → 나중에 복원."""
        self._last_inputs = {
            "response": self.fs_response.path(),
            "iteminfo": self.fs_iteminfo.path(),
            "cuts_file": self.fs_cuts.path(),
            "perform": self.fs_perform.path(),
            "perform_on": self.chk_perform.isChecked(),
            "ratio_p": self.spin_pencil_ratio.value(),
            "ratio_f": self.spin_perform_ratio.value(),
            "cuts": {lv: self.spin_cuts[lv].value() for lv in ["A","B","C","D","E"]},
        }

    def _revert_inputs(self):
        s = getattr(self, "_last_inputs", None)
        if not s:
            return
        self.fs_response.path_edit.setText(s["response"])
        self.fs_iteminfo.path_edit.setText(s["iteminfo"])
        self.fs_cuts.path_edit.setText(s["cuts_file"])
        self.fs_perform.path_edit.setText(s["perform"])
        self.chk_perform.setChecked(s["perform_on"])
        self.spin_pencil_ratio.spin.blockSignals(True)
        self.spin_pencil_ratio.setValue(s["ratio_p"]); self.spin_pencil_ratio.spin.blockSignals(False)
        self.spin_perform_ratio.spin.blockSignals(True)
        self.spin_perform_ratio.setValue(s["ratio_f"]); self.spin_perform_ratio.spin.blockSignals(False)
        for lv in ["A","B","C","D","E"]:
            self.spin_cuts[lv].setValue(s["cuts"][lv])
        self._update_cut_box_title()
        self.statusBar().showMessage("마지막 분석값으로 입력을 복원했습니다.", 4000)

    def _search_terms(self, text: str) -> list[str]:
        return [part.strip().lower() for part in text.replace("\n", ",").split(",") if part.strip()]

    def _student_matches_terms(self, st, terms: list[str]) -> bool:
        if not terms:
            return True
        haystack = " ".join([
            str(st.name or ""),
            str(st.class_no or ""),
            str(st.sid or ""),
            str(st.grade_class or ""),
        ]).lower()
        return any(term in haystack for term in terms)

    def _matching_student_indices(self, text: str) -> list[int]:
        if self.exam is None:
            return []
        terms = self._search_terms(text)
        if not terms:
            return []
        return [i for i, st in enumerate(self.exam.students) if self._student_matches_terms(st, terms)]

    def _counseling_mode_enabled(self) -> bool:
        return bool(
            hasattr(self, "btn_counsel_mode")
            and self.btn_counsel_mode.isChecked()
        )

    def _counseling_only_target_enabled(self) -> bool:
        return bool(
            hasattr(self, "chk_counsel_only_target")
            and self.chk_counsel_only_target.isChecked()
        )

    def _real_item_text(self, item: QTableWidgetItem | None) -> str:
        if item is None:
            return ""
        real = item.data(PRIVACY_REAL_TEXT_ROLE)
        return str(real if real is not None else item.text())

    def _data_privacy_target_indices(self, text: str | None = None) -> set[int]:
        if self.exam is None:
            return set()
        raw = self.le_search.text() if text is None and hasattr(self, "le_search") else (text or "")
        return set(self._matching_student_indices(raw))

    def _mask_identity_cell(self, item: QTableWidgetItem | None, real_text: str, masked_text: str, reveal: bool):
        if item is None:
            return
        if item.data(PRIVACY_REAL_TEXT_ROLE) is None:
            item.setData(PRIVACY_REAL_TEXT_ROLE, real_text)
        if reveal:
            item.setText(real_text or "상담 학생")
            item.setToolTip(real_text or "상담 학생")
        else:
            item.setText(masked_text)
            item.setToolTip("상담 모드에서 개인정보를 가렸습니다.")

    def _apply_data_privacy_masks(self):
        if not hasattr(self, "table_data") or self.table_data.rowCount() == 0:
            return
        privacy = self._counseling_mode_enabled()
        targets = self._data_privacy_target_indices()
        sorting_enabled = self.table_data.isSortingEnabled()
        self.table_data.setSortingEnabled(False)
        for r in range(self.table_data.rowCount()):
            cls_item = self.table_data.item(r, 0)
            name_item = self.table_data.item(r, 1)
            original_idx = cls_item.data(Qt.UserRole + 101) if cls_item is not None else None
            is_target = privacy and bool(targets) and original_idx in targets
            real_cls = self._real_item_text(cls_item)
            real_name = self._real_item_text(name_item)
            if privacy:
                self._mask_identity_cell(cls_item, real_cls, "비공개", is_target)
                self._mask_identity_cell(name_item, real_name, f"익명 {r + 1:03d}", is_target)
            else:
                self._mask_identity_cell(cls_item, real_cls, real_cls, True)
                self._mask_identity_cell(name_item, real_name, real_name or "-", True)
        self.table_data.setSortingEnabled(sorting_enabled)
        frozen = getattr(self.table_data, "frozenView", None)
        if frozen is not None:
            for r in range(self.table_data.rowCount()):
                frozen.setRowHidden(r, self.table_data.isRowHidden(r))
        refresh_frozen_columns(self.table_data)

    def _portfolio_privacy_target_rows(self) -> set[int]:
        if not hasattr(self, "table_portfolio"):
            return set()
        terms = self._search_terms(self.le_portfolio_search.text() if hasattr(self, "le_portfolio_search") else "")
        if not terms:
            return set()
        targets = set()
        for r in range(self.table_portfolio.rowCount()):
            cls = self._real_item_text(self.table_portfolio.item(r, 3)).lower()
            name = self._real_item_text(self.table_portfolio.item(r, 4)).lower()
            if any(term in cls or term in name for term in terms):
                targets.add(r)
        return targets

    def _apply_portfolio_privacy_masks(self):
        if not hasattr(self, "table_portfolio"):
            return
        privacy = self._counseling_mode_enabled()
        targets = self._portfolio_privacy_target_rows()
        sorting_enabled = self.table_portfolio.isSortingEnabled()
        self.table_portfolio.setSortingEnabled(False)
        for r in range(self.table_portfolio.rowCount()):
            cls_item = self.table_portfolio.item(r, 3)
            name_item = self.table_portfolio.item(r, 4)
            real_cls = self._real_item_text(cls_item)
            real_name = self._real_item_text(name_item)
            is_target = privacy and bool(targets) and r in targets
            if privacy:
                self._mask_identity_cell(cls_item, real_cls, "비공개", is_target)
                self._mask_identity_cell(name_item, real_name, f"익명 {r + 1:03d}", is_target)
            else:
                self._mask_identity_cell(cls_item, real_cls, real_cls, True)
                self._mask_identity_cell(name_item, real_name, real_name or "-", True)
        self.table_portfolio.setSortingEnabled(sorting_enabled)

    def _apply_privacy_to_visible_surfaces(self):
        self._apply_data_privacy_masks()
        self._apply_portfolio_privacy_masks()
        if hasattr(self, "le_search"):
            self._filter_data_table(self.le_search.text())
        if hasattr(self, "le_portfolio_search"):
            self._filter_portfolio_table(self.le_portfolio_search.text())

    def _on_counseling_mode_changed(self, enabled: bool):
        if hasattr(self, "chk_counsel_only_target"):
            self.chk_counsel_only_target.setEnabled(enabled)
        if hasattr(self, "btn_counsel_mode"):
            self.btn_counsel_mode.setToolTip(
                "상담 모드 켜짐: 검색 학생 외 이름과 반/번호를 가립니다. 다시 누르면 해제됩니다."
                if enabled else
                "상담 모드 (Ctrl+Shift+H): 검색한 학생 외 이름과 반/번호를 화면에서 가립니다."
            )
        self._apply_privacy_to_visible_surfaces()
        message = (
            "상담 모드 켜짐 · 검색한 학생 외 이름과 반/번호를 가립니다."
            if enabled else "상담 모드 꺼짐 · 학생 이름을 다시 표시합니다."
        )
        self.statusBar().showMessage(message, 6000)

    def _highlight_data_from_search(self, text: str) -> tuple[list[float], list[str]]:
        matches = self._matching_student_indices(text)
        scores, labels = [], []
        if self.exam is None:
            return scores, labels
        privacy = self._counseling_mode_enabled()
        for seq, idx in enumerate(matches[:10], start=1):
            st = self.exam.students[idx]
            name = f"상담 학생 {seq}" if privacy else (st.name or st.class_no or st.sid)
            scores.append(float(st.final_score))
            labels.append(f"{name} {st.final_score:.1f}")
        return scores, labels

    def _current_relative_grades(self) -> tuple[list[str], list[str]]:
        if self.exam is None:
            return [], []
        totals = [float(st.final_score) for st in self.exam.students]
        return (
            _relative_grade_labels(totals, GRADE9_CUMULATIVE),
            _relative_grade_labels(totals, GRADE5_CUMULATIVE),
        )

    def _current_grade_cut_points(self) -> list[dict]:
        if self.exam is None:
            return []
        totals = [float(st.final_score) for st in self.exam.students]
        return (
            _relative_grade_cut_points(totals, GRADE9_CUMULATIVE, "9등급")
            + _relative_grade_cut_points(totals, GRADE5_CUMULATIVE, "5등급")
        )

    def _chart_option_checked(self, attr: str, default: bool = True) -> bool:
        widget = getattr(self, attr, None)
        return bool(widget.isChecked()) if widget is not None else default

    def _visible_grade_cut_points(self) -> list[dict]:
        points = self._current_grade_cut_points()
        show_grade9 = self._chart_option_checked("chk_chart_grade9", True)
        show_grade5 = self._chart_option_checked("chk_chart_grade5", True)
        return [
            point for point in points
            if (point.get("kind") == "9등급" and show_grade9)
            or (point.get("kind") == "5등급" and show_grade5)
        ]

    def _refresh_score_histogram_options(self):
        self._render_score_histogram_for_search(self.le_search.text() if hasattr(self, "le_search") else "")

    def _render_score_histogram_for_search(self, text: str = ""):
        if self.exam is None or self.overall is None:
            return
        totals = [s.final_score for s in self.exam.students]
        scores, labels = self._highlight_data_from_search(text)
        if not self._chart_option_checked("chk_chart_student_markers", True):
            scores, labels = [], []
        elif not self._chart_option_checked("chk_chart_student_labels", False):
            labels = ["" for _ in scores]
        self.canvas_score_hist.set_figure(
            charts.fig_score_histogram_colored(
                totals,
                self.overall.levels_arr,
                achievement_cuts=self.exam.cut_scores if self._chart_option_checked("chk_chart_achievement_cuts", True) else None,
                grade_cut_points=self._visible_grade_cut_points(),
                highlight_scores=scores,
                highlight_labels=labels,
            )
        )

    def _sync_frozen_row_hidden(self, row: int, hidden: bool):
        frozen = getattr(self.table_data, "frozenView", None)
        if frozen is not None:
            frozen.setRowHidden(row, hidden)

    def _filter_data_table(self, text: str):
        """학생 응답표를 이름/반·번호로 즉시 필터링하고 그래프에 위치를 표시."""
        terms = self._search_terms(text)
        if not hasattr(self, "table_data") or self.table_data.rowCount() == 0:
            return
        level_filter = self.combo_filter_level.currentData() if hasattr(self, "combo_filter_level") else ""
        grade9_filter = self.combo_filter_grade9.currentData() if hasattr(self, "combo_filter_grade9") else ""
        grade5_filter = self.combo_filter_grade5.currentData() if hasattr(self, "combo_filter_grade5") else ""
        score_col = getattr(self, "_data_score_col", 3)
        level_col = getattr(self, "_data_level_col", 2)
        grade9_col = getattr(self, "_data_grade9_col", 3)
        grade5_col = getattr(self, "_data_grade5_col", 4)
        privacy = self._counseling_mode_enabled()
        targets = self._data_privacy_target_indices(text)
        target_only = privacy and self._counseling_only_target_enabled() and bool(terms)
        visible_rows = []
        for r in range(self.table_data.rowCount()):
            cls_item = self.table_data.item(r, 0)
            name_item = self.table_data.item(r, 1)
            cls = self._real_item_text(cls_item).lower()
            name = self._real_item_text(name_item).lower()
            original_idx = cls_item.data(Qt.UserRole + 101) if cls_item is not None else None
            level = self.table_data.item(r, level_col).text() if self.table_data.item(r, level_col) else ""
            grade9 = self.table_data.item(r, grade9_col).text() if self.table_data.item(r, grade9_col) else ""
            grade5 = self.table_data.item(r, grade5_col).text() if self.table_data.item(r, grade5_col) else ""
            text_match = True if not terms else any(term in cls or term in name for term in terms)
            level_match = not level_filter or level == level_filter
            grade9_match = not grade9_filter or grade9 == str(grade9_filter)
            grade5_match = not grade5_filter or grade5 == str(grade5_filter)
            target_match = not target_only or original_idx in targets
            visible = text_match and level_match and grade9_match and grade5_match and target_match
            if visible:
                self.table_data.setRowHidden(r, False)
                self._sync_frozen_row_hidden(r, False)
                visible_rows.append(r)
            else:
                self.table_data.setRowHidden(r, True)
                self._sync_frozen_row_hidden(r, True)
        if privacy and not terms:
            self.lbl_search_status.setText("상담 모드: 학생 이름·반번호를 가렸습니다. 상담 학생을 검색하면 해당 학생만 실제 이름으로 보입니다.")
        elif privacy:
            self.lbl_search_status.setText(
                f"상담 모드 · 표시 {len(visible_rows)}명 · 다른 학생 이름·반번호 비공개"
                if visible_rows else "상담 모드 · 검색 결과 없음"
            )
        elif not terms and not any([level_filter, grade9_filter, grade5_filter]):
            self.lbl_search_status.setText("")
        else:
            names = []
            for r in visible_rows[:4]:
                cls = self._real_item_text(self.table_data.item(r, 0))
                name = self._real_item_text(self.table_data.item(r, 1))
                score_item = self.table_data.item(r, score_col) if self.table_data.columnCount() > score_col else None
                score = f" {score_item.text()}점" if score_item is not None else ""
                names.append(f"{name}({cls}{score})")
            more = f" 외 {len(visible_rows) - 4}명" if len(visible_rows) > 4 else ""
            self.lbl_search_status.setText(
                f"검색 {len(visible_rows)}명: {', '.join(names)}{more}" if visible_rows else "검색 결과 없음"
            )
        if visible_rows:
            self.table_data.selectRow(visible_rows[0])
            first_item = self.table_data.item(visible_rows[0], 0)
            if first_item is not None:
                self.table_data.scrollToItem(first_item, QAbstractItemView.PositionAtCenter)
        self._apply_data_privacy_masks()
        self._render_score_histogram_for_search(text)

    def _portfolio_store_dir(self) -> Path:
        return self._ensure_writable_dir(self._ai_material_root_dir() / "subject_snapshots", "subject_snapshots")

    def _portfolio_student_key(self, row: dict) -> str:
        sid = str(row.get("sid", "") or "").strip()
        class_no = str(row.get("class_no", "") or "").strip()
        name = str(row.get("name", "") or "").strip()
        if sid:
            return f"sid:{sid}|class:{class_no}|name:{name}"
        return f"class:{class_no}|name:{name}"

    def _portfolio_student_label(self, row: dict, subject_count: int | None = None, record_count: int | None = None) -> str:
        name = str(row.get("name", "") or "이름 없음")
        class_no = str(row.get("class_no", "") or "반/번호 없음")
        tail = []
        if subject_count is not None:
            tail.append(f"{subject_count}과목")
        if record_count is not None:
            tail.append(f"{record_count}개 기록")
        return f"{name} ({class_no})" + (f" · {' · '.join(tail)}" if tail else "")

    def _portfolio_rows_for_key(self, key: str) -> list[dict]:
        rows = getattr(self, "_portfolio_rows_cache", None)
        if rows is None:
            rows = self._load_portfolio_rows()
        return [row for row in rows if self._portfolio_student_key(row) == key]

    def _current_subject_snapshot(self) -> dict | None:
        if self.exam is None or self.overall is None:
            return None
        grade9_labels, grade5_labels = self._current_relative_grades()
        students = []
        for idx, st in enumerate(self.exam.students):
            level = self.overall.levels_arr[idx] if idx < len(self.overall.levels_arr) else grade_level(st.final_score, self.exam.cut_scores)
            students.append({
                "class_no": st.class_no,
                "sid": st.sid,
                "name": st.name,
                "level": level,
                "grade9": grade9_labels[idx] if idx < len(grade9_labels) else "",
                "grade5": grade5_labels[idx] if idx < len(grade5_labels) else "",
                "final_score": round(float(st.final_score), 2),
                "rounded_score": round(float(st.final_score)),
                "pencil_total": round(float(st.total), 2),
                "perform_score": round(float(st.perform_score), 2),
            })
        return {
            "version": 1,
            "saved_at": datetime.now().isoformat(timespec="seconds"),
            "subject": self.exam.subject or "(과목 미상)",
            "grade": self.exam.grade,
            "semester": self.exam.semester,
            "n_students": len(self.exam.students),
            "n_items": len(self.exam.items),
            "cuts": {lv: round(float(self.exam.cut_scores.get(lv, 0)), 2) for lv in ["A", "B", "C", "D", "E"]},
            "students": students,
        }

    def save_current_subject_snapshot(self):
        snapshot = self._current_subject_snapshot()
        if snapshot is None:
            QMessageBox.information(self, "포트폴리오 저장", "먼저 분석을 실행해 주세요.")
            return
        store = self._portfolio_store_dir()
        safe_subject = re.sub(r"[\\/:*?\"<>|]+", "_", snapshot.get("subject", "subject")).strip() or "subject"
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = store / f"{stamp}_{safe_subject}.json"
        try:
            path.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")
        except OSError as exc:
            QMessageBox.warning(self, "포트폴리오 저장", f"저장하지 못했습니다.\n{exc}")
            return
        self.refresh_portfolio_tab()
        self.statusBar().showMessage(f"포트폴리오 스냅샷 저장 완료 · {path}", 7000)

    def _load_portfolio_rows(self) -> list[dict]:
        store = self._portfolio_store_dir()
        rows = []
        for path in sorted(store.glob("*.json"), reverse=True):
            try:
                data = json.loads(path.read_text(encoding="utf-8"))
            except Exception:
                continue
            saved_at = str(data.get("saved_at", ""))
            subject = str(data.get("subject", ""))
            term = " ".join(part for part in [str(data.get("semester", "")), str(data.get("grade", ""))] if part)
            for student in data.get("students", []):
                rows.append({
                    "saved_at": saved_at,
                    "subject": subject,
                    "term": term,
                    "class_no": str(student.get("class_no", "")),
                    "sid": str(student.get("sid", "")),
                    "name": str(student.get("name", "")),
                    "level": str(student.get("level", "")),
                    "grade9": str(student.get("grade9", "")),
                    "grade5": str(student.get("grade5", "")),
                    "score": float(student.get("final_score", 0) or 0),
                    "snapshot_path": str(path),
                })
        return rows

    def refresh_portfolio_tab(self):
        if not hasattr(self, "table_portfolio"):
            return
        rows = self._load_portfolio_rows()
        self._portfolio_rows_cache = rows
        previous_subject = (
            self.combo_portfolio_subject.currentData()
            if hasattr(self, "combo_portfolio_subject") else ""
        )
        previous_student = (
            self.combo_portfolio_student.currentData()
            if hasattr(self, "combo_portfolio_student") else ""
        )
        if hasattr(self, "combo_portfolio_subject"):
            subjects = sorted({row["subject"] for row in rows if row.get("subject")})
            self.combo_portfolio_subject.blockSignals(True)
            self.combo_portfolio_subject.clear()
            self.combo_portfolio_subject.addItem("전체 과목", "")
            for subject in subjects:
                self.combo_portfolio_subject.addItem(subject, subject)
            idx = self.combo_portfolio_subject.findData(previous_subject)
            self.combo_portfolio_subject.setCurrentIndex(idx if idx >= 0 else 0)
            self.combo_portfolio_subject.blockSignals(False)
        if hasattr(self, "combo_portfolio_student"):
            grouped: dict[str, dict] = {}
            for row in rows:
                key = self._portfolio_student_key(row)
                data = grouped.setdefault(key, {"row": row, "subjects": set(), "records": 0})
                if row.get("subject"):
                    data["subjects"].add(row["subject"])
                data["records"] += 1
            ordered = sorted(grouped.items(), key=lambda item: (
                str(item[1]["row"].get("class_no", "")),
                str(item[1]["row"].get("name", "")),
            ))
            self.combo_portfolio_student.blockSignals(True)
            self.combo_portfolio_student.clear()
            self.combo_portfolio_student.addItem("학생 선택", "")
            for key, data in ordered:
                self.combo_portfolio_student.addItem(
                    self._portfolio_student_label(data["row"], len(data["subjects"]), data["records"]),
                    key,
                )
            idx = self.combo_portfolio_student.findData(previous_student)
            self.combo_portfolio_student.setCurrentIndex(idx if idx >= 0 else 0)
            self.combo_portfolio_student.blockSignals(False)
        self.table_portfolio.setSortingEnabled(False)
        self.table_portfolio.setRowCount(len(rows))
        for r, row in enumerate(rows):
            student_key = self._portfolio_student_key(row)
            values = [
                row["saved_at"],
                row["subject"],
                row["term"],
                row["class_no"],
                row["name"],
                row["level"],
                row["grade9"],
                row["grade5"],
                f"{row['score']:.2f}",
            ]
            for c, value in enumerate(values):
                if c in (6, 7):
                    item = NaturalItem(value, int(value) if str(value).isdigit() else 99)
                    item.setTextAlignment(Qt.AlignCenter)
                    self.table_portfolio.setItem(r, c, item)
                elif c == 8:
                    item = NaturalItem(value, row["score"])
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    self.table_portfolio.setItem(r, c, item)
                else:
                    item = _set_item(self.table_portfolio, r, c, value, align_left=c in (1, 2, 4))
                    if c in (3, 4):
                        item.setData(PRIVACY_REAL_TEXT_ROLE, value)
                self.table_portfolio.item(r, c).setData(PORTFOLIO_STUDENT_KEY_ROLE, student_key)
        self.table_portfolio.setSortingEnabled(True)
        self._filter_portfolio_table(self.le_portfolio_search.text() if hasattr(self, "le_portfolio_search") else "")
        if hasattr(self, "lbl_portfolio_note"):
            self.lbl_portfolio_note.setText(
                f"저장 위치: {self._portfolio_store_dir()} · 저장 과목 {len(set(row['subject'] for row in rows))}개 · 현재 저장 행 {len(rows)}개. "
                "Data 탭의 '포트폴리오 저장'을 과목마다 누르면 다과목 포트폴리오가 누적됩니다."
            )

    def _filter_portfolio_table(self, text: str = ""):
        if not hasattr(self, "table_portfolio"):
            return
        terms = self._search_terms(text)
        subject_filter = (
            self.combo_portfolio_subject.currentData()
            if hasattr(self, "combo_portfolio_subject") else ""
        )
        selected_student = (
            self.combo_portfolio_student.currentData()
            if hasattr(self, "combo_portfolio_student") else ""
        )
        privacy = self._counseling_mode_enabled()
        target_only = privacy and self._counseling_only_target_enabled() and bool(terms)
        visible_rows = []
        for r in range(self.table_portfolio.rowCount()):
            subject = self._real_item_text(self.table_portfolio.item(r, 1))
            row_key = self.table_portfolio.item(r, 0).data(PORTFOLIO_STUDENT_KEY_ROLE) if self.table_portfolio.item(r, 0) else ""
            subject_match = not subject_filter or subject == str(subject_filter)
            selected_student_match = not selected_student or row_key == selected_student
            if not terms:
                visible = subject_match and selected_student_match
                self.table_portfolio.setRowHidden(r, not visible)
                if visible:
                    visible_rows.append(r)
                continue
            row_text = " ".join(
                self._real_item_text(self.table_portfolio.item(r, c)).lower()
                for c in range(self.table_portfolio.columnCount())
                if self.table_portfolio.item(r, c)
            )
            student_text = " ".join([
                self._real_item_text(self.table_portfolio.item(r, 3)).lower(),
                self._real_item_text(self.table_portfolio.item(r, 4)).lower(),
            ])
            general_match = any(term in row_text for term in terms)
            target_match = not target_only or any(term in student_text for term in terms)
            visible = subject_match and selected_student_match and general_match and target_match
            self.table_portfolio.setRowHidden(r, not visible)
            if visible:
                visible_rows.append(r)
        self._apply_portfolio_privacy_masks()
        self._update_portfolio_summary(visible_rows, terms)

    def _update_portfolio_summary(self, visible_rows: list[int], terms: list[str]):
        if not hasattr(self, "lbl_portfolio_summary") or not hasattr(self, "table_portfolio"):
            return
        if not visible_rows:
            self.lbl_portfolio_summary.setText("표시할 포트폴리오 기록이 없습니다.")
            return
        subjects = []
        scores = []
        levels = {}
        student_names = []
        for r in visible_rows:
            subject = self._real_item_text(self.table_portfolio.item(r, 1))
            name = self._real_item_text(self.table_portfolio.item(r, 4))
            level = self._real_item_text(self.table_portfolio.item(r, 5))
            score_text = self._real_item_text(self.table_portfolio.item(r, 8))
            if subject and subject not in subjects:
                subjects.append(subject)
            if name and name not in student_names:
                student_names.append(name)
            if level:
                levels[level] = levels.get(level, 0) + 1
            try:
                scores.append(float(score_text))
            except ValueError:
                pass
        subject_part = ", ".join(subjects[:5]) + (f" 외 {len(subjects) - 5}과목" if len(subjects) > 5 else "")
        level_part = ", ".join(f"{lv} {count}" for lv, count in sorted(levels.items()))
        score_part = f"평균 {sum(scores) / len(scores):.1f}점" if scores else "점수 없음"
        if terms and len(student_names) == 1:
            headline = f"{student_names[0]} · {len(subjects)}과목 · {score_part}"
        elif terms:
            headline = f"검색 결과 {len(student_names)}명 · {len(subjects)}과목 · {len(visible_rows)}개 기록"
        else:
            headline = f"전체 포트폴리오 · {len(subjects)}과목 · {len(visible_rows)}개 기록"
        self.lbl_portfolio_summary.setText(
            f"{headline} · 성취도 분포 {level_part or '-'} · 과목 {subject_part or '-'}"
        )

    def _on_portfolio_student_selected(self):
        key = self.combo_portfolio_student.currentData() if hasattr(self, "combo_portfolio_student") else ""
        if key:
            rows = self._portfolio_rows_for_key(key)
            if rows:
                row = rows[0]
                self.le_portfolio_search.blockSignals(True)
                self.le_portfolio_search.setText(str(row.get("name") or row.get("class_no") or ""))
                self.le_portfolio_search.blockSignals(False)
        elif hasattr(self, "le_portfolio_search"):
            self.le_portfolio_search.blockSignals(True)
            self.le_portfolio_search.setText("")
            self.le_portfolio_search.blockSignals(False)
        self._filter_portfolio_table(self.le_portfolio_search.text() if hasattr(self, "le_portfolio_search") else "")

    def _portfolio_key_from_visible_rows(self) -> str:
        if hasattr(self, "combo_portfolio_student"):
            key = self.combo_portfolio_student.currentData()
            if key:
                return str(key)
        keys = set()
        for r in range(self.table_portfolio.rowCount()):
            if self.table_portfolio.isRowHidden(r):
                continue
            item = self.table_portfolio.item(r, 0)
            key = item.data(PORTFOLIO_STUDENT_KEY_ROLE) if item else ""
            if key:
                keys.add(str(key))
        return next(iter(keys)) if len(keys) == 1 else ""

    def _portfolio_report_html(self, rows: list[dict]) -> str:
        ordered = sorted(rows, key=lambda row: str(row.get("saved_at", "")))
        first = ordered[0]
        name = html.escape(str(first.get("name") or "이름 없음"))
        class_no = html.escape(str(first.get("class_no") or "반/번호 없음"))
        scores = [float(row.get("score", 0.0)) for row in ordered]
        avg = sum(scores) / len(scores) if scores else 0.0
        best = max(ordered, key=lambda row: float(row.get("score", 0.0)))
        low = min(ordered, key=lambda row: float(row.get("score", 0.0)))
        level_counts: dict[str, int] = {}
        for row in ordered:
            lv = str(row.get("level", "") or "-")
            level_counts[lv] = level_counts.get(lv, 0) + 1
        level_text = ", ".join(f"{html.escape(k)} {v}" for k, v in sorted(level_counts.items()))
        rows_html = []
        for row in ordered:
            rows_html.append(
                "<tr>"
                f"<td>{html.escape(str(row.get('saved_at', ''))[:10])}</td>"
                f"<td>{html.escape(str(row.get('subject', '')))}</td>"
                f"<td>{html.escape(str(row.get('term', '')))}</td>"
                f"<td>{html.escape(str(row.get('level', '')))}</td>"
                f"<td>{html.escape(str(row.get('grade9', '')))}</td>"
                f"<td>{html.escape(str(row.get('grade5', '')))}</td>"
                f"<td style='text-align:right'>{float(row.get('score', 0.0)):.1f}</td>"
                "</tr>"
            )
        questions = [
            f"{html.escape(str(best.get('subject', '')))}에서 가장 높게 나온 이유가 무엇인지 학생의 학습 습관과 연결해 확인합니다.",
            f"{html.escape(str(low.get('subject', '')))}에서 막힌 단원, 문항 유형, 수행 과정 중 어디가 어려웠는지 구체적으로 묻습니다.",
            "다음 평가 전까지 한 과목을 어떻게 유지하고, 한 과목을 어떻게 회복할지 작은 실천 계획을 정합니다.",
        ]
        question_html = "".join(f"<li>{q}</li>" for q in questions)
        return f"""
        <html><head><style>
        body {{ font-family: sans-serif; line-height: 1.55; }}
        h1 {{ font-size: 22px; margin-bottom: 4px; }}
        h2 {{ font-size: 16px; margin-top: 18px; }}
        .muted {{ color: #64748b; }}
        .cards {{ display: flex; gap: 10px; flex-wrap: wrap; }}
        .card {{ border: 1px solid #cbd5e1; border-radius: 8px; padding: 10px; min-width: 140px; }}
        .value {{ font-size: 20px; font-weight: 700; }}
        table {{ border-collapse: collapse; width: 100%; margin-top: 8px; }}
        th, td {{ border: 1px solid #cbd5e1; padding: 6px 8px; }}
        th {{ background: #e2e8f0; }}
        </style></head><body>
        <h1>{name} 학생 포트폴리오</h1>
        <p class="muted">{class_no} · 저장된 과목 기록 {len(ordered)}개</p>
        <div class="cards">
          <div class="card"><div>과목 수</div><div class="value">{len({row.get('subject') for row in ordered})}</div></div>
          <div class="card"><div>평균 환산점수</div><div class="value">{avg:.1f}</div></div>
          <div class="card"><div>가장 높은 과목</div><div class="value">{html.escape(str(best.get('subject', '-')))}</div><div>{float(best.get('score', 0.0)):.1f}점</div></div>
          <div class="card"><div>점검 과목</div><div class="value">{html.escape(str(low.get('subject', '-')))}</div><div>{float(low.get('score', 0.0)):.1f}점</div></div>
        </div>
        <h2>성취도 흐름</h2>
        <p>{level_text or '-'}</p>
        <h2>과목별 기록</h2>
        <table>
          <thead><tr><th>저장일</th><th>과목</th><th>학기/학년</th><th>성취도</th><th>9등급</th><th>5등급</th><th>환산점수</th></tr></thead>
          <tbody>{''.join(rows_html)}</tbody>
        </table>
        <h2>상담 질문</h2>
        <ul>{question_html}</ul>
        <p class="muted">이 리포트는 저장된 과목 스냅샷을 바탕으로 한 상담 보조 자료입니다. 최신 분석 결과를 반영하려면 각 과목 분석 후 Data 탭에서 포트폴리오 저장을 눌러 주세요.</p>
        </body></html>
        """

    def show_selected_student_portfolio(self):
        if not hasattr(self, "table_portfolio"):
            return
        key = self._portfolio_key_from_visible_rows()
        if not key:
            QMessageBox.information(
                self,
                "학생 포트폴리오",
                "학생 선택에서 한 학생을 고르거나, 검색 결과가 한 학생만 남도록 좁힌 뒤 다시 눌러 주세요.",
            )
            return
        rows = self._portfolio_rows_for_key(key)
        if not rows:
            QMessageBox.information(self, "학생 포트폴리오", "선택한 학생의 저장 기록이 없습니다.")
            return
        dialog = QDialog(self)
        dialog.setWindowTitle("학생 포트폴리오 상담 리포트")
        dialog.resize(self._px(900), self._px(680))
        layout = QVBoxLayout(dialog)
        browser = QTextBrowser()
        browser.setOpenExternalLinks(False)
        browser.setHtml(self._portfolio_report_html(rows))
        layout.addWidget(browser, 1)
        btn_row = QHBoxLayout()
        btn_row.addStretch(1)
        btn_close = QPushButton("닫기")
        btn_close.clicked.connect(dialog.accept)
        btn_row.addWidget(btn_close)
        layout.addLayout(btn_row)
        dialog.exec()

    def show_portfolio_store_location(self):
        store = self._portfolio_store_dir()
        opened = QDesktopServices.openUrl(QUrl.fromLocalFile(str(store)))
        if opened:
            self.statusBar().showMessage(f"학생 포트폴리오 저장 폴더를 열었습니다 · {store}", 7000)
            return
        QMessageBox.information(
            self,
            "학생 포트폴리오 저장 위치",
            f"Finder에서 폴더를 열지 못했습니다. 아래 경로를 확인해 주세요.\n\n{store}",
        )

    def _show_monitoring_dialog(self):
        if self.exam is None or self.overall is None:
            QMessageBox.information(self, "모니터링 지표", "먼저 분석을 실행해 주세요.")
            return
        observed_a = float(self.overall.level_dist_pct.get("A", 0.0))
        default_national = max(0.0, min(100.0, observed_a))
        national_a, ok = QInputDialog.getDouble(
            self,
            "전국 평균 A 비율",
            "비교할 전국 평균 A 비율(%)을 입력하세요.",
            default_national,
            0.0,
            100.0,
            1,
        )
        if not ok:
            return
        threshold, ok = QInputDialog.getDouble(
            self,
            "점검 기준",
            "전국 평균과 몇 %p 이상 차이 날 때 점검으로 볼까요?",
            10.0,
            0.0,
            100.0,
            1,
        )
        if not ok:
            return

        totals = [s.final_score for s in self.exam.students]
        mean = sum(totals) / len(totals) if totals else 0.0
        std = (sum((x - mean) ** 2 for x in totals) / max(1, len(totals) - 1)) ** 0.5 if len(totals) > 1 else 0.0
        expected_a = (1 - _normal_cdf(float(self.exam.cut_scores["A"]), mean, std)) * 100 if std > 0 else 0.0
        diff_national = observed_a - national_a
        diff_normal = observed_a - expected_a
        status = "점검 필요" if abs(diff_national) >= threshold else "참고 범위"
        QMessageBox.information(
            self,
            "모니터링 지표",
            f"<h3>성취평가 모니터링 지표</h3>"
            f"<p><b>판정:</b> {status}</p>"
            f"<table cellspacing='4'>"
            f"<tr><td>학생 수</td><td align='right'>{self.overall.n_students}명</td></tr>"
            f"<tr><td>평균 / 표준편차</td><td align='right'>{mean:.2f} / {std:.2f}</td></tr>"
            f"<tr><td>실제 A 비율</td><td align='right'>{observed_a:.1f}%</td></tr>"
            f"<tr><td>입력한 전국 평균 A 비율</td><td align='right'>{national_a:.1f}%</td></tr>"
            f"<tr><td>전국 평균 대비 차이</td><td align='right'>{diff_national:+.1f}%p</td></tr>"
            f"<tr><td>정규분포 기대 A 비율</td><td align='right'>{expected_a:.1f}%</td></tr>"
            f"<tr><td>정규분포 기대 대비 차이</td><td align='right'>{diff_normal:+.1f}%p</td></tr>"
            f"<tr><td>미도달 비율</td><td align='right'>{self.overall.level_dist_pct.get('미도달', 0):.1f}%</td></tr>"
            f"</table>"
            f"<p>기준: 전국 평균과의 차이가 ±{threshold:.1f}%p 이상이면 점검 필요로 표시했습니다.</p>"
        )

    def _open_monitoring_tab(self):
        if hasattr(self, "tab_monitor"):
            self.tabs.setCurrentWidget(self.tab_monitor)
            self._render_monitor_tab()

    def _pick_folder_batch(self):
        """폴더를 선택하면 안의 .xlsx 파일들을 이름 패턴으로 자동 분류해 채운다.

        매칭 점수 방식:
          - 카테고리별 키워드 목록(가중치 부여) 중 하나라도 파일명에 포함되면 점수 누적.
          - 같은 파일이 여러 카테고리에 걸리면 더 높은 점수 카테고리에 배정.
          - 카테고리별 후보 중 최고 점수가 같으면 가장 최근 수정된 파일 우선.
        """
        directory = QFileDialog.getExistingDirectory(self, "성취평가 자료가 들어있는 폴더 선택")
        if not directory:
            return
        from pathlib import Path
        folder = Path(directory)
        files = sorted([p for p in folder.glob("*.xlsx") if not p.name.startswith("~")])
        if not files:
            QMessageBox.information(self, "폴더 일괄 불러오기",
                                    "선택한 폴더에 .xlsx 파일이 없습니다.")
            return

        # 카테고리별 키워드 (점수). 더 구체적/특이적인 키워드일수록 점수 높음.
        rules = {
            "response": [           # 학생답 정오표
                ("학생답 정오표", 100), ("학생답 정오", 90), ("정오표", 70),
                ("정오 data", 60), ("학생답안", 50), ("응답", 30), ("answer", 30),
            ],
            "iteminfo": [           # 문항정보표
                ("문항정보표", 100), ("문항 정보표", 90), ("문항정보", 70),
                ("문항 정보", 70), ("item info", 60), ("문항 출제", 50),
            ],
            "cuts": [               # 예상추정분할점수
                ("예상추정분할점수", 100), ("추정분할점수", 95),
                ("추정 분할점수", 90), ("예상 분할점수", 80),
                ("분할점수 조회", 70), ("분할점수", 50), ("cut score", 50),
            ],
            "perform": [            # 수행평가
                ("수행평가 결과", 100), ("수행평가 조회", 95),
                ("수행평가", 80), ("수행 평가", 75),
                ("수행 결과", 60), ("performance", 50),
            ],
            "grade5_report": [       # 2022 개정 1~5등급 컷 계산용 교과목별 일람표
                ("지필평가 교과목별 일람표", 110), ("교과목별 일람표", 100),
                ("일람표", 70),
            ],
        }

        # macOS Finder는 한글 파일명을 NFD(분해 형태)로 저장하지만 우리 키워드는 NFC다.
        # 양쪽을 NFC로 정규화한 뒤 비교한다. (Windows는 보통 NFC라 무관)
        import unicodedata
        def _norm(s: str) -> str:
            return unicodedata.normalize("NFC", s).lower().replace(" ", "").replace("_", "").replace("-", "")
        def score(path: Path, kws):
            name = _norm(path.name)
            best_score = 0
            for kw, w in kws:
                k = _norm(kw)
                if k in name:
                    best_score = max(best_score, w)
            return best_score

        # 카테고리별 후보 + 점수
        cands = {cat: [] for cat in rules}
        for p in files:
            for cat, kws in rules.items():
                s = score(p, kws)
                if s > 0:
                    cands[cat].append((s, p.stat().st_mtime, p))
        # 한 파일이 여러 카테고리에 걸리지 않도록 — 더 높은 점수 카테고리에 배정
        chosen: dict[str, Path] = {}
        used: set[Path] = set()
        # 카테고리 → 점수 내림차순으로 한 번에 처리
        order = ["response", "iteminfo", "cuts", "perform", "grade5_report"]
        # 첫 패스: 카테고리별 최고 점수 1개씩 잠정 배정
        prelim = {}
        for cat in order:
            lst = sorted(cands[cat], reverse=True)
            for s, mt, p in lst:
                prelim[cat] = (s, mt, p)
                break
        # 충돌(같은 파일이 여러 카테고리)이면 더 큰 점수 쪽이 가져감
        # → 점수 내림차순으로 카테고리 순회하며 안 쓰인 파일을 차지
        cat_by_score = sorted(prelim.items(), key=lambda kv: -kv[1][0])
        for cat, (s, mt, p) in cat_by_score:
            if p in used:
                # 다음 후보 시도
                lst = sorted(cands[cat], reverse=True)
                for s2, mt2, p2 in lst:
                    if p2 not in used:
                        chosen[cat] = p2; used.add(p2); break
            else:
                chosen[cat] = p; used.add(p)

        applied = []
        if "response" in chosen:
            self.fs_response.path_edit.setText(str(chosen["response"]))
            applied.append(("학생답 정오표 data", chosen["response"].name))
        if "iteminfo" in chosen:
            self.fs_iteminfo.path_edit.setText(str(chosen["iteminfo"]))
            applied.append(("문항정보표", chosen["iteminfo"].name))
        if "cuts" in chosen:
            self.fs_cuts.path_edit.setText(str(chosen["cuts"]))
            applied.append(("예상추정분할점수", chosen["cuts"].name))
        if "perform" in chosen:
            self.chk_perform.setChecked(True)
            self.fs_perform.path_edit.setText(str(chosen["perform"]))
            applied.append(("수행평가", chosen["perform"].name))
        if "grade5_report" in chosen:
            self.fs_grade5_report.path_edit.setText(str(chosen["grade5_report"]))
            applied.append(("교과목별 일람표", chosen["grade5_report"].name))
        if not applied:
            QMessageBox.information(
                self, "폴더 일괄 불러오기",
                f"폴더에 .xlsx 파일은 {len(files)}개 있지만 이름이 인식 가능한 패턴이 아닙니다.\n"
                "파일명에 다음 키워드 중 하나가 들어 있는지 확인해 주세요:\n"
                "  · 정오표 / 학생답\n  · 문항정보표 / 문항정보\n"
                "  · 추정분할점수 / 분할점수\n  · 수행평가\n  · 교과목별 일람표"
            )
            return

        msg_lines = [f"• {k} ← {v}" for k, v in applied]
        skipped = [p.name for p in files if p not in used]
        msg = "다음 파일을 자동으로 채웠습니다:\n\n" + "\n".join(msg_lines)
        if skipped:
            msg += "\n\n인식 안 된 파일:\n" + "\n".join(f"  · {n}" for n in skipped[:6])
            if len(skipped) > 6:
                msg += f"\n  · … 외 {len(skipped) - 6}개"
        if not ("response" in chosen and "iteminfo" in chosen):
            msg += "\n\n⚠ 정오표·문항정보표 중 일부가 인식되지 않아, 수동으로 지정해 주세요."
        self.statusBar().showMessage(f"폴더 일괄 불러오기 · {len(applied)}개 채움", 5000)
        QMessageBox.information(self, "폴더 일괄 불러오기", msg)

    def calculate_grade5_cuts_from_report(self):
        path = self.fs_grade5_report.path() if hasattr(self, "fs_grade5_report") else ""
        if not path:
            path, _ = QFileDialog.getOpenFileName(
                self,
                "지필평가 교과목별 일람표 선택",
                str(Path.home() / "Downloads"),
                "Excel 통합문서 (*.xlsx)",
            )
            if path:
                self.fs_grade5_report.path_edit.setText(path)
        if not path:
            return

        compact_name = Path(path).name.replace(" ", "").replace("_", "").replace("-", "")
        if "등급컷" in compact_name or ("등급" in compact_name and "계산기" in compact_name):
            QMessageBox.warning(
                self,
                "1-5등급 컷 계산",
                "등급 컷 계산기 파일은 사용하지 않습니다.\n"
                "지필평가 교과목별 일람표 파일을 선택해 주세요.",
            )
            return

        try:
            reports = _load_grade5_cut_reports(path)
        except Exception as exc:
            QMessageBox.warning(self, "1-5등급 컷 계산", str(exc))
            return

        self._last_grade5_cut_reports = reports
        first = reports[0]
        first_summary = _grade5_cut_summary(first["scores"])
        first_cut = first_summary["cut_rows"][0]["score"] if first_summary["cut_rows"] else 0
        source_label = "일람표 실제 점수 기준"
        self.lbl_grade5_cut_info.setText(
            f"{first['subject']} · {first_summary['n']}명 · {source_label} · 1등급 컷 {_format_score(first_cut)}점"
        )
        self.statusBar().showMessage(
            f"1-5등급 컷 계산 완료 · {first_summary['n']}명 · {source_label} · 1등급 컷 {_format_score(first_cut)}점",
            7000,
        )
        self._show_grade5_cut_dialog(path, reports)

    def _grade5_cut_report_html(self, path: str, reports: list[dict]) -> str:
        colors = self.theme.colors if self.theme is not None else {}
        bg = colors.get("panel", "#ffffff")
        card = colors.get("card", "#f8fbfb")
        shade = colors.get("shade", "#eef7f4")
        text = colors.get("text", "#202326")
        muted = colors.get("muted", "#6e7781")
        border = colors.get("border", "#dce3e6")
        accent = colors.get("accent", "#2a7770")
        sections = []
        for report in reports:
            summary = _grade5_cut_summary(report["scores"])
            notes_by_grade = {
                str(note["grade"]): " ".join(html.escape(str(message)) for message in note["messages"])
                for note in summary.get("boundary_notes", [])
            }
            cut_rows = "".join(
                "<tr>"
                f"<td>{row['grade']}등급</td>"
                f"<td><b>{_format_score(row['score'])}점</b></td>"
                f"<td>{row['boundary'] / 100:g}</td>"
                f"<td>{row['rank']}명</td>"
                f"<td class=\"note\">{notes_by_grade.get(str(row['grade']), '')}</td>"
                "</tr>"
                for row in summary["cut_rows"]
            )
            official_count_rows = "".join(
                f"<span>{grade}등급 {summary.get('official_counts', {}).get(grade, 0)}명</span>"
                for grade, _ in GRADE5_CUMULATIVE
            )
            source_label = "교과목별 일람표 실제 점수 기준"
            excluded_entries = report.get("excluded_entries") or []
            excluded_html = ""
            if excluded_entries:
                status_counts = {}
                for entry in excluded_entries:
                    status = str(entry.get("status") or "비점수").strip()
                    status_counts[status] = status_counts.get(status, 0) + 1
                status_text = ", ".join(f"{html.escape(k)} {v}건" for k, v in status_counts.items())
                examples = ", ".join(
                    f"{html.escape(str(entry.get('class') or '?'))}반 {html.escape(str(entry.get('number') or '?'))}번 {html.escape(str(entry.get('status') or ''))}"
                    for entry in excluded_entries[:8]
                )
                excluded_html = (
                    f"<p class=\"excluded\"><b>비점수 제외:</b> {len(excluded_entries)}건"
                    f" ({status_text})<br><span>{examples}</span></p>"
                )
            sections.append(f"""
            <section>
              <h2>{html.escape(str(report['subject']))}</h2>
              <p class="muted">시트: {html.escape(str(report['sheet']))} · {html.escape(str(report['source_note']))}</p>
              <p class="source-pill">현재 계산 기준: {html.escape(source_label)}</p>
              {excluded_html}
              <div class="summary">
                <b>응시 점수 {summary['n']}개</b>
                <span>최고 {_format_score(summary['max'])}점</span>
                <span>평균 {_format_score(summary['mean'])}점</span>
                <span>최저 {_format_score(summary['min'])}점</span>
              </div>
              <h3>컷 점수와 생활기록부 기준 점검</h3>
              <table>
                <thead><tr><th>구분</th><th>컷 점수</th><th>비율</th><th>조견표 인원</th><th>동점/근거 점검</th></tr></thead>
                <tbody>{cut_rows}</tbody>
              </table>
              <p class="muted">생활기록부 중간석차 기준 예상 인원: {official_count_rows}</p>
            </section>
            """)
        return f"""
        <html><head><style>
        html, body {{
            background: {bg};
            color: {text};
            font-family: -apple-system, BlinkMacSystemFont, 'Apple SD Gothic Neo', 'Malgun Gothic', sans-serif;
        }}
        body {{ margin: 0; }}
        h1 {{ margin: 0 0 8px 0; font-size: 22px; }}
        h2 {{ margin: 18px 0 4px 0; font-size: 18px; }}
        h3 {{ margin: 14px 0 6px 0; font-size: 15px; }}
        section {{ background: {bg}; color: {text}; }}
        .muted {{ color: {muted}; }}
        .muted span {{ margin-right: 8px; }}
        .summary {{ display: flex; gap: 12px; flex-wrap: wrap; margin: 8px 0 10px 0; }}
        .source-pill {{
            display: inline-block;
            margin: 4px 0 8px 0;
            padding: 5px 9px;
            border: 1px solid {border};
            border-radius: 6px;
            background: {shade};
            color: {text};
            font-weight: 700;
        }}
        .excluded {{
            margin: 4px 0 10px 0;
            padding: 8px 10px;
            border: 1px solid {border};
            border-radius: 6px;
            background: {card};
            color: {text};
            line-height: 1.45;
        }}
        .excluded span {{ color: {muted}; }}
        table {{ border-collapse: collapse; width: 100%; margin: 4px 0 10px 0; }}
        th, td {{ border: 1px solid {border}; padding: 7px 8px; text-align: center; color: {text}; }}
        th {{ background: {card}; color: {text}; font-weight: 700; }}
        td {{ background: {bg}; }}
        tr:nth-child(even) td {{ background: {shade}; }}
        td:first-child, th:first-child {{ text-align: left; }}
        .note {{ text-align: left; line-height: 1.45; }}
        b {{ color: {text}; }}
        a {{ color: {accent}; }}
        .small {{ width: 360px; max-width: 100%; }}
        </style></head><body>
        <h1>1-5등급 컷 계산 결과</h1>
        <p class="muted">파일: {html.escape(Path(path).name)}</p>
        {''.join(sections)}
        </body></html>
        """

    def _grade5_cut_example_text(self, reports: list[dict] | None = None) -> str:
        if reports:
            report = reports[0]
            summary = _grade5_cut_summary(report["scores"])
            if summary.get("cut_rows"):
                row = summary["cut_rows"][0]
                return (
                    f"현재 결과 예시\n"
                    f"- 기준 자료: {report.get('subject', '선택 자료')} / 교과목별 일람표 실제 점수\n"
                    f"- {summary['n']}명의 10%는 {row['raw_rank']:.1f}명이므로 "
                    f"INT({row['raw_rank']:.1f})={row['rank']}입니다.\n"
                    f"- 따라서 이 자료에서 1등급 컷 조견표 점수는 "
                    f"{row['rank']}번째로 큰 점수인 {_format_score(row['score'])}점입니다."
                )
        return (
            "예시\n"
            "- 총원이 382명이면 10%는 38.2명이므로 INT(38.2)=38입니다.\n"
            "- 컷 점수 조견표는 선택한 점수 범위에서 38번째로 큰 점수를 표시합니다.\n"
            "- 경계 점수에 동점자가 있으면 중간석차 기준으로 실제 등급을 함께 확인합니다."
        )

    def _grade5_cut_basis_text(self, reports: list[dict] | None = None) -> str:
        return (
            "근거\n"
            "- 학교생활기록부 종합지원포털 > 법령·규정 > 학교생활기록 작성 및 관리지침\n"
            "  [시행 2026.3.1.] [교육부훈령 제555호]\n"
            "  https://star.moe.go.kr/web/contents/m20103.do\n"
            "- 학교생활기록부 종합지원포털 > 자료실 > 학교생활기록부 기재요령\n"
            "  2026학년도 학교생활기록부 기재요령(고등학교)\n\n"
            "  https://star.moe.go.kr/web/contents/m21100.do\n\n"
            "5등급 누적 기준\n"
            "- 1등급: 10% 이하\n"
            "- 2등급: 10% 초과 ~ 34% 이하\n"
            "- 3등급: 34% 초과 ~ 66% 이하\n"
            "- 4등급: 66% 초과 ~ 90% 이하\n"
            "- 5등급: 90% 초과 ~ 100% 이하\n\n"
            "컷 점수 조견표 공식\n"
            "- 인원수 = INT(비율 × 총원)\n"
            "- 컷 점수 = LARGE(전체 점수 범위, 인원수)\n"
            "- 여기서 전체 점수 범위는 지필평가 교과목별 일람표의 실제 숫자 점수입니다.\n"
            "- 인정결·질병결·자퇴·전출처럼 점수가 아닌 값은 계산에서 제외합니다.\n\n"
            "생활기록부 등급 판정 점검\n"
            "- 실제 등급 판정은 누적인원 경계와 동점자를 함께 봅니다.\n"
            "- 경계 점수에 동점자가 걸리면 동석차의 중간석차 백분율로 등급을 확인합니다.\n"
            "- 그래서 결과표에는 컷 점수와 함께 동점/중간석차 점검을 표시합니다.\n\n"
            f"{self._grade5_cut_example_text(reports)}"
        )

    def _grade5_cut_report_text(self, reports: list[dict]) -> str:
        lines = [
            "1-5등급 컷 계산 결과",
            self._grade5_cut_basis_text(reports),
            "공식: 인원수 = INT(비율 × 총원), 컷 점수 = LARGE(전체 점수 범위, 인원수)",
        ]
        for report in reports:
            summary = _grade5_cut_summary(report["scores"])
            lines.append("")
            lines.append(f"[{report['subject']}] {summary['n']}명 · {report['source_note']}")
            if report.get("excluded_entries"):
                excluded = report["excluded_entries"]
                lines.append(
                    "비점수 제외: "
                    + ", ".join(
                        f"{entry.get('class', '?')}반 {entry.get('number', '?')}번 {entry.get('status', '')}"
                        for entry in excluded
                    )
                )
            for row in summary["cut_rows"]:
                note = next(
                    (n for n in summary.get("boundary_notes", []) if str(n.get("grade")) == str(row["grade"])),
                    None,
                )
                note_text = " ".join(str(message) for message in note["messages"]) if note else ""
                lines.append(
                    f"- {row['grade']}등급 컷: {_format_score(row['score'])}점 "
                    f"(비율 {row['boundary'] / 100:g}, 조견표 인원 {row['rank']}명)"
                    + (f" / {note_text}" if note_text else "")
                )
            official_counts = summary.get("official_counts", {})
            if official_counts:
                lines.append(
                    "생활기록부 중간석차 기준 예상 인원: "
                    + ", ".join(f"{grade}등급 {official_counts.get(grade, 0)}명" for grade, _ in GRADE5_CUMULATIVE)
                )
        return "\n".join(lines)

    def _show_grade5_basis_dialog(self, parent: QWidget, reports: list[dict]):
        dialog = QDialog(parent)
        dialog.setWindowTitle("1-5등급 컷 공식/근거")
        dialog.resize(self._px(720), self._px(560))
        layout = QVBoxLayout(dialog)
        browser = QTextBrowser()
        browser.setPlainText(self._grade5_cut_basis_text(reports))
        browser.setLineWrapMode(QTextBrowser.WidgetWidth)
        layout.addWidget(browser, 1)
        row = QHBoxLayout()
        row.addStretch(1)
        close_btn = QPushButton("확인")
        close_btn.clicked.connect(dialog.accept)
        row.addWidget(close_btn)
        layout.addLayout(row)
        dialog.exec()

    def _show_grade5_cut_dialog(self, path: str, reports: list[dict]):
        dialog = QDialog(self)
        dialog.setWindowTitle("1-5등급 컷 계산 결과")
        dialog.resize(self._px(820), self._px(620))
        layout = QVBoxLayout(dialog)
        browser = QTextBrowser()
        browser.setHtml(self._grade5_cut_report_html(path, reports))
        layout.addWidget(browser, 1)
        row = QHBoxLayout()
        row.addStretch(1)
        basis_btn = QPushButton("공식/근거 보기")
        basis_btn.clicked.connect(lambda: self._show_grade5_basis_dialog(dialog, reports))
        row.addWidget(basis_btn)
        copy_btn = QPushButton("결과 복사")
        copy_btn.clicked.connect(lambda: QApplication.clipboard().setText(self._grade5_cut_report_text(reports)))
        row.addWidget(copy_btn)
        close_btn = QPushButton("닫기")
        close_btn.clicked.connect(dialog.accept)
        row.addWidget(close_btn)
        layout.addLayout(row)
        dialog.exec()

    def _toggle_perform(self, on: bool):
        self.fs_perform.setEnabled(on)
        self.spin_perform_ratio.setEnabled(on)
        if not on:
            self.spin_perform_ratio.spin.blockSignals(True)
            self.spin_perform_ratio.setValue(0)
            self.spin_perform_ratio.spin.blockSignals(False)
            self.spin_pencil_ratio.spin.blockSignals(True)
            self.spin_pencil_ratio.setValue(100)
            self.spin_pencil_ratio.spin.blockSignals(False)

    def _sync_pencil_to_perform(self, v):
        if not self.chk_perform.isChecked():
            return
        self.spin_perform_ratio.spin.blockSignals(True)
        self.spin_perform_ratio.setValue(max(0, 100 - v))
        self.spin_perform_ratio.spin.blockSignals(False)

    def _sync_perform_to_pencil(self, v):
        self.spin_pencil_ratio.spin.blockSignals(True)
        self.spin_pencil_ratio.setValue(max(0, 100 - v))
        self.spin_pencil_ratio.spin.blockSignals(False)

    def _on_perform_path_changed(self, path: str):
        path = path.strip()
        if not path:
            self.lbl_perform_info.setText("(미설정)")
            return
        try:
            pd_obj = load_perform(path)
        except Exception as e:
            QMessageBox.warning(self, "수행평가 파일 오류", f"파일을 인식하지 못했습니다.\n{e}")
            self.lbl_perform_info.setText("(파일 인식 실패)")
            return
        areas_text = ", ".join(f"{a.name}({a.ratio_pct:.0f}%)" for a in pd_obj.areas)
        self.lbl_perform_info.setText(
            f"<b>{pd_obj.subject}</b><br>학생 {len(pd_obj.records)}명 · 영역 {len(pd_obj.areas)}개<br>{areas_text}"
        )
        # 영역 반영비율 합을 수행평가 반영비율 기본값으로 제안
        if pd_obj.ratio_total > 0:
            self.spin_perform_ratio.spin.blockSignals(True)
            self.spin_perform_ratio.setValue(round(pd_obj.ratio_total))
            self.spin_perform_ratio.spin.blockSignals(False)
            self.spin_pencil_ratio.spin.blockSignals(True)
            self.spin_pencil_ratio.setValue(max(0, 100 - round(pd_obj.ratio_total)))
            self.spin_pencil_ratio.spin.blockSignals(False)

    def _on_cuts_path_changed(self, path: str):
        path = path.strip()
        if not path:
            return
        try:
            cuts, extra = load_cut_scores(path)
        except Exception as e:
            QMessageBox.warning(self, "추정분할점수 파일 오류",
                                f"파일을 인식하지 못했습니다.\n{e}")
            return
        for lv, val in cuts.items():
            self.spin_cuts[lv].setValue(round(float(val), 2))
        self._update_cut_box_title()
        self.statusBar().showMessage(
            f"추정분할점수 자동 적용 · A {cuts['A']:.2f} / B {cuts['B']:.2f} / "
            f"C {cuts['C']:.2f} / D {cuts['D']:.2f} / E {cuts['E']:.2f}",
            8000
        )

    def _build_tabs(self) -> QWidget:
        self.tabs = QTabWidget()
        self.tabs.setMovable(True)
        self.tabs.setUsesScrollButtons(True)
        self.tabs.setElideMode(Qt.ElideRight)
        self.tabs.tabBar().setExpanding(False)
        self.tab_data = QWidget(); self._init_tab_data()
        self.tab_portfolio = QWidget(); self._init_tab_portfolio()
        self.tab_overview = QWidget(); self._init_tab_overview()
        self.tab_perform = QWidget(); self._init_tab_perform()
        self.tab_items = QWidget(); self._init_tab_items()
        self.tab_choice = QWidget(); self._init_tab_choice()
        self.tab_standard = QWidget(); self._init_tab_standard()
        self.tab_ai_review = QWidget(); self._init_tab_ai_review()
        self.tab_spliter = QWidget(); self._init_tab_spliter()
        self.tab_monitor = QWidget(); self._init_tab_monitor()
        self.tab_help = QWidget(); self._init_tab_help()
        self._tab_label_sets = [
            (self.tab_data, "Data", "Data", "Data"),
            (self.tab_portfolio, "학생 포트폴리오", "포트폴리오", "학생"),
            (self.tab_overview, "전체 성취도 분석", "성취도", "성취"),
            (self.tab_perform, "수행평가 분석", "수행평가", "수행"),
            (self.tab_items, "문항 분석", "문항", "문항"),
            (self.tab_choice, "성취수준별 답지반응 분포", "답지반응", "답지"),
            (self.tab_standard, "성취기준 분석 결과", "성취기준", "기준"),
            (self.tab_ai_review, "AI 문항 검토", "AI 검토", "AI"),
            (self.tab_spliter, "예상정답률 입력", "예상정답률", "정답률"),
            (self.tab_monitor, "모니터링", "모니터링", "모니터"),
            (self.tab_help, "도움말", "도움말", "도움"),
        ]
        for widget, full, _compact, _tiny in self._tab_label_sets:
            self.tabs.addTab(widget, full)
            self.tabs.setTabToolTip(self.tabs.indexOf(widget), full)
        self.tabs.currentChanged.connect(self._on_main_tab_changed)
        self._apply_responsive_tab_labels()
        return self.tabs

    def _apply_responsive_tab_labels(self):
        if not hasattr(self, "tabs") or not hasattr(self, "_tab_label_sets"):
            return
        width = self.width() or 1600
        label_index = 3 if width < 1120 else (2 if width < 1480 else 1)
        for labels in self._tab_label_sets:
            widget = labels[0]
            idx = self.tabs.indexOf(widget)
            if idx >= 0:
                self.tabs.setTabText(idx, labels[label_index])
                self.tabs.setTabToolTip(idx, labels[1])

    # ---- 예상정답률 입력 ----------------------------------------------
    def _init_tab_spliter(self):
        layout = QVBoxLayout(self.tab_spliter)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(8)

        self.spliter_toolbar = QFrame()
        self.spliter_toolbar.setProperty("role", "card")
        toolbar_layout = QHBoxLayout(self.spliter_toolbar)
        toolbar_layout.setContentsMargins(12, 10, 12, 10)
        toolbar_layout.setSpacing(8)

        self.lbl_spliter_status = QLabel("분석 전 · 계산기는 아래에서 바로 사용할 수 있습니다.")
        self.lbl_spliter_status.setProperty("role", "muted")
        self.lbl_spliter_status.setWordWrap(True)
        toolbar_layout.addWidget(self.lbl_spliter_status, 1)

        btn_send = QPushButton("분석자료 보내기")
        btn_send.setProperty("role", "primary")
        btn_send.clicked.connect(self.send_spliter_evidence_to_web)
        toolbar_layout.addWidget(btn_send)
        btn_neis = QPushButton("NEIS 입력표")
        btn_neis.setToolTip(
            "문항별 목표 성취수준을 교사가 수정한 뒤 5% 단위 NEIS 입력표를 만들거나 "
            "수정 목표를 예상정답률 계산기에 보냅니다."
        )
        btn_neis.clicked.connect(self.open_neis_expected_rate_dialog)
        toolbar_layout.addWidget(btn_neis)
        btn_import_paper = QPushButton("시험지 반영")
        btn_import_paper.setToolTip(
            "시험지 HWPX/PDF/자료를 읽어 문항 초안을 만들고 예상정답률 계산기에 바로 반영합니다. "
            "AI 없이 로컬 규칙으로 먼저 처리하며, AI 문항 검토 탭에서 보강할 수 있습니다."
        )
        btn_import_paper.clicked.connect(self.import_exam_paper_to_spliter)
        toolbar_layout.addWidget(btn_import_paper)
        btn_blueprint = QPushButton("문항 구성안")
        btn_blueprint.setToolTip(
            "문항 수를 입력하면 분석자료의 난이도·정답률·배점 분포를 참고해 "
            "성취수준 목표와 배점 초안을 예상정답률 계산기에 제안합니다."
        )
        btn_blueprint.clicked.connect(self.suggest_expected_rate_blueprint)
        toolbar_layout.addWidget(btn_blueprint)
        btn_export = QPushButton("근거 엑셀")
        btn_export.clicked.connect(self.export_spliter_evidence)
        toolbar_layout.addWidget(btn_export)
        btn_reload = QPushButton("새로고침")
        btn_reload.clicked.connect(lambda: self._load_spliter_web(force_recreate=True))
        toolbar_layout.addWidget(btn_reload)
        layout.addWidget(self.spliter_toolbar, 0)

        if QWebEngineView is None:
            self.lbl_spliter_web = QLabel(
                "이 Python 환경에는 Qt WebEngine이 없어 내장 계산기를 표시할 수 없습니다. "
                "근거 엑셀을 내보내 자료를 확인해 주세요."
            )
            self.lbl_spliter_web.setProperty("role", "muted")
            self.lbl_spliter_web.setWordWrap(True)
            layout.addWidget(self.lbl_spliter_web, 1)
        else:
            self.spliter_view = QWebEngineView()
            self._reset_spliter_web_page()
            layout.addWidget(self.spliter_view, 1)

    def _reset_spliter_web_page(self):
        if self.spliter_view is None:
            return
        if self._spliter_load_signal_connected:
            try:
                self.spliter_view.loadFinished.disconnect(self._on_spliter_loaded)
            except Exception:
                pass
            self._spliter_load_signal_connected = False
        if QWebEnginePage is not None:
            self.spliter_view.setPage(QWebEnginePage(self.spliter_view))
        self._install_spliter_web_bridge()
        self.spliter_view.loadFinished.connect(self._on_spliter_loaded)
        self._spliter_load_signal_connected = True

    def _on_main_tab_changed(self, _index: int):
        if not hasattr(self, "tabs") or not hasattr(self, "tab_spliter"):
            return
        if hasattr(self, "tab_ai_review") and self.tabs.currentWidget() is self.tab_ai_review:
            self._schedule_ai_ollama_mlx_help_popup()
        if self.tabs.currentWidget() is self.tab_spliter:
            QTimer.singleShot(0, self._ensure_spliter_tab_loaded)

    def _ensure_spliter_tab_loaded(self):
        if self.spliter_view is None:
            return
        if not self._spliter_load_requested:
            self._load_spliter_web(force_recreate=True)
            return
        if self._spliter_loaded:
            QTimer.singleShot(80, self._nudge_spliter_view)

    def _activate_spliter_tab_for_pending_payloads(self):
        if not hasattr(self, "tabs") or not hasattr(self, "tab_spliter"):
            return
        self.tabs.setCurrentWidget(self.tab_spliter)

        def send_or_load():
            if self.spliter_view is None:
                return
            if self._spliter_loaded:
                self._flush_spliter_project_payload()
                self._flush_spliter_payload()
            else:
                self._ensure_spliter_tab_loaded()

        QTimer.singleShot(0, send_or_load)

    def _nudge_spliter_view(self):
        if self.spliter_view is None:
            return
        self.spliter_view.update()
        viewport = getattr(self.spliter_view, "viewport", None)
        if callable(viewport):
            viewport().update()

    def _install_spliter_web_bridge(self):
        if self.spliter_view is None or QWebChannel is None:
            return
        self._spliter_bridge = SpliterWebBridge(self)
        self._spliter_channel = QWebChannel(self.spliter_view.page())
        self._spliter_channel.registerObject("goeduBridge", self._spliter_bridge)
        self.spliter_view.page().setWebChannel(self._spliter_channel)

    def _send_spliter_bridge(self):
        if self.spliter_view is None or not self._spliter_loaded:
            return
        script = r"""
(function(){
  if (window.__GOEDUSPLIT_QT_BRIDGE_INIT__) return;
  window.__GOEDUSPLIT_QT_BRIDGE_INIT__ = true;
  function attachBridge(){
    if (!window.qt || !window.qt.webChannelTransport || !window.QWebChannel) return false;
    new QWebChannel(window.qt.webChannelTransport, function(channel){
      window.goeduBridge = channel.objects.goeduBridge;
      window.goeduSplitBridge = window.goeduBridge;
      window.dispatchEvent(new Event('goedu-bridge-ready'));
    });
    return true;
  }
  if (attachBridge()) return;
  var script = document.createElement('script');
  script.src = 'qrc:///qtwebchannel/qwebchannel.js';
  script.onload = attachBridge;
  script.onerror = function(){ window.__GOEDUSPLIT_QT_BRIDGE_INIT__ = false; };
  document.head.appendChild(script);
})();
"""
        self.spliter_view.page().runJavaScript(script)

    def _send_spliter_teacher_presets(self, apply_current: bool = False):
        if self.spliter_view is None or not self._spliter_loaded:
            return
        presets_json = json.dumps(self._load_target_rate_presets(), ensure_ascii=False)
        fallback_cuts_json = json.dumps(self._current_cut_scores_from_inputs(), ensure_ascii=False)
        fallback_cuts_label_json = json.dumps(self._current_cut_scores_label(), ensure_ascii=False)
        apply_current_js = "true" if apply_current else "false"
        script = f"""
(function(){{
  const presets = {presets_json};
  const fallbackCuts = {fallback_cuts_json};
  const fallbackCutsLabel = {fallback_cuts_label_json};
  const applyCurrent = {apply_current_js};
  window.__GOEDUSPLIT_TARGET_RATE_PRESETS__ = presets;
  window.__GOEDUSPLIT_MANUAL_CUTS__ = fallbackCuts;
  window.__GOEDUSPLIT_MANUAL_CUTS_LABEL__ = fallbackCutsLabel;
  try {{ localStorage.setItem('goedu-target-rate-presets', JSON.stringify(presets)); }} catch (error) {{}}

  function callBridge(method, args) {{
    const bridge = window.goeduBridge || window.goeduSplitBridge;
    if (!bridge || typeof bridge[method] !== 'function') return;
    try {{ bridge[method](...(args || []), function(){{}}); }} catch (error) {{}}
  }}

  function installCompactStyle() {{
    document.getElementById('goedu-priority-ui-style')?.remove();
    const style = document.createElement('style');
    style.id = 'goedu-priority-ui-style';
    style.textContent = `
      html, body, #root {{ height:100% !important; overflow:hidden !important; }}
      .app-shell {{ padding:10px !important; }}
      .app-shell {{ height:100vh !important; min-height:0 !important; overflow:hidden !important; display:flex !important; flex-direction:column !important; }}
      .sticky-head {{ margin:-10px -10px 6px !important; padding:7px 10px !important; }}
      .sticky-head {{ flex:0 0 auto !important; position:relative !important; top:auto !important; z-index:30 !important; }}
      .topbar {{ align-items:center !important; gap:6px !important; flex-wrap:wrap !important; margin:0 auto 6px !important; }}
      .top-title {{ flex:1 1 240px !important; min-width:200px !important; }}
      .topbar h1 {{ font-size:19px !important; line-height:1.1 !important; margin:0 !important; letter-spacing:0 !important; }}
      .topbar .eyebrow {{ margin:0 0 2px !important; font-size:11px !important; }}
      .topbar .muted, .top-title span {{ font-size:11px !important; }}
      .top-actions {{ flex:1 1 420px !important; justify-content:flex-end !important; gap:6px !important; }}
      .top-actions button, .icon-button, .zoom-chip {{ flex:0 0 auto !important; min-width:34px !important; min-height:32px !important; border-radius:7px !important; white-space:nowrap !important; overflow:hidden !important; text-overflow:clip !important; }}
      .summary-grid {{ display:none !important; }}
      .workspace {{ display:flex !important; flex:1 1 auto !important; flex-direction:column !important; gap:8px !important; min-height:0 !important; overflow:hidden !important; }}
      .splitter {{ display:none !important; }}
      .table-panel {{ order:1 !important; display:flex !important; flex:1 1 auto !important; flex-direction:column !important; height:100% !important; min-height:0 !important; overflow:hidden !important; border-color:var(--accent) !important; }}
      .table-toolbar {{ position:relative !important; top:auto !important; flex:0 0 auto !important; z-index:12 !important; background:var(--panel) !important; align-items:center !important; gap:6px 10px !important; flex-wrap:wrap !important; padding:7px 9px !important; border-bottom:1px solid var(--line) !important; }}
      .table-toolbar > div:first-child {{ flex:1 1 280px !important; min-width:220px !important; }}
      .table-toolbar h2 {{ font-size:16px !important; line-height:1.15 !important; margin:0 0 2px !important; letter-spacing:0 !important; }}
      .table-toolbar p {{ max-width:680px !important; margin:0 !important; font-size:11px !important; line-height:1.25 !important; }}
      .table-controls {{ flex:2 1 560px !important; justify-content:flex-end !important; align-items:center !important; flex-wrap:wrap !important; gap:5px !important; min-width:360px !important; }}
      .table-controls button, .table-controls select, .table-controls label {{ flex:0 0 auto !important; font-size:12px !important; min-height:30px !important; white-space:nowrap !important; overflow:hidden !important; text-overflow:clip !important; }}
      .goedu-main-actions {{ display:inline-flex !important; flex-wrap:wrap !important; gap:6px !important; align-items:center !important; }}
      .goedu-main-action {{ flex:0 0 auto !important; min-height:30px !important; padding:0 9px !important; border-radius:7px !important; border:1px solid var(--line) !important; background:var(--panel) !important; color:var(--ink) !important; font-weight:800 !important; cursor:pointer !important; white-space:nowrap !important; overflow:hidden !important; text-overflow:clip !important; }}
      .goedu-main-action.primary {{ background:var(--accent) !important; border-color:var(--accent) !important; color:#fff !important; }}
      .goedu-main-action.active {{ outline:2px solid color-mix(in srgb, var(--accent), transparent 45%) !important; }}
      body.goedu-hide-web-head .sticky-head {{ display:none !important; }}
      body.goedu-hide-summary .table-toolbar > div:first-child {{ display:none !important; }}
      body.goedu-hide-summary .table-toolbar {{ padding:5px 8px !important; }}
      .goedu-score-strip {{ display:flex !important; flex:0 0 auto !important; flex-wrap:wrap !important; align-items:center !important; gap:8px !important; margin:0 0 5px !important; padding:9px 11px !important; border:2px solid color-mix(in srgb, var(--accent), transparent 38%) !important; border-radius:8px !important; background:linear-gradient(180deg, color-mix(in srgb, var(--accent), transparent 86%), color-mix(in srgb, var(--accent), transparent 94%)) !important; box-shadow:inset 4px 0 0 var(--accent) !important; font-size:14px !important; }}
      .goedu-score-title {{ color:var(--ink) !important; margin-right:4px !important; font-size:15px !important; font-weight:950 !important; letter-spacing:0 !important; }}
      .goedu-score-source {{ display:inline-flex !important; align-items:center !important; min-height:24px !important; padding:2px 7px !important; border-radius:999px !important; background:color-mix(in srgb, var(--accent), transparent 82%) !important; color:var(--accent) !important; font-size:11px !important; font-weight:950 !important; white-space:nowrap !important; }}
      .goedu-score-pill {{ display:inline-flex !important; align-items:baseline !important; gap:5px !important; border:1px solid color-mix(in srgb, var(--accent), transparent 50%) !important; border-radius:999px !important; padding:5px 10px !important; background:var(--panel) !important; color:var(--ink) !important; font-size:13px !important; font-weight:850 !important; white-space:nowrap !important; }}
      .goedu-score-pill span {{ color:var(--muted) !important; font-size:11px !important; font-weight:800 !important; }}
      .goedu-score-pill strong {{ color:var(--ink) !important; font-size:17px !important; line-height:1 !important; font-weight:950 !important; }}
      .goedu-score-pill em {{ color:var(--muted) !important; font-style:normal !important; font-size:11px !important; font-weight:800 !important; }}
      .goedu-score-note {{ flex:1 1 260px !important; min-width:220px !important; color:var(--muted) !important; font-size:11px !important; line-height:1.25 !important; font-weight:750 !important; }}
      .goedu-preset-strip {{ display:flex !important; flex:0 0 auto !important; flex-wrap:wrap !important; align-items:center !important; gap:4px !important; margin:0 0 5px !important; padding:5px 7px !important; border:1px solid var(--line) !important; border-radius:7px !important; background:var(--row) !important; font-size:11px !important; }}
      .goedu-preset-strip b {{ color:var(--ink) !important; margin-right:3px !important; }}
      .goedu-preset-strip button {{ flex:0 0 auto !important; min-height:26px !important; padding:0 8px !important; white-space:nowrap !important; overflow:hidden !important; text-overflow:clip !important; }}
      .goedu-preset-pill {{ color:var(--muted) !important; border:1px solid var(--line) !important; border-radius:999px !important; padding:2px 6px !important; background:var(--panel) !important; font-size:10px !important; white-space:nowrap !important; }}
      .item-table-wrap {{ flex:1 1 auto !important; height:auto !important; min-height:0 !important; overflow:auto !important; overscroll-behavior:contain !important; }}
      .item-table {{ width:100% !important; min-width:980px !important; table-layout:fixed !important; font-size:12px !important; }}
      .item-table th, .item-table td {{ white-space:nowrap !important; padding:5px 6px !important; }}
      .item-table thead th {{ position:sticky !important; top:0 !important; z-index:18 !important; box-shadow:0 1px 0 var(--line) !important; }}
      .item-table input, .item-table select, .rate-cell {{ font-size:12px !important; min-height:30px !important; }}
      .item-table .select-col {{ width:40px !important; min-width:40px !important; }}
      .item-table .number-input, .item-table .points-input {{ width:60px !important; min-width:60px !important; }}
      .item-table select {{ width:76px !important; min-width:76px !important; padding-left:6px !important; padding-right:18px !important; }}
      .sample-pill {{ min-width:58px !important; padding:6px 8px !important; }}
      .rate-cell {{ min-width:58px !important; padding:4px 5px !important; }}
      .rate-cell b {{ font-size:13px !important; }}
      .item-table th:nth-child(13), .item-table td:nth-child(13),
      .item-table th:nth-child(14), .item-table td:nth-child(14) {{ display:none !important; }}
      .side-panel, .detail-panel {{ order:2 !important; display:none !important; max-height:42vh !important; overflow:auto !important; }}
      body.goedu-show-left .side-panel, body.goedu-show-right .detail-panel {{ display:block !important; }}
      body.goedu-show-left .table-panel, body.goedu-show-right .table-panel {{ flex:0 0 56vh !important; min-height:0 !important; }}
      .goedu-side-rail {{ display:none !important; }}
      @media (max-width: 1180px) {{
        .app-shell {{ padding:8px !important; }}
        .sticky-head {{ margin:-8px -8px 6px !important; padding:7px 8px !important; }}
        .topbar h1 {{ font-size:18px !important; }}
        .top-actions {{ justify-content:flex-start !important; }}
        .table-controls {{ min-width:0 !important; justify-content:flex-start !important; }}
        .table-panel {{ min-height:0 !important; }}
        .item-table {{ width:100% !important; min-width:900px !important; table-layout:fixed !important; }}
        .item-table th:nth-child(4), .item-table td:nth-child(4),
        .item-table th:nth-child(7), .item-table td:nth-child(7) {{ display:none !important; }}
      }}
      @media (max-width: 860px) {{
        .topbar h1 {{ font-size:17px !important; }}
        .top-actions {{ flex-basis:100% !important; }}
        .table-toolbar {{ padding:6px 7px !important; }}
        .table-toolbar h2 {{ font-size:15px !important; }}
        .table-toolbar p {{ display:none !important; }}
        .table-controls {{ justify-content:flex-start !important; }}
        .goedu-preset-strip .goedu-preset-pill {{ display:none !important; }}
        .goedu-score-strip {{ padding:6px 8px !important; gap:5px !important; border-width:1px !important; }}
        .goedu-score-title {{ font-size:13px !important; margin-right:2px !important; }}
        .goedu-score-source {{ min-height:22px !important; padding:2px 6px !important; }}
        .goedu-score-pill {{ padding:4px 7px !important; }}
        .goedu-score-pill strong {{ font-size:15px !important; }}
        .goedu-score-note {{ flex-basis:100% !important; min-width:0 !important; }}
        .item-table th, .item-table td {{ padding:4px !important; }}
        .item-table .number-input, .item-table .points-input {{ width:52px !important; min-width:52px !important; }}
        .item-table select {{ width:68px !important; min-width:68px !important; }}
        .sample-pill {{ min-width:50px !important; padding:5px 6px !important; }}
        .rate-cell {{ min-width:50px !important; }}
      }}
    `;
    document.head.appendChild(style);
  }}

  function formatPresets() {{
    return Object.entries(presets).map(([target, row]) => {{
      const values = ['A','B','C','D','E'].map(level => `${{level}}${{row[level] ?? ''}}`).join('/');
      return `<span class="goedu-preset-pill"><b>목표 ${{target}}</b>${{values}}</span>`;
    }}).join('');
  }}

  function getCutScoreSource() {{
    const live = getLiveTableCuts();
    if (live) return live;
    const sources = [
      {{ data: window.__GOEDUSPLIT_EVIDENCE__, label: '분석' }},
      {{ data: window.__GOEDUSPLIT_PROJECT__?.evidence, label: '분석' }},
      {{ data: window.__GOEDUSPLIT_PROJECT__?.evidenceData, label: '분석' }},
      {{ data: window.__GOEDUSPLIT_PROJECT__?.analysis, label: '분석' }}
    ];
    for (const source of sources) {{
      const cuts = source.data?.cuts;
      if (cuts && typeof cuts === 'object' && Object.keys(cuts).length) {{
        return {{ cuts, label: source.label }};
      }}
    }}
    if (window.__GOEDUSPLIT_MANUAL_CUTS__ && typeof window.__GOEDUSPLIT_MANUAL_CUTS__ === 'object') {{
      return {{ cuts: window.__GOEDUSPLIT_MANUAL_CUTS__, label: window.__GOEDUSPLIT_MANUAL_CUTS_LABEL__ || '기본' }};
    }}
    return null;
  }}

  function numberFromText(text) {{
    const match = String(text || '').replace(',', '.').match(/-?\\d+(?:\\.\\d+)?/);
    return match ? Number(match[0]) : NaN;
  }}

  function getLiveTableCuts() {{
    const rows = Array.from(document.querySelectorAll('.item-table tbody tr'));
    if (!rows.length) return null;
    const totals = {{ A: 0, B: 0, C: 0, D: 0, E: 0 }};
    let totalPoints = 0;
    let itemCount = 0;
    rows.forEach((row) => {{
      const cells = row.querySelectorAll('td');
      const points = Number(cells[2]?.querySelector('input')?.value ?? cells[2]?.textContent ?? 0);
      if (!Number.isFinite(points) || points <= 0) return;
      totalPoints += points;
      itemCount += 1;
      ['A', 'B', 'C', 'D', 'E'].forEach((level, index) => {{
        const cell = cells[7 + index];
        const pctText = cell?.querySelector('.rate-cell span')?.textContent || cell?.textContent || '';
        const pct = numberFromText(pctText);
        if (Number.isFinite(pct)) totals[level] += points * pct / 100;
      }});
    }});
    if (totalPoints <= 0) return null;
    const cuts = {{}};
    ['A', 'B', 'C', 'D', 'E'].forEach((level) => {{
      cuts[level] = totals[level] / totalPoints * 100;
    }});
    return {{
      cuts,
      label: `계산 · 전체 ${{itemCount}}문항 ${{formatScore(totalPoints)}}점`,
      note: '선택 체크와 무관하게 표 전체 문항의 배점 × A~D 예상정답률로 계산합니다.'
    }};
  }}

  function formatScore(value) {{
    const number = Number(value);
    if (!Number.isFinite(number)) return '';
    if (Math.abs(number - Math.round(number)) < 0.01) return String(Math.round(number));
    return number.toFixed(1);
  }}

  function formatCutScores() {{
    const source = getCutScoreSource();
    if (!source?.cuts) return '<span class="goedu-score-pill"><span>기본</span><strong>90</strong><em>점</em></span>';
    const cuts = source.cuts;
    const pairs = [['A/B', 'A'], ['B/C', 'B'], ['C/D', 'C'], ['D/E', 'D']];
    const prefix = `<span class="goedu-score-source">${{source.label || '기준'}}</span>`;
    const pills = pairs.map(([label, key]) => {{
      const score = formatScore(cuts[key]);
      return `<span class="goedu-score-pill"><span>${{label}}</span><strong>${{score || '-'}}</strong><em>점</em></span>`;
    }}).join('');
    const note = source.note ? `<span class="goedu-score-note">${{source.note}}</span>` : '';
    return prefix + pills + note;
  }}

  function polishText(root) {{
    (root || document).querySelectorAll('.panel-title').forEach((el) => {{
      const text = (el.textContent || '').trim();
      if (text === '교사 입력') el.textContent = '검토안';
      if (text.includes('예상정답 판단')) el.textContent = text.replace('예상정답 판단', 'A~E 수준별 조정');
      if (text === '교사별 A/B') el.textContent = '검토안별 A/B';
    }});
    (root || document).querySelectorAll('label.field-label').forEach((el) => {{
      if ((el.firstChild && el.firstChild.nodeType === Node.TEXT_NODE) && el.firstChild.nodeValue.includes('선택 교사명')) {{
        el.firstChild.nodeValue = '선택 검토안명';
      }}
    }});
    (root || document).querySelectorAll('button').forEach((el) => {{
      if (el.textContent === '교사 추가') el.textContent = '검토안 추가';
    }});
    (root || document).querySelectorAll('p').forEach((el) => {{
      if ((el.textContent || '').includes('입력값을 편집 중입니다.')) {{
        el.textContent = el.textContent.replace('입력값을 편집 중입니다.', '검토안을 편집 중입니다.');
      }}
      if ((el.textContent || '').includes('교사들이 입력한 예상정답률')) {{
        el.textContent = '검토안별 예상정답률의 평균과 표준편차입니다. 표준편차가 클수록 재논의가 필요합니다.';
      }}
    }});
  }}

  function installPresetStrip() {{
    const panel = document.querySelector('.table-panel');
    const wrap = document.querySelector('.item-table-wrap');
    if (!panel || !wrap) return;
    let scoreStrip = document.getElementById('goedu-score-strip');
    if (!scoreStrip) {{
      scoreStrip = document.createElement('div');
      scoreStrip.id = 'goedu-score-strip';
      scoreStrip.className = 'goedu-score-strip';
      panel.insertBefore(scoreStrip, wrap);
    }}
    const scoreHtml = `<b class="goedu-score-title">전체 예상 분할점수</b>${{formatCutScores()}}`;
    if (scoreStrip.dataset.html !== scoreHtml) {{
      scoreStrip.innerHTML = scoreHtml;
      scoreStrip.dataset.html = scoreHtml;
    }}
    let strip = document.getElementById('goedu-preset-strip');
    if (!strip) {{
      strip = document.createElement('div');
      strip.id = 'goedu-preset-strip';
      strip.className = 'goedu-preset-strip';
      panel.insertBefore(strip, wrap);
    }}
    const html = `<b>{TARGET_RATE_PRESET_TITLE}</b>${{formatPresets()}}<button type="button" data-open-presets>설정</button>`;
    if (strip.dataset.html !== html) {{
      strip.innerHTML = html;
      strip.dataset.html = html;
    }}
    const button = strip.querySelector('[data-open-presets]');
    if (button) button.onclick = () => callBridge('openTargetRatePresets');
  }}

  function installMainActions() {{
    const controls = document.querySelector('.table-controls');
    if (!controls || controls.querySelector('[data-goedu-main-actions]')) return;
    const group = document.createElement('span');
    group.dataset.goeduMainActions = '1';
    group.className = 'goedu-main-actions';
    group.innerHTML = `
      <button type="button" class="goedu-main-action" data-proxy-load>작업 불러오기</button>
      <button type="button" class="goedu-main-action" data-proxy-save>작업 저장</button>
      <button type="button" class="goedu-main-action primary" data-proxy-presets>목표 설정</button>
      <button type="button" class="goedu-main-action" data-toggle-head>제목줄 숨기기</button>
      <button type="button" class="goedu-main-action" data-toggle-summary>요약 숨기기</button>
      <button type="button" class="goedu-main-action" data-toggle-left>검토안·근거</button>
      <button type="button" class="goedu-main-action" data-toggle-right>선택 문항</button>
    `;
    controls.insertBefore(group, controls.firstChild);
    group.querySelector('[data-proxy-load]').onclick = () => document.querySelector('button[title="불러오기"]')?.click();
    group.querySelector('[data-proxy-save]').onclick = () => document.querySelector('button[title="작업 저장"]')?.click();
    group.querySelector('[data-proxy-presets]').onclick = () => callBridge('openTargetRatePresets');
    const headButton = group.querySelector('[data-toggle-head]');
    const summaryButton = group.querySelector('[data-toggle-summary]');
    const leftButton = group.querySelector('[data-toggle-left]');
    const rightButton = group.querySelector('[data-toggle-right]');
    try {{
      if (localStorage.getItem('goedu-hide-web-head') === '1') document.body.classList.add('goedu-hide-web-head');
      if (localStorage.getItem('goedu-hide-summary') === '1') document.body.classList.add('goedu-hide-summary');
    }} catch (error) {{}}
    const savePref = (key, enabled) => {{
      try {{ localStorage.setItem(key, enabled ? '1' : '0'); }} catch (error) {{}}
    }};
    const sync = () => {{
      const headHidden = document.body.classList.contains('goedu-hide-web-head');
      const summaryHidden = document.body.classList.contains('goedu-hide-summary');
      headButton.textContent = headHidden ? '제목줄 보이기' : '제목줄 숨기기';
      summaryButton.textContent = summaryHidden ? '요약 보이기' : '요약 숨기기';
      headButton.classList.toggle('active', headHidden);
      summaryButton.classList.toggle('active', summaryHidden);
      leftButton.classList.toggle('active', document.body.classList.contains('goedu-show-left'));
      rightButton.classList.toggle('active', document.body.classList.contains('goedu-show-right'));
    }};
    headButton.onclick = () => {{
      document.body.classList.toggle('goedu-hide-web-head');
      savePref('goedu-hide-web-head', document.body.classList.contains('goedu-hide-web-head'));
      sync();
    }};
    summaryButton.onclick = () => {{
      document.body.classList.toggle('goedu-hide-summary');
      savePref('goedu-hide-summary', document.body.classList.contains('goedu-hide-summary'));
      sync();
    }};
    leftButton.onclick = () => {{
      document.body.classList.toggle('goedu-show-left');
      sync();
    }};
    rightButton.onclick = () => {{
      document.body.classList.toggle('goedu-show-right');
      sync();
    }};
    sync();
  }}

  function apply() {{
    document.querySelectorAll('.goedu-side-rail').forEach((el) => el.remove());
    document.querySelectorAll('.goedu-collapsed').forEach((el) => el.classList.remove('goedu-collapsed'));
    const workspace = document.querySelector('.workspace');
    if (workspace) workspace.style.gridTemplateColumns = '';
    installCompactStyle();
    polishText(document);
    installPresetStrip();
    installMainActions();
  }}
  let attempts = 0;
  function applyWhenReady() {{
    apply();
    attempts += 1;
    if (!document.querySelector('.table-controls') && attempts < 10) {{
      setTimeout(applyWhenReady, 120);
    }}
  }}
  if (!window.__GOEDUSPLIT_PRIORITY_UI_EVENTS__) {{
    window.__GOEDUSPLIT_PRIORITY_UI_EVENTS__ = true;
    let refreshTimer = null;
    const scheduleApply = () => {{
      window.clearTimeout(refreshTimer);
      refreshTimer = window.setTimeout(apply, 80);
    }};
    window.addEventListener('message', (event) => {{
      if (event.data?.type === 'goedusplit-evidence' || event.data?.type === 'goedusplit-project') {{
        setTimeout(apply, 0);
      }}
    }});
    window.addEventListener('goedu-evidence-updated', () => setTimeout(apply, 0));
    document.addEventListener('input', (event) => {{
      if (event.target?.closest?.('.table-panel')) scheduleApply();
    }}, true);
    document.addEventListener('change', (event) => {{
      if (event.target?.closest?.('.table-panel')) scheduleApply();
    }}, true);
    document.addEventListener('click', (event) => {{
      if (event.target?.closest?.('.table-panel')) scheduleApply();
    }}, true);
    const root = document.querySelector('#root');
    if (root) {{
      const observer = new MutationObserver((mutations) => {{
        if (mutations.some((mutation) => mutation.target?.closest?.('.table-panel') || Array.from(mutation.addedNodes || []).some((node) => node.nodeType === 1 && node.closest?.('.table-panel')))) {{
          scheduleApply();
        }}
      }});
      observer.observe(root, {{ childList: true, subtree: true, characterData: true }});
    }}
  }}
  applyWhenReady();
  window.dispatchEvent(new CustomEvent('goedu-target-presets-updated', {{ detail: {{ presets, applyCurrent }} }}));
}})();
"""
        self.spliter_view.page().runJavaScript(script)

    def _toggle_spliter_summary(self):
        if not hasattr(self, "spliter_summary_body"):
            return
        visible = self.spliter_summary_body.isVisible()
        self.spliter_summary_body.setVisible(not visible)
        self._apply_tool_icon(
            self.btn_spliter_summary_toggle,
            "chevron-right" if visible else "chevron-down",
        )

    def _spliter_web_index(self) -> Path | None:
        here = Path(__file__).resolve().parent
        bundle_base = Path(getattr(sys, "_MEIPASS", here))
        candidates = [
            here / "spliter_ox_web" / "index.html",
            bundle_base / "app" / "spliter_ox_web" / "index.html",
            bundle_base / "spliter_ox_web" / "index.html",
            here.parents[1] / "spliter-ox" / "dist" / "index.html",
        ]
        for candidate in candidates:
            if candidate.exists():
                return candidate
        return None

    def _load_spliter_web(self, force_recreate: bool = False):
        if self.spliter_view is None:
            return
        if self._spliter_load_requested and not force_recreate and not self._spliter_loaded:
            return
        index = self._spliter_web_index()
        if index is None:
            self.lbl_spliter_status.setText(
                "예상정답률 계산기 파일을 찾지 못했습니다. "
                "앱을 다시 빌드하거나 예상정답률 계산기 웹 파일이 포함되어 있는지 확인해 주세요."
            )
            return
        if self.exam is not None and self.overall is not None:
            self._spliter_pending_payload = self._build_spliter_evidence_payload()
        if force_recreate:
            self._reset_spliter_web_page()
        self._spliter_loaded = False
        self._spliter_load_requested = True
        self.spliter_view.setUrl(QUrl.fromLocalFile(str(index)))

    def _on_spliter_loaded(self, ok: bool):
        self._spliter_loaded = ok
        if ok:
            self._send_spliter_bridge()
            self._send_spliter_teacher_presets()
            self._send_spliter_theme()
            self._send_spliter_zoom()
            QTimer.singleShot(450, self._send_spliter_zoom)
            QTimer.singleShot(1200, self._send_spliter_zoom)
            QTimer.singleShot(120, self._nudge_spliter_view)
            QTimer.singleShot(650, self._nudge_spliter_view)
            self._flush_spliter_project_payload()
            self._flush_spliter_payload()

    def _send_spliter_theme(self):
        if self.spliter_view is None or not self._spliter_loaded:
            return
        theme = self.theme.effective if self.theme is not None else "light"
        data = json.dumps(theme, ensure_ascii=False)
        script = (
            f"window.__GOEDUSPLIT_THEME__ = {data};"
            "window.postMessage({ type: 'goedusplit-theme', theme: window.__GOEDUSPLIT_THEME__ }, '*');"
        )
        self.spliter_view.page().runJavaScript(script)

    def _send_spliter_zoom(self):
        if self.spliter_view is None or not self._spliter_loaded:
            return
        zoom = self._zoom_percent()
        data = json.dumps(zoom, ensure_ascii=False)
        script = (
            f"window.__GOEDUSPLIT_ZOOM__ = {data};"
            "window.postMessage({ type: 'goedusplit-zoom', zoom: window.__GOEDUSPLIT_ZOOM__ }, '*');"
        )
        self.spliter_view.page().runJavaScript(script)

    def _send_spliter_manual_cuts(self):
        if self.spliter_view is None or not self._spliter_loaded:
            return
        cuts = json.dumps(self._current_cut_scores_from_inputs(), ensure_ascii=False)
        label = json.dumps(self._current_cut_scores_label(), ensure_ascii=False)
        script = (
            f"window.__GOEDUSPLIT_MANUAL_CUTS__ = {cuts};"
            f"window.__GOEDUSPLIT_MANUAL_CUTS_LABEL__ = {label};"
            "window.dispatchEvent(new CustomEvent('goedu-evidence-updated'));"
        )
        self.spliter_view.page().runJavaScript(script)

    def _flush_spliter_payload(self):
        if self.spliter_view is None or not self._spliter_loaded or self._spliter_pending_payload is None:
            return
        payload = self._spliter_pending_payload
        self._spliter_pending_payload = None
        data = json.dumps(payload, ensure_ascii=False)
        script = (
            f"window.__GOEDUSPLIT_EVIDENCE__ = {data};"
            "window.postMessage({ type: 'goedusplit-evidence', payload: window.__GOEDUSPLIT_EVIDENCE__ }, '*');"
            "window.dispatchEvent(new CustomEvent('goedu-evidence-updated'));"
        )
        self.spliter_view.page().runJavaScript(script)

    def _flush_spliter_project_payload(self):
        if self.spliter_view is None or not self._spliter_loaded or self._spliter_pending_project_payload is None:
            return
        payload = self._spliter_pending_project_payload
        self._spliter_pending_project_payload = None
        data = json.dumps(payload, ensure_ascii=False)
        script = (
            f"window.__GOEDUSPLIT_PROJECT__ = {data};"
            "window.postMessage({ type: 'goedusplit-project', payload: window.__GOEDUSPLIT_PROJECT__ }, '*');"
        )
        self.spliter_view.page().runJavaScript(script)

    def send_spliter_evidence_to_web(self):
        if self.exam is None or self.overall is None:
            QMessageBox.warning(self, "예상정답률 계산기", "먼저 분석을 실행해 주세요.")
            return
        payload = self._build_spliter_evidence_payload()
        self._spliter_pending_payload = payload
        self._activate_spliter_tab_for_pending_payloads()
        self.statusBar().showMessage("예상정답률 계산기에 현재 분석자료를 전달했습니다.", 6000)

    def import_exam_paper_to_spliter(self):
        if self.spliter_view is None:
            QMessageBox.information(self, "예상정답률 계산기", "이 환경에서는 내장 예상정답률 계산기를 열 수 없습니다.")
            return
        paths, _ = QFileDialog.getOpenFileNames(
            self,
            "시험지 또는 문항 자료 불러와 예상정답률에 반영",
            "",
            "시험지/문항 자료 (*.hwp *.hwpx *.pdf *.docx *.txt *.md *.csv *.json *.xlsx *.xlsm);;모든 파일 (*.*)",
        )
        if not paths:
            return
        existing = self._ai_review_source_text().strip()
        append_existing = False
        if existing:
            choice = QMessageBox.question(
                self,
                "시험지 자료 반영",
                f"AI 문항 검토 탭에 이미 문항 자료가 들어 있습니다.\n"
                f"선택한 {len(paths)}개 자료를 기존 자료 뒤에 추가할까요?\n\n"
                "예: 추가해서 함께 반영\n아니오: 기존 자료를 지우고 새 시험지만 반영",
                QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel,
                QMessageBox.No,
            )
            if choice == QMessageBox.Cancel:
                return
            append_existing = choice == QMessageBox.Yes
        parts = []
        failures = []
        for path_str in paths:
            path = Path(path_str)
            try:
                saved = self._store_ai_source_file(path)
                text = self._extract_text_from_review_file(str(path))
            except Exception as exc:
                failures.append(f"{path.name}: {exc}")
                continue
            parts.append(
                f"{self._format_ai_material_header(f'시험지/문항 자료: {path.name}', saved)}\n\n"
                f"{text}"
            )
        if not parts:
            QMessageBox.warning(
                self,
                "시험지 자동 반영",
                "선택한 시험지 자료를 읽지 못했습니다.\n" + "\n".join(failures[:5]),
            )
            return
        text = "\n\n".join(parts)
        if append_existing:
            text = f"{existing}\n\n{text}"
        self._set_ai_review_source_text(text, target="source")
        self._save_ai_source_cache(text)
        self._load_ai_reference_cache_if_empty()
        self._generate_ai_review_draft()
        self._send_ai_review_to_spliter()
        message = (
            f"시험지 자료 {len(parts)}개를 읽어 예상정답률 계산기에 반영했습니다. "
            "필요하면 AI 문항 검토 탭에서 성취기준·난이도·예상 O/X를 보강하세요."
        )
        if failures:
            message += f" 읽기 실패 {len(failures)}개."
            QMessageBox.warning(
                self,
                "일부 시험지 자료 읽기 실패",
                "다음 자료는 불러오지 못했습니다.\n" + "\n".join(failures[:8]),
            )
        self.statusBar().showMessage(message, 8000)

    def suggest_expected_rate_blueprint(self):
        if self.spliter_view is None:
            QMessageBox.information(self, "예상정답률 계산기", "이 환경에서는 내장 예상정답률 계산기를 열 수 없습니다.")
            return
        default_count = len(self.exam.items) if self.exam is not None and self.exam.items else 18
        count, ok = QInputDialog.getInt(
            self,
            "문항 구성안 제안",
            "새 시험의 문항 수를 입력하세요.",
            max(1, min(80, default_count)),
            1,
            80,
            1,
        )
        if not ok:
            return
        project = self._build_expected_rate_blueprint_project(count)
        self._spliter_pending_project_payload = project
        if self.exam is not None and self.overall is not None:
            self._spliter_pending_payload = self._build_spliter_evidence_payload()
        self._activate_spliter_tab_for_pending_payloads()
        basis = "현재 분석자료" if self.exam is not None and self.item_stats else "기본 100점 구성"
        self.statusBar().showMessage(f"{basis}를 바탕으로 {count}문항 구성안을 예상정답률 계산기에 제안했습니다.", 7000)

    def _build_expected_rate_blueprint_project(self, count: int) -> dict:
        count = max(1, min(80, int(count)))
        total_score = self._expected_rate_blueprint_total_score()
        levels = self._expected_rate_blueprint_level_sequence(count)
        points = self._expected_rate_blueprint_points(levels, total_score)
        source_items = self._expected_rate_blueprint_source_items(levels)
        items = []
        for idx, target in enumerate(levels, start=1):
            source = source_items[idx - 1] if idx - 1 < len(source_items) else None
            difficulty = self._expected_rate_blueprint_difficulty(target, source)
            item_type = getattr(source, "item_type", "") if source is not None else ""
            item_type = item_type if item_type in {"선택형", "서답형"} else "선택형"
            standard = ""
            content = ""
            if source is not None:
                standard = " ".join(
                    part for part in [getattr(source, "standard_code", ""), getattr(source, "standard", "")]
                    if part
                ).strip()
                content = getattr(source, "content_area", "") or ""
            expected = self._ai_default_ox_expected_values(target, difficulty)
            counts = {}
            sample_size = 3
            for lv in LEVELS_AE:
                count_value, denominator = self._parse_expected_count(expected.get(f"{lv} 예상", ""), sample_size)
                sample_size = denominator
                counts[lv] = 0 if count_value is None else count_value
            judgments = self._judgments_from_expected_counts(counts, sample_size)
            items.append({
                "id": f"blueprint-{idx}",
                "number": idx,
                "title": f"{idx}번 · {target} 수준 제안",
                "standard": standard,
                "points": points[idx - 1],
                "sampleSize": sample_size,
                "type": item_type,
                "difficulty": difficulty,
                "targetLevel": target,
                "judgmentsByJudge": {
                    "teacher-1": judgments,
                    "teacher-2": judgments,
                },
                "evidence": ["구성안 제안"],
                "note": self._expected_rate_blueprint_note(target, difficulty, points[idx - 1], content),
            })
        project = {
            "version": 1,
            "judges": [
                {"id": "teacher-1", "name": "검토안 1"},
                {"id": "teacher-2", "name": "검토안 2"},
            ],
            "activeJudgeId": "teacher-1",
            "items": items,
            "evidenceMode": "difficultyAverage",
        }
        if self.exam is not None and self.overall is not None:
            project["evidenceData"] = self._build_spliter_evidence_payload()
        return project

    def _expected_rate_blueprint_total_score(self) -> float:
        if self.exam is not None and self.exam.items:
            total = sum(float(getattr(item, "score", 0.0) or 0.0) for item in self.exam.items)
            if total > 0:
                return round(total, 2)
        return 100.0

    def _expected_rate_blueprint_level_sequence(self, count: int) -> list[str]:
        weights = {lv: 0 for lv in LEVELS_AE}
        if self.item_stats:
            for stat in self.item_stats:
                p_value = float(getattr(stat, "p_value", 0.0) or 0.0)
                if p_value >= 0.80:
                    target = "E"
                elif p_value >= 0.65:
                    target = "D"
                elif p_value >= 0.45:
                    target = "C"
                elif p_value >= 0.30:
                    target = "B"
                else:
                    target = "A"
                weights[target] += 1
        if not any(weights.values()):
            weights = {"A": 1, "B": 3, "C": 5, "D": 5, "E": 4}
        scaled = self._scale_counts_to_total(weights, count, ["E", "D", "C", "B", "A"])
        sequence = []
        for lv in ["E", "D", "C", "B", "A"]:
            sequence.extend([lv] * scaled.get(lv, 0))
        return sequence[:count] or ["C"] * count

    @staticmethod
    def _scale_counts_to_total(weights: dict[str, int | float], total: int, order: list[str]) -> dict[str, int]:
        total = max(1, int(total))
        weight_sum = sum(max(0.0, float(weights.get(key, 0.0))) for key in order)
        if weight_sum <= 0:
            weight_sum = float(len(order))
            weights = {key: 1.0 for key in order}
        raw = {key: max(0.0, float(weights.get(key, 0.0))) / weight_sum * total for key in order}
        counts = {key: int(math.floor(raw[key])) for key in order}
        remaining = total - sum(counts.values())
        for key in sorted(order, key=lambda k: raw[k] - counts[k], reverse=True):
            if remaining <= 0:
                break
            counts[key] += 1
            remaining -= 1
        while sum(counts.values()) > total:
            for key in reversed(order):
                if counts.get(key, 0) > 0 and sum(counts.values()) > total:
                    counts[key] -= 1
        return counts

    def _expected_rate_blueprint_points(self, levels: list[str], total_score: float) -> list[float]:
        if not levels:
            return []
        level_weights = {"E": 0.9, "D": 1.0, "C": 1.1, "B": 1.2, "A": 1.3}
        weights = [level_weights.get(lv, 1.0) for lv in levels]
        weight_sum = sum(weights) or 1.0
        raw = [max(1.0, total_score * weight / weight_sum) for weight in weights]
        points = [max(1, int(round(value))) for value in raw]
        diff = int(round(total_score - sum(points)))
        hard_order = sorted(range(len(levels)), key=lambda i: LEVELS_AE.index(levels[i]))
        easy_order = list(reversed(hard_order))
        guard = 0
        while diff != 0 and guard < 1000:
            order = hard_order if diff > 0 else easy_order
            changed = False
            for idx in order:
                if diff > 0:
                    points[idx] += 1
                    diff -= 1
                    changed = True
                elif points[idx] > 1:
                    points[idx] -= 1
                    diff += 1
                    changed = True
                if diff == 0:
                    break
            if not changed:
                break
            guard += 1
        return [float(value) for value in points]

    def _expected_rate_blueprint_source_items(self, levels: list[str]) -> list:
        if self.exam is None or not self.exam.items:
            return [None for _ in levels]
        source_items = sorted(
            self.exam.items,
            key=lambda item: (
                {"쉬움": 0, "보통": 1, "어려움": 2}.get(getattr(item, "difficulty", ""), 1),
                getattr(item, "number", 0),
            ),
        )
        if not source_items:
            return [None for _ in levels]
        return [source_items[(idx - 1) % len(source_items)] for idx in range(1, len(levels) + 1)]

    @staticmethod
    def _expected_rate_blueprint_difficulty(target: str, source) -> str:
        if source is not None and getattr(source, "difficulty", "") in {"쉬움", "보통", "어려움"}:
            return getattr(source, "difficulty", "")
        if target in {"A", "B"}:
            return "어려움"
        if target == "C":
            return "보통"
        return "쉬움"

    @staticmethod
    def _expected_rate_blueprint_note(target: str, difficulty: str, points: float, content: str) -> str:
        base = f"{target} 수준 문항 제안 · {difficulty} · {points:g}점"
        if content:
            base += f" · 참고 내용영역: {content}"
        return base

    def _render_spliter_tab(self):
        if not hasattr(self, "lbl_spliter_status"):
            return
        if self.exam is None or self.overall is None:
            self.lbl_spliter_status.setText("분석 전 · 계산기는 아래에서 바로 사용할 수 있습니다.")
            return
        select_count = len(self.exam.select_items)
        serdap_count = len(self.exam.serdap_items)
        serdap_text = f" · 서답형 {serdap_count}문항" if serdap_count else ""
        self.lbl_spliter_status.setText(
            f"<b>분석자료</b>: {self.exam.subject or '(과목 미상)'} · "
            f"학생 {len(self.exam.students)}명 · 선택형 근거 문항 {select_count}개{serdap_text} · "
            f"분할점수 A≥{self.exam.cut_scores['A']:.0f} B≥{self.exam.cut_scores['B']:.0f} "
            f"C≥{self.exam.cut_scores['C']:.0f} D≥{self.exam.cut_scores['D']:.0f} E≥{self.exam.cut_scores['E']:.0f}"
        )

    # ---- Data 탭 -----------------------------------------------------
    def _init_tab_data(self):
        layout = QVBoxLayout(self.tab_data)
        # KPI 카드 — 좁은 창에서 자동으로 2열/1열로 접힘
        self.kpi_grid = QGridLayout(); self.kpi_grid.setSpacing(8)
        self.kpi_n = self._kpi_card("전체 학생 수", "-")
        self.kpi_n_items = self._kpi_card("문항 수", "-")
        self.kpi_subject = self._kpi_card("과목", "-")
        self._kpis = [self.kpi_n, self.kpi_n_items, self.kpi_subject]
        self._relayout_kpis(cols=3)
        layout.addLayout(self.kpi_grid)

        self.data_empty_state = QFrame()
        self.data_empty_state.setProperty("role", "card")
        empty_layout = QVBoxLayout(self.data_empty_state)
        empty_layout.setContentsMargins(14, 12, 14, 12)
        empty_layout.setSpacing(8)
        self.lbl_data_empty_state = QLabel(build_data_empty_state_html())
        self.lbl_data_empty_state.setWordWrap(True)
        self.lbl_data_empty_state.setTextFormat(Qt.RichText)
        empty_layout.addWidget(self.lbl_data_empty_state)
        empty_buttons = QHBoxLayout()
        btn_show_input = QPushButton("입력 패널 열기")
        btn_show_input.setToolTip("왼쪽 입력 패널이 접혀 있으면 다시 펼칩니다.")
        btn_show_input.clicked.connect(self._show_sidebar)
        empty_buttons.addWidget(btn_show_input)
        btn_help = QPushButton("도움말 보기")
        btn_help.clicked.connect(lambda: self.tabs.setCurrentWidget(self.tab_help) if hasattr(self, "tabs") else None)
        empty_buttons.addWidget(btn_help)
        empty_buttons.addStretch(1)
        empty_layout.addLayout(empty_buttons)
        layout.addWidget(self.data_empty_state)

        # 학생 검색창 (이름·반/번호로 즉시 필터)
        search_box = QVBoxLayout()
        search_box.setSpacing(6)
        search_row = QHBoxLayout()
        search_row.addWidget(QLabel("학생 검색:"))
        self.le_search = QLineEdit()
        self.le_search.setPlaceholderText("이름 또는 반/번호 입력 · 여러 명은 쉼표로 구분 (예: 강예서, 신태빈, 1/3)")
        self.le_search.setMinimumWidth(self._px(160))
        self.le_search.textChanged.connect(self._filter_data_table)
        search_row.addWidget(self.le_search, 1)
        search_box.addLayout(search_row)

        filter_row = QHBoxLayout()
        filter_row.setSpacing(8)
        self.combo_filter_level = QComboBox()
        self.combo_filter_level.addItem("성취도 전체", "")
        for lv in ["A", "B", "C", "D", "E", "미도달"]:
            self.combo_filter_level.addItem(lv, lv)
        self.combo_filter_level.currentIndexChanged.connect(lambda _: self._filter_data_table(self.le_search.text()))
        filter_row.addWidget(self.combo_filter_level)
        self.combo_filter_grade9 = QComboBox()
        self.combo_filter_grade9.addItem("9등급 전체", "")
        for grade in range(1, 10):
            self.combo_filter_grade9.addItem(f"9등급 {grade}", str(grade))
        self.combo_filter_grade9.currentIndexChanged.connect(lambda _: self._filter_data_table(self.le_search.text()))
        filter_row.addWidget(self.combo_filter_grade9)
        self.combo_filter_grade5 = QComboBox()
        self.combo_filter_grade5.addItem("5등급 전체", "")
        for grade in range(1, 6):
            self.combo_filter_grade5.addItem(f"5등급 {grade}", str(grade))
        self.combo_filter_grade5.currentIndexChanged.connect(lambda _: self._filter_data_table(self.le_search.text()))
        filter_row.addWidget(self.combo_filter_grade5)
        self.chk_counsel_only_target = QCheckBox("검색 학생만 보기")
        self.chk_counsel_only_target.setToolTip("상담 모드에서 검색한 학생만 표에 남깁니다. 검색어가 없으면 전체를 익명으로 보여줍니다.")
        self.chk_counsel_only_target.setEnabled(False)
        self.chk_counsel_only_target.toggled.connect(lambda _: self._apply_privacy_to_visible_surfaces())
        filter_row.addWidget(self.chk_counsel_only_target)
        self.lbl_search_status = QLabel("")
        self.lbl_search_status.setProperty("role", "muted")
        filter_row.addWidget(self.lbl_search_status, 1)
        btn_save_portfolio = QPushButton("포트폴리오 저장")
        btn_save_portfolio.setToolTip("현재 과목 분석 결과를 학생 포트폴리오 스냅샷으로 저장합니다.")
        btn_save_portfolio.clicked.connect(self.save_current_subject_snapshot)
        filter_row.addWidget(btn_save_portfolio)
        btn_monitor = QPushButton("모니터링 탭")
        btn_monitor.clicked.connect(self._open_monitoring_tab)
        filter_row.addWidget(btn_monitor)
        self.btn_clear_search = QPushButton("✕")
        self.btn_clear_search.setFixedWidth(self._px(32))
        self.btn_clear_search.clicked.connect(lambda: self.le_search.setText(""))
        filter_row.addWidget(self.btn_clear_search)
        search_box.addLayout(filter_row)
        layout.addLayout(search_box)

        chart_option_row = QHBoxLayout()
        chart_option_row.setSpacing(10)
        chart_option_row.addWidget(QLabel("그래프 표시:"))
        self.chk_chart_achievement_cuts = QCheckBox("성취도선")
        self.chk_chart_achievement_cuts.setChecked(True)
        self.chk_chart_grade9 = QCheckBox("9등급선")
        self.chk_chart_grade9.setChecked(True)
        self.chk_chart_grade5 = QCheckBox("5등급선")
        self.chk_chart_grade5.setChecked(True)
        self.chk_chart_student_markers = QCheckBox("학생 위치")
        self.chk_chart_student_markers.setChecked(True)
        self.chk_chart_student_labels = QCheckBox("학생 이름")
        self.chk_chart_student_labels.setChecked(False)
        for checkbox in (
            self.chk_chart_achievement_cuts,
            self.chk_chart_grade9,
            self.chk_chart_grade5,
            self.chk_chart_student_markers,
            self.chk_chart_student_labels,
        ):
            checkbox.toggled.connect(lambda _: self._refresh_score_histogram_options())
            chart_option_row.addWidget(checkbox)
        chart_option_row.addStretch(1)
        layout.addLayout(chart_option_row)

        # 차트 ↔ 표 splitter
        data_split = QSplitter(Qt.Vertical)
        self.score_chart_tabs = QTabWidget()
        self.score_chart_tabs.setMinimumHeight(max(190, self._px(220)))
        self.canvas_score_hist = CanvasHolder()
        self.score_chart_tabs.addTab(self.canvas_score_hist, "환산점수 분포")

        normal_page = QWidget()
        normal_layout = QVBoxLayout(normal_page); normal_layout.setContentsMargins(0, 0, 0, 0)
        self.canvas_score_normal = CanvasHolder()
        normal_layout.addWidget(self.canvas_score_normal, 1)
        self.lbl_normal_note = QLabel(
            "정규분포 곡선은 현재 학급 점수의 평균과 표준편차를 기준으로 한 참고선입니다. "
            "현장 모니터링에서는 실제 학교·교과 상황과 함께 해석하세요."
        )
        self.lbl_normal_note.setProperty("role", "muted"); self.lbl_normal_note.setWordWrap(True)
        normal_layout.addWidget(self.lbl_normal_note)
        self.score_chart_tabs.addTab(normal_page, "정규분포·점검")
        data_split.addWidget(self.score_chart_tabs)

        self.table_data = QTableWidget(0, 0)
        _setup_table(self.table_data, word_wrap=False, horizontal_scroll=True)
        self.table_data.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.table_data.setSortingEnabled(True)  # 헤더 클릭으로 정렬
        data_split.addWidget(self.table_data)

        data_split.setStretchFactor(0, 1); data_split.setStretchFactor(1, 2)
        data_split.setSizes([240, 380])
        layout.addWidget(data_split, 1)

        self.lbl_data_note = QLabel(
            "▸ 표 머리글을 클릭하면 정렬됩니다 ▸ 학생 검색과 성취도/9등급/5등급 필터를 함께 사용할 수 있습니다 ▸ 그래프 위 마우스 휠로 확대/축소\n"
            "성취수준은 환산점수를 반올림한 원점수 기준, 9등급/5등급은 현재 분석 집단 안의 상대 석차 기준입니다. "
            "상담 모드(Ctrl+Shift+H)를 켜면 검색 학생 외 이름과 반/번호를 가립니다."
        )
        self.lbl_data_note.setProperty("role", "muted"); self.lbl_data_note.setWordWrap(True)
        layout.addWidget(self.lbl_data_note)

    # ---- 학생 포트폴리오 ----------------------------------------------
    def _init_tab_portfolio(self):
        layout = QVBoxLayout(self.tab_portfolio)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(8)
        top = QHBoxLayout()
        title = QLabel("학생 포트폴리오")
        title.setProperty("role", "title")
        top.addWidget(title)
        top.addStretch(1)
        self.le_portfolio_search = QLineEdit()
        self.le_portfolio_search.setPlaceholderText("학생 이름 또는 반/번호 검색")
        self.le_portfolio_search.textChanged.connect(self._filter_portfolio_table)
        top.addWidget(self.le_portfolio_search, 1)
        self.combo_portfolio_student = QComboBox()
        self.combo_portfolio_student.addItem("학생 선택", "")
        self.combo_portfolio_student.currentIndexChanged.connect(lambda _: self._on_portfolio_student_selected())
        top.addWidget(self.combo_portfolio_student)
        self.combo_portfolio_subject = QComboBox()
        self.combo_portfolio_subject.addItem("전체 과목", "")
        self.combo_portfolio_subject.currentIndexChanged.connect(
            lambda _: self._filter_portfolio_table(self.le_portfolio_search.text())
        )
        top.addWidget(self.combo_portfolio_subject)
        btn_report = QPushButton("상담 리포트 보기")
        btn_report.setToolTip("선택한 학생의 여러 과목 기록을 상담용 리포트로 정리합니다.")
        btn_report.clicked.connect(self.show_selected_student_portfolio)
        top.addWidget(btn_report)
        btn_reload = QPushButton("저장자료 불러오기")
        btn_reload.clicked.connect(self.refresh_portfolio_tab)
        top.addWidget(btn_reload)
        btn_location = QPushButton("저장 위치")
        btn_location.clicked.connect(self.show_portfolio_store_location)
        top.addWidget(btn_location)
        layout.addLayout(top)

        self.lbl_portfolio_note = QLabel(
            "Data 탭의 '포트폴리오 저장'을 누르면 과목별 학생 결과가 이곳에 쌓입니다. "
            "이후 학생 선택에서 한 학생을 고르면 여러 과목 흐름이 자동으로 정리됩니다. "
            "상담 모드에서는 검색 대상 외 학생 이름과 반/번호가 가려집니다."
        )
        self.lbl_portfolio_note.setProperty("role", "muted")
        self.lbl_portfolio_note.setWordWrap(True)
        layout.addWidget(self.lbl_portfolio_note)

        self.lbl_portfolio_summary = QLabel("학생 선택에서 한 학생을 고르면 저장된 과목 흐름을 요약해 보여줍니다.")
        self.lbl_portfolio_summary.setProperty("role", "muted")
        self.lbl_portfolio_summary.setWordWrap(True)
        layout.addWidget(self.lbl_portfolio_summary)

        self.table_portfolio = QTableWidget(0, 9)
        self.table_portfolio.setHorizontalHeaderLabels([
            "저장일", "과목", "학기/학년", "반/번호", "이름", "성취도", "9등급", "5등급", "원점수",
        ])
        _setup_table(self.table_portfolio, word_wrap=False, horizontal_scroll=True)
        self.table_portfolio.setSortingEnabled(True)
        layout.addWidget(self.table_portfolio, 1)
        self.refresh_portfolio_tab()

    # ---- 모니터링 ------------------------------------------------------
    def _make_collapsible_panel(self, title: str, body: QWidget, *, opened: bool = True) -> QFrame:
        panel = QFrame()
        panel.setProperty("role", "card")
        panel.body_widget = body
        panel.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)
        body.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(10, 8, 10, 10)
        layout.setSpacing(8)
        layout.setSizeConstraint(QLayout.SetMinimumSize)
        if body.layout() is not None:
            body.layout().setSizeConstraint(QLayout.SetMinimumSize)

        toggle = QToolButton()
        panel.toggle_button = toggle
        toggle.setText(title)
        toggle.setCheckable(True)
        toggle.setChecked(opened)
        toggle.setArrowType(Qt.DownArrow if opened else Qt.RightArrow)
        toggle.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)
        toggle.setProperty("role", "collapsebtn")
        toggle.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        def _sync(opened: bool):
            body.setVisible(opened)
            toggle.setArrowType(Qt.DownArrow if opened else Qt.RightArrow)
            self._refresh_collapsible_panel_size(panel)
            QTimer.singleShot(0, lambda: self._refresh_collapsible_panel_size(panel))

        toggle.toggled.connect(_sync)
        layout.addWidget(toggle)
        layout.addWidget(body)
        body.setVisible(opened)
        self._refresh_collapsible_panel_size(panel)
        return panel

    def _refresh_collapsible_panel_size(self, panel: QFrame):
        body = getattr(panel, "body_widget", None)
        if body is not None:
            if body.layout() is not None:
                body.layout().activate()
                body.setMinimumHeight(body.layout().minimumSize().height() if not body.isHidden() else 0)
            body.updateGeometry()
        if panel.layout() is not None:
            panel.layout().activate()
        panel.setMinimumHeight(max(self._px(52), panel.sizeHint().height() + self._px(8)))
        panel.setMaximumHeight(16777215)
        panel.updateGeometry()

    def _init_tab_monitor(self):
        layout = QVBoxLayout(self.tab_monitor)
        head_row = QHBoxLayout()
        intro = QLabel(
            "현재 과목의 A 비율, 과목 평균, A/B 분할점수를 "
            "동일 학교유형·교과군의 전국 평균±표준편차 및 전년도 값과 비교합니다. "
            "전국 기준값은 외부 모니터링 자료가 있을 때 입력하세요."
        )
        intro.setProperty("role", "muted")
        intro.setWordWrap(True)
        head_row.addWidget(intro, 1)
        btn_refresh = QPushButton("새로고침")
        btn_refresh.clicked.connect(self._refresh_monitoring_tab)
        head_row.addWidget(btn_refresh)
        layout.addLayout(head_row)

        top_area = QWidget()
        top_layout = QVBoxLayout(top_area)
        top_layout.setContentsMargins(0, 0, 0, 0)
        top_layout.setSpacing(8)
        current_grid = QGridLayout(); current_grid.setSpacing(8)
        self.monitor_current_cards = {
            "a_pct": self._kpi_card("현재 A 비율", "-"),
            "mean": self._kpi_card("현재 과목 평균", "-"),
            "std": self._kpi_card("현재 표준편차", "-"),
            "cut_a": self._kpi_card("현재 A/B 분할점수", "-"),
        }
        for i, card in enumerate(self.monitor_current_cards.values()):
            current_grid.addWidget(card, 0, i)
        top_layout.addLayout(current_grid)

        self.monitor_spins: dict[str, StepperSpinBox] = {}
        input_row = QSplitter(Qt.Horizontal)
        national_body = QWidget()
        national_form = QFormLayout(national_body); national_form.setLabelAlignment(Qt.AlignRight)
        self._add_monitor_spin(national_form, "ref_a_mean", "A 비율 평균", 0.0, suffix=" %")
        self._add_monitor_spin(national_form, "ref_a_sd", "A 비율 표준편차", 0.0, suffix=" %p")
        self._add_monitor_spin(national_form, "ref_mean_mean", "과목 평균", 0.0, suffix=" 점")
        self._add_monitor_spin(national_form, "ref_mean_sd", "과목 평균 표준편차", 0.0, suffix=" 점")
        self._add_monitor_spin(national_form, "ref_cut_mean", "A/B 분할점수 평균", 0.0, suffix=" 점")
        self._add_monitor_spin(national_form, "ref_cut_sd", "A/B 분할점수 표준편차", 0.0, suffix=" 점")
        national_box = self._make_collapsible_panel("동일 학교유형·교과군 전국 기준", national_body)
        input_row.addWidget(national_box)

        history_body = QWidget()
        history_form = QFormLayout(history_body); history_form.setLabelAlignment(Qt.AlignRight)
        self._add_monitor_spin(history_form, "prev_a_pct", "전년도 A 비율", 0.0, suffix=" %")
        self._add_monitor_spin(history_form, "prev_mean", "전년도 과목 평균", 0.0, suffix=" 점")
        self._add_monitor_spin(history_form, "prev_cut_a", "전년도 A/B 분할점수", 0.0, suffix=" 점")
        self._add_monitor_spin(history_form, "th_a_delta", "A 비율 증가 기준", 10.0, suffix=" %p")
        self._add_monitor_spin(history_form, "th_mean_stable", "평균 큰 변동 아님", 5.0, suffix=" 점")
        self._add_monitor_spin(history_form, "th_cut_delta", "분할점수 감소 기준", 10.0, suffix=" 점")
        history_box = self._make_collapsible_panel("전년도·판정 기준", history_body)
        input_row.addWidget(history_box)
        input_row.setStretchFactor(0, 1); input_row.setStretchFactor(1, 1)
        top_layout.addWidget(input_row, 1)

        body = QSplitter(Qt.Vertical)
        self.canvas_monitor = CanvasHolder()
        body.addWidget(self.canvas_monitor)

        flow_note = QLabel(
            "점검 흐름: 1단계 A 비율 변화 확인 → 2단계 대상교·학생 특성 확인 → "
            "3단계 지필평가 특성 확인 → 4단계 분할점수 재산출 → 5단계 A 비율 재산정"
        )
        flow_note.setProperty("role", "muted")
        flow_note.setWordWrap(True)

        table_wrap = QWidget()
        table_layout = QVBoxLayout(table_wrap); table_layout.setContentsMargins(0, 0, 0, 0)
        table_layout.addWidget(flow_note)
        self.table_monitor = QTableWidget(0, 5)
        self.table_monitor.setHorizontalHeaderLabels(["지표", "현재/변화", "기준", "판정", "해석·질문"])
        _setup_table(self.table_monitor, word_wrap=True, horizontal_scroll=True, row_height=42)
        self.table_monitor.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        table_layout.addWidget(self.table_monitor, 1)
        body.addWidget(table_wrap)
        body.setStretchFactor(0, 1); body.setStretchFactor(1, 2)
        body.setSizes([260, 360])
        monitor_split = QSplitter(Qt.Vertical)
        monitor_split.addWidget(top_area)
        monitor_split.addWidget(body)
        monitor_split.setStretchFactor(0, 0)
        monitor_split.setStretchFactor(1, 1)
        monitor_split.setSizes([250, 620])
        layout.addWidget(monitor_split, 1)
        self._render_monitor_tab()

    def _refresh_monitoring_tab(self):
        self._render_monitor_tab()
        self.statusBar().showMessage("모니터링 탭을 새로고침했습니다.", 2500)

    def _add_monitor_spin(self, form: QFormLayout, key: str, label: str, value: float,
                          *, suffix: str = ""):
        spin = StepperSpinBox(value=value, minimum=-100.0, maximum=200.0,
                              step=1.0, decimals=1, suffix=suffix)
        spin.valueChanged.connect(lambda *_: self._render_monitor_tab())
        self.monitor_spins[key] = spin
        form.addRow(label, spin)
        return spin

    def _monitor_value(self, key: str) -> float:
        spin = getattr(self, "monitor_spins", {}).get(key)
        return float(spin.value()) if spin is not None else 0.0

    def _monitor_z_status(self, value: float, ref: float, sd: float, *, low_is_risk: bool = False):
        if sd <= 0:
            return "기준 입력 필요", "전국 평균과 표준편차를 입력하면 Ⅰ/Ⅱ/Ⅲ 수준을 자동 판정합니다.", 0
        z = (ref - value) / sd if low_is_risk else (value - ref) / sd
        if z >= 2:
            return "Ⅲ 점검", f"전국 평균에서 {'아래로' if low_is_risk else '위로'} {z:.2f}σ 벗어났습니다.", 3
        if z >= 1:
            return "Ⅱ 주의", f"전국 평균에서 {'아래로' if low_is_risk else '위로'} {z:.2f}σ 벗어났습니다.", 2
        return "참고 범위", f"전국 평균±1σ 안쪽 또는 위험 방향이 아닙니다. z={z:.2f}", 1

    def _monitor_rows_and_benchmarks(self):
        if self.exam is None or self.overall is None:
            return [], []
        a_pct = float(self.overall.level_dist_pct.get("A", 0.0))
        mean = float(self.overall.mean)
        std = float(self.overall.std)
        cut_a = float(self.exam.cut_scores.get("A", 0.0))
        prev_a = self._monitor_value("prev_a_pct")
        prev_mean = self._monitor_value("prev_mean")
        prev_cut = self._monitor_value("prev_cut_a")
        th_a = self._monitor_value("th_a_delta")
        th_mean = self._monitor_value("th_mean_stable")
        th_cut = self._monitor_value("th_cut_delta")

        rows = []
        status, detail, sev = self._monitor_z_status(
            a_pct, self._monitor_value("ref_a_mean"), self._monitor_value("ref_a_sd")
        )
        rows.append((
            "필수 1 · 성취수준 A 비율이 높음",
            f"{a_pct:.1f}%",
            "전국 평균+1σ 이상: Ⅱ · +2σ 이상: Ⅲ",
            status,
            f"{detail} A 비율만으로 단정하지 않고 평균·분할점수와 함께 봅니다.",
            sev,
        ))

        if mean >= cut_a:
            status, detail, sev = "Ⅲ 점검", "과목 평균이 A/B 분할점수 이상입니다.", 3
        elif cut_a - mean <= 3:
            status, detail, sev = "Ⅱ 주의", "과목 평균이 A/B 분할점수에 매우 가깝습니다.", 2
        else:
            status, detail, sev = "참고 범위", "과목 평균이 A/B 분할점수보다 낮습니다.", 1
        rows.append((
            "필수 2 · 과목 평균이 A/B 분할점수보다 높음",
            f"평균 {mean:.1f}점 / A/B {cut_a:.1f}점",
            "평균 ≥ A/B 분할점수이면 점검",
            status,
            f"{detail} 평균 학생이 A 수준 경계 위에 놓이는지 보는 지표입니다.",
            sev,
        ))

        status, detail, sev = self._monitor_z_status(
            mean, self._monitor_value("ref_mean_mean"), self._monitor_value("ref_mean_sd")
        )
        rows.append((
            "특성 3 · 과목 평균이 높음",
            f"{mean:.1f}점",
            "전국 평균+1σ 이상: 주의 · +2σ 이상: 점검",
            status,
            f"{detail} 평가도구가 쉬웠는지, 학생 특성이 높은지 함께 봅니다.",
            sev,
        ))

        status, detail, sev = self._monitor_z_status(
            cut_a, self._monitor_value("ref_cut_mean"), self._monitor_value("ref_cut_sd"),
            low_is_risk=True,
        )
        rows.append((
            "특성 4 · A/B 분할점수가 낮음",
            f"{cut_a:.1f}점",
            "전국 평균-1σ 이하: 주의 · -2σ 이하: 점검",
            status,
            f"{detail} A 비율이 높은 원인이 낮은 분할점수인지 확인합니다.",
            sev,
        ))

        if prev_a <= 0 or prev_mean <= 0:
            status, detail, sev = "기준 입력 필요", "전년도 A 비율과 과목 평균을 입력하면 판정합니다.", 0
            current = "-"
        else:
            delta_a = a_pct - prev_a
            delta_mean = mean - prev_mean
            current = f"A {delta_a:+.1f}%p / 평균 {delta_mean:+.1f}점"
            if delta_a >= th_a and abs(delta_mean) <= th_mean:
                status, detail, sev = "특성 발생", "평균은 크게 변하지 않았는데 A 비율이 증가했습니다.", 3
            elif delta_a >= th_a:
                status, detail, sev = "Ⅱ 주의", "A 비율은 증가했지만 평균도 함께 변했습니다.", 2
            else:
                status, detail, sev = "참고 범위", "A 비율 증가 폭이 기준보다 작습니다.", 1
        rows.append((
            "특성 5 · 평균 큰 변동 없이 A 비율 증가",
            current,
            f"A 증가 ≥ {th_a:.1f}%p, 평균 변화 ≤ {th_mean:.1f}점",
            status,
            f"{detail} 역동적 성적 과대평가 가능성을 묻는 지표입니다.",
            sev,
        ))

        if prev_a <= 0 or prev_cut <= 0:
            status, detail, sev = "기준 입력 필요", "전년도 A 비율과 A/B 분할점수를 입력하면 판정합니다.", 0
            current = "-"
        else:
            delta_a = a_pct - prev_a
            cut_drop = prev_cut - cut_a
            current = f"A {delta_a:+.1f}%p / A/B {cut_a - prev_cut:+.1f}점"
            if delta_a >= th_a and cut_drop >= th_cut:
                status, detail, sev = "특성 발생", "A 비율 증가와 A/B 분할점수 감소가 함께 나타났습니다.", 3
            elif delta_a >= th_a and cut_drop > 0:
                status, detail, sev = "Ⅱ 주의", "A 비율 증가와 A/B 분할점수 감소가 함께 보입니다.", 2
            else:
                status, detail, sev = "참고 범위", "두 조건이 동시에 충분히 나타나지 않았습니다.", 1
        rows.append((
            "특성 6 · A 비율 증가 + A/B 분할점수 감소",
            current,
            f"A 증가 ≥ {th_a:.1f}%p, A/B 감소 ≥ {th_cut:.1f}점",
            status,
            f"{detail} 쉬워진 평가도구인지, 낮아진 분할점수인지 추가 확인합니다.",
            sev,
        ))

        benchmarks = [
            {
                "label": "A 비율",
                "value": a_pct,
                "ref": self._monitor_value("ref_a_mean") if self._monitor_value("ref_a_sd") > 0 else None,
                "sd": self._monitor_value("ref_a_sd") if self._monitor_value("ref_a_sd") > 0 else None,
                "color": "#ef4444",
            },
            {
                "label": "과목 평균",
                "value": mean,
                "ref": self._monitor_value("ref_mean_mean") if self._monitor_value("ref_mean_sd") > 0 else None,
                "sd": self._monitor_value("ref_mean_sd") if self._monitor_value("ref_mean_sd") > 0 else None,
                "color": "#f59e0b",
            },
            {
                "label": "A/B 분할점수",
                "value": cut_a,
                "ref": self._monitor_value("ref_cut_mean") if self._monitor_value("ref_cut_sd") > 0 else None,
                "sd": self._monitor_value("ref_cut_sd") if self._monitor_value("ref_cut_sd") > 0 else None,
                "color": "#2a7770",
            },
        ]
        return rows, benchmarks

    def _render_monitor_tab(self):
        if not hasattr(self, "table_monitor"):
            return
        if self.exam is None or self.overall is None:
            for card in getattr(self, "monitor_current_cards", {}).values():
                card.value_label.setText("-")
            self.table_monitor.setRowCount(0)
            if hasattr(self, "canvas_monitor"):
                self.canvas_monitor.set_figure(charts.fig_monitoring_benchmarks([]))
            return

        a_pct = float(self.overall.level_dist_pct.get("A", 0.0))
        self.monitor_current_cards["a_pct"].value_label.setText(f"{a_pct:.1f}%")
        self.monitor_current_cards["mean"].value_label.setText(f"{self.overall.mean:.1f}점")
        self.monitor_current_cards["std"].value_label.setText(f"{self.overall.std:.1f}점")
        self.monitor_current_cards["cut_a"].value_label.setText(f"{self.exam.cut_scores.get('A', 0):.1f}점")

        rows, benchmarks = self._monitor_rows_and_benchmarks()
        self.canvas_monitor.set_figure(charts.fig_monitoring_benchmarks(benchmarks))

        self.table_monitor.setSortingEnabled(False)
        self.table_monitor.setRowCount(len(rows))
        colors = self.theme.colors if self.theme else {}
        shade = QColor(colors.get("shade", "#e1e6ee"))
        warn = QColor("#f8dfa3"); warn.setAlpha(150)
        danger = QColor("#f5b5b5"); danger.setAlpha(165)
        ok = QColor(colors.get("card", "#f4f7f8"))
        for r, row in enumerate(rows):
            bg = danger if row[5] >= 3 else warn if row[5] == 2 else shade if row[5] == 0 else ok
            for c, value in enumerate(row[:5]):
                _set_item(
                    self.table_monitor,
                    r,
                    c,
                    value,
                    align_left=(c in (0, 4)),
                    bg=bg if c == 3 else None,
                    bold=(c == 3),
                    tooltip=str(value),
                )
        self.table_monitor.setSortingEnabled(False)
        for col, width in enumerate([250, 180, 260, 120, 520]):
            self._set_scaled_column_width(self.table_monitor, col, width)
        self._apply_zoom_to_surfaces()

    def _relayout_kpis(self, cols: int):
        """KPI 카드들을 cols 개수만큼 한 행에 배치."""
        # 기존 위치 모두 제거 (위젯은 보존)
        for k in self._kpis:
            self.kpi_grid.removeWidget(k)
        for i, k in enumerate(self._kpis):
            r, c = divmod(i, max(1, cols))
            self.kpi_grid.addWidget(k, r, c)

    def _kpi_card(self, label: str, value: str) -> QFrame:
        frame = QFrame()
        frame.setProperty("role", "kpi")
        v = QVBoxLayout(frame); v.setContentsMargins(14, 10, 14, 10); v.setSpacing(2)
        l1 = QLabel(label); l1.setProperty("role", "kpi-label")
        l2 = QLabel(value); l2.setProperty("role", "kpi-value")
        v.addWidget(l1); v.addWidget(l2)
        frame.value_label = l2
        return frame

    # ---- 전체 성취도 ----------------------------------------------------
    def _init_tab_overview(self):
        layout = QVBoxLayout(self.tab_overview)

        ov_split = QSplitter(Qt.Vertical)

        # 위쪽: 도넛 + 평균±표준편차
        top_widget = QWidget(); top_l = QHBoxLayout(top_widget); top_l.setContentsMargins(0,0,0,0)
        self.canvas_level_stack = CanvasHolder()
        self.canvas_level_means = CanvasHolder()
        top_l.addWidget(self.canvas_level_stack, 1)
        top_l.addWidget(self.canvas_level_means, 1)
        ov_split.addWidget(top_widget)

        # 가운데: 성취수준 표
        self.table_level = QTableWidget(0, 5)
        self.table_level.setHorizontalHeaderLabels(["성취수준", "학생수", "비율(%)", "평균(원점수)", "표준편차"])
        _setup_table(self.table_level)
        ov_split.addWidget(self.table_level)

        # 아래: 학급별 평균
        self.canvas_class = CanvasHolder()
        ov_split.addWidget(self.canvas_class)

        ov_split.setSizes([340, 200, 280])
        ov_split.setStretchFactor(0, 3); ov_split.setStretchFactor(1, 1); ov_split.setStretchFactor(2, 2)
        layout.addWidget(ov_split, 1)

    # ---- 수행평가 분석 --------------------------------------------------
    def _init_tab_perform(self):
        layout = QVBoxLayout(self.tab_perform)

        head_row = QHBoxLayout()
        self.lbl_perform_tab_note = QLabel(
            "수행평가 영역별 점수율, 지필총점과의 관계, 학생별 차이를 함께 봅니다. "
            "왼쪽 입력 패널에서 수행평가 파일을 지정하고 체크한 뒤 분석을 실행하세요."
        )
        self.lbl_perform_tab_note.setProperty("role", "muted")
        self.lbl_perform_tab_note.setWordWrap(True)
        head_row.addWidget(self.lbl_perform_tab_note, 1)
        btn_refresh = QPushButton("새로고침")
        btn_refresh.clicked.connect(self._refresh_perform_tab)
        head_row.addWidget(btn_refresh)
        layout.addLayout(head_row)

        grid = QGridLayout(); grid.setSpacing(8)
        self.perform_cards = {
            "ratio": self._kpi_card("수행 반영비율", "-"),
            "areas": self._kpi_card("수행 영역", "-"),
            "matched": self._kpi_card("매칭 학생", "-"),
            "corr": self._kpi_card("지필-수행 상관", "-"),
        }
        for i, card in enumerate(self.perform_cards.values()):
            grid.addWidget(card, 0, i)
        layout.addLayout(grid)

        split = QSplitter(Qt.Vertical)

        chart_row = QWidget()
        chart_layout = QHBoxLayout(chart_row); chart_layout.setContentsMargins(0, 0, 0, 0)
        self.canvas_perform_area = CanvasHolder()
        self.canvas_perform_scatter = CanvasHolder()
        chart_layout.addWidget(self.canvas_perform_area, 1)
        chart_layout.addWidget(self.canvas_perform_scatter, 1)
        split.addWidget(chart_row)

        table_tabs = QTabWidget()
        self.perform_table_tabs = table_tabs
        self.table_perform_areas = QTableWidget(0, 14)
        self.table_perform_areas.setHorizontalHeaderLabels([
            "영역", "만점", "반영비율", "평균", "표준편차", "평균점수율",
            "만점자", "50% 미만", "A", "B", "C", "D", "E", "미도달",
        ])
        _setup_table(self.table_perform_areas, word_wrap=False, horizontal_scroll=True, row_height=30)
        self.table_perform_areas.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        self._perform_area_bar_delegate = ItemBarDelegate(self.table_perform_areas)
        self.table_perform_areas.setItemDelegateForColumn(5, self._perform_area_bar_delegate)
        table_tabs.addTab(self.table_perform_areas, "영역별 요약")

        self.table_perform_students = QTableWidget(0, 7)
        self.table_perform_students.setHorizontalHeaderLabels([
            "반/번호", "이름", "성취도", "지필총점", "수행환산", "수행-지필", "해석",
        ])
        _setup_table(self.table_perform_students, word_wrap=False, horizontal_scroll=True, row_height=30)
        self.table_perform_students.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        table_tabs.addTab(self.table_perform_students, "학생별 차이")

        recalc_page = QWidget()
        recalc_layout = QVBoxLayout(recalc_page)
        recalc_layout.setContentsMargins(8, 8, 8, 8)
        recalc_layout.setSpacing(8)

        recalc_head = QHBoxLayout()
        self.lbl_perform_recalc_note = QLabel(
            "평가요소별로 A~E 최소능력자가 받을 것으로 예상되는 점수를 넣으면 "
            "수행평가 분할점수가 자동 계산됩니다."
        )
        self.lbl_perform_recalc_note.setProperty("role", "muted")
        self.lbl_perform_recalc_note.setWordWrap(True)
        recalc_head.addWidget(self.lbl_perform_recalc_note, 1)
        btn_perf_recommend = QPushButton("자료 기준 딸깍 추천")
        btn_perf_recommend.clicked.connect(self._recommend_perform_recalc_rows)
        recalc_head.addWidget(btn_perf_recommend)
        btn_perf_default = QPushButton("기본안")
        btn_perf_default.clicked.connect(self._reset_perform_recalc_rows)
        recalc_head.addWidget(btn_perf_default)
        btn_perf_add = QPushButton("평가요소 추가")
        btn_perf_add.clicked.connect(self._add_perform_recalc_row)
        recalc_head.addWidget(btn_perf_add)
        btn_perf_del = QPushButton("선택 삭제")
        btn_perf_del.clicked.connect(self._delete_selected_perform_recalc_rows)
        recalc_head.addWidget(btn_perf_del)
        recalc_layout.addLayout(recalc_head)

        cut_grid = QGridLayout(); cut_grid.setSpacing(8)
        self.perform_recalc_cards = {
            "max": self._kpi_card("수행 만점", "-"),
            "A": self._kpi_card("A/B", "-"),
            "B": self._kpi_card("B/C", "-"),
            "C": self._kpi_card("C/D", "-"),
            "D": self._kpi_card("D/E", "-"),
            "E": self._kpi_card("E/미도달", "-"),
            "check": self._kpi_card("점검", "-"),
        }
        for i, card in enumerate(self.perform_recalc_cards.values()):
            cut_grid.addWidget(card, i // 4, i % 4)
        recalc_layout.addLayout(cut_grid)

        recalc_split = QSplitter(Qt.Vertical)
        self.table_perform_recalc = QTableWidget(0, 8)
        self.table_perform_recalc.setHorizontalHeaderLabels([
            "평가요소", "만점", "A 예상", "B 예상", "C 예상", "D 예상", "E 예상", "근거/메모",
        ])
        _setup_table(self.table_perform_recalc, word_wrap=False, horizontal_scroll=True, row_height=32)
        self.table_perform_recalc.setEditTriggers(
            QAbstractItemView.DoubleClicked
            | QAbstractItemView.EditKeyPressed
            | QAbstractItemView.AnyKeyPressed
        )
        self.table_perform_recalc.itemChanged.connect(self._render_perform_recalc_results)
        recalc_split.addWidget(self.table_perform_recalc)

        self.table_perform_recalc_results = QTableWidget(0, 4)
        self.table_perform_recalc_results.setHorizontalHeaderLabels(["경계", "원점수", "100점 환산", "의미"])
        _setup_table(self.table_perform_recalc_results, word_wrap=False, horizontal_scroll=True, row_height=30)
        recalc_split.addWidget(self.table_perform_recalc_results)
        recalc_split.setStretchFactor(0, 3)
        recalc_split.setStretchFactor(1, 1)
        recalc_split.setSizes([330, 170])
        recalc_layout.addWidget(recalc_split, 1)
        table_tabs.addTab(recalc_page, "분할점수 재산정")

        split.addWidget(table_tabs)
        split.setStretchFactor(0, 1)
        split.setStretchFactor(1, 2)
        split.setSizes([300, 430])
        layout.addWidget(split, 1)
        self._reset_perform_recalc_rows()
        self._render_perform_tab()

    def _refresh_perform_tab(self):
        self._render_perform_tab()
        self.statusBar().showMessage("수행평가 분석 탭을 새로고침했습니다.", 2500)

    def _set_perform_recalc_item(self, row: int, col: int, value, *, align_right=False):
        item = QTableWidgetItem("" if value is None else str(value))
        item.setTextAlignment((Qt.AlignRight if align_right else Qt.AlignLeft) | Qt.AlignVCenter)
        self.table_perform_recalc.setItem(row, col, item)
        return item

    def _add_perform_recalc_row(self, values: dict | None = None):
        if not hasattr(self, "table_perform_recalc"):
            return
        user_added = values is None
        values = values or {}
        if user_added:
            self._perform_recalc_dirty = True
        table = self.table_perform_recalc
        table.blockSignals(True)
        row = table.rowCount()
        table.insertRow(row)
        max_score = float(values.get("max_score", 10.0))
        name = values.get("name", f"평가요소 {row + 1}")
        self._set_perform_recalc_item(row, 0, name)
        self._set_perform_recalc_item(row, 1, f"{max_score:.1f}", align_right=True)
        for idx, lv in enumerate(LEVELS_AE, start=2):
            default = max_score * PERFORM_DEFAULT_RATES[lv]
            score = float(values.get(lv, default))
            self._set_perform_recalc_item(row, idx, f"{score:.1f}", align_right=True)
        self._set_perform_recalc_item(row, 7, values.get("memo", ""))
        table.blockSignals(False)
        self._render_perform_recalc_results()

    def _reset_perform_recalc_rows(self):
        if not hasattr(self, "table_perform_recalc"):
            return
        self._perform_recalc_dirty = False
        self.table_perform_recalc.blockSignals(True)
        self.table_perform_recalc.setRowCount(0)
        self.table_perform_recalc.blockSignals(False)
        if self.perform_data is not None and getattr(self.perform_data, "areas", None):
            self._recommend_perform_recalc_rows()
            return
        self._add_perform_recalc_row({"name": "평가요소 1", "max_score": 10.0, "memo": "성취율 기준 기본안"})
        self._add_perform_recalc_row({"name": "평가요소 2", "max_score": 10.0, "memo": "성취율 기준 기본안"})
        self._add_perform_recalc_row({"name": "평가요소 3", "max_score": 10.0, "memo": "성취율 기준 기본안"})

    def _delete_selected_perform_recalc_rows(self):
        if not hasattr(self, "table_perform_recalc"):
            return
        rows = sorted({idx.row() for idx in self.table_perform_recalc.selectedIndexes()}, reverse=True)
        if not rows:
            return
        self._perform_recalc_dirty = True
        self.table_perform_recalc.blockSignals(True)
        for row in rows:
            self.table_perform_recalc.removeRow(row)
        self.table_perform_recalc.blockSignals(False)
        self._render_perform_recalc_results()

    def _recommended_perform_scores_by_area(self):
        recommendations = []
        if self.perform_data is None or not getattr(self.perform_data, "areas", None):
            return recommendations
        matches = self._perform_matches()
        for area in self.perform_data.areas:
            values = {"name": area.name, "max_score": float(area.max_score)}
            for lv in LEVELS_AE:
                candidates = []
                for _, student, level, rec in matches:
                    if level != lv:
                        continue
                    candidates.append((float(student.final_score), float(rec.scores.get(area.name, 0.0))))
                candidates.sort(key=lambda item: item[0])
                if candidates:
                    chosen = candidates[: min(3, len(candidates))]
                    score = sum(score for _, score in chosen) / len(chosen)
                else:
                    score = float(area.max_score) * PERFORM_DEFAULT_RATES[lv]
                values[lv] = max(0.0, min(float(area.max_score), score))
            prev = float(area.max_score)
            for lv in LEVELS_AE:
                values[lv] = min(values[lv], prev)
                prev = values[lv]
            values["memo"] = "현재 자료의 성취수준별 하위 대표 3명 평균"
            recommendations.append(values)
        return recommendations

    def _recommend_perform_recalc_rows(self):
        if not hasattr(self, "table_perform_recalc"):
            return
        rows = self._recommended_perform_scores_by_area()
        if not rows:
            rows = [
                {"name": "평가요소 1", "max_score": 10.0, "memo": "자료 없음 · 성취율 기준"},
                {"name": "평가요소 2", "max_score": 10.0, "memo": "자료 없음 · 성취율 기준"},
                {"name": "평가요소 3", "max_score": 10.0, "memo": "자료 없음 · 성취율 기준"},
            ]
        self._perform_recalc_dirty = False
        table = self.table_perform_recalc
        table.blockSignals(True)
        table.setRowCount(0)
        table.blockSignals(False)
        for row in rows:
            self._add_perform_recalc_row(row)
        self.statusBar().showMessage("수행평가 분할점수 재산정 기본값을 적용했습니다.", 3000)

    def _perform_recalc_rows(self):
        if not hasattr(self, "table_perform_recalc"):
            return [], []
        rows = []
        warnings = []
        table = self.table_perform_recalc
        for r in range(table.rowCount()):
            name_item = table.item(r, 0)
            name = name_item.text().strip() if name_item else f"평가요소 {r + 1}"
            try:
                max_score = float((table.item(r, 1).text() if table.item(r, 1) else "0").replace(",", ""))
            except Exception:
                max_score = 0.0
                warnings.append(f"{name}: 만점 숫자 확인")
            level_scores = {}
            for idx, lv in enumerate(LEVELS_AE, start=2):
                try:
                    score = float((table.item(r, idx).text() if table.item(r, idx) else "0").replace(",", ""))
                except Exception:
                    score = 0.0
                    warnings.append(f"{name}: {lv} 예상점수 숫자 확인")
                if score < 0 or score > max_score:
                    warnings.append(f"{name}: {lv} 예상점수 범위 확인")
                level_scores[lv] = max(0.0, min(max_score, score))
            rows.append({"name": name, "max_score": max_score, "scores": level_scores})
        return rows, warnings

    def _render_perform_recalc_results(self, *changed_items):
        if not hasattr(self, "table_perform_recalc_results"):
            return
        if changed_items:
            self._perform_recalc_dirty = True
        rows, warnings = self._perform_recalc_rows()
        max_total = sum(row["max_score"] for row in rows)
        totals = {
            lv: sum(row["scores"].get(lv, 0.0) for row in rows)
            for lv in LEVELS_AE
        }
        for left, right in zip(LEVELS_AE, LEVELS_AE[1:]):
            if totals[left] + 1e-9 < totals[right]:
                warnings.append(f"{left}/{right} 역전")

        if hasattr(self, "perform_recalc_cards"):
            self.perform_recalc_cards["max"].value_label.setText(f"{max_total:.1f}")
            for lv in LEVELS_AE:
                self.perform_recalc_cards[lv].value_label.setText(f"{totals[lv]:.1f}")
            self.perform_recalc_cards["check"].value_label.setText(str(len(warnings)))

        result_rows = [
            ("A/B", totals["A"], "A 최소능력자 예상점수 합"),
            ("B/C", totals["B"], "B 최소능력자 예상점수 합"),
            ("C/D", totals["C"], "C 최소능력자 예상점수 합"),
            ("D/E", totals["D"], "D 최소능력자 예상점수 합"),
            ("E/미도달", totals["E"], "E 최소능력자 예상점수 합"),
        ]
        table = self.table_perform_recalc_results
        table.setSortingEnabled(False)
        table.setRowCount(len(result_rows) + (1 if warnings else 0))
        for r, (label, raw, meaning) in enumerate(result_rows):
            pct = raw / max_total * 100.0 if max_total > 0 else 0.0
            _set_item(table, r, 0, label)
            _set_item(table, r, 1, f"{raw:.2f}", align_right=True)
            _set_item(table, r, 2, f"{pct:.1f}", align_right=True)
            _set_item(table, r, 3, meaning, align_left=True, tooltip=meaning)
        if warnings:
            warn_text = " · ".join(dict.fromkeys(warnings))
            bg = QColor("#f8dfa3"); bg.setAlpha(150)
            r = len(result_rows)
            _set_item(table, r, 0, "점검", bg=bg, bold=True)
            _set_item(table, r, 1, len(warnings), align_right=True, bg=bg)
            _set_item(table, r, 2, "-", bg=bg)
            _set_item(table, r, 3, warn_text, align_left=True, bg=bg, tooltip=warn_text)
        table.setSortingEnabled(False)
        for col, width in enumerate([92, 90, 100, 360]):
            self._set_scaled_column_width(table, col, width)
        for col, width in enumerate([210, 76, 82, 82, 82, 82, 82, 260]):
            self._set_scaled_column_width(self.table_perform_recalc, col, width)

    # ---- 문항 분석 -----------------------------------------------------
    def _init_tab_items(self):
        layout = QVBoxLayout(self.tab_items)
        self.lbl_alpha = QLabel("지필평가 신뢰도 (Cronbach's alpha): -")
        f = self.lbl_alpha.font(); f.setPointSize(f.pointSize()+3); f.setBold(True); self.lbl_alpha.setFont(f)
        layout.addWidget(self.lbl_alpha)

        sub = QLabel("선다형 문항 분석 결과 (음영: 성취수준별 정답률 2/3 미만)")
        sub.setStyleSheet("color:#888;")
        layout.addWidget(sub)

        # 통합 문항분석 테이블 — 컬럼이 16개라 가로스크롤 허용
        headers = ["문항", "예상난이도", "정답률(%)", "변별도",
                   "1", "2", "3", "4", "5", "무응답",
                   "A", "B", "C", "D", "E", "미도달"]
        self.table_items = QTableWidget(0, len(headers))
        self.table_items.setHorizontalHeaderLabels(headers)
        _setup_table(self.table_items, word_wrap=False, horizontal_scroll=True, row_height=30)
        self.table_items.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        # 정답률(idx 2), 변별도(idx 3) 컬럼에 막대 그리기
        self._item_bar_delegate = ItemBarDelegate(self.table_items)
        self.table_items.setItemDelegateForColumn(2, self._item_bar_delegate)
        self.table_items.setItemDelegateForColumn(3, self._item_bar_delegate)

        # 서답형 분석 표 (별도) — 원본 웹앱과 동일 형식
        sub2 = QLabel("서답형 문항 분석 결과")
        f = sub2.font(); f.setBold(True); f.setPointSize(f.pointSize()+1); sub2.setFont(f)
        layout.addWidget(sub2)
        sd_headers = ["서답형", "최솟값", "최댓값", "평균", "표준편차", "정답률(%)", "변별도",
                      "A", "B", "C", "D", "E", "미도달"]
        self.table_serdap = QTableWidget(0, len(sd_headers))
        self.table_serdap.setHorizontalHeaderLabels(sd_headers)
        _setup_table(self.table_serdap, word_wrap=False, row_height=30)
        self.table_serdap.setMaximumHeight(self._px(110))
        # 정답률, 변별도 컬럼에 막대 표시
        self._serdap_bar_delegate = ItemBarDelegate(self.table_serdap)
        self.table_serdap.setItemDelegateForColumn(5, self._serdap_bar_delegate)
        self.table_serdap.setItemDelegateForColumn(6, self._serdap_bar_delegate)
        layout.addWidget(self.table_serdap)

        # 표와 보조 차트 사이를 splitter로 감싸 비율 조절 가능
        body = QSplitter(Qt.Vertical)
        body.addWidget(self.table_items)
        chart_row = QWidget(); chr_l = QHBoxLayout(chart_row); chr_l.setContentsMargins(0,0,0,0)
        self.canvas_pvalue = CanvasHolder()
        self.canvas_discr = CanvasHolder()
        chr_l.addWidget(self.canvas_pvalue, 1)
        chr_l.addWidget(self.canvas_discr, 1)
        body.addWidget(chart_row)
        body.setStretchFactor(0, 2); body.setStretchFactor(1, 1)
        body.setSizes([400, 280])
        layout.addWidget(body, 1)

    # ---- 답지반응분포 ---------------------------------------------------
    def _init_tab_choice(self):
        layout = QVBoxLayout(self.tab_choice)
        top = QHBoxLayout()
        top.addWidget(QLabel("답지반응분석 문항:"))
        self.combo_item = QComboBox()
        self.combo_item.currentIndexChanged.connect(self._refresh_choice)
        top.addWidget(self.combo_item, 1)
        layout.addLayout(top)

        self.lbl_choice_info = QLabel("")
        self.lbl_choice_info.setWordWrap(True)
        self.lbl_choice_info.setProperty("role", "muted")
        layout.addWidget(self.lbl_choice_info)

        # 표↔차트 사이 splitter (사용자가 비율 조절)
        choice_split = QSplitter(Qt.Vertical)
        # 매트릭스 표
        self.table_choice = QTableWidget(0, 8)
        self.table_choice.setHorizontalHeaderLabels(
            ["성취수준", "정답률(%)", "1", "2", "3", "4", "5", "무응답"])
        _setup_table(self.table_choice)
        choice_split.addWidget(self.table_choice)
        # 답지반응분포 차트
        self.canvas_choice = CanvasHolder()
        choice_split.addWidget(self.canvas_choice)
        choice_split.setStretchFactor(0, 1); choice_split.setStretchFactor(1, 2)
        choice_split.setSizes([240, 360])
        layout.addWidget(choice_split, 1)

    # ---- 성취기준 분석 --------------------------------------------------
    def _init_tab_standard(self):
        layout = QVBoxLayout(self.tab_standard)
        sub = QLabel("성취기준별 평균 정답률 (음영: 성취율 50% 이하 · 셀 안 막대 = 정답률)")
        sub.setProperty("role", "muted"); layout.addWidget(sub)
        headers = ["성취기준", "문항수", "전체(%)", "A", "B", "C", "D", "E", "미도달"]
        self.table_std = QTableWidget(0, len(headers))
        self.table_std.setHorizontalHeaderLabels(headers)
        # 첫 컬럼은 Stretch, 나머지는 Fixed
        _setup_table(self.table_std, word_wrap=False, row_height=30)
        self.table_std.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        hh = self.table_std.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.Stretch)
        for i in range(1, len(headers)):
            hh.setSectionResizeMode(i, QHeaderView.Fixed)
            self._set_scaled_column_width(self.table_std, i, 72)
        self.table_std.setSortingEnabled(True)
        # 전체 성취율 컬럼에 막대 그리기
        self._std_bar_delegate = ItemBarDelegate(self.table_std)
        self.table_std.setItemDelegateForColumn(2, self._std_bar_delegate)
        # 표/차트를 splitter로 묶어서 사용자가 비율 조절 + 차트는 ScrollArea로 잘림 방지
        std_split = QSplitter(Qt.Vertical)
        std_split.addWidget(self.table_std)
        self.canvas_std = CanvasHolder()
        std_chart_scroll = QScrollArea(); std_chart_scroll.setWidgetResizable(True)
        std_chart_scroll.setWidget(self.canvas_std)
        std_split.addWidget(std_chart_scroll)
        std_split.setStretchFactor(0, 2); std_split.setStretchFactor(1, 3)
        std_split.setSizes([280, 420])
        layout.addWidget(std_split, 1)

    # ---- AI 문항 검토 --------------------------------------------------
    def _init_tab_ai_review(self):
        layout = QVBoxLayout(self.tab_ai_review)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(8)

        head = QHBoxLayout()
        head.addStretch(1)
        self.btn_ai_load_source = QPushButton("문항 자료")
        self.btn_ai_load_source.setToolTip(
            "시험 문제 HWPX/PDF, 문항정보표, 수행평가 채점기준표를 문항 자료 칸으로 불러옵니다."
        )
        self.btn_ai_load_source.clicked.connect(self._load_ai_review_file)
        head.addWidget(self.btn_ai_load_source)
        btn_reference = QPushButton("성취기준·수준 자료")
        btn_reference.setToolTip("성취기준, 성취수준, 최소능력자 설명 자료를 참고자료 칸에 불러옵니다.")
        btn_reference.clicked.connect(self._load_ai_reference_file)
        head.addWidget(btn_reference)
        btn_written_example = QPushButton("지필 예시")
        btn_written_example.setToolTip("지필평가 문항 예시를 원문 칸에 넣어 흐름을 테스트합니다.")
        btn_written_example.clicked.connect(lambda: self._insert_ai_review_example("written"))
        head.addWidget(btn_written_example)
        btn_perform_example = QPushButton("수행 예시")
        btn_perform_example.setToolTip("수행평가 채점기준표 예시를 원문 칸에 넣어 흐름을 테스트합니다.")
        btn_perform_example.clicked.connect(lambda: self._insert_ai_review_example("perform"))
        head.addWidget(btn_perform_example)
        btn_exam = QPushButton("현재 문항정보표")
        btn_exam.clicked.connect(self._load_ai_review_from_exam)
        head.addWidget(btn_exam)
        self.btn_ai_generate = QPushButton("검토 초안 생성")
        self.btn_ai_generate.setProperty("role", "primary")
        self.btn_ai_generate.clicked.connect(self._generate_ai_review_draft)
        head.addWidget(self.btn_ai_generate)
        self.btn_ai_enrich = QPushButton("AI로 보강")
        self.btn_ai_enrich.setToolTip("AI 설정에 지정한 로컬/클라우드 모델로 검토 초안을 보강합니다.")
        self.btn_ai_enrich.clicked.connect(self._run_ai_review_completion)
        head.addWidget(self.btn_ai_enrich)
        self.btn_ai_review_stop = QPushButton("보강 중지")
        self.btn_ai_review_stop.setToolTip("현재 응답 중인 묶음이 끝난 뒤 AI 보강을 멈춥니다.")
        self.btn_ai_review_stop.clicked.connect(self._request_ai_review_stop)
        self.btn_ai_review_stop.setEnabled(False)
        head.addWidget(self.btn_ai_review_stop)
        self.btn_ai_review_resume = QPushButton("이어하기")
        self.btn_ai_review_resume.setToolTip("중지된 AI 보강을 다음 묶음부터 이어서 진행합니다.")
        self.btn_ai_review_resume.clicked.connect(self._resume_ai_review_completion)
        self.btn_ai_review_resume.setEnabled(False)
        head.addWidget(self.btn_ai_review_resume)
        btn_to_spliter = QPushButton("지필→예상정답률")
        btn_to_spliter.clicked.connect(self._send_ai_review_to_spliter)
        head.addWidget(btn_to_spliter)
        btn_to_perform = QPushButton("수행→재산정")
        btn_to_perform.clicked.connect(self._send_ai_review_to_perform_recalc)
        head.addWidget(btn_to_perform)
        btn_export = QPushButton("CSV 내보내기…")
        btn_export.clicked.connect(self._export_ai_review_csv)
        head.addWidget(btn_export)
        layout.addLayout(head)

        card_grid = QGridLayout(); card_grid.setSpacing(8)
        self.ai_review_cards = {
            "rows": self._kpi_card("검토 항목", "-"),
            "standards": self._kpi_card("성취기준", "-"),
            "warnings": self._kpi_card("점검", "-"),
            "mode": self._kpi_card("AI 상태", "로컬 초안"),
        }
        for i, card in enumerate(self.ai_review_cards.values()):
            card_grid.addWidget(card, 0, i)
        layout.addLayout(card_grid)

        split = QSplitter(Qt.Horizontal)
        left = QWidget()
        left_layout = QVBoxLayout(left); left_layout.setContentsMargins(0, 0, 0, 0); left_layout.setSpacing(6)
        left_title = QLabel("입력 자료")
        left_title.setProperty("role", "title")
        left_layout.addWidget(left_title)

        self.ai_review_input_tabs = QTabWidget()
        source_panel = QWidget()
        source_layout = QVBoxLayout(source_panel); source_layout.setContentsMargins(0, 0, 0, 0)
        source_tools = QHBoxLayout()
        self.lbl_ai_source_store = QLabel()
        self.lbl_ai_source_store.setProperty("role", "muted")
        self.lbl_ai_source_store.setWordWrap(False)
        source_tools.addWidget(self.lbl_ai_source_store, 1)
        btn_source_store_location = QPushButton("위치")
        btn_source_store_location.setToolTip("문항 자료가 저장되는 앱 자료 폴더 위치를 확인합니다.")
        btn_source_store_location.clicked.connect(lambda: self._show_ai_store_location("source"))
        source_tools.addWidget(btn_source_store_location)
        btn_load_saved_source = QPushButton("저장자료 불러오기")
        btn_load_saved_source.setToolTip("앱 자료 폴더에 저장된 문항 자료를 선택해 다시 불러옵니다.")
        btn_load_saved_source.clicked.connect(self._load_saved_ai_source_text)
        source_tools.addWidget(btn_load_saved_source)
        btn_save_source_text = QPushButton("현재 내용 저장")
        btn_save_source_text.setToolTip("현재 문항 자료 칸의 내용을 앱 자료 폴더에 저장합니다.")
        btn_save_source_text.clicked.connect(lambda: self._save_current_ai_text("source"))
        source_tools.addWidget(btn_save_source_text)
        source_layout.addLayout(source_tools)
        self.ai_source_view_tabs = QTabWidget()
        self.txt_ai_review_source = QPlainTextEdit()
        self.txt_ai_review_source.setPlaceholderText(
            "시험 문제 HWPX/PDF에서 추출한 문항, 문항정보표, 수행평가 채점기준표를 붙여 넣거나 '문항 자료'로 불러오세요."
        )
        self.ai_source_view_tabs.addTab(self.txt_ai_review_source, "편집")
        self.browser_ai_review_source = QTextBrowser()
        self.browser_ai_review_source.setReadOnly(True)
        self.browser_ai_review_source.setOpenExternalLinks(False)
        self._style_ai_markdown_browser(self.browser_ai_review_source)
        self.ai_source_view_tabs.addTab(self.browser_ai_review_source, "미리보기")
        self.txt_ai_review_source.textChanged.connect(
            lambda: self._update_ai_markdown_preview("source")
        )
        source_layout.addWidget(self.ai_source_view_tabs, 1)
        self.ai_review_input_tabs.addTab(source_panel, "문항 자료")

        reference_panel = QWidget()
        reference_layout = QVBoxLayout(reference_panel); reference_layout.setContentsMargins(0, 0, 0, 0)
        reference_tools = QHBoxLayout()
        self.lbl_ai_reference_store = QLabel()
        self.lbl_ai_reference_store.setProperty("role", "muted")
        self.lbl_ai_reference_store.setWordWrap(False)
        reference_tools.addWidget(self.lbl_ai_reference_store, 1)
        btn_reference_store_location = QPushButton("위치")
        btn_reference_store_location.setToolTip("성취기준·수준 자료가 저장되는 앱 자료 폴더 위치를 확인합니다.")
        btn_reference_store_location.clicked.connect(lambda: self._show_ai_store_location("reference"))
        reference_tools.addWidget(btn_reference_store_location)
        btn_load_saved_reference = QPushButton("저장자료 불러오기")
        btn_load_saved_reference.setToolTip("앱 자료 폴더에 저장된 성취기준·수준 자료를 선택해 다시 불러옵니다.")
        btn_load_saved_reference.clicked.connect(self._load_saved_ai_reference_text)
        reference_tools.addWidget(btn_load_saved_reference)
        btn_save_reference_text = QPushButton("현재 내용 저장")
        btn_save_reference_text.setToolTip("현재 성취기준·수준 칸의 내용을 앱 자료 폴더에 저장합니다.")
        btn_save_reference_text.clicked.connect(lambda: self._save_current_ai_text("reference"))
        reference_tools.addWidget(btn_save_reference_text)
        reference_layout.addLayout(reference_tools)
        self.txt_ai_review_reference = QPlainTextEdit()
        self.txt_ai_review_reference.setPlaceholderText(
            "성취기준, 성취수준 A~E 설명, 최소능력자 특성, 평가기준 자료를 붙여 넣거나 '성취기준·수준 자료'로 불러오세요."
        )
        self.ai_reference_view_tabs = QTabWidget()
        self.ai_reference_view_tabs.addTab(self.txt_ai_review_reference, "편집")
        self.browser_ai_review_reference = QTextBrowser()
        self.browser_ai_review_reference.setReadOnly(True)
        self.browser_ai_review_reference.setOpenExternalLinks(False)
        self._style_ai_markdown_browser(self.browser_ai_review_reference)
        self.ai_reference_view_tabs.addTab(self.browser_ai_review_reference, "미리보기")
        self.txt_ai_review_reference.textChanged.connect(
            lambda: self._update_ai_markdown_preview("reference")
        )
        reference_layout.addWidget(self.ai_reference_view_tabs, 1)
        self.ai_review_input_tabs.addTab(reference_panel, "성취기준·수준")
        left_layout.addWidget(self.ai_review_input_tabs, 1)
        split.addWidget(left)

        right = QWidget()
        right_layout = QVBoxLayout(right); right_layout.setContentsMargins(0, 0, 0, 0); right_layout.setSpacing(6)
        self.ai_review_tabs = QTabWidget()
        self.table_ai_review = QTableWidget(0, len(AI_REVIEW_HEADERS))
        self.table_ai_review.setHorizontalHeaderLabels(AI_REVIEW_HEADERS)
        _setup_table(self.table_ai_review, word_wrap=False, horizontal_scroll=True, row_height=32)
        self.table_ai_review.setEditTriggers(
            QAbstractItemView.DoubleClicked
            | QAbstractItemView.EditKeyPressed
            | QAbstractItemView.AnyKeyPressed
        )
        self.ai_review_tabs.addTab(self.table_ai_review, "검토 초안")
        self.txt_ai_review_prompt = QPlainTextEdit()
        self.txt_ai_review_prompt.setReadOnly(True)
        self.txt_ai_review_prompt.setPlaceholderText("검토 초안을 만들면 AI 연결용 프롬프트가 여기에 생성됩니다.")
        self.ai_review_tabs.addTab(self.txt_ai_review_prompt, "AI 프롬프트")
        self.txt_ai_progress = QPlainTextEdit()
        self.txt_ai_progress.setReadOnly(True)
        self.txt_ai_progress.setPlaceholderText("AI 연결, 서버 탐색, 보강 요청의 진행 상황이 여기에 표시됩니다.")
        self.ai_review_tabs.addTab(self.txt_ai_progress, "진행 로그")
        self.ai_review_tabs.addTab(self._build_ai_usage_panel(), "사용 예시")
        self.ai_review_tabs.addTab(self._build_ai_settings_panel(), "AI 설정")
        self.ai_review_tabs.currentChanged.connect(self._on_ai_review_subtab_changed)
        right_layout.addWidget(self.ai_review_tabs, 1)
        split.addWidget(right)
        split.setStretchFactor(0, 1)
        split.setStretchFactor(1, 2)
        split.setSizes([420, 780])
        layout.addWidget(split, 1)
        self.lbl_ai_review_footer = QLabel(
            "AI 문항 검토는 교사 검토를 돕는 초안입니다. 분석 실행 자료가 있으면 이전 시험의 성취수준별 정답률을 목표수준 후보와 A~E 예상값에 먼저 반영합니다."
        )
        self.lbl_ai_review_footer.setProperty("role", "muted")
        self.lbl_ai_review_footer.setWordWrap(True)
        layout.addWidget(self.lbl_ai_review_footer, 0)
        self._refresh_ai_source_store_label()
        self._refresh_ai_reference_store_label()
        self._set_ai_review_summary(0, 0, 0)

    @staticmethod
    def _style_ai_markdown_browser(browser: QTextBrowser):
        browser.document().setDefaultStyleSheet("""
            body {
                line-height: 1.55;
                font-size: 13px;
            }
            h1 {
                font-size: 20px;
                margin: 8px 0 10px;
            }
            h2 {
                font-size: 16px;
                margin: 14px 0 8px;
            }
            h3 {
                font-size: 14px;
                margin: 12px 0 6px;
            }
            p {
                margin: 5px 0;
            }
            ul, ol {
                margin-top: 4px;
                margin-bottom: 8px;
            }
            li {
                margin: 3px 0;
            }
            code {
                padding: 1px 4px;
            }
            pre {
                padding: 8px;
                white-space: pre-wrap;
            }
            table {
                border-collapse: collapse;
            }
            th, td {
                padding: 4px 6px;
                border: 1px solid #789;
            }
        """)

    def _update_ai_markdown_preview(self, target: str):
        if target == "reference":
            if not hasattr(self, "browser_ai_review_reference"):
                return
            text = self.txt_ai_review_reference.toPlainText()
            browser = self.browser_ai_review_reference
        else:
            if not hasattr(self, "browser_ai_review_source"):
                return
            text = self.txt_ai_review_source.toPlainText()
            browser = self.browser_ai_review_source
        try:
            browser.setMarkdown(text)
        except Exception:
            browser.setPlainText(text)

    def _build_ai_usage_panel(self) -> QWidget:
        panel = QWidget()
        layout = QVBoxLayout(panel); layout.setContentsMargins(10, 10, 10, 10)
        browser = QTextBrowser()
        browser.setReadOnly(True)
        browser.setOpenExternalLinks(False)
        browser.setHtml("""
        <style>
          body { line-height: 1.55; }
          h2 { margin: 0 0 8px; font-size: 18px; }
          h3 { margin: 16px 0 6px; font-size: 15px; }
          p { margin: 5px 0; }
          ol, ul { margin-top: 5px; }
          li { margin: 4px 0; }
          pre { padding: 10px; border: 1px solid #789; border-radius: 6px; white-space: pre-wrap; }
          code { padding: 1px 4px; border-radius: 4px; }
        </style>
        <h2>AI 문항 검토 사용 흐름</h2>
        <p><b>목적</b>: AI가 최종 판단을 대신하는 것이 아니라, 교사가 검토할 표를 먼저 채워 분할점수 산정의 출발점을 빠르게 만드는 것입니다.</p>
        <ol>
          <li><b>문항 자료</b> 칸에 시험 문제 PDF에서 추출한 문항, 문항정보표, 수행평가 채점기준표를 넣습니다.</li>
          <li><b>성취기준·수준</b> 칸에 성취기준, A~E 성취수준 설명, 최소능력자 특성 자료를 넣습니다. 여러 PDF를 한 번에 선택할 수 있고, 불러온 원본은 앱 자료 폴더에 저장됩니다.</li>
          <li>성취기준·수준 PDF는 원문 전체가 아니라 성취기준 코드와 A~E 성취수준 설명 중심으로 정리되어 표시됩니다.</li>
          <li><b>검토 초안 생성</b>을 누르면 앱 내부 규칙으로 문항을 먼저 나누고, 참고자료와 대조해 성취기준, 평가유형, 목표수준, 난이도, A~E 예상값을 만듭니다. 이미 분석을 실행한 이전 시험 자료가 있으면 성취수준별 정답률을 우선 근거로 씁니다.</li>
          <li>로컬 AI를 연결했다면 <b>AI로 보강</b>을 눌러 근거와 판단을 더 정교하게 보강합니다.</li>
          <li>지필 문항은 <b>지필→예상정답률</b>, 수행평가는 <b>수행→재산정</b>으로 보냅니다.</li>
          <li>교사가 문항을 직접 보며 A~E 예상값을 수정하고 최종 분할점수를 확인합니다.</li>
        </ol>

        <h3>지필평가 예시 해석</h3>
        <p><code>C 예상 2/3</code>은 C 수준 최소능력자 3명 중 2명 정도가 맞힐 것 같다는 뜻입니다. 이전 시험의 유사 문항에서 C 수준 학생 정답률이 약 2/3 이상이면 목표수준 후보가 C 쪽으로 잡힙니다.</p>
        <p><code>A 수준 문항</code>은 A 학생만 맞히는 문항이라는 뜻이 아니라, A 수준 최소능력자 3명 중 약 2명이 해결할 수 있는 문항이라는 뜻으로 봅니다.</p>
        <p>AI 보강은 문항 자료만 보지 않고, 성취기준·수준 자료를 기준표로 삼아 어떤 성취기준과 성취수준에 가까운지 다시 판단합니다.</p>

        <h3>수행평가 예시 해석</h3>
        <p>수행평가는 O/X가 아니라 평가요소별 예상점수로 봅니다. 예를 들어 <code>B 예상 7.5점</code>은 B 수준 최소능력자가 해당 평가요소에서 대략 7.5점을 받을 것으로 본다는 뜻입니다.</p>

        <h3>Ollama 로컬 연결 예시</h3>
        <pre>ollama --version
ollama pull gemma4:e4b
ollama serve</pre>
        <p>그 뒤 <b>AI 설정</b>에서 다음처럼 입력합니다.</p>
        <pre>AI 제공자: Ollama 로컬
엔드포인트: http://127.0.0.1:11434/api/chat
모델: gemma4:e4b</pre>
        <p>Ollama 0.30.x 이상에서는 Apple Silicon 환경에서 지원 모델이 MLX 엔진 최적화를 내부적으로 사용할 수 있습니다. Goedu-Split은 별도 MLX 서버나 MLX 입력칸을 요구하지 않고 Ollama API만 호출합니다.</p>
        <p>모델이 오래 로딩되거나 응답하지 않으면 <code>ollama ps</code>로 실행 상태를 확인하고, 더 작은 채팅 모델로 연결 테스트를 먼저 통과시키세요.</p>
        <h3>Codex CLI OAuth 처음 설치</h3>
        <p>Windows 선생님은 터미널을 관리자 권한이 아닌 일반 권한으로 열고 아래 순서대로 입력합니다.</p>
        <pre>winget install OpenJS.NodeJS.LTS
npm install -g @openai/codex
where codex
codex --version
codex login
codex login status</pre>
        <p>Node.js 설치 직후 <code>npm</code>을 못 찾으면 터미널을 닫았다가 다시 열어 주세요. macOS는 <code>brew install codex</code> 또는 <code>npm install -g @openai/codex</code>를 사용할 수 있습니다.</p>
        <p><b>Codex CLI 클라우드 (OAuth)</b>는 API Key를 입력하지 않습니다. 로컬 <code>codex</code> CLI의 ChatGPT OAuth 로그인 상태로만 실행합니다.</p>
        <p>터미널에서는 되는데 앱에서 못 찾는 경우가 있어, Windows에서는 <code>codex.cmd</code>와 npm/Node 경로를, macOS에서는 <code>/opt/homebrew/bin/codex</code>와 <code>/usr/local/bin/codex</code>도 직접 확인합니다.</p>
        """)
        layout.addWidget(browser, 1)
        return panel

    def _build_ai_settings_panel(self) -> QWidget:
        panel = QWidget()
        layout = QVBoxLayout(panel); layout.setContentsMargins(10, 10, 10, 10); layout.setSpacing(10)

        note = QLabel("AI 보강을 누를 때만 선택한 제공자를 호출합니다.")
        note.setProperty("role", "muted")
        note.setWordWrap(True)
        layout.addWidget(note)

        form_card = QFrame()
        form_card.setProperty("role", "card")
        form = QFormLayout(form_card)
        form.setContentsMargins(14, 12, 14, 12)
        form.setSpacing(8)
        form.setFieldGrowthPolicy(QFormLayout.ExpandingFieldsGrow)

        self.cmb_ai_provider = QComboBox()
        self.cmb_ai_provider.addItem("로컬 초안만 사용", "local_draft")
        self.cmb_ai_provider.addItem("Ollama 로컬", "ollama")
        self.cmb_ai_provider.addItem("Codex CLI 클라우드 (OAuth)", "codex_cli")
        self.cmb_ai_provider.currentIndexChanged.connect(self._sync_ai_provider_defaults)
        form.addRow("AI 제공자", self.cmb_ai_provider)

        self.lbl_ai_endpoint_field = QLabel("엔드포인트")
        self.edit_ai_endpoint = QLineEdit()
        self.edit_ai_endpoint.setPlaceholderText("예: http://127.0.0.1:11434/api/chat")
        form.addRow(self.lbl_ai_endpoint_field, self.edit_ai_endpoint)

        self.lbl_ai_model_field = QLabel("모델")
        self.cmb_ai_model = QComboBox()
        self.cmb_ai_model.setEditable(True)
        self.cmb_ai_model.setInsertPolicy(QComboBox.NoInsert)
        self.cmb_ai_model.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.cmb_ai_model.lineEdit().setPlaceholderText("예: gemma4:e4b, qwen3.6:latest")
        self.edit_ai_model = self.cmb_ai_model.lineEdit()
        form.addRow(self.lbl_ai_model_field, self.cmb_ai_model)

        self.lbl_ai_model_hint_field = QLabel("추천")
        self.lbl_ai_model_hint = QLabel("모델 새로고침을 누르면 감지된 Ollama 모델과 추천값이 표시됩니다.")
        self.lbl_ai_model_hint.setProperty("role", "muted")
        self.lbl_ai_model_hint.setWordWrap(True)
        form.addRow(self.lbl_ai_model_hint_field, self.lbl_ai_model_hint)

        self.edit_ai_api_key = QLineEdit()
        self.edit_ai_api_key.setEchoMode(QLineEdit.Password)
        self.edit_ai_api_key.hide()

        self.lbl_ai_timeout_field = QLabel("대기 시간")
        self.spin_ai_timeout = QDoubleSpinBox()
        self.spin_ai_timeout.setRange(5, 600)
        self.spin_ai_timeout.setDecimals(0)
        self.spin_ai_timeout.setSingleStep(15)
        self.spin_ai_timeout.setSuffix(" 초")
        form.addRow(self.lbl_ai_timeout_field, self.spin_ai_timeout)

        self.lbl_ai_privacy_field = QLabel("개인정보")
        self.chk_ai_scrub = QCheckBox("클라우드/외부 서버로 보낼 때 학생 이름·반번호를 가능한 한 제거")
        form.addRow(self.lbl_ai_privacy_field, self.chk_ai_scrub)

        layout.addWidget(form_card)

        guide = QLabel("자세한 연결 방법은 '사용 예시' 탭에서 확인하세요.")
        guide.setProperty("role", "muted")
        guide.setWordWrap(True)
        layout.addWidget(guide)

        row = QHBoxLayout()
        btn_save = QPushButton("설정 저장")
        btn_save.setProperty("role", "primary")
        btn_save.clicked.connect(self._save_ai_settings)
        row.addWidget(btn_save)
        self.btn_ai_test = QPushButton("연결 테스트")
        self.btn_ai_test.clicked.connect(self._test_ai_connection)
        row.addWidget(self.btn_ai_test)
        self.btn_refresh_ai_models = QPushButton("모델 새로고침")
        self.btn_refresh_ai_models.setToolTip("현재 AI 제공자의 모델 목록을 다시 읽고 추천 모델을 선택합니다.")
        self.btn_refresh_ai_models.clicked.connect(self._refresh_ai_model_list)
        row.addWidget(self.btn_refresh_ai_models)
        self.btn_ai_ollama_mlx_help = QPushButton("AI 연결 안내")
        self.btn_ai_ollama_mlx_help.setToolTip("Ollama/MLX와 Codex CLI OAuth 연결 방법을 다시 봅니다.")
        self.btn_ai_ollama_mlx_help.clicked.connect(lambda: self._show_ai_ollama_mlx_help_popup(force=True))
        row.addWidget(self.btn_ai_ollama_mlx_help)
        self.btn_find_mlx = QPushButton("MLX 찾기")
        self.btn_find_mlx.setToolTip("실행 중인 MLX-LM/LM Studio 로컬 서버를 자동으로 찾아 설정합니다.")
        self.btn_find_mlx.clicked.connect(self._find_local_mlx_server)
        row.addWidget(self.btn_find_mlx)
        self.btn_start_mlx = QPushButton("MLX 서버 시작")
        self.btn_start_mlx.setToolTip("로컬에 설치된 mlx_lm.server를 실행합니다. 모델이 없으면 다운로드가 필요할 수 있습니다.")
        self.btn_start_mlx.clicked.connect(self._start_local_mlx_server)
        row.addWidget(self.btn_start_mlx)
        row.addStretch(1)
        layout.addLayout(row)

        self.lbl_ai_settings_status = QLabel("AI 설정을 저장하면 다음 실행 때도 유지됩니다.")
        self.lbl_ai_settings_status.setProperty("role", "muted")
        self.lbl_ai_settings_status.setWordWrap(True)
        layout.addWidget(self.lbl_ai_settings_status)
        layout.addStretch(1)
        self._load_ai_settings()
        return panel

    def _provider_from_combo(self) -> str:
        if not hasattr(self, "cmb_ai_provider"):
            return "local_draft"
        data = self.cmb_ai_provider.currentData()
        return str(data or "local_draft")

    @staticmethod
    def _settings_truthy(value) -> bool:
        if isinstance(value, bool):
            return value
        return str(value).strip().lower() in {"1", "true", "yes", "on"}

    @staticmethod
    def _version_tuple(version: str) -> tuple[int, int, int]:
        parts = [int(part) for part in re.findall(r"\d+", version or "")[:3]]
        while len(parts) < 3:
            parts.append(0)
        return tuple(parts[:3])

    def _on_ai_review_subtab_changed(self, _index: int):
        if not hasattr(self, "ai_review_tabs"):
            return
        if self.ai_review_tabs.tabText(self.ai_review_tabs.currentIndex()) == "AI 설정":
            self._schedule_ai_ollama_mlx_help_popup()

    def _schedule_ai_ollama_mlx_help_popup(self):
        if self._settings_truthy(self.settings.value("ai/connection_help_seen_v3", False)):
            return
        if getattr(self, "_ai_ollama_mlx_help_pending", False):
            return
        self._ai_ollama_mlx_help_pending = True

        def show_later():
            self._ai_ollama_mlx_help_pending = False
            if hasattr(self, "tabs") and hasattr(self, "tab_ai_review") and self.tabs.currentWidget() is not self.tab_ai_review:
                return
            self._show_ai_ollama_mlx_help_popup(force=False)

        QTimer.singleShot(220, show_later)

    @staticmethod
    def _ollama_mlx_help_html() -> str:
        return """
        <style>
          body { line-height: 1.55; font-size: 13px; }
          h2 { margin: 0 0 10px; font-size: 18px; }
          h3 { margin: 14px 0 6px; font-size: 15px; }
          p { margin: 5px 0; }
          ol, ul { margin-top: 5px; margin-bottom: 8px; }
          li { margin: 4px 0; }
          code { padding: 1px 4px; border-radius: 4px; }
          pre { padding: 10px; border: 1px solid #789; border-radius: 6px; white-space: pre-wrap; }
        </style>
        <h2>AI 연결 안내</h2>
        <p><b>핵심</b>: 기본값은 외부 전송이 없는 로컬 초안입니다. AI 보강을 쓸 때만 <b>Ollama 로컬</b> 또는 <b>Codex CLI 클라우드 (OAuth)</b>를 선택합니다.</p>

        <h3>Ollama와 MLX</h3>
        <p>Goedu-Split에서 MLX를 따로 설정하지 않습니다. Ollama 0.30.x 이상이 Apple Silicon에서 지원 모델을 실행할 때 MLX engine 최적화를 내부적으로 사용할 수 있습니다.</p>
        <p>앱에서는 <b>Ollama 로컬</b>만 선택하고, 실제 모델 사용 가능 여부는 <b>모델 새로고침</b>과 <b>연결 테스트</b>로 확인합니다.</p>

        <h3>처음 설정</h3>
        <ol>
          <li>터미널에서 <code>ollama --version</code>이 0.30.x 이상인지 확인합니다. server가 낮게 나오고 client만 0.30.x라면 Ollama 앱이나 <code>ollama serve</code>를 완전히 재시작합니다.</li>
          <li><code>ollama serve</code>를 실행합니다. 이미 Ollama 앱이 떠 있으면 생략할 수 있습니다.</li>
          <li><code>ollama list</code>로 채팅 모델이 있는지 확인합니다. embedding 전용 모델은 AI 검토에 쓰지 않습니다.</li>
          <li>모델이 없으면 <code>ollama pull gemma4:e4b</code>처럼 작은 채팅 모델부터 준비합니다.</li>
          <li>Goedu-Split에서 <b>AI 검토 -> AI 설정 -> Ollama 로컬 -> 모델 새로고침 -> 연결 테스트</b> 순서로 확인합니다.</li>
        </ol>

        <h3>Ollama가 느리거나 실패할 때</h3>
        <ul>
          <li><code>ollama ps</code>로 모델이 로딩 중인지 확인합니다.</li>
          <li><code>ollama --version</code>에서 server/client 버전이 다르면 업데이트된 Ollama server가 아직 떠 있지 않은 상태일 수 있습니다.</li>
          <li>큰 모델이 계속 timeout이면 더 작은 채팅 모델을 먼저 테스트합니다.</li>
          <li>MLX 사용 여부는 Goedu-Split이 강제로 켜는 스위치가 아니라 Ollama 런타임과 모델 지원 범위에 따라 결정됩니다.</li>
        </ul>

        <h3>Codex CLI 클라우드 OAuth</h3>
        <p>Codex 클라우드는 API Key를 입력하지 않습니다. PC에 설치된 <code>codex</code> CLI의 ChatGPT OAuth 로그인만 사용합니다.</p>
        <p><b>Windows 처음 설치</b>: 시작 메뉴에서 <b>터미널</b>을 열고 아래 명령을 한 줄씩 입력합니다.</p>
        <pre>winget install OpenJS.NodeJS.LTS
npm install -g @openai/codex
where codex
codex --version
codex login
codex login status</pre>
        <ol>
          <li><code>winget</code>으로 Node.js LTS를 설치합니다. 이미 설치되어 있으면 건너뛰어도 됩니다.</li>
          <li>설치 직후 <code>npm</code>을 못 찾으면 터미널을 완전히 닫았다가 다시 엽니다.</li>
          <li><code>npm install -g @openai/codex</code>로 Codex CLI를 설치하거나 갱신합니다.</li>
          <li><code>where codex</code>와 <code>codex --version</code>으로 설치 위치와 실행 여부를 확인합니다.</li>
          <li><code>codex login</code>을 실행해 브라우저에서 ChatGPT 계정으로 로그인합니다.</li>
          <li><code>codex login status</code>가 <code>Logged in using ChatGPT</code>인지 확인합니다.</li>
          <li>Goedu-Split에서 <b>AI 검토 -> AI 설정 -> Codex CLI 클라우드 (OAuth) -> 연결 테스트</b>를 누릅니다.</li>
        </ol>
        <p><b>macOS</b>는 <code>brew install codex</code> 또는 <code>npm install -g @openai/codex</code> 후 <code>codex login</code>을 실행합니다.</p>
        <p>터미널에서는 되는데 앱에서 못 찾는 경우를 줄이기 위해, Goedu-Split은 Windows의 <code>codex.cmd</code>, npm/Node 경로, <code>~/.codex/config.toml</code>의 <code>CODEX_CLI_PATH</code>, macOS Homebrew 경로를 함께 확인합니다. 그래도 실패하면 <code>CODEX_CLI_PATH</code>에 codex 실행 파일 전체 경로를 지정할 수 있습니다.</p>
        <p>Codex provider는 파일을 고치지 않고 제공된 문항 자료, 성취기준·수준 자료, 이전시험 근거를 바탕으로 JSON 검토표만 반환하도록 호출됩니다.</p>

        <h3>공식 근거</h3>
        <p>Ollama 0.30은 Apple Silicon의 MLX engine을 보강했고, 0.30.10은 일부 모델군이 Apple Silicon에서 MLX engine으로 실행된다고 공지했습니다. Codex provider는 로컬 Codex CLI OAuth 세션을 사용합니다.</p>
        <p><a href="https://github.com/ollama/ollama/releases/tag/v0.30.0">Ollama v0.30.0 release</a> · <a href="https://github.com/ollama/ollama/releases/tag/v0.30.10">Ollama v0.30.10 release</a></p>
        """

    def _show_ai_ollama_mlx_help_popup(self, force: bool = False):
        if getattr(self, "_ai_ollama_mlx_help_open", False):
            return
        if not force and self._settings_truthy(self.settings.value("ai/connection_help_seen_v3", False)):
            return
        if not force:
            self.settings.setValue("ai/connection_help_seen_v3", True)

        self._ai_ollama_mlx_help_open = True
        dialog = QDialog(self)
        dialog.setWindowTitle("AI 연결 안내")
        dialog.resize(680, 640)
        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(14, 14, 14, 14)
        layout.setSpacing(10)

        browser = QTextBrowser(dialog)
        browser.setReadOnly(True)
        browser.setOpenExternalLinks(True)
        browser.setHtml(self._ollama_mlx_help_html())
        layout.addWidget(browser, 1)

        foot = QLabel("이 안내는 처음 AI 검토 탭에 들어올 때 한 번만 표시됩니다. 다시 보려면 AI 설정의 'AI 연결 안내'를 누르세요.")
        foot.setProperty("role", "muted")
        foot.setWordWrap(True)
        layout.addWidget(foot)

        row = QHBoxLayout()
        row.addStretch(1)
        close = QPushButton("확인")
        close.setDefault(True)
        close.clicked.connect(dialog.accept)
        row.addWidget(close)
        layout.addLayout(row)

        dialog.finished.connect(lambda _result: setattr(self, "_ai_ollama_mlx_help_open", False))
        dialog.exec()

    def _provider_setting_value(self, provider: str, key: str, default: str = "") -> str:
        provider_key = f"ai/providers/{provider}/{key}"
        value = self.settings.value(provider_key, None)
        if value is None and str(self.settings.value("ai/provider", "")) == provider:
            value = self.settings.value(f"ai/{key}", default)
        if value is None:
            value = default
        return str(value)

    def _store_current_ai_provider_settings(self, provider: str | None = None):
        if not hasattr(self, "cmb_ai_provider"):
            return
        provider = provider or self._provider_from_combo()
        endpoint = normalize_endpoint(provider, self.edit_ai_endpoint.text())
        model = self._ai_model_text() or default_model(provider)
        model = self._sanitize_ai_model_for_provider(provider, model)
        self.settings.setValue(f"ai/providers/{provider}/endpoint", endpoint)
        self.settings.setValue(f"ai/providers/{provider}/model", model)
        self.settings.setValue(f"ai/providers/{provider}/api_key", "")

    def _provider_model_note(self, provider: str) -> str:
        if provider == "local_draft":
            return "로컬 초안은 모델 선택이 필요하지 않습니다."
        if provider == "ollama":
            return "Ollama 0.30.x 이상은 Apple Silicon에서 지원 모델의 MLX engine 최적화를 내부적으로 사용할 수 있습니다. 앱에서는 Ollama 모델만 선택하면 됩니다."
        if provider == "mlx_compatible":
            return "고급 로컬 서버용 숨김 경로입니다. 기본 로컬 AI는 Ollama를 사용하세요."
        if provider == "codex_cli":
            return "API Key 없이 로컬 Codex CLI의 ChatGPT OAuth 로그인으로 실행합니다. 앱은 /opt/homebrew/bin/codex 같은 Homebrew 경로도 직접 확인합니다."
        return "모델 새로고침을 누르면 서버의 모델 목록을 선택할 수 있습니다."

    @staticmethod
    def _is_ai_chat_model_candidate(model: str) -> bool:
        lowered = (model or "").lower()
        return not any(term in lowered for term in ("embed", "embedding", "rerank", "whisper"))

    @staticmethod
    def _sanitize_ai_model_for_provider(provider: str, model: str) -> str:
        model = (model or "").strip()
        lowered = model.lower()
        if not model:
            return default_model(provider)
        if provider == "ollama" and any(term in lowered for term in ("mlx-community/", "lmstudio-community/")):
            return default_model(provider)
        if provider == "mlx_compatible" and model == default_model("ollama"):
            return default_model(provider)
        if provider == "openai_cloud":
            return default_model("codex_cli")
        if provider == "codex_cli" and any(term in lowered for term in ("mlx-community/", "lmstudio-community/", "qwen", "ollama", "gemma")):
            return default_model(provider)
        return model

    def _set_ai_provider_field_state(self, provider: str):
        if not hasattr(self, "edit_ai_endpoint"):
            return
        if provider in {"openai_cloud", "openai_compatible", "mlx_compatible"}:
            provider = "ollama" if provider == "mlx_compatible" else "codex_cli"
        endpoint_visible = provider == "ollama"
        model_visible = provider in {"ollama", "codex_cli"}
        active_visible = provider != "local_draft"
        for attr in ("lbl_ai_endpoint_field", "edit_ai_endpoint"):
            widget = getattr(self, attr, None)
            if widget is not None:
                widget.setVisible(endpoint_visible)
        for attr in ("lbl_ai_model_field", "cmb_ai_model", "lbl_ai_model_hint_field", "lbl_ai_model_hint"):
            widget = getattr(self, attr, None)
            if widget is not None:
                widget.setVisible(model_visible)
        for attr in ("lbl_ai_timeout_field", "spin_ai_timeout", "lbl_ai_privacy_field", "chk_ai_scrub"):
            widget = getattr(self, attr, None)
            if widget is not None:
                widget.setVisible(active_visible)
        self.edit_ai_endpoint.setEnabled(provider == "ollama")
        self.edit_ai_endpoint.setPlaceholderText(
            "예: http://127.0.0.1:11434/api/chat"
        )
        self.edit_ai_api_key.setText("")
        if provider in {"codex_cli", "local_draft"}:
            self.edit_ai_endpoint.setText("")
        if hasattr(self, "btn_find_mlx"):
            self.btn_find_mlx.setVisible(False)
            self.btn_find_mlx.setEnabled(False)
        if hasattr(self, "btn_start_mlx"):
            self.btn_start_mlx.setVisible(False)
            self.btn_start_mlx.setEnabled(False)
        if hasattr(self, "btn_refresh_ai_models"):
            self.btn_refresh_ai_models.setVisible(provider == "ollama")
            self.btn_refresh_ai_models.setEnabled(provider == "ollama")
        if hasattr(self, "btn_ai_test"):
            self.btn_ai_test.setVisible(active_visible)
            self.btn_ai_test.setEnabled(active_visible)

    def _update_ai_provider_help(self, provider: str):
        self._set_ai_provider_field_state(provider)
        if not hasattr(self, "lbl_ai_settings_status"):
            return
        if provider == "local_draft":
            status = "로컬 초안은 외부 AI를 호출하지 않습니다."
        elif provider == "ollama":
            status = "Ollama 로컬 서버를 사용합니다. Apple Silicon MLX 최적화는 Ollama 런타임 내부에서 처리됩니다."
        elif provider == "mlx_compatible":
            status = "고급 로컬 서버 경로는 기본 화면에서 숨겼습니다. 일반 사용은 Ollama 로컬을 선택하세요."
        elif provider == "codex_cli":
            status = "Codex CLI 클라우드는 API Key 없이 로컬 `codex` ChatGPT OAuth 로그인으로 실행합니다."
        else:
            status = "OpenAI 호환 서버의 chat completions 엔드포인트를 입력하세요."
        self.lbl_ai_settings_status.setText(status)
        if hasattr(self, "lbl_ai_model_hint"):
            self.lbl_ai_model_hint.setText(self._provider_model_note(provider))

    def _load_ai_settings(self):
        if not hasattr(self, "cmb_ai_provider"):
            return
        provider = str(self.settings.value("ai/provider", "local_draft"))
        if provider == "openai_cloud":
            provider = "codex_cli"
        elif provider in {"mlx_compatible", "openai_compatible"}:
            provider = "ollama"
        idx = self.cmb_ai_provider.findData(provider)
        self._loading_ai_settings = True
        self.cmb_ai_provider.blockSignals(True)
        self.cmb_ai_provider.setCurrentIndex(idx if idx >= 0 else 0)
        self.cmb_ai_provider.blockSignals(False)
        provider = self._provider_from_combo()
        endpoint = self._provider_setting_value(provider, "endpoint", default_endpoint(provider))
        model = self._sanitize_ai_model_for_provider(
            provider,
            self._provider_setting_value(provider, "model", default_model(provider)),
        )
        self.edit_ai_endpoint.setText(normalize_endpoint(provider, endpoint))
        self._set_ai_model_choices([], provider, recommended=model, select_recommended=True, note=self._provider_model_note(provider))
        self._set_ai_model_text(model or default_model(provider))
        self.edit_ai_api_key.setText("")
        try:
            self.spin_ai_timeout.setValue(float(self.settings.value("ai/timeout", 180)))
        except Exception:
            self.spin_ai_timeout.setValue(180)
        scrub = self.settings.value("ai/scrub_personal_data", True)
        self.chk_ai_scrub.setChecked(str(scrub).lower() not in {"false", "0", "no"})
        self._ai_active_provider = provider
        self._loading_ai_settings = False
        self._update_ai_provider_help(provider)

    def _ai_model_text(self) -> str:
        if hasattr(self, "cmb_ai_model"):
            return self.cmb_ai_model.currentText().strip() or self.edit_ai_model.text().strip()
        return self.edit_ai_model.text().strip() if hasattr(self, "edit_ai_model") else ""

    def _set_ai_model_text(self, model: str):
        model = (model or "").strip()
        if hasattr(self, "cmb_ai_model"):
            idx = self.cmb_ai_model.findText(model)
            if idx >= 0:
                self.cmb_ai_model.setCurrentIndex(idx)
            else:
                self.cmb_ai_model.setEditText(model)
        elif hasattr(self, "edit_ai_model"):
            self.edit_ai_model.setText(model)

    @staticmethod
    def _model_size_hint(model: str) -> float:
        lowered = model.lower()
        matches = re.findall(r"(\d+(?:\.\d+)?)\s*b", lowered)
        if not matches:
            return 0.0
        try:
            return max(float(value) for value in matches)
        except Exception:
            return 0.0

    @classmethod
    def _recommend_ai_model(cls, models: list[str], provider: str) -> str:
        clean = [model for model in dict.fromkeys(m.strip() for m in models if m and m.strip())]
        if not clean:
            return default_model(provider)

        def score(model: str) -> tuple[float, str]:
            lowered = model.lower()
            value = 0.0
            if any(term in lowered for term in ("embed", "rerank", "vision", "whisper")):
                value -= 200
            if any(term in lowered for term in ("qwen", "gemma", "llama", "mistral")):
                value += 25
            if provider == "ollama":
                if lowered.startswith("gemma4:e4b") or lowered == "gemma4":
                    value += 120
                elif lowered.startswith("qwen3.6"):
                    value += 90
                elif lowered.startswith("gemma4:26b"):
                    value += 82
            if any(term in lowered for term in ("instruct", "chat", "it")):
                value += 18
            if "coder" in lowered:
                value -= 8
            if any(term in lowered for term in ("optiq", "mlx-community")):
                value += 8
            if "4bit" in lowered or "q4" in lowered:
                value += 5
            size = cls._model_size_hint(model)
            if 7 <= size <= 10:
                value += 70
            elif 11 <= size <= 15:
                value += 62
            elif 20 <= size <= 32:
                value += 48
            elif 3 <= size < 7:
                value += 38
            elif 1 <= size < 3:
                value += 18
            elif 0 < size < 1:
                value -= 25
            if provider == "ollama" and any(term in lowered for term in ("qwen3", "qwen2.5", "gemma4", "gemma", "llama3")):
                value += 10
            return value, model

        return max(clean, key=score)

    def _set_ai_model_choices(
        self,
        models: list[str],
        provider: str,
        recommended: str = "",
        select_recommended: bool = True,
        note: str = "",
    ):
        if not hasattr(self, "cmb_ai_model"):
            return
        raw_clean = [model for model in dict.fromkeys(m.strip() for m in models if m and m.strip())]
        clean = [model for model in raw_clean if self._is_ai_chat_model_candidate(model)]
        removed_count = len(raw_clean) - len(clean)
        if raw_clean and not clean:
            self.cmb_ai_model.blockSignals(True)
            self.cmb_ai_model.clear()
            self.cmb_ai_model.setEditText("")
            self.cmb_ai_model.blockSignals(False)
            self.lbl_ai_model_hint.setText(
                f"감지된 모델 {len(raw_clean)}개가 모두 임베딩/비채팅 계열이라 숨겼습니다. "
                "채팅 모델을 설치하거나 모델명을 직접 입력하세요."
            )
            return
        recommended = recommended or self._recommend_ai_model(clean, provider)
        if recommended and recommended not in clean and clean:
            recommended = self._recommend_ai_model(clean, provider)
        current = self._ai_model_text()
        self.cmb_ai_model.blockSignals(True)
        self.cmb_ai_model.clear()
        for model in clean:
            self.cmb_ai_model.addItem(model)
        target = recommended if select_recommended and recommended else current
        if target and target not in clean:
            self.cmb_ai_model.addItem(target)
        self.cmb_ai_model.blockSignals(False)
        self._set_ai_model_text(target or default_model(provider))
        if clean:
            hint = f"감지 모델 {len(clean)}개"
            if recommended:
                size = self._model_size_hint(recommended)
                size_note = " · 균형 추천" if 7 <= size <= 15 else (" · 고품질/느림" if size >= 20 else " · 빠른 확인용")
                hint += f" · 추천: {recommended}{size_note}"
            if note:
                hint += f" · {note}"
            if removed_count:
                hint += f" · 비채팅 모델 {removed_count}개 숨김"
            self.lbl_ai_model_hint.setText(hint)
        else:
            self.lbl_ai_model_hint.setText(note or "감지된 모델이 없습니다. 모델명을 직접 입력할 수 있습니다.")

    def _save_ai_settings(self):
        if not hasattr(self, "cmb_ai_provider"):
            return
        provider = self._provider_from_combo()
        if provider in {"openai_cloud", "openai_compatible", "mlx_compatible"}:
            provider = "ollama"
        endpoint = normalize_endpoint(provider, self.edit_ai_endpoint.text()) if provider == "ollama" else ""
        self.edit_ai_endpoint.setText(endpoint)
        model = self._ai_model_text() or default_model(provider)
        model = self._sanitize_ai_model_for_provider(provider, model)
        self._set_ai_model_text(model)
        api_key = ""
        self.edit_ai_api_key.setText("")
        self.settings.setValue("ai/provider", provider)
        self.settings.setValue("ai/endpoint", endpoint)
        self.settings.setValue("ai/model", model)
        self.settings.setValue("ai/api_key", api_key)
        self.settings.setValue("ai/timeout", int(self.spin_ai_timeout.value()))
        self.settings.setValue("ai/scrub_personal_data", self.chk_ai_scrub.isChecked())
        self.settings.setValue(f"ai/providers/{provider}/endpoint", endpoint)
        self.settings.setValue(f"ai/providers/{provider}/model", model)
        self.settings.setValue(f"ai/providers/{provider}/api_key", api_key)
        self.settings.sync()
        self._ai_active_provider = provider
        self.lbl_ai_settings_status.setText("AI 설정을 저장했습니다.")
        self.statusBar().showMessage("AI 설정 저장 완료", 3000)

    def _sync_ai_provider_defaults(self, *_args):
        provider = self._provider_from_combo()
        if provider in {"openai_cloud", "openai_compatible", "mlx_compatible"}:
            provider = "ollama"
        if getattr(self, "_loading_ai_settings", False):
            self._update_ai_provider_help(provider)
            return
        previous = getattr(self, "_ai_active_provider", None)
        if previous and previous != provider:
            self._store_current_ai_provider_settings(previous)
        endpoint = self._provider_setting_value(provider, "endpoint", default_endpoint(provider))
        model = self._sanitize_ai_model_for_provider(
            provider,
            self._provider_setting_value(provider, "model", default_model(provider)),
        )
        self.edit_ai_endpoint.setText(normalize_endpoint(provider, endpoint))
        self._set_ai_model_choices([], provider, recommended=model, select_recommended=True, note=self._provider_model_note(provider))
        self._set_ai_model_text(model or default_model(provider))
        self.edit_ai_api_key.setText("")
        self._ai_active_provider = provider
        self.settings.setValue("ai/provider", provider)
        self.settings.sync()
        self._update_ai_provider_help(provider)

    def _ai_provider_config(self) -> AIProviderConfig:
        if not hasattr(self, "cmb_ai_provider"):
            return AIProviderConfig()
        provider = self._provider_from_combo()
        return AIProviderConfig(
            provider=provider,
            endpoint=normalize_endpoint(provider, self.edit_ai_endpoint.text()),
            model=self._sanitize_ai_model_for_provider(provider, self._ai_model_text() or default_model(provider)),
            api_key="",
            timeout=int(self.spin_ai_timeout.value()),
        )

    def _append_ai_progress(self, message: str):
        stamp = datetime.now().strftime("%H:%M:%S")
        line = f"[{stamp}] {message}"
        if hasattr(self, "txt_ai_progress"):
            self.txt_ai_progress.appendPlainText(line)
            self.ai_review_tabs.setCurrentWidget(self.txt_ai_progress)
        if hasattr(self, "lbl_ai_settings_status"):
            self.lbl_ai_settings_status.setText(message)
        self.statusBar().showMessage(message, 5000)

    def _set_ai_review_control_state(
        self,
        *,
        running: bool = False,
        can_resume: bool = False,
        stop_requested: bool = False,
    ):
        if hasattr(self, "btn_ai_review_stop"):
            self.btn_ai_review_stop.setEnabled(running and not stop_requested)
            self.btn_ai_review_stop.setText("중지 요청됨" if stop_requested else "보강 중지")
        if hasattr(self, "btn_ai_review_resume"):
            self.btn_ai_review_resume.setEnabled((not running) and can_resume)

    def _request_ai_review_stop(self):
        event = getattr(self, "_ai_review_cancel_event", None)
        if not getattr(self, "_ai_worker_active", False) or event is None:
            QMessageBox.information(self, "AI 문항 검토", "현재 중지할 AI 보강 작업이 없습니다.")
            return
        event.set()
        self._set_ai_review_control_state(running=True, can_resume=False, stop_requested=True)
        self._append_ai_progress("AI 보강 중지 요청 · 현재 응답 중인 묶음이 끝난 뒤 멈춥니다.")

    def _resume_ai_review_completion(self):
        state = getattr(self, "_ai_review_resume_state", None)
        if not state:
            QMessageBox.information(self, "AI 문항 검토", "이어갈 AI 보강 작업이 없습니다.")
            return
        self._run_ai_review_completion(resume_state=state)

    def _set_ai_busy(self, busy: bool, label: str = ""):
        for attr in (
            "btn_ai_enrich",
            "btn_ai_generate",
            "btn_ai_test",
            "btn_find_mlx",
            "btn_start_mlx",
            "btn_refresh_ai_models",
            "btn_ai_load_source",
        ):
            widget = getattr(self, attr, None)
            if widget is not None:
                widget.setEnabled(not busy)
        if busy:
            QApplication.setOverrideCursor(Qt.WaitCursor)
        else:
            while QApplication.overrideCursor() is not None:
                QApplication.restoreOverrideCursor()
        if busy and label:
            self._append_ai_progress(label)

    def _run_ai_background_task(self, title: str, func, on_success):
        if getattr(self, "_ai_worker_active", False):
            QMessageBox.information(self, title, "이미 AI 작업이 진행 중입니다. 진행 로그를 확인해 주세요.")
            return
        self._set_ai_busy(True, f"{title} 시작")
        events: queue.Queue = queue.Queue()
        self._ai_worker_active = True
        self._ai_worker_queue = events

        def progress(message: str):
            events.put(("progress", str(message)))

        def runner():
            try:
                events.put(("success", func(progress)))
            except Exception:
                events.put(("failure", traceback.format_exc()))

        worker = threading.Thread(target=runner, daemon=True, name=f"GoeduSplit-{title}")
        self._ai_worker_thread = worker
        worker.start()

        timer = QTimer(self)
        timer.setInterval(150)
        self._ai_worker_timer = timer

        def cleanup():
            timer.stop()
            timer.deleteLater()
            self._ai_worker_active = False
            self._ai_worker_queue = None
            self._ai_worker_thread = None
            self._ai_worker_timer = None
            self._set_ai_busy(False)
            self._set_ai_provider_field_state(self._provider_from_combo())

        def handle_success(result):
            try:
                on_success(result)
            except Exception:
                trace = traceback.format_exc()
                short = trace.strip().splitlines()[-1] if trace.strip() else "알 수 없는 오류"
                self._append_ai_progress(f"{title} 결과 반영 실패: {short}")
                if hasattr(self, "txt_ai_progress"):
                    self.txt_ai_progress.appendPlainText(trace)
                QMessageBox.warning(
                    self,
                    title,
                    f"결과를 화면에 반영하는 중 오류가 발생했습니다.\n{short}\n\n자세한 내용은 진행 로그를 확인해 주세요.",
                )
            finally:
                cleanup()

        def handle_failure(trace: str):
            try:
                short = trace.strip().splitlines()[-1] if trace.strip() else "알 수 없는 오류"
                self._append_ai_progress(f"{title} 실패: {short}")
                if hasattr(self, "txt_ai_progress"):
                    self.txt_ai_progress.appendPlainText(trace)
                QMessageBox.warning(
                    self,
                    title,
                    f"작업 중 오류가 발생했습니다.\n{short}\n\n자세한 내용은 진행 로그를 확인해 주세요.",
                )
            finally:
                if title == "AI 문항 검토":
                    self._ai_review_cancel_event = None
                    self._set_ai_review_control_state(
                        running=False,
                        can_resume=bool(getattr(self, "_ai_review_resume_state", None)),
                    )
                cleanup()

        def drain_events():
            while True:
                try:
                    kind, payload = events.get_nowait()
                except queue.Empty:
                    break
                if kind == "progress":
                    self._append_ai_progress(payload)
                elif kind == "success":
                    handle_success(payload)
                    break
                elif kind == "failure":
                    handle_failure(payload)
                    break
            if getattr(self, "_ai_worker_active", False) and not worker.is_alive() and events.empty():
                self._append_ai_progress(f"{title} 종료 상태를 확인하지 못했습니다. 다시 시도해 주세요.")
                cleanup()

        timer.timeout.connect(drain_events)
        timer.start()

    def _candidate_mlx_endpoints(self) -> list[str]:
        current = self.edit_ai_endpoint.text().strip() if hasattr(self, "edit_ai_endpoint") else ""
        raw = [
            "http://127.0.0.1:18080/v1/chat/completions",
            current,
            "http://127.0.0.1:8080/v1/chat/completions",
            "http://127.0.0.1:8000/v1/chat/completions",
            "http://127.0.0.1:1234/v1/chat/completions",
            "http://localhost:18080/v1/chat/completions",
            "http://localhost:8080/v1/chat/completions",
            "http://localhost:8000/v1/chat/completions",
            "http://localhost:1234/v1/chat/completions",
        ]
        endpoints = []
        for endpoint in raw:
            normalized = normalize_endpoint("mlx_compatible", endpoint)
            if normalized and normalized not in endpoints:
                endpoints.append(normalized)
        return endpoints

    def _probe_mlx_models(self) -> tuple[str, list[str], list[str]]:
        errors = []
        timeout = int(self.spin_ai_timeout.value()) if hasattr(self, "spin_ai_timeout") else 20
        for endpoint in self._candidate_mlx_endpoints():
            try:
                models = list_openai_compatible_models(
                    endpoint,
                    "",
                    "mlx_compatible",
                    min(timeout, 15),
                )
            except Exception as exc:
                errors.append(f"{endpoint}: {exc}")
                continue
            if models:
                return endpoint, models, errors
            errors.append(f"{endpoint}: 모델 목록이 비어 있습니다.")
        return "", [], errors

    def _refresh_ai_model_list(self):
        config = self._ai_provider_config()
        self._save_ai_settings()
        provider = config.provider
        if provider == "local_draft":
            self.lbl_ai_model_hint.setText("로컬 초안은 모델 선택이 필요하지 않습니다.")
            return
        if provider == "codex_cli":
            self._set_ai_model_choices([], provider, recommended=config.model or default_model(provider), select_recommended=True, note=self._provider_model_note(provider))
            self._append_ai_progress("Codex CLI 클라우드는 모델 목록을 새로고침하지 않습니다. 연결 테스트로 OAuth 상태를 확인하세요.")
            return
        endpoints = self._candidate_mlx_endpoints()

        def work(progress):
            progress("모델 목록 새로고침 시작")
            if provider == "ollama":
                progress("Ollama 모델 목록 확인 중")
                models = list_ollama_models(config.endpoint, min(config.timeout, 30))
                return {"provider": provider, "endpoint": config.endpoint, "models": models}
            if provider == "mlx_compatible":
                errors = []
                for endpoint in endpoints:
                    progress(f"MLX 모델 목록 확인 중: {endpoint}")
                    try:
                        models = list_openai_compatible_models(
                            endpoint,
                            "",
                            "mlx_compatible",
                            min(config.timeout, 30),
                        )
                    except Exception as exc:
                        errors.append(f"{endpoint}: {exc}")
                        continue
                    if models:
                        return {
                            "provider": provider,
                            "endpoint": endpoint,
                            "models": models,
                            "errors": errors,
                        }
                    errors.append(f"{endpoint}: 모델 목록이 비어 있습니다.")
                detail = "\n".join(errors[:5])
                raise ValueError(f"MLX/LM Studio 모델 목록을 읽지 못했습니다.\n{detail}")
            progress("OpenAI 호환 모델 목록 확인 중")
            models = list_openai_compatible_models(
                config.endpoint,
                config.api_key,
                provider,
                min(config.timeout, 30),
            )
            return {"provider": provider, "endpoint": config.endpoint, "models": models}

        def success(result):
            models = result.get("models") or []
            result_provider = result.get("provider") or provider
            endpoint = result.get("endpoint") or config.endpoint
            recommended = self._recommend_ai_model(models, result_provider)
            if endpoint:
                self.edit_ai_endpoint.setText(normalize_endpoint(result_provider, endpoint))
            self._set_ai_model_choices(
                models,
                result_provider,
                recommended=recommended,
                select_recommended=True,
                note="직접 다른 모델을 선택할 수 있습니다.",
            )
            self._save_ai_settings()
            message = f"모델 {len(models)}개 감지 · 추천 모델: {recommended}"
            self._append_ai_progress(message)

        self._run_ai_background_task("모델 새로고침", work, success)

    @staticmethod
    def _is_local_port_open(port: int, timeout: float = 0.25) -> bool:
        try:
            with socket.create_connection(("127.0.0.1", int(port)), timeout=timeout):
                return True
        except OSError:
            return False

    def _preferred_mlx_port(self) -> int:
        for port in (18080, 8080, 8000, 1234):
            if not self._is_local_port_open(port):
                return port
        return 18080

    def _find_local_mlx_server(self):
        config = AIProviderConfig(
            provider="mlx_compatible",
            endpoint=normalize_endpoint("mlx_compatible", self.edit_ai_endpoint.text()),
            model="",
            timeout=int(self.spin_ai_timeout.value()),
        )
        endpoints = self._candidate_mlx_endpoints()

        def work(progress):
            progress("MLX/LM Studio 로컬 서버 자동 탐색 시작")
            prepared, note = self._prepare_ai_connection_config_for_worker(config, endpoints, progress)
            models = list_openai_compatible_models(
                prepared.endpoint,
                "",
                "mlx_compatible",
                min(prepared.timeout, 30),
            )
            recommended = self._recommend_ai_model(models, "mlx_compatible")
            return {"config": prepared, "note": note, "models": models, "recommended": recommended}

        def success(result):
            prepared = result["config"]
            note = result.get("note", "")
            models = result.get("models") or []
            self._apply_prepared_ai_config(prepared, note)
            self._set_ai_model_choices(
                models,
                "mlx_compatible",
                recommended=result.get("recommended") or prepared.model,
                select_recommended=True,
                note="AI로 보강 품질을 높이려면 7B~15B급 모델을 권장합니다.",
            )
            message = f"MLX 서버 감지 완료 · {prepared.endpoint} · 모델 {prepared.model}"
            self.lbl_ai_settings_status.setText(message)
            self.statusBar().showMessage(message, 6000)

        self._run_ai_background_task("MLX 서버 찾기", work, success)

    def _start_local_mlx_server(self):
        exe = shutil.which("mlx_lm.server")
        if not exe:
            for version in ("3.14", "3.13", "3.12", "3.11"):
                known = Path(f"/Library/Frameworks/Python.framework/Versions/{version}/bin/mlx_lm.server")
                if known.exists():
                    exe = str(known)
                    break
        if not exe:
            QMessageBox.warning(
                self,
                "MLX 서버 시작",
                "mlx_lm.server 명령을 찾지 못했습니다.\n"
                "터미널에서 `pip install mlx-lm` 설치 후 다시 시도하세요.",
            )
            return
        current_model = self._ai_model_text() if hasattr(self, "edit_ai_model") else ""
        default_model_id = (
            current_model
            if current_model and current_model not in {"local-model", default_model("openai_compatible")}
            else default_model("mlx_compatible")
        )
        model, ok = QInputDialog.getText(
            self,
            "MLX 서버 시작",
            "사용할 MLX 모델 ID를 입력하세요.\n"
            "처음 사용하는 모델은 다운로드가 필요할 수 있습니다.\n"
            "연결 확인이 목적이면 작은 모델(Qwen3-0.6B-4bit)을 권장합니다.",
            text=default_model_id,
        )
        if not ok or not model.strip():
            return
        model = model.strip()
        open_ports = [port for port in (18080, 8080, 8000, 1234) if self._is_local_port_open(port)]
        if open_ports:
            port = open_ports[0]
            endpoint = f"http://127.0.0.1:{port}/v1/chat/completions"
            idx = self.cmb_ai_provider.findData("mlx_compatible")
            if idx >= 0:
                self.cmb_ai_provider.setCurrentIndex(idx)
            self.edit_ai_endpoint.setText(endpoint)
            self._set_ai_model_text(model)
            self._save_ai_settings()
            self._append_ai_progress(
                f"이미 열린 로컬 서버 포트 {port}를 발견했습니다. 새 서버를 띄우지 않고 연결 확인을 진행합니다."
            )
            QTimer.singleShot(300, self._find_local_mlx_server)
            return
        port = self._preferred_mlx_port()
        log_path = self._ai_reference_store_dir().parent / "mlx_server.log"
        try:
            log_file = open(log_path, "a", encoding="utf-8")
            process = subprocess.Popen(
                [exe, "--model", model, "--host", "127.0.0.1", "--port", str(port)],
                stdout=log_file,
                stderr=log_file,
                start_new_session=True,
            )
        except Exception as exc:
            QMessageBox.warning(self, "MLX 서버 시작", f"MLX 서버를 시작하지 못했습니다.\n{exc}")
            return
        self.mlx_server_process = process
        self.mlx_server_log_file = log_file
        idx = self.cmb_ai_provider.findData("mlx_compatible")
        if idx >= 0:
            self.cmb_ai_provider.setCurrentIndex(idx)
        self.edit_ai_endpoint.setText(f"http://127.0.0.1:{port}/v1/chat/completions")
        self._set_ai_model_text(model)
        self._save_ai_settings()
        self.lbl_ai_settings_status.setText(
            f"MLX 서버 시작 요청 완료. 5~20초 뒤 'MLX 찾기' 또는 '연결 테스트'를 누르세요. 로그: {log_path}"
        )
        self._append_ai_progress(f"MLX 서버 시작 요청 완료 · PID {process.pid} · 포트 {port} · 로그 {log_path}")
        QTimer.singleShot(7000, self._find_local_mlx_server)

    @staticmethod
    def _prepare_ai_connection_config_for_worker(
        config: AIProviderConfig,
        mlx_endpoints: list[str] | None = None,
        progress=None,
    ) -> tuple[AIProviderConfig, str]:
        def tell(message: str):
            if progress:
                progress(message)

        note = ""
        if config.provider in {"openai_cloud", "openai_compatible", "mlx_compatible"}:
            config.provider = "ollama"
            config.endpoint = normalize_endpoint("ollama", config.endpoint)
            config.model = MainWindow._sanitize_ai_model_for_provider("ollama", config.model or default_model("ollama"))
            config.api_key = ""
        if config.provider == "codex_cli":
            codex_path = find_codex_cli()
            if not codex_path:
                raise ValueError(
                    "codex CLI를 찾지 못했습니다. Windows에서는 `where codex`, macOS에서는 `which codex`를 확인하고 `codex login` 후 앱을 다시 여세요.\n"
                    "앱에서만 실패하면 CODEX_CLI_PATH 환경변수에 codex 실행 파일 전체 경로를 지정할 수 있습니다."
                )
            config.endpoint = ""
            config.api_key = ""
            config.model = MainWindow._sanitize_ai_model_for_provider("codex_cli", config.model or default_model("codex_cli"))
            tell(f"Codex CLI 설치 확인 완료: {codex_path}")
            tell("OAuth 로그인 상태는 실제 짧은 요청으로 확인합니다.")
            return config, "Codex CLI OAuth provider를 사용합니다."
        if config.provider == "ollama":
            tell("Ollama 서버 모델 목록 확인 중")
            note_parts: list[str] = []
            try:
                server_version = get_ollama_version(config.endpoint, min(config.timeout, 8))
            except Exception as exc:
                tell(f"Ollama 서버 버전 확인 실패: {exc}")
                server_version = ""
            if server_version:
                tell(f"Ollama 서버 버전 확인: {server_version}")
                if MainWindow._version_tuple(server_version) < (0, 30, 0):
                    note_parts.append(
                        f"Ollama 서버가 {server_version}입니다. Apple Silicon MLX engine 안내는 0.30.x 이상 기준이므로 Ollama 앱/serve를 재시작하거나 업데이트 반영을 확인하세요."
                    )
            models = list_ollama_models(config.endpoint, min(config.timeout, 20))
            if models:
                recommended = MainWindow._recommend_ai_model(models, "ollama")
                ordered = []
                for name in (config.model, recommended, *models):
                    name = (name or "").strip()
                    lowered = name.lower()
                    if not name or name in ordered:
                        continue
                    if name not in models:
                        continue
                    if any(term in lowered for term in ("embed", "rerank", "whisper")):
                        continue
                    ordered.append(name)
                probe_errors = []
                selected = ""
                for name in ordered[:5]:
                    tell(f"Ollama 짧은 실제 응답 확인 중: {name}")
                    probe_config = AIProviderConfig(
                        provider="ollama",
                        endpoint=config.endpoint,
                        model=name,
                        api_key=config.api_key,
                        timeout=min(max(config.timeout, 60), 300),
                    )
                    try:
                        output = run_completion(
                            '다음 JSON만 반환하세요: {"status":"ok"}',
                            probe_config,
                            max_tokens=96,
                        )
                    except Exception as exc:
                        probe_errors.append(f"{name}: {exc}")
                        continue
                    if output.strip():
                        selected = name
                        break
                    probe_errors.append(f"{name}: 빈 응답")
                if not selected:
                    detail = "\n".join(probe_errors[:5])
                    raise ValueError(
                        "Ollama 서버와 모델 목록은 확인했지만 실제 답변 가능한 채팅 모델을 찾지 못했습니다.\n"
                        "임베딩 모델은 문항 검토에 사용할 수 없습니다. `ollama pull gemma4:e4b` 같은 "
                        "채팅 모델을 받은 뒤 다시 시도하세요.\n\n"
                        f"점검 내용:\n{detail}"
                    )
                if config.model != selected:
                    note_parts.append(f"설치된 Ollama 모델을 감지해 모델을 {selected}(으)로 맞췄습니다.")
                config.model = selected
            else:
                note_parts.append("Ollama 서버는 응답했지만 설치된 모델 목록이 비어 있습니다.")
            if note_parts:
                note = " ".join(note_parts)
        elif config.provider == "mlx_compatible":
            errors = []
            endpoint = ""
            models = []
            selected_model = ""
            for candidate in mlx_endpoints or []:
                tell(f"MLX 서버 확인 중: {candidate}")
                try:
                    found = list_openai_compatible_models(
                        candidate,
                        "",
                        "mlx_compatible",
                        min(config.timeout, 15),
                    )
                except Exception as exc:
                    errors.append(f"{candidate}: {exc}")
                    continue
                found = [name for name in found if MainWindow._is_ai_chat_model_candidate(name)]
                if not found:
                    errors.append(f"{candidate}: 사용할 수 있는 채팅 모델이 없습니다.")
                    continue
                models = found
                recommended = MainWindow._recommend_ai_model(found, "mlx_compatible")
                ordered = []
                for name in (config.model, recommended, *found):
                    name = (name or "").strip()
                    if name and name in found and name not in ordered:
                        ordered.append(name)
                tell(f"모델 목록 확인 성공: {', '.join(ordered[:4])}")
                probe_config_timeout = min(max(config.timeout, 45), 120)
                probe_errors = []
                for preferred in ordered[:5]:
                    tell(f"짧은 실제 응답 확인 중: {preferred}")
                    try:
                        probe_openai_compatible_chat(
                            candidate,
                            "",
                            "mlx_compatible",
                            preferred,
                            probe_config_timeout,
                        )
                    except Exception as exc:
                        probe_errors.append(f"{preferred}: {exc}")
                        continue
                    endpoint = candidate
                    selected_model = preferred
                    tell(f"실제 응답 확인 완료: {candidate} · {preferred}")
                    break
                if endpoint:
                    break
                errors.append(
                    f"{candidate}: 모델 목록은 확인됐지만 실제 응답 가능한 모델을 찾지 못했습니다. "
                    + " / ".join(probe_errors[:3])
                )
            if not endpoint:
                detail = "\n".join(errors[:4])
                raise ValueError(
                    "MLX-LM/LM Studio 로컬 서버가 실제 응답 가능한 상태인지 확인하지 못했습니다.\n"
                    "서버 주소만 열려 있어도 큰 모델이 로딩 중이면 시간초과가 날 수 있습니다.\n"
                    "작은 모델로 먼저 확인하려면 터미널에서 다음처럼 실행하세요.\n"
                    "mlx_lm.server --model mlx-community/Qwen3-0.6B-4bit --host 127.0.0.1 --port 18080\n\n"
                    f"점검 내용:\n{detail}"
                )
            config.endpoint = endpoint
            if selected_model:
                if config.model != selected_model:
                    note = f"MLX 서버 실제 응답 모델을 감지해 모델을 {selected_model}(으)로 맞췄습니다."
                config.model = selected_model
            elif models and (not config.model or config.model not in models):
                config.model = MainWindow._recommend_ai_model(models, "mlx_compatible")
                note = f"MLX 서버 모델 목록을 감지해 모델을 {config.model}(으)로 맞췄습니다."
        elif config.provider == "openai_compatible":
            tell("OpenAI 호환 서버 모델 목록 확인 중")
            try:
                models = list_openai_compatible_models(
                    config.endpoint,
                    config.api_key,
                    config.provider,
                    min(config.timeout, 20),
                )
            except Exception:
                models = []
            if models and (not config.model or config.model not in models):
                config.model = models[0]
                note = f"서버 모델 목록을 감지해 모델을 {config.model}(으)로 맞췄습니다."
        return config, note

    def _apply_prepared_ai_config(self, config: AIProviderConfig, note: str = ""):
        if not hasattr(self, "cmb_ai_provider"):
            return
        idx = self.cmb_ai_provider.findData(config.provider)
        if idx >= 0:
            self.cmb_ai_provider.setCurrentIndex(idx)
        self.edit_ai_endpoint.setText(config.endpoint)
        self._set_ai_model_text(config.model)
        self.settings.setValue("ai/provider", config.provider)
        self.settings.setValue("ai/endpoint", config.endpoint)
        self.settings.setValue("ai/model", config.model)
        self.settings.setValue(f"ai/providers/{config.provider}/endpoint", config.endpoint)
        self.settings.setValue(f"ai/providers/{config.provider}/model", config.model)
        self.settings.setValue(f"ai/providers/{config.provider}/api_key", "")
        self.settings.sync()
        self._ai_active_provider = config.provider
        if note:
            self._append_ai_progress(note)

    def _prepare_ai_connection_config(self, config: AIProviderConfig) -> tuple[AIProviderConfig, str]:
        prepared, note = self._prepare_ai_connection_config_for_worker(
            config,
            self._candidate_mlx_endpoints(),
            self._append_ai_progress,
        )
        self._apply_prepared_ai_config(prepared, note)
        return prepared, note

    def _student_names_for_privacy(self) -> list[str]:
        if self.exam is None:
            return []
        return [st.name for st in self.exam.students if getattr(st, "name", "")]

    def _test_ai_connection(self):
        config = self._ai_provider_config()
        self._save_ai_settings()
        if config.provider == "local_draft":
            QMessageBox.information(self, "AI 연결 테스트", "로컬 초안 모드는 외부 연결 없이 바로 사용할 수 있습니다.")
            return
        endpoints = self._candidate_mlx_endpoints()
        prompt = '다음 JSON만 반환하세요: {"status":"ok"}'

        def work(progress):
            prepared, note = self._prepare_ai_connection_config_for_worker(config, endpoints, progress)
            progress(f"{prepared.label}에 최종 짧은 테스트 요청 전송 중")
            if prepared.provider == "codex_cli":
                check = check_codex_cli_oauth(prepared.model, min(prepared.timeout, 120))
                output = check["smoke"]
                progress(f"Codex CLI 버전 확인: {check['version']}")
                progress(f"Codex CLI OAuth 상태: {check['status']}")
            else:
                output = run_completion(prompt, prepared, max_tokens=96)
            if not output.strip():
                raise ValueError("AI 서버가 빈 응답을 보냈습니다. 모델 새로고침으로 다른 채팅 모델을 선택해 주세요.")
            progress("AI 응답 수신 완료")
            return {"config": prepared, "note": note, "output": output}

        def success(result):
            prepared = result["config"]
            note = result.get("note", "")
            self._apply_prepared_ai_config(prepared, note)
            status = f"{prepared.label} 연결 확인 완료. 모델: {prepared.model}"
            if note:
                status += f" · {note}"
            self.lbl_ai_settings_status.setText(status)
            self._append_ai_progress(status)
            QMessageBox.information(
                self,
                "AI 연결 테스트",
                f"{prepared.label} 응답을 정상적으로 읽었습니다.\n"
                f"모델: {prepared.model}\n"
                f"응답 예시: {(result.get('output') or '').strip()[:120]}\n"
                f"{note}".strip(),
            )

        self._run_ai_background_task("AI 연결 테스트", work, success)

    def _ai_review_example_text(self, kind: str) -> str:
        if kind == "perform":
            return (
                "평가요소: 수학적 모델링\n"
                "만점 10점\n"
                "성취기준 [10공수1-03-01] 함수의 그래프를 해석하고 상황에 맞게 모델링할 수 있다.\n"
                "목표수준 A\n"
                "난이도 어려움\n"
                "A: 주어진 상황을 함수로 일반화하고 그래프의 의미를 정확히 해석한다.\n"
                "B: 상황을 함수로 표현하고 그래프의 주요 특징을 설명한다.\n"
                "C: 기본적인 함수식과 그래프를 연결한다.\n"
                "D: 일부 값의 대응 관계를 찾는다.\n"
                "E: 안내를 받아 함수식의 의미를 확인한다.\n"
                "\n"
                "평가요소: 풀이 과정 설명\n"
                "만점 8점\n"
                "성취기준 [10공수1-02-02] 이차방정식의 실근과 허근을 이해하고 판별식을 활용할 수 있다.\n"
                "목표수준 B\n"
                "난이도 보통\n"
                "A: 판별식의 의미를 근의 개수와 연결해 논리적으로 설명한다.\n"
                "B: 판별식을 계산하고 근의 종류를 정확히 판별한다.\n"
                "C: 판별식 계산 절차를 수행한다.\n"
                "D: 일부 항을 대입해 계산한다.\n"
                "E: 안내를 받아 판별식의 형태를 확인한다."
            )
        return (
            "문항 1 | 유형 선택형 | 난이도 쉬움 | 목표수준 E | 배점 4 | "
            "성취기준 [10공수1-01-01] 다항식의 사칙연산 원리를 설명하고 계산할 수 있다.\n"
            "문항 내용: 두 다항식의 합을 계산하는 기본 문항이다. 보기 5개 중 정답을 고른다.\n"
            "\n"
            "문항 2 | 유형 서답형 | 난이도 보통 | 목표수준 C | 배점 5 | "
            "성취기준 [10공수1-02-03] 이차방정식의 근과 계수의 관계를 설명할 수 있다.\n"
            "문항 내용: 이차방정식 x^2 - 5x + 6 = 0의 두 근을 alpha, beta라 할 때 "
            "alpha+beta와 alpha beta를 구하고 근과 계수의 관계를 설명한다.\n"
            "\n"
            "문항 3 | 유형 서답형 | 난이도 어려움 | 목표수준 A | 배점 6 | "
            "성취기준 [10공수1-03-01] 함수의 그래프를 해석하고 상황에 맞게 모델링할 수 있다.\n"
            "문항 내용: 실생활 자료를 함수식으로 모델링하고 그래프의 변화율을 해석하여 결론을 정당화한다."
        )

    def _ai_review_reference_example_text(self, kind: str) -> str:
        if kind == "perform":
            return (
                "[10공수1-03-01] 함수의 그래프를 해석하고 상황에 맞게 모델링할 수 있다.\n"
                "A 수준: 주어진 상황을 수학적 모델로 일반화하고, 그래프의 의미와 한계를 근거와 함께 설명한다.\n"
                "B 수준: 상황을 함수식이나 그래프로 표현하고 주요 특징을 설명한다.\n"
                "C 수준: 기본적인 함수식과 그래프의 대응 관계를 설명한다.\n"
                "D 수준: 안내된 절차에 따라 일부 값의 대응 관계를 찾는다.\n"
                "E 수준: 도움을 받아 함수식 또는 그래프의 기본 의미를 확인한다.\n"
                "\n"
                "[10공수1-02-02] 이차방정식의 실근과 허근을 이해하고 판별식을 활용할 수 있다.\n"
                "A 수준: 판별식을 근의 구조와 연결하고 이유를 논리적으로 설명한다.\n"
                "B 수준: 판별식을 계산하여 근의 종류를 정확히 판별한다.\n"
                "C 수준: 판별식 계산 절차를 알고 기본 문항에 적용한다.\n"
                "D 수준: 안내에 따라 계수를 대입하고 일부 계산을 수행한다.\n"
                "E 수준: 도움을 받아 판별식의 형태와 의미를 확인한다."
            )
        return (
            "[10공수1-01-01] 다항식의 사칙연산 원리를 설명하고 계산할 수 있다.\n"
            "A 수준: 다항식 연산 원리를 일반화하고 여러 표현을 연결해 설명한다.\n"
            "B 수준: 다항식의 사칙연산 원리를 설명하고 복합 계산을 수행한다.\n"
            "C 수준: 기본적인 다항식 덧셈, 뺄셈, 곱셈을 수행한다.\n"
            "D 수준: 안내된 절차에 따라 간단한 다항식 계산을 수행한다.\n"
            "E 수준: 도움을 받아 동류항과 기본 계산 절차를 확인한다.\n"
            "\n"
            "[10공수1-02-03] 이차방정식의 근과 계수의 관계를 설명할 수 있다.\n"
            "A 수준: 근과 계수의 관계를 여러 상황에 적용하고 식의 의미를 정당화한다.\n"
            "B 수준: 근과 계수의 관계를 설명하고 표준적인 문항에 적용한다.\n"
            "C 수준: 두 근의 합과 곱을 계수와 연결해 구한다.\n"
            "D 수준: 안내에 따라 두 근의 합 또는 곱 일부를 구한다.\n"
            "E 수준: 도움을 받아 근과 계수 관계의 기본 형태를 확인한다.\n"
            "\n"
            "[10공수1-03-01] 함수의 그래프를 해석하고 상황에 맞게 모델링할 수 있다.\n"
            "A 수준: 실생활 자료를 함수식으로 모델링하고 그래프의 변화율을 해석하여 결론을 정당화한다.\n"
            "B 수준: 자료의 관계를 함수식이나 그래프로 표현하고 주요 특징을 설명한다.\n"
            "C 수준: 기본적인 그래프 정보를 읽고 함수식과 연결한다.\n"
            "D 수준: 안내에 따라 그래프의 일부 정보를 찾는다.\n"
            "E 수준: 도움을 받아 좌표, 증가·감소 등 기본 정보를 확인한다."
        )

    def _insert_ai_review_example(self, kind: str):
        if not hasattr(self, "txt_ai_review_source"):
            return
        source_has_text = self.txt_ai_review_source.toPlainText().strip()
        reference_has_text = (
            hasattr(self, "txt_ai_review_reference")
            and self.txt_ai_review_reference.toPlainText().strip()
        )
        if source_has_text or reference_has_text:
            answer = QMessageBox.question(
                self,
                "예시 입력",
                "현재 문항 자료와 성취기준·수준 자료를 예시로 바꿀까요?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            if answer != QMessageBox.Yes:
                return
        self.txt_ai_review_source.setPlainText(self._ai_review_example_text(kind))
        if hasattr(self, "txt_ai_review_reference"):
            self.txt_ai_review_reference.setPlainText(self._ai_review_reference_example_text(kind))
        label = "수행평가" if kind == "perform" else "지필평가"
        self.statusBar().showMessage(f"{label} 문항 자료와 성취기준·수준 예시를 입력했습니다. '검토 초안 생성'을 누르세요.", 5000)

    def _extract_text_from_review_file(self, path: str) -> str:
        p = Path(path)
        suffix = p.suffix.lower()
        if suffix in {".txt", ".md", ".csv", ".json"}:
            return self._format_ai_extracted_markdown(p.read_text(encoding="utf-8", errors="replace"))
        if suffix in {".xlsx", ".xlsm"}:
            import openpyxl
            wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
            parts = []
            for ws in wb.worksheets[:8]:
                parts.append(f"## 시트: {ws.title}")
                for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    if r_idx > 120:
                        parts.append("...")
                        break
                    values = [str(v).strip() for v in row[:24] if v not in (None, "")]
                    if values:
                        parts.append(" | ".join(values))
            return "\n".join(parts)
        if suffix == ".docx":
            return self._format_ai_extracted_markdown(self._extract_docx_text(p))
        if suffix == ".hwpx":
            return self._format_ai_extracted_markdown(self._extract_hwpx_text(p))
        if suffix == ".hwp":
            return self._format_ai_extracted_markdown(self._extract_hwp_text_with_external_tools(p))
        if suffix == ".pdf":
            try:
                import pypdf
            except Exception as exc:
                raise ValueError("PDF 텍스트 추출을 위해 pypdf가 필요합니다. requirements.txt를 반영해 설치 후 다시 시도하세요.") from exc
            reader = pypdf.PdfReader(str(p))
            pages = []
            for i, page in enumerate(reader.pages[:80], start=1):
                text = self._normalize_pdf_math_text(page.extract_text() or "")
                if text.strip():
                    pages.append(f"## PDF {i}쪽\n\n{self._format_ai_extracted_markdown(text)}")
            return "\n".join(pages)
        raise ValueError("지원 형식: .txt, .md, .csv, .json, .xlsx, .xlsm, .docx, .hwp, .hwpx, .pdf")

    @staticmethod
    def _xml_local_name(tag: str) -> str:
        return tag.rsplit("}", 1)[-1] if "}" in tag else tag

    def _extract_docx_text(self, path: Path) -> str:
        try:
            with zipfile.ZipFile(path) as zf:
                names = [name for name in zf.namelist() if name == "word/document.xml"]
                if not names:
                    raise ValueError("DOCX 본문 XML을 찾지 못했습니다.")
                root = ET.fromstring(zf.read(names[0]))
        except zipfile.BadZipFile as exc:
            raise ValueError("DOCX 파일 구조를 읽지 못했습니다.") from exc
        paragraphs = []
        for para in root.iter():
            if self._xml_local_name(para.tag) != "p":
                continue
            texts = []
            for node in para.iter():
                name = self._xml_local_name(node.tag)
                if name == "t" and node.text:
                    texts.append(node.text)
                elif name in {"tab", "br"}:
                    texts.append("\t" if name == "tab" else "\n")
            joined = "".join(texts).strip()
            if joined:
                paragraphs.append(joined)
        return "\n".join(paragraphs)

    def _extract_hwpx_text(self, path: Path) -> str:
        try:
            with zipfile.ZipFile(path) as zf:
                xml_names = [
                    name for name in zf.namelist()
                    if name.lower().endswith(".xml")
                    and (
                        "/section" in name.lower()
                        or name.lower().startswith("contents/")
                        or name.lower().startswith("bodytext/")
                    )
                ]
                section_names = [
                    name for name in xml_names
                    if "section" in Path(name).name.lower()
                ] or xml_names
                section_names = sorted(section_names)[:120]
                parts = []
                for name in section_names:
                    try:
                        root = ET.fromstring(zf.read(name))
                    except Exception:
                        continue
                    parts.extend(self._extract_text_from_hwpx_xml_root(root))
        except zipfile.BadZipFile as exc:
            raise ValueError("HWPX 파일 구조를 읽지 못했습니다.") from exc
        text = "\n".join(part for part in parts if part.strip())
        if not text.strip():
            raise ValueError("HWPX에서 본문 텍스트를 찾지 못했습니다.")
        return text

    def _extract_text_from_hwpx_xml_root(self, root: ET.Element) -> list[str]:
        paragraphs: list[str] = []
        for elem in root.iter():
            if self._xml_local_name(elem.tag) not in {"p", "para"}:
                continue
            chunks = []
            for node in elem.iter():
                name = self._xml_local_name(node.tag)
                if name in {"t", "text"} and node.text:
                    chunks.append(node.text)
                elif name in {"lineBreak", "br"}:
                    chunks.append("\n")
                elif name in {"tab"}:
                    chunks.append("\t")
            line = re.sub(r"[ \t]+", " ", "".join(chunks)).strip()
            if line:
                paragraphs.append(line)
        return paragraphs

    @staticmethod
    def _is_useful_converter_output(text: str) -> bool:
        stripped = (text or "").strip()
        if len(stripped) < 20:
            return False
        lowered = stripped.lower()
        if "usage:" in lowered and len(stripped) < 800:
            return False
        if "not found" in lowered and len(stripped) < 400:
            return False
        return True

    def _extract_hwp_text_with_external_tools(self, path: Path) -> str:
        attempts: list[tuple[str, list[str]]] = []
        if shutil.which("hwp5txt"):
            attempts.append(("hwp5txt", ["hwp5txt", str(path)]))
        if shutil.which("kordoc"):
            attempts.extend([
                ("kordoc", ["kordoc", str(path)]),
                ("kordoc parse", ["kordoc", "parse", str(path)]),
                ("kordoc parse-document", ["kordoc", "parse-document", str(path)]),
                ("kordoc to-markdown", ["kordoc", "to-markdown", str(path)]),
            ])
        errors = []
        for label, command in attempts:
            try:
                completed = subprocess.run(
                    command,
                    capture_output=True,
                    text=True,
                    timeout=45,
                    check=False,
                )
            except Exception as exc:
                errors.append(f"{label}: {exc}")
                continue
            output = (completed.stdout or "").strip()
            if self._is_useful_converter_output(output):
                return output
            err = (completed.stderr or output or f"종료코드 {completed.returncode}").strip()
            errors.append(f"{label}: {err[:180]}")
        detail = "\n".join(errors[:5])
        raise ValueError(
            "HWP 파일은 현재 앱 내장 파서만으로 안정적으로 읽기 어렵습니다. "
            "HWPX로 저장해 불러오거나, 로컬에 kordoc 또는 hwp5txt를 설치한 뒤 다시 시도해 주세요."
            + (f"\n\n시도 결과:\n{detail}" if detail else "")
        )

    @staticmethod
    def _format_ai_material_header(title: str, path: Path | str) -> str:
        return f"# {title}\n\n**저장 위치:** `{path}`"

    @staticmethod
    def _format_ai_extracted_markdown(text: str) -> str:
        if not text:
            return ""
        text = text.replace("\r\n", "\n").replace("\r", "\n")
        text = re.sub(r"[ \t]+", " ", text)
        text = re.sub(r"\n{3,}", "\n\n", text)
        text = re.sub(r"(?<!\n)\s+(?=##\s*PDF\s*\d+쪽)", "\n\n", text)
        text = re.sub(r"(?<!^)(?<!\n)\s+(?=(?:[1-9]|[1-9]\d)\.\s)", "\n\n", text)
        text = re.sub(r"(?<!^)(?<!\n)\s+(?=문항\s*\d{1,3}\b)", "\n\n", text)
        text = re.sub(r"\s+(?=(?:①|②|③|④|⑤|⑥|⑦|⑧|⑨|⑩))", " ", text)
        lines = []
        for raw in text.splitlines():
            line = raw.strip()
            if not line:
                lines.append("")
                continue
            if re.match(r"^(?:[1-9]|[1-9]\d)\.\s", line):
                lines.append(f"### {line}")
            elif re.match(r"^문항\s*\d{1,3}\b", line):
                lines.append(f"### {line}")
            elif line.startswith(("과목:", "학기/학년:", "시험일자", "교사")):
                lines.append(f"**{line}**")
            elif line.startswith("※"):
                lines.append(f"> {line}")
            else:
                lines.append(line)
        formatted = "\n".join(lines)
        formatted = re.sub(r"\n{3,}", "\n\n", formatted)
        return formatted.strip()

    @staticmethod
    def _normalize_pdf_math_text(text: str) -> str:
        if not text:
            return ""
        special_replacements = {
            "\ue05c\ue06d\ue046\ue034": r"\sqrt{-1}",
            "\ue06d\ue0fe": r"\overline{z}",
            "\ue06d\ue0fb": r"\overline{w}",
            "\ue06d\ue0fc": r"\overline{x}",
            "\ue06d\ue0e5": r"\overline{a}",
            "\ue06d\ue0e6": r"\overline{b}",
        }
        for old, new in special_replacements.items():
            text = text.replace(old, new)
        glyphs = {
            "\ue034": "1",
            "\ue035": "2",
            "\ue036": "3",
            "\ue037": "4",
            "\ue038": "5",
            "\ue039": "6",
            "\ue03a": "7",
            "\ue03b": "8",
            "\ue03c": "9",
            "\ue03d": "0",
            "\ue000": "A",
            "\ue001": "B",
            "\ue002": "C",
            "\ue003": "D",
            "\ue010": "Q",
            "\ue011": "R",
            "\ue044": "(",
            "\ue045": ")",
            "\ue046": "-",
            "\ue047": "=",
            "\ue048": "+",
            "\ue052": ",",
            "\ue055": r"\le ",
            "\ue05c": r"\sqrt",
            "\ue06d": "",
            "\ue0e5": "a",
            "\ue0e6": "b",
            "\ue0e7": "c",
            "\ue0e8": "d",
            "\ue0ed": "i",
            "\ue0fb": "w",
            "\ue0fc": "x",
            "\ue0fd": "y",
            "\ue0fe": "z",
        }
        text = "".join(glyphs.get(ch, ch) for ch in text)
        text = re.sub(r"([xyzwabcd])([0-9])", r"\1^{\2}", text)
        text = re.sub(r"\\sqrt\s*\{?-1\}?", r"\\sqrt{-1}", text)
        text = re.sub(r"\s+([①②③④⑤])", r" \1", text)
        return text

    def _compact_ai_reference_text(self, text: str) -> str:
        entries = self._ai_reference_entries(text)
        level_parts = []
        standard_parts = []
        seen = set()
        for entry in entries:
            raw = str(entry.get("text", "")).strip()
            code = str(entry.get("code", "")).strip()
            if not raw:
                continue
            if not AI_STANDARD_CODE_RE.search(raw):
                continue
            if any(token in raw for token in ("예시 답안", "채점 시 고려", "피드백 방안", "문항 개발", "수행 과제")):
                continue
            cleaned = self._clean_ai_reference_standard_text(raw, code)
            if not cleaned:
                continue
            key = cleaned[:120]
            if key in seen:
                continue
            seen.add(key)
            levels = self._extract_ai_reference_levels(raw)
            if levels:
                level_parts.append(cleaned + "\n" + "\n".join(levels))
            else:
                standard_parts.append(cleaned)
        if level_parts:
            return "\n\n".join(level_parts)
        if standard_parts:
            return "\n\n".join(standard_parts)
        relevant = []
        for line in text.splitlines():
            compact = re.sub(r"\s+", " ", line).strip()
            if not compact:
                continue
            if AI_STANDARD_CODE_RE.search(compact) or any(
                token in compact for token in ("성취기준", "성취수준", "평가기준", "최소능력자")
            ):
                relevant.append(compact)
        return "\n".join(relevant[:500]) if relevant else text[:AI_REFERENCE_TEXT_LIMIT]

    @staticmethod
    def _extract_ai_reference_levels(text: str) -> list[str]:
        compact = re.sub(r"\s+", " ", text or "").strip()
        levels = []
        for lv in LEVELS_AE:
            match = re.search(rf"(?:^|\s){lv}\s+(.+?)(?=\s[ABCDE]\s+|$)", compact)
            if not match:
                continue
            body = match.group(1).strip()
            body = re.sub(r"\s+", " ", body)
            if len(body) < 12 or not re.search(r"[가-힣]", body):
                continue
            body = body[:260].strip()
            levels.append(f"{lv}: {body}")
        return levels

    def _ai_material_root_dir(self) -> Path:
        if sys.platform == "darwin":
            root = Path.home() / "Library" / "Application Support" / "Goedu-Split"
        elif sys.platform.startswith("win"):
            root = Path(os.environ.get("APPDATA", Path.home() / "AppData" / "Roaming")) / "Goedu-Split"
        else:
            root = Path(os.environ.get("XDG_DATA_HOME", Path.home() / ".local" / "share")) / "Goedu-Split"
        try:
            root.mkdir(parents=True, exist_ok=True)
        except OSError:
            root = Path(tempfile.gettempdir()) / "Goedu-Split"
            root.mkdir(parents=True, exist_ok=True)
        return root

    @staticmethod
    def _ensure_writable_dir(path: Path, fallback_name: str) -> Path:
        try:
            path.mkdir(parents=True, exist_ok=True)
            probe = path / ".__goedusplit_write_test__"
            probe.write_text("ok", encoding="utf-8")
            probe.unlink(missing_ok=True)
            return path
        except OSError:
            fallback = Path(tempfile.gettempdir()) / "Goedu-Split" / fallback_name
            fallback.mkdir(parents=True, exist_ok=True)
            return fallback

    def _ai_source_store_dir(self) -> Path:
        return self._ensure_writable_dir(self._ai_material_root_dir() / "ai_source_files", "ai_source_files")

    def _ai_reference_store_dir(self) -> Path:
        return self._ensure_writable_dir(self._ai_material_root_dir() / "ai_reference_files", "ai_reference_files")

    def _ai_source_cache_path(self) -> Path:
        return self._ai_source_store_dir() / AI_SOURCE_CACHE_NAME

    def _ai_reference_cache_path(self) -> Path:
        return self._ai_reference_store_dir() / AI_REFERENCE_CACHE_NAME

    def _refresh_ai_source_store_label(self):
        if not hasattr(self, "lbl_ai_source_store"):
            return
        store = self._ai_source_store_dir()
        stored_files = [
            path for path in store.iterdir()
            if path.is_file() and path.name != AI_SOURCE_CACHE_NAME
        ]
        self.lbl_ai_source_store.setText(f"저장자료 {len(stored_files)}개")
        self.lbl_ai_source_store.setToolTip(f"저장 위치: {store}")

    def _refresh_ai_reference_store_label(self):
        if not hasattr(self, "lbl_ai_reference_store"):
            return
        store = self._ai_reference_store_dir()
        stored_files = [
            path for path in store.iterdir()
            if path.is_file() and path.name != AI_REFERENCE_CACHE_NAME
        ]
        self.lbl_ai_reference_store.setText(f"저장자료 {len(stored_files)}개")
        self.lbl_ai_reference_store.setToolTip(f"저장 위치: {store}")

    def _show_ai_store_location(self, kind: str):
        if kind == "reference":
            title = "성취기준·수준 자료 저장 위치"
            store = self._ai_reference_store_dir()
        else:
            title = "문항 자료 저장 위치"
            store = self._ai_source_store_dir()
        QMessageBox.information(
            self,
            title,
            f"불러온 자료와 직접 저장한 텍스트는 아래 폴더에 보관됩니다.\n\n{store}",
        )

    @staticmethod
    def _unique_store_path(directory: Path, filename: str) -> Path:
        safe_name = re.sub(r"[\\/:*?\"<>|]+", "_", filename).strip() or "reference"
        target = directory / safe_name
        if not target.exists():
            return target
        stem = target.stem
        suffix = target.suffix
        for i in range(2, 1000):
            candidate = directory / f"{stem}_{i}{suffix}"
            if not candidate.exists():
                return candidate
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return directory / f"{stem}_{stamp}{suffix}"

    def _store_ai_reference_file(self, path: Path) -> Path:
        store = self._ai_reference_store_dir()
        target = self._unique_store_path(store, path.name)
        shutil.copy2(path, target)
        self._refresh_ai_reference_store_label()
        return target

    def _store_ai_source_file(self, path: Path) -> Path:
        store = self._ai_source_store_dir()
        target = self._unique_store_path(store, path.name)
        shutil.copy2(path, target)
        self._refresh_ai_source_store_label()
        return target

    def _save_ai_source_cache(self, text: str):
        cache = self._ai_source_cache_path()
        cache.write_text(text[:AI_SOURCE_TEXT_LIMIT], encoding="utf-8")
        self._refresh_ai_source_store_label()

    def _save_ai_reference_cache(self, text: str):
        cache = self._ai_reference_cache_path()
        cache.write_text(text[:AI_REFERENCE_TEXT_LIMIT], encoding="utf-8")
        self._refresh_ai_reference_store_label()

    def _load_ai_reference_cache_if_empty(self):
        if self._ai_review_reference_text():
            return
        cache = self._ai_reference_cache_path()
        if not cache.exists():
            return
        try:
            text = cache.read_text(encoding="utf-8", errors="replace")
        except OSError:
            return
        if text.strip():
            self._set_ai_review_source_text(text, target="reference")

    def _save_current_ai_text(self, kind: str):
        if kind == "source":
            text = self._ai_review_source_text()
            store = self._ai_source_store_dir()
            label = "문항 자료"
            suffix = "source"
            limit = AI_SOURCE_TEXT_LIMIT
        else:
            text = self._ai_review_reference_text()
            store = self._ai_reference_store_dir()
            label = "성취기준·수준 자료"
            suffix = "reference"
            limit = AI_REFERENCE_TEXT_LIMIT
        if not text:
            QMessageBox.information(self, "현재 내용 저장", f"저장할 {label}가 없습니다.")
            return
        default = store / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{suffix}.txt"
        path, _ = QFileDialog.getSaveFileName(
            self,
            f"{label} 저장",
            str(default),
            "텍스트 파일 (*.txt);;모든 파일 (*.*)",
        )
        if not path:
            return
        try:
            Path(path).write_text(text[:limit], encoding="utf-8")
            if kind == "source":
                self._save_ai_source_cache(text)
            else:
                self._save_ai_reference_cache(text)
        except Exception as exc:
            QMessageBox.warning(self, "현재 내용 저장", f"저장하지 못했습니다.\n{exc}")
            return
        self.statusBar().showMessage(f"{label} 저장 완료 · {path}", 5000)

    def _choose_saved_ai_files(self, kind: str) -> list[str]:
        if kind == "source":
            store = self._ai_source_store_dir()
            title = "저장된 문항 자료 선택"
        else:
            store = self._ai_reference_store_dir()
            title = "저장된 성취기준·수준 자료 선택"
        paths, _ = QFileDialog.getOpenFileNames(
            self,
            title,
            str(store),
            "저장 자료 (*.txt *.md *.csv *.json *.xlsx *.xlsm *.docx *.hwp *.hwpx *.pdf);;모든 파일 (*.*)",
        )
        return paths

    def _load_saved_ai_source_text(self):
        paths = self._choose_saved_ai_files("source")
        if not paths:
            return
        parts = []
        failures = []
        for path_str in paths:
            path = Path(path_str)
            try:
                text = self._extract_text_from_review_file(str(path))
            except Exception as exc:
                failures.append(f"{path.name}: {exc}")
                continue
            parts.append(f"{self._format_ai_material_header(f'문항 자료: {path.name}', path)}\n\n{text}")
        if not parts:
            QMessageBox.warning(self, "저장자료 불러오기", "선택한 문항 자료를 읽지 못했습니다.\n" + "\n".join(failures[:5]))
            return
        text = "\n\n".join(parts)
        self._set_ai_review_source_text(text, target="source")
        self._save_ai_source_cache(text)
        self._refresh_ai_source_store_label()
        self.statusBar().showMessage(f"저장된 문항 자료 {len(parts)}개를 불러왔습니다.", 5000)
        if failures:
            QMessageBox.warning(self, "일부 문항 자료 읽기 실패", "\n".join(failures[:8]))

    def _load_saved_ai_reference_text(self):
        paths = self._choose_saved_ai_files("reference")
        if not paths:
            return
        parts = []
        failures = []
        for path_str in paths:
            path = Path(path_str)
            try:
                text = self._compact_ai_reference_text(self._extract_text_from_review_file(str(path)))
            except Exception as exc:
                failures.append(f"{path.name}: {exc}")
                continue
            parts.append(f"{self._format_ai_material_header(f'성취기준·수준 자료: {path.name}', path)}\n\n{text}")
        if not parts:
            QMessageBox.warning(
                self,
                "저장자료 불러오기",
                "선택한 성취기준·수준 자료를 읽지 못했습니다.\n" + "\n".join(failures[:5]),
            )
            return
        text = "\n\n".join(parts)
        self._set_ai_review_source_text(text, target="reference")
        self._save_ai_reference_cache(text)
        self._refresh_ai_reference_store_label()
        self.statusBar().showMessage(f"저장된 성취기준·수준 자료 {len(parts)}개를 불러왔습니다.", 5000)
        if failures:
            QMessageBox.warning(self, "일부 참고자료 읽기 실패", "\n".join(failures[:8]))

    def _set_ai_review_source_text(self, text: str, *, target: str = "source"):
        if target == "reference" and hasattr(self, "txt_ai_review_reference"):
            self.txt_ai_review_reference.setPlainText(text[:AI_REFERENCE_TEXT_LIMIT])
            self._update_ai_markdown_preview("reference")
            if hasattr(self, "ai_review_input_tabs"):
                self.ai_review_input_tabs.setCurrentIndex(1)
            if hasattr(self, "ai_reference_view_tabs"):
                self.ai_reference_view_tabs.setCurrentIndex(1)
            return
        self.txt_ai_review_source.setPlainText(text[:AI_SOURCE_TEXT_LIMIT])
        self._update_ai_markdown_preview("source")
        if hasattr(self, "ai_review_input_tabs"):
            self.ai_review_input_tabs.setCurrentIndex(0)
        if hasattr(self, "ai_source_view_tabs"):
            self.ai_source_view_tabs.setCurrentIndex(1)

    def _ai_review_source_text(self) -> str:
        if not hasattr(self, "txt_ai_review_source"):
            return ""
        return self.txt_ai_review_source.toPlainText().strip()

    def _ai_review_reference_text(self) -> str:
        if not hasattr(self, "txt_ai_review_reference"):
            return ""
        return self.txt_ai_review_reference.toPlainText().strip()

    def _load_ai_review_file(self):
        paths, _ = QFileDialog.getOpenFileNames(
            self,
            "시험 문제 또는 문항 자료 불러오기",
            "",
            "검토 자료 (*.txt *.md *.csv *.json *.xlsx *.xlsm *.docx *.hwp *.hwpx *.pdf);;모든 파일 (*.*)",
        )
        if not paths:
            return
        existing = self._ai_review_source_text().strip()
        append_existing = False
        if existing:
            choice = QMessageBox.question(
                self,
                "문항 자료 추가",
                f"이미 문항 자료가 들어 있습니다.\n선택한 {len(paths)}개 자료를 기존 자료 뒤에 추가할까요?\n\n"
                "예: 추가\n아니오: 기존 자료를 지우고 교체",
                QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel,
                QMessageBox.Yes,
            )
            if choice == QMessageBox.Cancel:
                return
            append_existing = choice == QMessageBox.Yes
        parts = []
        failures = []
        for path_str in paths:
            path = Path(path_str)
            try:
                saved = self._store_ai_source_file(path)
                text = self._extract_text_from_review_file(str(path))
            except Exception as e:
                failures.append(f"{path.name}: {e}")
                continue
            parts.append(
                f"{self._format_ai_material_header(f'문항 자료: {path.name}', saved)}\n\n"
                f"{text}"
            )
        if not parts:
            QMessageBox.warning(
                self,
                "AI 문항 검토",
                "선택한 문항 자료를 읽지 못했습니다.\n" + "\n".join(failures[:5]),
            )
            return
        text = "\n\n".join(parts)
        if append_existing:
            text = f"{existing}\n\n{text}"
        self._set_ai_review_source_text(text, target="source")
        self._save_ai_source_cache(text)
        message = (
            f"문항 자료 {len(parts)}개 불러오기 완료 · "
            f"저장 위치: {self._ai_source_store_dir()}"
        )
        if failures:
            message += f" · 실패 {len(failures)}개"
            QMessageBox.warning(
                self,
                "일부 문항 자료 읽기 실패",
                "다음 자료는 불러오지 못했습니다.\n" + "\n".join(failures[:8]),
            )
        self.statusBar().showMessage(message, 7000)

    def _load_ai_reference_file(self):
        paths, _ = QFileDialog.getOpenFileNames(
            self,
            "성취기준·성취수준 참고자료 불러오기",
            "",
            "참고 자료 (*.txt *.md *.csv *.json *.xlsx *.xlsm *.docx *.hwp *.hwpx *.pdf);;모든 파일 (*.*)",
        )
        if not paths:
            return
        existing = self._ai_review_reference_text().strip()
        append_existing = False
        if existing:
            choice = QMessageBox.question(
                self,
                "성취기준·수준 자료 추가",
                f"이미 참고자료가 들어 있습니다.\n선택한 {len(paths)}개 자료를 기존 자료 뒤에 추가할까요?\n\n"
                "예: 추가\n아니오: 기존 자료를 지우고 교체",
                QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel,
                QMessageBox.Yes,
            )
            if choice == QMessageBox.Cancel:
                return
            append_existing = choice == QMessageBox.Yes
        loaded_parts = []
        failures = []
        for path_str in paths:
            path = Path(path_str)
            try:
                saved = self._store_ai_reference_file(path)
                text = self._compact_ai_reference_text(self._extract_text_from_review_file(str(path)))
            except Exception as e:
                failures.append(f"{path.name}: {e}")
                continue
            loaded_parts.append(
                f"{self._format_ai_material_header(f'성취기준·수준 자료: {path.name}', saved)}\n\n"
                f"{text}"
            )
        if not loaded_parts:
            QMessageBox.warning(
                self,
                "AI 문항 검토",
                "선택한 참고자료를 읽지 못했습니다.\n" + "\n".join(failures[:5]),
            )
            return
        text = "\n\n".join(loaded_parts)
        if append_existing:
            text = f"{existing}\n\n{text}"
        self._set_ai_review_source_text(text, target="reference")
        self._save_ai_reference_cache(text)
        message = (
            f"성취기준·수준 자료 {len(loaded_parts)}개 불러오기 완료 · "
            f"저장 위치: {self._ai_reference_store_dir()}"
        )
        if failures:
            message += f" · 실패 {len(failures)}개"
            QMessageBox.warning(
                self,
                "일부 참고자료 읽기 실패",
                "다음 자료는 불러오지 못했습니다.\n" + "\n".join(failures[:8]),
            )
        self.statusBar().showMessage(message, 7000)

    def _load_ai_review_from_exam(self):
        if self.exam is None:
            QMessageBox.warning(self, "AI 문항 검토", "먼저 문항정보표를 불러와 분석을 실행해 주세요.")
            return
        lines = [
            "# 현재 문항정보표",
            "",
            f"**과목:** {self.exam.subject or '(과목 미상)'}",
            f"**학기/학년:** {self.exam.semester} {self.exam.grade}",
            "",
        ]
        type_order = {"선택형": 0, "서답형": 1}
        for item in sorted(self.exam.items, key=lambda it: (it.number, type_order.get(it.item_type, 9))):
            standard = " ".join(part for part in [item.standard_code, item.standard] if part).strip()
            lines.append(
                f"### 문항 {item.number} | 유형 {item.item_type} | 난이도 {item.difficulty or '-'} | "
                f"배점 {item.score:g} | 내용영역 {item.content_area or '-'} | 성취기준 {standard or '-'}"
            )
        self._set_ai_review_source_text("\n".join(lines), target="source")
        self.statusBar().showMessage("현재 문항정보표를 AI 문항 검토 원문으로 가져왔습니다.", 3500)

    def _split_ai_review_blocks(self, text: str) -> list[dict]:
        text = self._mark_ai_review_item_boundaries(text)
        normalized_lines = []
        for line in text.splitlines():
            line = re.sub(r"^#{1,6}\s*", "", line.strip())
            line = re.sub(r"^\s*[-*]\s+", "", line)
            line = line.replace("**", "").replace("__", "").replace("`", "")
            normalized_lines.append(re.sub(r"\s+", " ", line).strip())
        lines = normalized_lines
        lines = [line for line in lines if line]
        blocks = []
        current = None
        section_type = ""
        item_re = re.compile(r"^(?:문항\s*)?(\d{1,3})\s*(?:번|[.)]|[|])?\s*(.*)$")
        for raw_line in lines:
            line = raw_line
            for marker, marker_type in (
                ("[선택형 문제]", "선택형"),
                ("[선택형]", "선택형"),
                ("[서답형 문제]", "서답형"),
                ("[서답형]", "서답형"),
                ("[논술형 문제]", "서답형"),
                ("[논술형]", "서답형"),
            ):
                if marker in line:
                    section_type = marker_type
                    before, after = line.split(marker, 1)
                    line = after.strip() or before.strip()
                    break
            if not line:
                continue
            m = item_re.match(line)
            looks_like_item = False
            if m:
                no = int(m.group(1))
                looks_like_item = no <= 150 and (
                    "문항" in line
                    or re.match(r"^\d{1,3}\s*(?:번|[.)])", line)
                    or "성취기준" in line
                    or "배점" in line
                    or any(mark in line for mark in ("①", "②", "③", "④", "⑤", "보기"))
                )
            if looks_like_item:
                label = f"{int(m.group(1))}번"
                starts_new_item_record = bool(re.match(r"^문항\s*\d{1,3}\b", line))
                if current and current.get("kind") == "문항" and current.get("label") == label and not starts_new_item_record:
                    current["text"] += "\n" + line
                else:
                    if current:
                        blocks.append(current)
                    current = {"kind": "문항", "label": label, "text": line, "section_type": section_type}
            elif any(key in line for key in ("평가요소", "채점기준", "수행 수준", "수행수준")):
                if current:
                    blocks.append(current)
                current = {"kind": "수행평가", "label": line[:32], "text": line, "section_type": "수행평가"}
            elif current:
                current["text"] += "\n" + line
            elif re.search(r"\[[^\]\n]{4,40}\]", line):
                blocks.append({"kind": "성취기준", "label": "성취기준", "text": line})
        if current:
            blocks.append(current)
        if not blocks and text.strip():
            blocks.append({"kind": "전체 자료", "label": "전체", "text": text.strip()[:4000]})
        return blocks[:120]

    @staticmethod
    def _mark_ai_review_item_boundaries(text: str) -> str:
        """Put likely item labels on their own line when PDF extraction joins them."""
        if not text:
            return ""
        marked = re.sub(
            r"(?<!^)(?<!\n)(?<=\S)\s+(?=(?:문항\s*)?\d{1,3}\s*(?:번|[.)])\s*)",
            "\n",
            text,
        )
        marked = re.sub(
            r"(?<!^)(?<!\n)(?<=\S)\s+(?=문항\s*\d{1,3}\b)",
            "\n",
            marked,
        )
        marked = re.sub(
            r"(?<!^)(?<!\n)(?<=\S)(?=문항\s*\d{1,3}\b)",
            "\n",
            marked,
        )
        return marked

    def _supplement_ai_review_rows_from_exam(self, rows: list[dict]) -> tuple[list[dict], int]:
        if self.exam is None:
            return self._sort_ai_review_rows(rows), 0
        existing_keys: set[tuple[str, int]] = set()
        existing_unknown_numbers: set[int] = set()
        for row in rows:
            raw_type = row.get("review_type") or row.get("평가유형") or ""
            number = self._ai_review_number(row.get("label") or row.get("번호/요소") or "", -1)
            if number <= 0:
                continue
            if "수행" in str(raw_type):
                continue
            if "서답" in str(raw_type) or "서술" in str(raw_type) or "논술" in str(raw_type):
                existing_keys.add(("서답형", number))
            elif "선택" in str(raw_type) or "객관" in str(raw_type):
                existing_keys.add(("선택형", number))
            else:
                existing_unknown_numbers.add(number)

        supplemented = list(rows)
        added = 0
        type_order = {"선택형": 0, "서답형": 1}
        for item in sorted(self.exam.items, key=lambda it: (type_order.get(it.item_type, 9), it.number)):
            key = (item.item_type, int(item.number))
            if key in existing_keys:
                continue
            if item.item_type == "선택형" and item.number in existing_unknown_numbers:
                continue
            standard = " ".join(part for part in [item.standard_code, item.standard] if part).strip()
            block = {
                "kind": "문항",
                "label": f"{item.number}번",
                "section_type": item.item_type,
                "text": (
                    f"{item.number}번 유형 {item.item_type} 난이도 {item.difficulty or '-'} "
                    f"배점 {float(item.score):g}점 내용영역 {item.content_area or '-'} "
                    f"성취기준 {standard or '-'}"
                ),
            }
            row = self._infer_ai_review_row(block, [])
            row["evidence"] = " · ".join(
                part for part in [row.get("evidence", ""), "문항정보표 보충"] if part
            )
            row["next_step"] = " / ".join(
                part for part in [row.get("next_step", ""), "문항 원문 확인"] if part
            )
            supplemented.append(row)
            added += 1
        return self._sort_ai_review_rows(supplemented), added

    @staticmethod
    def _ai_reference_tokens(text: str) -> set[str]:
        stopwords = {
            "성취기준", "성취수준", "수준", "문항", "자료", "설명", "확인",
            "있다", "한다", "대한", "통해", "기본", "도움", "안내",
            "서로", "다른", "대하여", "다음", "값은", "경우", "단", "각각",
            "옳은", "것은", "실수", "자연수", "만족", "때의",
        }
        tokens = set(re.findall(r"[가-힣A-Za-z0-9]{2,}", text or ""))
        return {token for token in tokens if token not in stopwords}

    def _ai_reference_entries(self, reference_text: str) -> list[dict[str, str | set[str]]]:
        reference_text = re.sub(r"(?=\[(?=[^\]\n]*\d)[^\]\n]{4,40}\])", "\n", reference_text)
        normalized_lines = []
        for line in reference_text.splitlines():
            line = re.sub(r"^#{1,6}\s*", "", line.strip())
            line = re.sub(r"^\s*[-*]\s+", "", line)
            line = line.replace("**", "").replace("__", "").replace("`", "")
            normalized_lines.append(re.sub(r"\s+", " ", line).strip())
        lines = normalized_lines
        lines = [line for line in lines if line]
        entries: list[dict[str, str | set[str]]] = []
        current: list[str] = []
        current_code = ""

        def flush():
            if not current:
                return
            text = " ".join(current).strip()
            if not text:
                return
            entries.append({
                "code": current_code,
                "text": text[:900],
                "tokens": self._ai_reference_tokens(text),
            })

        for line in lines:
            codes = AI_STANDARD_CODE_RE.findall(line)
            starts_standard = bool(codes) or line.startswith(("성취기준", "평가기준"))
            if starts_standard and current:
                flush()
                current = []
                current_code = ""
            if codes and not current_code:
                current_code = codes[0]
            current.append(line)
        flush()
        if not entries and reference_text.strip():
            text = reference_text.strip()[:3000]
            entries.append({"code": "", "text": text, "tokens": self._ai_reference_tokens(text)})
        return entries[:120]

    @staticmethod
    def _clean_ai_reference_standard_text(reference_text: str, reference_code: str = "") -> str:
        text = re.sub(r"\s+", " ", reference_text or "").strip()
        if not text:
            return ""
        if reference_code and reference_code in text:
            text = text[text.find(reference_code):]
        cut_positions = []
        for pattern in (
            r"\sA[:\s]",
            r"\sB[:\s]",
            r"\sC[:\s]",
            r"\sD[:\s]",
            r"\sE[:\s]",
            r"\s프로젝트\s",
            r"\s공통수학\d+-\d+",
            r"\s예시문항\s",
            r"\s평가도구\s",
        ):
            match = re.search(pattern, text)
            if match and match.start() > 0:
                cut_positions.append(match.start())
        if cut_positions:
            text = text[:min(cut_positions)].strip()
        if reference_code and reference_code not in text:
            text = f"{reference_code} {text}".strip()
        return text[:240].strip()

    def _match_ai_reference_entry(self, compact_text: str, entries: list[dict[str, str | set[str]]]) -> dict[str, str | set[str]] | None:
        if not entries:
            return None
        codes = AI_STANDARD_CODE_RE.findall(compact_text)
        for code in codes:
            for entry in entries:
                if code and code == entry.get("code"):
                    return entry
                if code and code in str(entry.get("text", "")):
                    return entry
        keyword_match = self._match_ai_reference_by_keywords(compact_text, entries)
        if keyword_match:
            return keyword_match
        tokens = self._ai_reference_tokens(compact_text)
        if not tokens:
            return None
        best = None
        best_score = 0
        for entry in entries:
            code = str(entry.get("code", ""))
            text = str(entry.get("text", ""))
            if "예시 답안" in text:
                continue
            entry_tokens = entry.get("tokens", set())
            if not isinstance(entry_tokens, set):
                continue
            score = len(tokens & entry_tokens)
            if score > best_score:
                best = entry
                best_score = score
        return best if best_score >= 2 else None

    @staticmethod
    def _entry_for_standard_code(entries: list[dict[str, str | set[str]]], code: str) -> dict[str, str | set[str]] | None:
        target = code if code.startswith("[") else f"[{code}]"
        for entry in entries:
            if entry.get("code") == target or target in str(entry.get("text", "")):
                return entry
        return None

    def _match_ai_reference_by_keywords(self, compact_text: str, entries: list[dict[str, str | set[str]]]) -> dict[str, str | set[str]] | None:
        text = compact_text or ""
        keyword_rules = [
            ("10공수1-04-02", ("행렬", "행과 열", "")),
            ("10공수1-02-01", ("복소수", "허수", "켤레복소수", "")),
            ("10공수1-03-03", ("조합", "C", "nCr")),
            ("10공수1-03-02", ("순열", "일렬", "나열")),
            ("10공수1-03-01", ("경우의 수", "주사위", "메뉴판", "동시에 던져", "선택하는 경우", "자연수의 개수")),
            ("10공수1-01-03", ("인수분해",)),
            ("10공수1-01-02", ("나머지정리", "인수정리", "조립제법", "나누었을 때", "나누어떨어")),
            ("10공수1-01-01", ("다항식", "동류항", "사칙연산")),
            ("10공수1-02-06", ("최대", "최소")),
            ("10공수1-02-05", ("그래프와 직선", "위치 관계")),
            ("10공수1-02-04", ("이차방정식과 이차함수", "판별식", "그래프")),
            ("10공수1-02-03", ("근과 계수",)),
            ("10공수1-02-02", ("실근", "허근", "이차방정식")),
        ]
        for code, keywords in keyword_rules:
            if any(keyword in text for keyword in keywords):
                entry = self._entry_for_standard_code(entries, code)
                if entry:
                    return entry
        return None

    def _ai_previous_exam_expected_values(
        self,
        compact_text: str,
        standard: str,
        review_type: str,
        difficulty: str,
    ) -> tuple[dict[str, str] | None, str, str]:
        """Use loaded exam statistics as the first local evidence for A~E judgment."""
        if not self.item_stats:
            return None, "", ""
        expected_type = self._ai_review_expected_type(review_type)
        candidates = []
        text_tokens = self._ai_reference_tokens(compact_text)
        standard_codes = set(AI_STANDARD_CODE_RE.findall(f"{standard} {compact_text}"))
        for stat in self.item_stats:
            item = getattr(stat, "item", None)
            if item is None:
                continue
            p_by_level = getattr(stat, "p_by_level", {}) or {}
            if not any(lv in p_by_level for lv in LEVELS_AE):
                continue
            item_type = getattr(item, "item_type", "")
            if expected_type in {"선택형", "서답형"} and item_type and item_type != expected_type:
                continue
            score = 0
            item_code = getattr(item, "standard_code", "") or ""
            if item_code and item_code in standard_codes:
                score += 12
            item_text = " ".join(
                str(part or "")
                for part in (
                    item_code,
                    getattr(item, "standard", ""),
                    getattr(item, "content_area", ""),
                )
            )
            item_tokens = self._ai_reference_tokens(item_text)
            score += min(8, len(text_tokens & item_tokens))
            item_difficulty = getattr(item, "difficulty", "") or ""
            if difficulty and item_difficulty and self._ai_review_difficulty(item_difficulty) == difficulty:
                score += 2
            content_area = str(getattr(item, "content_area", "") or "").strip()
            if content_area and content_area in compact_text:
                score += 3
            candidates.append((score, stat))

        if not candidates:
            return None, "", ""
        candidates.sort(key=lambda pair: pair[0], reverse=True)
        best_score = candidates[0][0]
        if best_score < 3:
            return None, "", ""
        selected = [stat for score, stat in candidates[:3] if score >= max(3, best_score - 2)]
        basis = f"유사 문항 {len(selected)}개"

        rates: dict[str, float] = {}
        for lv in LEVELS_AE:
            values = []
            for stat in selected:
                p_by_level = getattr(stat, "p_by_level", {}) or {}
                if lv not in p_by_level or p_by_level.get(lv) is None:
                    continue
                try:
                    values.append(float(p_by_level.get(lv)))
                except Exception:
                    continue
            if values:
                rates[lv] = sum(values) / len(values)
        if not rates:
            return None, "", ""
        expected, target, _ = ai_review_expected_values_from_rates(rates, sample_size=3)
        summary = ai_review_rate_summary(rates)
        evidence = f"이전시험 {basis} 성취수준별 정답률({summary})"
        return expected, target, evidence

    def _infer_ai_review_row(self, block: dict, reference_entries: list[dict[str, str | set[str]]] | None = None) -> dict:
        text = block["text"]
        compact = re.sub(r"\s+", " ", text)
        codes = AI_STANDARD_CODE_RE.findall(compact)
        standard = codes[0] if codes else ""
        if standard:
            pos = compact.find(standard)
            standard = compact[pos:pos + 130].strip()
        matched_reference = self._match_ai_reference_entry(compact, reference_entries or [])
        if matched_reference:
            reference_text = str(matched_reference.get("text", "")).strip()
            reference_code = str(matched_reference.get("code", "")).strip()
            cleaned_reference = self._clean_ai_reference_standard_text(reference_text, reference_code)
            if reference_text and (not standard or standard == reference_code or (reference_code and reference_code in standard)):
                standard = cleaned_reference or reference_text[:130].strip()

        explicit_level = ""
        for lv in LEVELS_AE:
            if re.search(rf"(목표|판별|성취수준|수준)\s*[:：]?\s*{lv}\b", compact):
                explicit_level = lv
                break

        high_terms = ["복합", "추론", "증명", "정당화", "일반화", "모델링", "탐구", "분석", "해석", "실생활", "최대", "최소"]
        mid_terms = ["활용", "설명", "해결", "관계", "그래프", "원리", "성질", "이해"]
        low_terms = ["간단", "기본", "안내", "계산", "대입", "구하", "알고", "따라"]
        high = sum(1 for term in high_terms if term in compact)
        mid = sum(1 for term in mid_terms if term in compact)
        low = sum(1 for term in low_terms if term in compact)
        score = high * 2 + mid - low
        point_values = [
            float(match.group(1))
            for match in re.finditer(r"(\d+(?:\.\d+)?)\s*점", compact)
        ]
        inferred_points = max(point_values) if point_values else 0.0

        if "어려" in compact or re.search(r"\b상\b", compact):
            difficulty = "어려움"
        elif "쉬움" in compact or re.search(r"\b하\b", compact):
            difficulty = "쉬움"
        elif "보통" in compact or re.search(r"\b중\b", compact):
            difficulty = "보통"
        elif inferred_points >= 8:
            difficulty = "어려움"
        elif inferred_points >= 7:
            difficulty = "보통"
        elif inferred_points > 0:
            difficulty = "쉬움"
        elif score >= 4:
            difficulty = "어려움"
        elif score <= 0:
            difficulty = "쉬움"
        else:
            difficulty = "보통"

        if explicit_level:
            target = explicit_level
        elif inferred_points >= 8:
            target = "B"
        elif inferred_points >= 7:
            target = "C"
        elif inferred_points >= 5:
            target = "C" if score >= 1 else "D"
        elif inferred_points > 0:
            target = "D" if score >= 0 else "E"
        elif score >= 5:
            target = "A"
        elif score >= 3:
            target = "B"
        elif score >= 1:
            target = "C"
        elif score <= -2:
            target = "E"
        else:
            target = "D"

        block_section_type = str(block.get("section_type", ""))
        if block_section_type in {"선택형", "서답형"}:
            review_type = block_section_type
        elif any(term in compact for term in ("수행평가", "평가요소", "채점기준", "수행 수준", "수행수준")):
            review_type = "수행평가"
        elif "유형 선택형" in compact:
            review_type = "선택형"
        elif "유형 서답형" in compact:
            review_type = "서답형"
        elif any(term in compact for term in ("서답", "서술", "논술", "풀이과정", "증명하", "구하시오", "설명하시오", "쓰시오", "답하시오", "계산하시오")):
            review_type = "서답형"
        elif any(term in compact for term in ("선택형", "객관식", "보기", "①", "②", "③", "④", "⑤")):
            review_type = "선택형"
        else:
            review_type = "지필 문항"

        previous_expected = None
        previous_evidence = ""
        if review_type != "수행평가":
            previous_expected, previous_target, previous_evidence = self._ai_previous_exam_expected_values(
                compact, standard, review_type, difficulty
            )
            if previous_expected and previous_target:
                target = previous_target

        evidence_terms = [term for term in high_terms + mid_terms + low_terms if term in compact][:4]
        if inferred_points > 0:
            evidence_terms.insert(0, f"배점 {inferred_points:g}점")
        if matched_reference:
            evidence_terms.append("참고자료 매칭")
        if previous_evidence:
            evidence_terms.append(previous_evidence)
        evidence = " · ".join(evidence_terms) if evidence_terms else compact[:90]
        warnings = []
        if not codes and not matched_reference:
            warnings.append("성취기준 코드 확인")
        if reference_entries and not matched_reference:
            warnings.append("성취기준·수준 자료 대조 확인")
        if block["kind"] == "전체 자료":
            warnings.append("문항/평가요소 분리 확인")
        if review_type == "지필 문항" and "배점" not in compact:
            warnings.append("배점 확인")
        next_step = " / ".join(warnings) if warnings else "성취기준 진술과 문항 요구 행동을 비교"
        if review_type == "수행평가":
            expected = self._ai_default_perform_expected_values(target, difficulty)
        else:
            expected = previous_expected or self._ai_default_ox_expected_values(target, difficulty)
        return {
            "kind": block["kind"],
            "label": block["label"],
            "standard": standard or "(후보 없음)",
            "review_type": review_type,
            "target": target,
            "difficulty": difficulty,
            **expected,
            "evidence": evidence,
            "next_step": next_step,
        }

    def _ai_default_ox_expected_values(self, target: str, difficulty: str) -> dict[str, str]:
        target = self._ai_review_level(target, "C")
        difficulty = self._ai_review_difficulty(difficulty)
        base = {
            "A": [2, 1, 0, 0, 0],
            "B": [3, 2, 1, 0, 0],
            "C": [3, 3, 2, 1, 0],
            "D": [3, 3, 3, 2, 1],
            "E": [3, 3, 3, 3, 2],
        }[target]
        if difficulty == "쉬움":
            counts = [min(3, value + 1) for value in base]
        elif difficulty == "어려움":
            target_index = LEVELS_AE.index(target)
            counts = [
                value if idx <= target_index else max(0, value - 1)
                for idx, value in enumerate(base)
            ]
        else:
            counts = base
        target_index = LEVELS_AE.index(target)
        counts = list(counts)
        counts[target_index] = 2
        for idx in range(target_index):
            counts[idx] = max(counts[idx], 2)
        for idx in range(target_index + 1, len(counts)):
            counts[idx] = min(counts[idx], 2)
        prev = 3
        normalized = []
        for value in counts:
            clipped = max(0, min(3, min(prev, int(value))))
            normalized.append(clipped)
            prev = clipped
        return {f"{lv} 예상": f"{count}/3" for lv, count in zip(LEVELS_AE, normalized)}

    def _ai_default_perform_expected_values(self, target: str, difficulty: str) -> dict[str, str]:
        target = self._ai_review_level(target, "C")
        target_index = LEVELS_AE.index(target)
        base_rates = {"A": 90, "B": 80, "C": 70, "D": 60, "E": 40}
        if difficulty == "쉬움":
            delta = 5
        elif difficulty == "어려움":
            delta = -5
        else:
            delta = 0
        values = {}
        prev = 100
        for idx, lv in enumerate(LEVELS_AE):
            level_adjust = 4 if idx < target_index else (-4 if idx > target_index else 0)
            rate = max(0, min(100, base_rates[lv] + delta + level_adjust))
            rate = min(prev, rate)
            values[f"{lv} 예상"] = f"{rate}%"
            prev = rate
        return values

    def _make_ai_review_prompt(self, rows: list[dict], source_text: str, reference_text: str = "") -> str:
        preview = source_text.strip()
        if len(preview) > 12000:
            preview = preview[:12000] + "\n...(이하 생략)"
        reference_preview = reference_text.strip()
        if len(reference_preview) > 10000:
            reference_preview = reference_preview[:10000] + "\n...(이하 생략)"
        row_lines = "\n".join(
            f"- {row['label']}: 유형={row['review_type']}, 목표={row['target']}, 난이도={row['difficulty']}, 성취기준={row['standard']}"
            for row in rows[:60]
        )
        return (
            "너는 고등학교 성취평가 현장지원단의 문항 검토 보조자다.\n"
            "먼저 시험 문제 자료를 문항 단위로 분석한 뒤, 성취기준·성취수준 참고자료와 대조하여 "
            "각 문항 또는 수행평가 평가요소의 성취기준, 목표 성취수준 후보, 난이도 후보를 제안하라.\n"
            "중요 원칙: 자동 확정하지 말고, 교사가 검토할 수 있도록 근거 문장을 함께 제시한다.\n"
            "A 수준 문항은 A 수준 최소능력자 3명 중 약 2명이 해결할 수 있는 문항이라는 기준으로 판단한다.\n"
            "목표수준 후보가 X라면 X 예상은 원칙적으로 2/3으로 둔다. X보다 높은 수준은 2/3~3/3, "
            "X보다 낮은 수준은 0/3~2/3 범위에서 문항 난이도에 맞춰 조정한다.\n"
            "수행평가는 평가요소별로 A~E 최소능력자의 예상점수를 산출할 수 있도록 채점기준표의 행동 표현을 분석한다.\n\n"
            "출력 형식은 표로 한다: 번호/요소 | 성취기준 후보 | 평가유형 | 목표수준 후보 | 난이도 후보 | "
            "A 예상 | B 예상 | C 예상 | D 예상 | E 예상 | 근거 | 추가 확인 질문.\n"
            "지필 문항의 A~E 예상은 3명 기준 O/X 개수처럼 3/3, 2/3, 1/3, 0/3으로 쓰고, "
            "수행평가는 만점 대비 예상점수 또는 90% 같은 비율로 쓴다.\n\n"
            "[로컬 초안]\n"
            f"{row_lines}\n\n"
            "[시험 문제 자료]\n"
            f"{preview}\n\n"
            "[성취기준·성취수준 참고자료]\n"
            f"{reference_preview or '(참고자료 없음)'}"
        )

    def _make_structured_ai_review_prompt(
        self,
        rows: list[dict],
        source_text: str,
        reference_text: str = "",
        *,
        source_limit: int = 18_000,
        reference_limit: int = 16_000,
        row_limit: int = 80,
        scope_note: str = "",
    ) -> str:
        preview = source_text.strip()
        if len(preview) > source_limit:
            preview = preview[:source_limit] + "\n...(이하 생략)"
        reference_preview = reference_text.strip()
        if len(reference_preview) > reference_limit:
            reference_preview = reference_preview[:reference_limit] + "\n...(이하 생략)"
        draft = "\n".join(
            f"- {row.get('label') or row.get('번호/요소')}: 유형={row.get('review_type') or row.get('평가유형')}, "
            f"목표={row.get('target') or row.get('목표수준 후보')}, 난이도={row.get('difficulty') or row.get('난이도 후보')}, "
            f"예상={', '.join(str(row.get(f'{lv} 예상', '')) for lv in LEVELS_AE)}, "
            f"성취기준={row.get('standard') or row.get('성취기준 후보')}, "
            f"근거={row.get('evidence') or row.get('근거') or ''}, "
            f"다음확인={row.get('next_step') or row.get('다음 확인') or ''}"
            for row in rows[:row_limit]
        )
        scope = ""
        if scope_note:
            scope = (
                f"이번 요청 범위: {scope_note}\n"
                "반드시 이번 범위에 있는 항목만 출력하고, 빠진 항목 없이 같은 순서로 JSON 배열을 작성한다.\n\n"
            )
        return (
            "너는 고등학교 성취평가 현장지원단의 문항 검토 보조자다.\n"
            "이 요청은 앱 파일 수정이나 코드 작성 요청이 아니다. 제공된 시험 문제 자료, 로컬 초안, "
            "성취기준·성취수준 참고자료만 읽고 교사용 검토표 JSON을 작성한다.\n"
            "아래 시험 문제 자료와 로컬 초안, 성취기준·성취수준 참고자료를 바탕으로 문항/수행평가 평가요소를 다시 검토하라.\n"
            f"{scope}"
            "작업 순서는 반드시 1) 시험 문제 자료를 문항 단위로 읽기, 2) 참고자료에서 가장 가까운 성취기준과 성취수준 설명 찾기, "
            "3) 로컬 초안의 이전시험 성취수준별 정답률 근거가 있으면 우선 검토하기, "
            "4) 그 기준에 맞춰 목표수준·난이도·A~E 예상값 제안하기 순서로 한다.\n"
            "출력 행 수와 문항 순서는 반드시 로컬 초안과 같아야 한다. 번호를 재정렬하거나 누락하지 마라.\n"
            "번호/요소 값은 로컬 초안의 항목 번호를 그대로 유지한다.\n"
            "성취기준 후보는 참고자료의 코드나 진술이 직접 뒷받침할 때만 바꾸고, 근거가 약하면 로컬 초안을 유지하라.\n"
            "참고자료에 없는 성취기준명이나 코드를 새로 만들지 마라.\n"
            "목표수준 후보는 C/D/E 기본값으로 몰지 말고, 배점·문항 요구 행동·성취수준 설명·이전시험 정답률이 뒷받침할 때 A/B도 검토하라.\n"
            "로컬 초안과 다른 목표수준을 제안하려면 근거 열에 왜 바꾸는지 문항 단서와 참고자료 단서를 함께 써라.\n"
            "성취기준이나 목표수준을 바꿀 수 있을 만큼 근거가 충분하지 않으면 바꾸지 말고 다음 확인에 `근거 부족`을 남긴다.\n"
            "AI 판정은 최종 확정이 아니라 교사가 검토할 초안이다.\n"
            "A 수준 문항은 A 수준 최소능력자 3명 중 약 2명이 해결할 수 있다는 기준으로 본다.\n"
            "목표수준 후보가 X라면 X 예상은 원칙적으로 2/3으로 둔다. X보다 높은 수준은 2/3~3/3, "
            "X보다 낮은 수준은 0/3~2/3 범위에서 문항 난이도에 맞춰 조정한다.\n"
            "수행평가는 평가요소별로 A~E 최소능력자의 예상점수 산정에 도움이 되도록 근거를 쓴다.\n\n"
            "반드시 JSON 배열만 출력하라. 설명 문장, 마크다운, 코드블록을 붙이지 마라.\n"
            "각 객체의 키는 반드시 다음 13개만 사용한다:\n"
            '["구분","번호/요소","성취기준 후보","평가유형","목표수준 후보","난이도 후보",'
            '"A 예상","B 예상","C 예상","D 예상","E 예상","근거","다음 확인"]\n'
            "평가유형은 선택형, 서답형, 수행평가 중 하나를 우선 사용한다.\n"
            "목표수준 후보는 A/B/C/D/E 중 하나를 사용한다.\n"
            "난이도 후보는 쉬움/보통/어려움 중 하나를 사용한다.\n\n"
            "지필 문항의 A~E 예상은 각 수준 대표학생 3명 중 몇 명이 맞힐지 3/3, 2/3, 1/3, 0/3으로 쓴다.\n"
            "수행평가의 A~E 예상은 만점 대비 점수나 비율로 쓴다. 예: 9점, 80%, 7.5/10.\n"
            "A에서 E로 갈수록 예상값은 같거나 낮아야 한다.\n\n"
            "근거 열에는 가능한 한 `문항: ... / 기준: ... / 수준: ...` 형태로 쓴다. "
            "문항 단서, 성취기준 코드 또는 진술, A~E 성취수준 설명 또는 이전시험 정답률 중 최소 2가지를 포함하라.\n"
            "성취기준 후보에 코드가 있으면 근거 열에도 같은 코드나 해당 성취기준 핵심어를 포함한다.\n"
            "로컬 초안의 근거에 `이전시험 ... 성취수준별 정답률`이 있으면 근거 열에 그 핵심을 유지하고, "
            "근거가 부족한 문항은 로컬 초안의 목표수준과 A~E 예상값을 유지하고 다음 확인에 기본값 추정임을 남긴다.\n"
            "근거와 다음 확인은 각각 140자 이내로 간결하게 쓴다.\n\n"
            "[로컬 초안]\n"
            f"{draft}\n\n"
            "[시험 문제 자료]\n"
            f"{preview}\n\n"
            "[성취기준·성취수준 참고자료]\n"
            f"{reference_preview or '(참고자료 없음)'}"
        )

    @staticmethod
    def _ai_review_source_for_blocks(blocks: list[dict], limit_per_block: int = 2400) -> str:
        parts = []
        for block in blocks:
            label = str(block.get("label", "")).strip()
            kind = str(block.get("kind", "문항")).strip()
            text = str(block.get("text", "")).strip()
            if len(text) > limit_per_block:
                text = text[:limit_per_block] + "\n...(문항 일부 생략)"
            parts.append(f"### {kind} {label}\n{text}")
        return "\n\n".join(parts)

    def _ai_reference_text_for_review_chunk(
        self,
        rows: list[dict],
        reference_entries: list[dict[str, str | set[str]]],
        fallback_reference_text: str,
        source_text: str = "",
    ) -> str:
        if not reference_entries:
            return fallback_reference_text[:AI_REVIEW_CHUNK_REFERENCE_LIMIT]
        row_text = " ".join(
            " ".join(str(row.get(key, "")) for key in ("label", "번호/요소", "standard", "성취기준 후보", "evidence", "근거"))
            for row in rows
        )
        query_text = f"{row_text}\n{source_text or ''}"
        query_tokens = self._ai_reference_tokens(query_text)
        query_codes = set(AI_STANDARD_CODE_RE.findall(query_text))
        scored: list[tuple[int, int, str]] = []
        for idx, entry in enumerate(reference_entries):
            text = str(entry.get("text", ""))
            code = str(entry.get("code", ""))
            tokens = entry.get("tokens", set())
            token_overlap = len(query_tokens & tokens) if isinstance(tokens, set) else 0
            score = token_overlap
            if code and code in query_codes:
                score += 30
            elif code and code in query_text:
                score += 24
            entry_codes = set(AI_STANDARD_CODE_RE.findall(text))
            if query_codes & entry_codes:
                score += 18
            if any(term in text for term in ("성취수준", "평가기준", "최소능력자", "A:", "B:", "C:", "D:", "E:")):
                score += 3
            if "예시 답안" in text or "피드백 방안" in text:
                score -= 8
            if score >= 2:
                scored.append((score, idx, text))
        scored.sort(key=lambda item: (-item[0], item[1]))
        selected = [text for _score, _idx, text in scored[:10]]
        if not selected:
            selected = [str(entry.get("text", "")) for entry in reference_entries[:8]]
        if fallback_reference_text and len("\n\n".join(selected)) < 2500:
            selected.append(fallback_reference_text[:2500])
        joined = "\n\n".join(part for part in selected if part)
        return joined[:AI_REVIEW_CHUNK_REFERENCE_LIMIT] or fallback_reference_text[:AI_REVIEW_CHUNK_REFERENCE_LIMIT]

    @staticmethod
    def _ai_review_type_rank(row: dict) -> int:
        raw_type = str(
            row.get("review_type")
            or row.get("평가유형")
            or row.get("section_type")
            or row.get("kind")
            or row.get("구분")
            or ""
        )
        if "수행" in raw_type:
            return 2
        if "서답" in raw_type or "서술" in raw_type or "논술" in raw_type:
            return 1
        if "선택" in raw_type or "객관" in raw_type or "지필" in raw_type or "문항" in raw_type:
            return 0
        return 9

    @staticmethod
    def _ai_review_row_number(row: dict, fallback_index: int = 0) -> int:
        label = str(row.get("label") or row.get("번호/요소") or row.get("번호") or "").strip()
        number = re.search(r"\d+", label)
        if number:
            return int(number.group(0))
        return 10_000 + fallback_index

    @classmethod
    def _ai_review_row_sort_key(cls, row: dict, fallback_index: int = 0) -> tuple[int, int, int]:
        return (
            cls._ai_review_type_rank(row),
            cls._ai_review_row_number(row, fallback_index),
            fallback_index,
        )

    @classmethod
    def _sort_ai_review_rows(cls, rows: list[dict]) -> list[dict]:
        return [
            row
            for _idx, row in sorted(
                enumerate(rows),
                key=lambda pair: cls._ai_review_row_sort_key(pair[1], pair[0]),
            )
        ]

    @classmethod
    def _ai_review_row_key(cls, row: dict, fallback_index: int = 0) -> str:
        number = cls._ai_review_row_number(row, fallback_index)
        if number < 10_000:
            return f"{cls._ai_review_type_rank(row)}:{number}"
        return f"idx:{fallback_index}"

    def _ai_review_merge_chunk_rows(self, local_rows: list[dict], ai_rows: list[dict]) -> list[dict]:
        if len(ai_rows) == len(local_rows) and all(
            self._ai_review_row_key(ai_row, idx) == self._ai_review_row_key(local_row, idx)
            for idx, (ai_row, local_row) in enumerate(zip(ai_rows, local_rows))
        ):
            return [
                self._normalize_ai_review_output_row(ai_row, local_row)
                for ai_row, local_row in zip(ai_rows, local_rows)
            ]
        by_key = {
            self._ai_review_row_key(row, idx): row
            for idx, row in enumerate(ai_rows)
        }
        by_number: dict[int, list[dict]] = {}
        for idx, row in enumerate(ai_rows):
            number = self._ai_review_row_number(row, idx)
            if number < 10_000:
                by_number.setdefault(number, []).append(row)
        merged = []
        used_ids = set()
        for idx, local in enumerate(local_rows):
            key = self._ai_review_row_key(local, idx)
            candidate = by_key.get(key)
            if candidate is None:
                number_matches = [
                    row for row in by_number.get(self._ai_review_row_number(local, idx), [])
                    if id(row) not in used_ids
                ]
                if len(number_matches) == 1:
                    candidate = number_matches[0]
            if candidate is not None:
                merged.append(self._normalize_ai_review_output_row(candidate, local))
                used_ids.add(id(candidate))
            else:
                fallback = dict(local)
                next_step = fallback.get("next_step") or fallback.get("다음 확인") or ""
                fallback["next_step"] = (next_step + " / AI 보강 항목 확인").strip(" /")
                merged.append(fallback)
        for row in ai_rows:
            if id(row) not in used_ids:
                merged.append(self._normalize_ai_review_output_row(row, {}))
        return merged

    @staticmethod
    def _ai_review_standard_has_code(value: str) -> bool:
        return bool(re.search(r"\[[^\]\n]{4,40}\]", value or ""))

    @staticmethod
    def _ai_review_evidence_is_weak(value: str) -> bool:
        text = re.sub(r"\s+", " ", value or "").strip()
        if len(text) < 10:
            return True
        if AI_STANDARD_CODE_RE.search(text):
            return False
        strong_terms = (
            "이전시험", "성취수준별 정답률", "배점", "문항정보표", "[10",
            "다항식", "인수분해", "나머지정리", "복소수", "실수부", "허수부",
            "이차방정식", "근과 계수", "이차함수", "그래프", "최대", "최소",
            "경우의 수", "순열", "조합", "행렬", "보기", "서답형",
            "문항:", "기준:", "수준:", "최소능력자", "2/3",
        )
        if "기본 원리" in text and len(text) < 28 and "[10" not in text:
            return True
        return not any(term in text for term in strong_terms)

    @staticmethod
    def _ai_review_standard_codes(value: str) -> set[str]:
        return set(AI_STANDARD_CODE_RE.findall(value or ""))

    def _ai_review_evidence_supports_standard_change(
        self,
        evidence: str,
        ai_standard: str,
        local_standard: str,
    ) -> bool:
        if self._ai_review_evidence_is_weak(evidence):
            return False
        ai_codes = self._ai_review_standard_codes(ai_standard)
        local_codes = self._ai_review_standard_codes(local_standard)
        evidence_codes = self._ai_review_standard_codes(evidence)
        if ai_codes:
            if ai_codes & evidence_codes:
                return True
            if not (ai_codes & local_codes):
                return False
        standard_tokens = self._ai_reference_tokens(ai_standard)
        evidence_tokens = self._ai_reference_tokens(evidence)
        return len(standard_tokens & evidence_tokens) >= 2

    def _ai_review_evidence_supports_target_change(self, evidence: str) -> bool:
        if self._ai_review_evidence_is_weak(evidence):
            return False
        signals = 0
        for group in (
            ("성취수준", "수준:", "최소능력자", "2/3", "A 수준", "B 수준", "C 수준", "D 수준", "E 수준"),
            ("이전시험", "정답률", "성취수준별 정답률"),
            ("배점", "고난도", "어려움", "추론", "복합", "모델링", "일반화", "증명", "해석"),
            ("문항:", "기준:", "보기", "서답형"),
        ):
            if any(term in evidence for term in group):
                signals += 1
        return signals >= 2

    @staticmethod
    def _append_ai_review_note(existing: str, note: str) -> str:
        existing = str(existing or "").strip()
        note = str(note or "").strip()
        if not note or note in existing:
            return existing
        return (existing + " / " + note).strip(" /") if existing else note

    def _normalize_ai_review_output_row(self, ai_row: dict, local_row: dict | None = None) -> dict:
        row = dict(ai_row or {})
        local_row = local_row or {}
        if not (row.get("구분") or row.get("kind")) and local_row:
            row["구분"] = local_row.get("kind") or local_row.get("구분") or "문항"
        if not (row.get("번호/요소") or row.get("label")) and local_row:
            row["번호/요소"] = local_row.get("label") or local_row.get("번호/요소") or ""
        raw_type = row.get("평가유형") or row.get("review_type") or ""
        local_type = local_row.get("review_type") or local_row.get("평가유형") or ""
        if "수행" in str(raw_type):
            review_type = "수행평가"
        elif "서답" in str(raw_type) or "서술" in str(raw_type) or "논술" in str(raw_type):
            review_type = "서답형"
        elif "선택" in str(raw_type) or "객관" in str(raw_type):
            review_type = "선택형"
        elif local_type in {"선택형", "서답형", "수행평가"}:
            review_type = local_type
        else:
            review_type = "선택형"
        row["평가유형"] = review_type
        ai_standard = str(row.get("성취기준 후보") or row.get("standard") or "").strip()
        local_standard = str(local_row.get("standard") or local_row.get("성취기준 후보") or "").strip()
        ai_evidence = str(row.get("근거") or row.get("evidence") or "").strip()
        local_evidence = str(local_row.get("evidence") or local_row.get("근거") or "").strip()
        local_next_step = str(local_row.get("next_step") or local_row.get("다음 확인") or "").strip()
        valid_local_standard = bool(local_standard and local_standard != "(후보 없음)")
        standard_changed = (
            valid_local_standard
            and ai_standard
            and ai_standard != "(후보 없음)"
            and ai_standard != local_standard
        )
        standard_locked_to_local = False
        if standard_changed and not self._ai_review_evidence_supports_standard_change(ai_evidence, ai_standard, local_standard):
            row["성취기준 후보"] = local_standard
            row["다음 확인"] = self._append_ai_review_note(row.get("다음 확인") or row.get("next_step") or "", "AI 성취기준 근거 확인")
            standard_locked_to_local = True
        if not standard_locked_to_local and local_standard and (
            not ai_standard
            or ai_standard == "(후보 없음)"
            or (
                self._ai_review_standard_has_code(local_standard)
                and not self._ai_review_standard_has_code(ai_standard)
            )
        ):
            row["성취기준 후보"] = local_standard
        elif not standard_locked_to_local and ai_standard:
            row["성취기준 후보"] = ai_standard
        raw_target = row.get("목표수준 후보") or row.get("target") or ""
        target = self._ai_review_level(raw_target, "")
        local_target = self._ai_review_level(local_row.get("target") or local_row.get("목표수준 후보") or "", "")
        weak_evidence = self._ai_review_evidence_is_weak(ai_evidence)
        if not target:
            target = local_target or "C"
        elif local_target and target != local_target and (
            weak_evidence
            or not self._ai_review_evidence_supports_target_change(ai_evidence)
            or "이전시험" in local_evidence
            or "성취수준별 정답률" in local_evidence
        ):
            row["다음 확인"] = self._append_ai_review_note(row.get("다음 확인") or row.get("next_step") or "", "AI 목표수준 근거 확인")
            target = local_target
        raw_difficulty = str(row.get("난이도 후보") or row.get("difficulty") or "")
        if any(term in raw_difficulty for term in ("쉬", "보통", "중", "어려", "상", "하")):
            difficulty = self._ai_review_difficulty(raw_difficulty)
        else:
            difficulty = self._ai_review_difficulty(local_row.get("difficulty") or local_row.get("난이도 후보") or "보통")
        local_difficulty_raw = local_row.get("difficulty") or local_row.get("난이도 후보") or ""
        local_difficulty = self._ai_review_difficulty(local_difficulty_raw) if local_difficulty_raw else ""
        if local_difficulty and difficulty != local_difficulty and weak_evidence:
            difficulty = local_difficulty
        row["목표수준 후보"] = target
        row["난이도 후보"] = difficulty
        if weak_evidence and local_evidence:
            row["근거"] = local_evidence
            row["다음 확인"] = self._append_ai_review_note(row.get("다음 확인") or row.get("next_step") or "", "AI 근거 보강 필요")
        elif ai_evidence:
            row["근거"] = ai_evidence
        elif local_evidence:
            row["근거"] = local_evidence
        if local_next_step and not (row.get("다음 확인") or row.get("next_step")):
            row["다음 확인"] = local_next_step
        if review_type == "수행평가":
            return row
        counts = {}
        valid_count = 0
        for lv in LEVELS_AE:
            count, denominator = self._parse_expected_count(str(row.get(f"{lv} 예상", "")), 3)
            if count is not None and denominator > 0:
                counts[lv] = max(0, min(3, int(round(count / denominator * 3))))
                valid_count += 1
        if valid_count < len(LEVELS_AE):
            defaults = self._ai_default_ox_expected_values(target, difficulty)
            for lv in LEVELS_AE:
                count, _ = self._parse_expected_count(defaults.get(f"{lv} 예상", ""), 3)
                counts[lv] = 0 if count is None else count
        target_index = LEVELS_AE.index(target)
        threshold = ai_review_target_threshold(3)
        counts[target] = threshold
        for idx, lv in enumerate(LEVELS_AE):
            if idx < target_index:
                counts[lv] = max(counts.get(lv, 0), threshold)
            elif idx > target_index:
                counts[lv] = min(counts.get(lv, 0), threshold)
        counts = ai_review_normalize_counts(counts, 3)
        for lv in LEVELS_AE:
            value = counts[lv]
            row[f"{lv} 예상"] = f"{value}/3"
        return row

    @staticmethod
    def _ai_review_failed_chunk_rows(local_rows: list[dict], reason: str) -> list[dict]:
        failed = []
        short_reason = str(reason).splitlines()[0][:90]
        for row in local_rows:
            fallback = dict(row)
            next_step = fallback.get("next_step") or fallback.get("다음 확인") or ""
            fallback["next_step"] = (next_step + f" / AI 보강 실패: {short_reason}").strip(" /")
            failed.append(fallback)
        return failed

    @staticmethod
    def _ai_review_signature(source_text: str, reference_text: str) -> str:
        digest = hashlib.sha256()
        digest.update((source_text or "").encode("utf-8", errors="replace"))
        digest.update(b"\0")
        digest.update((reference_text or "").encode("utf-8", errors="replace"))
        return digest.hexdigest()

    def _set_ai_review_summary(self, rows: int, standards: int, warnings: int, mode: str = "로컬 초안"):
        if not hasattr(self, "ai_review_cards"):
            return
        self.ai_review_cards["rows"].value_label.setText(str(rows))
        self.ai_review_cards["standards"].value_label.setText(str(standards))
        self.ai_review_cards["warnings"].value_label.setText(str(warnings))
        self.ai_review_cards["mode"].value_label.setText(mode)

    def _populate_ai_review_table(self, rows: list[dict], *, mode: str = "로컬 초안"):
        table = self.table_ai_review
        table.setSortingEnabled(False)
        table.setRowCount(len(rows))
        warnings = 0
        standards = set()
        for r, row in enumerate(rows):
            standard = row.get("standard") or row.get("성취기준 후보") or "(후보 없음)"
            next_step = row.get("next_step") or row.get("다음 확인") or ""
            if standard != "(후보 없음)":
                standards.add(standard.split("]")[0] + "]" if "]" in standard else standard)
            if "확인" in next_step or "점검" in next_step:
                warnings += 1
            values = [
                row.get("kind") or row.get("구분") or "",
                row.get("label") or row.get("번호/요소") or "",
                standard,
                row.get("review_type") or row.get("평가유형") or "",
                row.get("target") or row.get("목표수준 후보") or "",
                row.get("difficulty") or row.get("난이도 후보") or "",
                *[
                    row.get(f"{lv} 예상") or row.get(lv) or ""
                    for lv in LEVELS_AE
                ],
                row.get("evidence") or row.get("근거") or "",
                next_step,
            ]
            for c, value in enumerate(values):
                bg = None
                if c == len(values) - 1 and ("확인" in value or "점검" in value):
                    bg = QColor("#f8dfa3"); bg.setAlpha(140)
                _set_item(table, r, c, value, align_left=c in (1, 2, 11, 12), bg=bg, tooltip=value)
        table.setSortingEnabled(False)
        for col, width in enumerate([86, 120, 260, 92, 100, 100, 76, 76, 76, 76, 76, 260, 240]):
            self._set_scaled_column_width(table, col, width)
        self._set_ai_review_summary(len(rows), len(standards), warnings, mode)

    def _generate_ai_review_draft(self):
        if not hasattr(self, "txt_ai_review_source"):
            return
        text = self._ai_review_source_text()
        reference_text = self._ai_review_reference_text()
        if not text:
            QMessageBox.information(self, "AI 문항 검토", "먼저 문항 자료를 붙여 넣거나 시험 문제 PDF를 불러와 주세요.")
            return
        reference_entries = self._ai_reference_entries(reference_text)
        blocks = self._split_ai_review_blocks(text)
        rows = [self._infer_ai_review_row(block, reference_entries) for block in blocks]
        rows, supplemented = self._supplement_ai_review_rows_from_exam(rows)
        self._populate_ai_review_table(rows, mode="로컬 초안")
        self.txt_ai_review_prompt.setPlainText(self._make_ai_review_prompt(rows, text, reference_text))
        suffix = f" · 문항정보표 보충 {supplemented}개" if supplemented else ""
        self.statusBar().showMessage(f"AI 문항 검토 초안 생성 완료 · {len(rows)}개 항목{suffix}", 3500)

    def _run_ai_review_completion(self, resume_state: dict | None = None):
        if not hasattr(self, "txt_ai_review_source"):
            return
        text = self._ai_review_source_text()
        reference_text = self._ai_review_reference_text()
        if not text:
            QMessageBox.information(self, "AI 문항 검토", "먼저 문항 자료를 붙여 넣거나 시험 문제 PDF를 불러와 주세요.")
            return
        source_signature = self._ai_review_signature(text, reference_text)
        if resume_state and resume_state.get("signature") != source_signature:
            QMessageBox.warning(
                self,
                "AI 문항 검토",
                "문항 자료나 성취기준·수준 자료가 바뀌어 이어하기를 할 수 없습니다.\n처음부터 다시 AI로 보강해 주세요.",
            )
            self._ai_review_resume_state = None
            self._set_ai_review_control_state(running=False, can_resume=False)
            return
        config = self._ai_provider_config()
        self._save_ai_settings()
        reference_entries = self._ai_reference_entries(reference_text)
        blocks = self._split_ai_review_blocks(text)
        local_rows = [self._infer_ai_review_row(block, reference_entries) for block in blocks]
        local_rows, supplemented = self._supplement_ai_review_rows_from_exam(local_rows)
        resume_state = resume_state or None
        start_index = int(resume_state.get("next_index", 0)) if resume_state else 0
        seeded_rows = list(resume_state.get("merged_rows", [])) if resume_state else []
        seeded_outputs = list(resume_state.get("outputs", [])) if resume_state else []
        seeded_prompts = list(resume_state.get("prompts", [])) if resume_state else []
        seeded_failures = list(resume_state.get("failures", [])) if resume_state else []
        if resume_state:
            display_rows = seeded_rows + local_rows[start_index:]
            self._populate_ai_review_table(display_rows, mode="AI 보강 이어하기 준비")
        else:
            self._populate_ai_review_table(local_rows, mode="로컬 초안")
            self._ai_review_resume_state = None
            self._set_ai_review_control_state(running=False, can_resume=False)
        if config.provider == "local_draft":
            self.txt_ai_review_prompt.setPlainText(self._make_ai_review_prompt(local_rows, text, reference_text))
            suffix = f" · 문항정보표 보충 {supplemented}개" if supplemented else ""
            self.statusBar().showMessage(f"로컬 초안으로 검토표를 갱신했습니다{suffix}.", 3500)
            return

        preview_prompt = self._make_structured_ai_review_prompt(
            local_rows[: min(5, len(local_rows))],
            self._ai_review_source_for_blocks(blocks[: min(5, len(blocks))]),
            reference_text,
            source_limit=AI_REVIEW_CHUNK_SOURCE_LIMIT,
            reference_limit=AI_REVIEW_CHUNK_REFERENCE_LIMIT,
            row_limit=5,
            scope_note="미리보기용 일부 항목",
        )
        if hasattr(self, "chk_ai_scrub") and self.chk_ai_scrub.isChecked():
            preview_prompt = scrub_personal_data(preview_prompt, self._student_names_for_privacy())
        self.txt_ai_review_prompt.setPlainText(
            preview_prompt
            + "\n\n※ 실제 AI 보강은 시간초과를 줄이기 위해 문항을 작은 묶음으로 나누어 전송합니다."
        )
        endpoints = self._candidate_mlx_endpoints()
        scrub_enabled = hasattr(self, "chk_ai_scrub") and self.chk_ai_scrub.isChecked()
        student_names = self._student_names_for_privacy()
        cancel_event = threading.Event()
        self._ai_review_cancel_event = cancel_event
        self._set_ai_review_control_state(running=True, can_resume=False)

        def work(progress):
            prepared, note = self._prepare_ai_connection_config_for_worker(config, endpoints, progress)
            if prepared.provider in {"mlx_compatible", "ollama"} and prepared.timeout < 240:
                prepared.timeout = 240
                progress("문항 검토는 묶음별 요청으로 나누고, 대기 시간을 이번 요청에 한해 240초로 적용합니다.")
            if prepared.provider == "codex_cli" and prepared.timeout < 240:
                prepared.timeout = 240
                progress("Codex CLI 문항 검토는 OAuth 클라우드 응답 대기를 이번 요청에 한해 240초로 적용합니다.")
            size = self._model_size_hint(prepared.model)
            if 0 < size < 3:
                progress("현재 선택 모델은 빠른 확인용에 가까워 검토 품질이 낮을 수 있습니다. 품질이 중요하면 7B~15B급 모델을 선택해 보세요.")
            elif size >= 20:
                progress("현재 선택 모델은 큰 모델이라 품질은 좋아질 수 있지만 응답이 오래 걸릴 수 있습니다.")
            chunk_size = AI_REVIEW_CLOUD_CHUNK_SIZE
            if prepared.provider in {"mlx_compatible", "ollama"}:
                chunk_size = AI_REVIEW_LOCAL_CHUNK_SIZE
                if size >= 20:
                    chunk_size = 2
            chunks = [
                (idx, local_rows[idx: idx + chunk_size], blocks[idx: idx + chunk_size])
                for idx in range(start_index, len(local_rows), chunk_size)
            ]
            reference_entries_for_chunks = reference_entries
            progress(
                f"{prepared.label}에 문항 검토 요청 전송 중 · "
                f"{len(local_rows)}개 항목 중 {start_index + 1}번째부터 {len(chunks)}개 묶음으로 처리합니다."
            )
            merged_rows = list(seeded_rows)
            outputs = list(seeded_outputs)
            sent_prompts = list(seeded_prompts)
            failures = list(seeded_failures)
            for chunk_no, (start, chunk_rows, chunk_blocks) in enumerate(chunks, start=1):
                if cancel_event.is_set():
                    progress(f"중지 요청 반영 · {start + 1}번째 항목부터 이어하기 가능")
                    return {
                        "config": prepared,
                        "note": note,
                        "output": "\n\n".join(outputs),
                        "prompts": "\n\n".join(sent_prompts),
                        "rows": merged_rows + local_rows[start:],
                        "merged_rows": merged_rows,
                        "failures": failures,
                        "chunk_count": len(chunks),
                        "canceled": True,
                        "next_index": start,
                        "signature": source_signature,
                        "outputs_list": outputs,
                        "prompts_list": sent_prompts,
                    }
                end = start + len(chunk_rows)
                source_chunk = self._ai_review_source_for_blocks(chunk_blocks)
                reference_chunk = self._ai_reference_text_for_review_chunk(
                    chunk_rows,
                    reference_entries_for_chunks,
                    reference_text,
                    source_chunk,
                )
                scope = f"{start + 1}-{end}번째 항목"
                chunk_prompt = self._make_structured_ai_review_prompt(
                    chunk_rows,
                    source_chunk,
                    reference_chunk,
                    source_limit=AI_REVIEW_CHUNK_SOURCE_LIMIT,
                    reference_limit=AI_REVIEW_CHUNK_REFERENCE_LIMIT,
                    row_limit=len(chunk_rows),
                    scope_note=scope,
                )
                if scrub_enabled:
                    chunk_prompt = scrub_personal_data(chunk_prompt, student_names)
                sent_prompts.append(f"===== 묶음 {chunk_no}/{len(chunks)} · {scope} =====\n{chunk_prompt}")
                progress(f"묶음 {chunk_no}/{len(chunks)} 전송 중 · {scope}")
                chunk_error = ""
                try:
                    output = run_completion(
                        chunk_prompt,
                        prepared,
                        max_tokens=max(1200, min(2600, len(chunk_rows) * 750)),
                    )
                except Exception as exc:
                    chunk_error = str(exc)
                    progress(f"묶음 {chunk_no}/{len(chunks)} 실패 · 1문항씩 재시도 · {exc}")
                    outputs.append(f"===== 묶음 {chunk_no} 실패 =====\n{exc}")
                    ai_rows = []
                else:
                    outputs.append(f"===== 묶음 {chunk_no}/{len(chunks)} 응답 =====\n{output}")
                    ai_rows = parse_review_rows(output)
                if not ai_rows or len(ai_rows) != len(chunk_rows):
                    if chunk_error:
                        progress(f"묶음 {chunk_no}/{len(chunks)} 실패 후 1문항씩 재시도")
                    elif ai_rows:
                        progress(
                            f"묶음 {chunk_no}/{len(chunks)} 응답 수 불일치 "
                            f"({len(ai_rows)}/{len(chunk_rows)}) · 1문항씩 재시도"
                        )
                    else:
                        progress(f"묶음 {chunk_no}/{len(chunks)} 응답 해석 실패 · 1문항씩 재시도")
                    single_rows = []
                    single_failures = []
                    for offset, (single_row, single_block) in enumerate(zip(chunk_rows, chunk_blocks), start=0):
                        if cancel_event.is_set():
                            next_index = start + offset
                            progress(f"중지 요청 반영 · {next_index + 1}번째 항목부터 이어하기 가능")
                            merged_rows.extend(single_rows)
                            return {
                                "config": prepared,
                                "note": note,
                                "output": "\n\n".join(outputs),
                                "prompts": "\n\n".join(sent_prompts),
                                "rows": merged_rows + local_rows[next_index:],
                                "merged_rows": merged_rows,
                                "failures": failures + single_failures,
                                "chunk_count": len(chunks),
                                "canceled": True,
                                "next_index": next_index,
                                "signature": source_signature,
                                "outputs_list": outputs,
                                "prompts_list": sent_prompts,
                            }
                        single_scope = f"{start + offset + 1}번째 항목"
                        single_prompt = self._make_structured_ai_review_prompt(
                            [single_row],
                            self._ai_review_source_for_blocks([single_block], limit_per_block=1800),
                            self._ai_reference_text_for_review_chunk(
                                [single_row],
                                reference_entries_for_chunks,
                                reference_text,
                                self._ai_review_source_for_blocks([single_block], limit_per_block=1800),
                            ),
                            source_limit=3600,
                            reference_limit=5000,
                            row_limit=1,
                            scope_note=single_scope,
                        )
                        if scrub_enabled:
                            single_prompt = scrub_personal_data(single_prompt, student_names)
                        sent_prompts.append(f"===== 재시도 · {single_scope} =====\n{single_prompt}")
                        progress(f"재시도 전송 중 · {single_scope}")
                        try:
                            single_output = run_completion(single_prompt, prepared, max_tokens=1000)
                        except Exception as exc:
                            message = f"재시도 실패 · {single_scope} · {exc}"
                            progress(message)
                            single_failures.append(message)
                            outputs.append(f"===== 재시도 실패 · {single_scope} =====\n{exc}")
                            single_rows.extend(self._ai_review_failed_chunk_rows([single_row], str(exc)))
                            continue
                        outputs.append(f"===== 재시도 응답 · {single_scope} =====\n{single_output}")
                        parsed_single = parse_review_rows(single_output)
                        if parsed_single:
                            single_rows.extend(self._ai_review_merge_chunk_rows([single_row], parsed_single))
                            progress(f"재시도 완료 · {single_scope}")
                        else:
                            message = f"재시도 응답 해석 실패 · {single_scope}"
                            progress(message)
                            single_failures.append(message)
                            single_rows.extend(self._ai_review_failed_chunk_rows([single_row], "응답 해석 실패"))
                    merged_rows.extend(single_rows)
                    failures.extend(single_failures)
                    if single_failures:
                        failures.append(f"묶음 {chunk_no}/{len(chunks)} 일부 재시도 실패")
                    if chunk_error and not single_failures:
                        failures.append(f"묶음 {chunk_no}/{len(chunks)}는 실패했지만 1문항 재시도는 완료")
                    continue
                merged_rows.extend(self._ai_review_merge_chunk_rows(chunk_rows, ai_rows))
                progress(f"묶음 {chunk_no}/{len(chunks)} 완료 · {len(ai_rows)}개 응답")
            progress("AI 응답 수신, 검토표로 변환 중")
            return {
                "config": prepared,
                "note": note,
                "output": "\n\n".join(outputs),
                "prompts": "\n\n".join(sent_prompts),
                "rows": merged_rows,
                "merged_rows": merged_rows,
                "failures": failures,
                "chunk_count": len(chunks),
                "canceled": False,
                "next_index": len(local_rows),
                "signature": source_signature,
                "outputs_list": outputs,
                "prompts_list": sent_prompts,
            }

        def success(result):
            prepared = result["config"]
            note = result.get("note", "")
            output = result.get("output") or ""
            prompts = result.get("prompts") or ""
            ai_rows = result.get("rows") or []
            failures = result.get("failures") or []
            canceled = bool(result.get("canceled"))
            self._apply_prepared_ai_config(prepared, note)
            if not ai_rows:
                self.txt_ai_review_prompt.setPlainText(
                    prompts[:30000] + "\n\n[AI 원문 출력]\n" + (output or "(빈 응답)")[:20000]
                )
                self._ai_review_resume_state = None
                self._set_ai_review_control_state(running=False, can_resume=False)
                self._ai_review_cancel_event = None
                QMessageBox.information(
                    self,
                    "AI 문항 검토",
                    "AI 응답은 받았지만 검토표로 해석하지 못했습니다.\nAI 프롬프트 탭의 원문 출력을 확인해 주세요.",
                )
                return
            if canceled:
                mode = f"{prepared.label} 보강 중지"
            else:
                mode = f"{prepared.label} 보강" if not failures else f"{prepared.label} 일부 보강"
            self._populate_ai_review_table(ai_rows, mode=mode)
            self.txt_ai_review_prompt.setPlainText(
                prompts[:30000] + "\n\n[AI 원문 출력]\n" + output[:30000]
            )
            self.ai_review_tabs.setCurrentWidget(self.table_ai_review)
            if canceled:
                self._ai_review_resume_state = {
                    "signature": result.get("signature"),
                    "next_index": int(result.get("next_index", 0)),
                    "merged_rows": list(result.get("merged_rows") or []),
                    "outputs": list(result.get("outputs_list") or []),
                    "prompts": list(result.get("prompts_list") or []),
                    "failures": list(failures),
                }
                self._set_ai_review_control_state(running=False, can_resume=True)
                done_message = (
                    f"AI 보강 중지 · {result.get('next_index', 0) + 1}번째 항목부터 이어하기 가능"
                )
            else:
                self._ai_review_resume_state = None
                self._set_ai_review_control_state(running=False, can_resume=False)
                done_message = f"AI 보강 완료 · {len(ai_rows)}개 항목 · 묶음 {result.get('chunk_count', 1)}개"
            if failures:
                done_message += f" · 일부 로컬 초안 유지 {len(failures)}개"
            if note:
                done_message += f" · {note}"
            self._append_ai_progress(done_message)
            self.ai_review_tabs.setCurrentWidget(self.table_ai_review)
            self._ai_review_cancel_event = None
            if failures and not canceled:
                QMessageBox.information(
                    self,
                    "AI 문항 검토",
                    "일부 묶음은 시간초과 또는 응답 해석 실패로 로컬 초안을 유지했습니다.\n"
                    "표의 '다음 확인' 열과 진행 로그를 확인해 주세요.",
                )

        self._run_ai_background_task("AI 문항 검토", work, success)

    def _export_ai_review_csv(self):
        if not hasattr(self, "table_ai_review") or self.table_ai_review.rowCount() == 0:
            QMessageBox.information(self, "AI 문항 검토", "먼저 검토 초안을 생성해 주세요.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "AI 문항 검토 CSV 저장", "ai_item_review.csv", "CSV (*.csv)"
        )
        if not path:
            return
        if not path.lower().endswith(".csv"):
            path += ".csv"
        try:
            with open(path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(AI_REVIEW_HEADERS)
                for r in range(self.table_ai_review.rowCount()):
                    writer.writerow([
                        self.table_ai_review.item(r, c).text() if self.table_ai_review.item(r, c) else ""
                        for c in range(self.table_ai_review.columnCount())
                    ])
        except Exception as e:
            QMessageBox.critical(self, "AI 문항 검토", f"CSV 저장 중 오류가 발생했습니다.\n{e}")
            return
        self.statusBar().showMessage(f"AI 문항 검토 CSV 저장 완료 · {path}", 5000)

    def _ai_review_table_rows(self) -> list[dict]:
        if not hasattr(self, "table_ai_review") or self.table_ai_review.rowCount() == 0:
            return []
        rows = []
        for r in range(self.table_ai_review.rowCount()):
            values = [
                self.table_ai_review.item(r, c).text().strip() if self.table_ai_review.item(r, c) else ""
                for c in range(min(self.table_ai_review.columnCount(), len(AI_REVIEW_HEADERS)))
            ]
            if not any(values):
                continue
            rows.append(dict(zip(AI_REVIEW_HEADERS, values)))
        return rows

    @staticmethod
    def _ai_review_number(label: str, fallback: int) -> int:
        m = re.search(r"(\d{1,3})", label or "")
        if not m:
            return fallback
        try:
            return max(1, int(m.group(1)))
        except Exception:
            return fallback

    @staticmethod
    def _ai_review_level(value: str, fallback: str = "C") -> str:
        value = (value or "").strip().upper()
        return value if value in LEVELS_AE else fallback

    @staticmethod
    def _ai_review_difficulty(value: str) -> str:
        text = value or ""
        if "어려" in text or re.search(r"\b상\b", text):
            return "어려움"
        if "쉬" in text or re.search(r"\b하\b", text):
            return "쉬움"
        if "보통" in text or re.search(r"\b중\b", text):
            return "보통"
        return "보통"

    @staticmethod
    def _ai_review_expected_type(value: str) -> str:
        text = value or ""
        if "서답" in text or "서술" in text or "논술" in text:
            return "서답형"
        return "선택형"

    def _ai_review_points_for_item(self, number: int, item_type: str, fallback: float = 5.0) -> float:
        if self.exam is None:
            return fallback
        candidates = [
            item for item in self.exam.items
            if int(item.number) == int(number) and item.item_type == item_type
        ]
        if not candidates:
            candidates = [item for item in self.exam.items if int(item.number) == int(number)]
        if not candidates:
            return fallback
        try:
            return float(candidates[0].score) or fallback
        except Exception:
            return fallback

    def _ai_review_points_for_row(self, number: int, item_type: str, row: dict, fallback: float = 5.0) -> float:
        joined = " ".join(
            str(row.get(key, ""))
            for key in ("번호/요소", "근거", "다음 확인", "성취기준 후보")
        )
        patterns = [
            r"배점\s*(\d+(?:\.\d+)?)\s*점",
            r"\[(\d+(?:\.\d+)?)\s*점\]",
            r"(\d+(?:\.\d+)?)\s*점",
        ]
        for pattern in patterns:
            matches = re.findall(pattern, joined)
            if not matches:
                continue
            try:
                value = float(matches[0])
            except Exception:
                continue
            if 0 < value <= 100:
                return value
        return self._ai_review_points_for_item(number, item_type, fallback)

    @staticmethod
    def _parse_expected_count(value: str, default_sample: int = 3) -> tuple[int | None, int]:
        text = (value or "").strip()
        if not text:
            return None, default_sample
        ox_text = text.upper().replace("○", "O").replace("×", "X")
        if any(mark in ox_text for mark in ("O", "X")):
            marks = [ch for ch in ox_text if ch in {"O", "X"}]
            if marks:
                return marks.count("O"), max(1, min(10, len(marks)))
        frac = re.search(r"(\d+(?:\.\d+)?)\s*/\s*(\d+(?:\.\d+)?)", text)
        if frac:
            numerator = float(frac.group(1))
            denominator = max(1, min(10, round(float(frac.group(2)))))
            return max(0, min(denominator, round(numerator))), denominator
        pct = re.search(r"(-?\d+(?:\.\d+)?)\s*%", text)
        if pct:
            rate = max(0.0, min(100.0, float(pct.group(1))))
            return round((rate / 100.0) * default_sample), default_sample
        number = re.search(r"-?\d+(?:\.\d+)?", text)
        if number:
            value_num = float(number.group(0))
            if 0 <= value_num <= 1:
                return round(value_num * default_sample), default_sample
            if 0 <= value_num <= default_sample:
                return round(value_num), default_sample
            if 0 <= value_num <= 100:
                return round((value_num / 100.0) * default_sample), default_sample
        return None, default_sample

    def _ai_expected_counts_for_row(self, row: dict) -> tuple[dict[str, int], int]:
        parsed: dict[str, int] = {}
        sample_size = 3
        for lv in LEVELS_AE:
            count, denominator = self._parse_expected_count(row.get(f"{lv} 예상", ""), sample_size)
            if count is not None:
                sample_size = denominator
                parsed[lv] = max(0, min(sample_size, count))
        if len(parsed) != len(LEVELS_AE):
            fallback = self._ai_default_ox_expected_values(
                row.get("목표수준 후보", "C"),
                row.get("난이도 후보", "보통"),
            )
            for lv in LEVELS_AE:
                if lv not in parsed:
                    count, denominator = self._parse_expected_count(fallback.get(f"{lv} 예상", ""), sample_size)
                    sample_size = denominator
                    parsed[lv] = 0 if count is None else count
        target = self._ai_review_level(row.get("목표수준 후보", "C"), "C")
        parsed = ai_review_normalize_counts(
            parsed,
            sample_size,
            target=target,
            enforce_target=True,
        )
        return parsed, sample_size

    @staticmethod
    def _judgments_from_expected_counts(counts: dict[str, int], sample_size: int) -> dict:
        judgments = {}
        sample_size = max(1, min(10, int(sample_size)))
        for lv in LEVELS_AE:
            count = max(0, min(sample_size, int(counts.get(lv, 0))))
            rate = round((count / sample_size) * 100, 2)
            judgments[lv] = {
                "correct": [idx < count for idx in range(sample_size)],
                "targetRate": rate,
                "overrideRate": None,
            }
        return judgments

    @staticmethod
    def _parse_expected_score(value: str, max_score: float) -> float | None:
        text = (value or "").strip()
        if not text:
            return None
        frac = re.search(r"(\d+(?:\.\d+)?)\s*/\s*(\d+(?:\.\d+)?)", text)
        if frac:
            denominator = float(frac.group(2))
            if denominator > 0:
                return max_score * (float(frac.group(1)) / denominator)
        pct = re.search(r"(-?\d+(?:\.\d+)?)\s*%", text)
        if pct:
            return max_score * (float(pct.group(1)) / 100.0)
        number = re.search(r"-?\d+(?:\.\d+)?", text)
        if not number:
            return None
        value_num = float(number.group(0))
        if 0 <= value_num <= 1:
            return max_score * value_num
        if 0 <= value_num <= max_score:
            return value_num
        if 0 <= value_num <= 100:
            return max_score * (value_num / 100.0)
        return None

    def _ai_review_expected_scores_for_perform_row(self, row: dict, max_score: float) -> dict:
        parsed = {}
        for lv in LEVELS_AE:
            score = self._parse_expected_score(row.get(f"{lv} 예상", ""), max_score)
            if score is not None:
                parsed[lv] = max(0.0, min(float(max_score), score))
        if not parsed:
            return {}
        prev = float(max_score)
        for lv in LEVELS_AE:
            if lv not in parsed:
                parsed[lv] = prev
            parsed[lv] = max(0.0, min(prev, parsed[lv]))
            prev = parsed[lv]
        return parsed

    def _build_ai_review_spliter_project(self) -> dict | None:
        rows = self._ai_review_table_rows()
        items = []
        for idx, row in enumerate(rows, start=1):
            review_type = str(row.get("평가유형", "") or "").strip()
            if "수행" in review_type:
                continue
            if not any(token in review_type for token in ("선택", "객관", "서답", "서술", "논술", "지필", "문항")):
                review_type = "선택형"
            number = self._ai_review_number(row.get("번호/요소", ""), idx)
            item_type = self._ai_review_expected_type(review_type)
            difficulty = self._ai_review_difficulty(row.get("난이도 후보", ""))
            target = self._ai_review_level(row.get("목표수준 후보", ""), "C")
            standard = row.get("성취기준 후보", "")
            evidence = row.get("근거", "")
            next_step = row.get("다음 확인", "")
            counts, sample_size = self._ai_expected_counts_for_row(row)
            judgments = self._judgments_from_expected_counts(counts, sample_size)
            items.append({
                "id": f"ai-review-{idx}-{number}",
                "number": number,
                "title": row.get("번호/요소", "") or f"{number}번",
                "standard": "" if standard == "(후보 없음)" else standard,
                "points": self._ai_review_points_for_row(number, item_type, row, 5.0),
                "sampleSize": sample_size,
                "type": item_type,
                "difficulty": difficulty,
                "targetLevel": target,
                "judgmentsByJudge": {
                    "teacher-1": judgments,
                    "teacher-2": judgments,
                },
                "evidence": ["AI 검토 초안"],
                "note": " · ".join(part for part in ["AI 검토 초안", "교사 확인 필요", evidence, next_step] if part),
            })
        if not items:
            return None
        project = {
            "version": 1,
            "judges": [
                {"id": "teacher-1", "name": "검토안 1"},
                {"id": "teacher-2", "name": "검토안 2"},
            ],
            "activeJudgeId": "teacher-1",
            "items": items,
            "evidenceMode": "difficultyAverage",
        }
        if self.exam is not None and self.overall is not None:
            project["evidenceData"] = self._build_spliter_evidence_payload()
        return project

    def _send_ai_review_to_spliter(self):
        project = self._build_ai_review_spliter_project()
        if project is None:
            QMessageBox.information(
                self,
                "AI 문항 검토",
                "예상정답률 계산기로 보낼 지필 문항 초안이 없습니다.\n먼저 검토 초안을 생성하거나 평가유형을 선택형/서답형으로 수정해 주세요.",
            )
            return
        if self.spliter_view is None:
            QMessageBox.information(self, "AI 문항 검토", "이 환경에서는 내장 예상정답률 계산기를 열 수 없습니다.")
            return
        self._spliter_pending_project_payload = project
        if self.exam is not None and self.overall is not None:
            self._spliter_pending_payload = self._build_spliter_evidence_payload()
        self._activate_spliter_tab_for_pending_payloads()
        self.statusBar().showMessage(f"AI 검토 지필 초안 {len(project['items'])}개를 예상정답률 계산기로 보냈습니다.", 5000)

    def _extract_perform_max_score_from_review(self, *texts: str) -> float:
        joined = " ".join(texts)
        for area in getattr(self.perform_data, "areas", []) if self.perform_data is not None else []:
            if area.name and area.name in joined:
                return float(area.max_score)
        numbers = re.findall(r"(\d+(?:\.\d+)?)\s*점", joined)
        if numbers:
            try:
                return max(1.0, float(numbers[-1]))
            except Exception:
                pass
        return 10.0

    def _send_ai_review_to_perform_recalc(self):
        rows = [
            row for row in self._ai_review_table_rows()
            if "수행" in row.get("평가유형", "") or "수행" in row.get("구분", "")
        ]
        if not rows:
            QMessageBox.information(
                self,
                "AI 문항 검토",
                "수행평가 재산정으로 보낼 초안이 없습니다.\n수행평가 채점기준표를 불러와 초안을 만들거나 평가유형을 수행평가로 수정해 주세요.",
            )
            return
        self.table_perform_recalc.blockSignals(True)
        self.table_perform_recalc.setRowCount(0)
        self.table_perform_recalc.blockSignals(False)
        self._perform_recalc_dirty = True
        for row in rows:
            name = row.get("번호/요소", "") or f"평가요소 {self.table_perform_recalc.rowCount() + 1}"
            max_score = self._extract_perform_max_score_from_review(
                name, row.get("성취기준 후보", ""), row.get("근거", ""), row.get("다음 확인", "")
            )
            values = {
                "name": name,
                "max_score": max_score,
                "memo": "AI 검토 초안 · " + " · ".join(
                    part for part in [row.get("목표수준 후보", ""), row.get("난이도 후보", ""), row.get("근거", "")]
                    if part
                ),
            }
            values.update(self._ai_review_expected_scores_for_perform_row(row, max_score))
            self._add_perform_recalc_row(values)
        self._render_perform_recalc_results()
        self.tabs.setCurrentWidget(self.tab_perform)
        if hasattr(self, "perform_table_tabs"):
            self.perform_table_tabs.setCurrentIndex(2)
        self.statusBar().showMessage(f"AI 검토 수행 초안 {len(rows)}개를 수행평가 재산정 표로 보냈습니다.", 5000)

    # ---- 도움말 -------------------------------------------------------
    def _init_tab_help(self):
        layout = QVBoxLayout(self.tab_help)
        txt = QTextBrowser(); txt.setReadOnly(True); txt.setOpenExternalLinks(True)
        txt.setHtml(f"""
        <style>
          body {{ line-height: 1.62; }}
          h1 {{ margin: 0 0 4px; font-size: 22px; }}
          h2 {{ margin: 22px 0 8px; font-size: 17px; }}
          p {{ margin: 6px 0; }}
          ul, ol {{ margin-top: 6px; margin-bottom: 10px; }}
          li {{ margin: 5px 0; }}
          code {{ padding: 2px 5px; border-radius: 4px; }}
          .muted {{ opacity: .78; }}
          .card {{ border: 1px solid #8aa; border-radius: 8px; padding: 12px 14px; margin: 10px 0; }}
        </style>
        <h1>성취평가 결과 분석 Goedu-Split</h1>
        <p class="muted">v{APP_VERSION} · 제작자: {APP_AUTHOR} · {APP_COPYRIGHT}</p>

        <div class="card">
          <b>이 앱은 무엇을 하나요?</b>
          <p>정오표와 문항정보표를 불러와 학생 성취수준, 문항 정답률, 변별도, 답지반응,
          성취기준별 정답률을 한 번에 확인하는 로컬 분석 도구입니다. 자료는 사용자의 컴퓨터 안에서만 처리됩니다.</p>
        </div>

        <h2>빠른 시작</h2>
        <ol>
          <li>왼쪽의 <b>폴더에서 한 번에 불러오기</b>를 누르고 성취평가 자료 폴더를 선택합니다.</li>
          <li>정오표, 문항정보표, 추정분할점수, 수행평가 파일이 자동으로 채워졌는지 확인합니다.</li>
          <li>분할점수를 확인한 뒤 <b>분석 실행</b>을 누릅니다.</li>
        </ol>

        <h2>왼쪽 입력 패널</h2>
        <ul>
          <li><b>학생답 정오표</b>: 학생별 답, 선택형/서답형 점수, 과목총점이 들어 있는 파일입니다.</li>
          <li><b>문항정보표</b>: 문항번호, 성취기준, 난이도, 배점, 정답을 읽습니다.</li>
          <li><b>분할점수</b>: 90, 80, 70, 60, 40이면 <b>고정 분할 방식</b>, 그 밖의 값이면 <b>추정 분할 방식</b>으로 표시합니다.</li>
          <li><b>1-5등급 컷 계산</b>: 지필평가 교과목별 일람표를 넣으면 2022 개정 5등급 누적비율 기준으로 1등급, 2등급, 3등급, 4등급 컷 점수를 바로 계산합니다. 동점자가 경계에 걸리면 실제 인원이 기준 비율과 달라질 수 있음을 함께 표시합니다.</li>
          <li><b>수행평가</b>: 체크하면 지필과 수행 반영비율을 합산해 최종 환산점수를 계산합니다.</li>
        </ul>

        <h2>Data 탭</h2>
        <ul>
          <li><b>환산점수 분포</b>: 학생 점수가 어디에 모여 있는지 성취수준 색으로 보여주고, A/B/C/D/E 분할점수와 9등급·5등급 경계 점수를 함께 표시합니다. 그래프 표시 옵션에서 성취도선, 9등급선, 5등급선, 학생 위치, 학생 이름을 각각 켜고 끌 수 있습니다.</li>
          <li><b>학생 검색·필터</b>: 이름이나 반/번호를 입력하고, 성취도·9등급·5등급 콤보를 함께 써서 표를 좁혀 봅니다. 여러 명은 쉼표로 구분합니다.</li>
          <li>검색된 학생은 그래프 위에 세로선으로 표시되어, 분포 안 위치를 바로 볼 수 있습니다. 이름 라벨은 겹침을 줄이기 위해 기본은 꺼져 있고 필요할 때 켭니다.</li>
          <li><b>상담 모드</b>: 상단 방패 아이콘 또는 Ctrl+Shift+H로 켭니다. 검색한 학생 외 다른 학생의 이름과 반/번호는 화면에서 가려지며, <b>검색 학생만 보기</b>를 켜면 상담 대상만 표에 남깁니다.</li>
          <li><b>포트폴리오 저장</b>: 현재 과목 결과를 저장해 다른 과목 분석 결과와 함께 학생별 이력으로 쌓을 수 있습니다.</li>
          <li><b>정규분포·점검</b>: 평균, ±1표준편차, ±2표준편차와 정규분포 곡선을 함께 보여줍니다.</li>
          <li><b>모니터링 탭</b>: 성취평가 외부점검 지표를 별도 탭에서 더 자세히 확인합니다.</li>
        </ul>

        <h2>학생 포트폴리오 탭</h2>
        <p>과목별 분석이 끝난 뒤 Data 탭의 <b>포트폴리오 저장</b>을 누르면 학생별 성취도, 9등급, 5등급, 환산점수가 저장됩니다.
        이후 다른 과목을 분석하고 다시 저장하면 <b>학생 선택</b>에서 한 학생을 고르는 것만으로 여러 과목 결과가 정리됩니다.
        <b>상담 리포트 보기</b>를 누르면 과목별 흐름, 강점/점검 과목, 상담 질문을 한 화면으로 볼 수 있습니다.
        새로 저장한 기록은 <b>저장자료 불러오기</b>로 즉시 갱신합니다.
        상담 모드가 켜져 있으면 포트폴리오에서도 검색 대상 외 학생 이름과 반/번호를 가립니다.</p>

        <h2>모니터링 탭</h2>
        <ul>
          <li><b>현재 A 비율, 과목 평균, 표준편차, A/B 분할점수</b>는 분석 결과에서 자동으로 들어옵니다.</li>
          <li><b>전국 기준</b>에는 동일 학교유형·교과군의 전국 평균과 표준편차를 입력합니다.</li>
          <li><b>전년도 기준</b>에는 같은 과목의 전년도 A 비율, 과목 평균, A/B 분할점수를 입력합니다.</li>
          <li><b>필수 지표</b>는 A 비율이 전국 평균보다 높은지, 과목 평균이 A/B 분할점수보다 높은지를 봅니다.</li>
          <li><b>특성 지표</b>는 과목 평균, A/B 분할점수, 전년도 대비 변화가 어떤 방향으로 움직였는지 함께 해석합니다.</li>
          <li>Ⅰ/Ⅱ/Ⅲ 판정은 자동 결론이 아니라 “어디부터 질문할지”를 정하는 점검 신호로 사용하세요.</li>
        </ul>

        <h2>전체 성취도 분석</h2>
        <ul>
          <li><b>성취수준 비율</b>: A부터 미도달까지 학생 비율과 인원을 도넛 차트로 보여줍니다.</li>
          <li><b>평균±표준편차</b>: 각 성취수준 안의 점수 평균과 흩어짐을 확인합니다.</li>
          <li><b>학급별 평균</b>: 학급 간 평균 차이를 살펴봅니다.</li>
        </ul>

        <h2>수행평가 분석</h2>
        <ul>
          <li><b>영역별 요약</b>: 수행평가 각 영역의 평균, 표준편차, 평균점수율, 만점자 비율, 50% 미만 비율을 봅니다.</li>
          <li><b>지필총점 × 수행평가 환산</b>: 지필과 수행이 비슷하게 움직이는지, 특정 학생군이 수행에서 강점/보완점을 보이는지 확인합니다.</li>
          <li><b>학생별 차이</b>: 수행환산점수와 지필총점의 차이가 큰 학생을 먼저 보여줍니다. 수행 강점·수행 보완 학생을 빠르게 찾을 수 있습니다.</li>
          <li><b>분할점수 재산정</b>: 평가요소별로 A~E 최소능력자가 받을 것으로 예상되는 점수를 입력하면 A/B, B/C, C/D, D/E, E/미도달 분할점수를 계산합니다.</li>
          <li><b>자료 기준 딸깍 추천</b>: 수행평가 자료가 있으면 각 성취수준에서 경계에 가까운 대표 학생들의 영역별 점수를 바탕으로 초기값을 제안합니다.</li>
          <li>영역별 2/3 기준선은 “해당 수준 학생이 어느 정도 안정적으로 해결 가능한가”를 보는 참고선입니다. 최종 판단은 채점기준표와 문항지를 함께 보며 교사가 조정합니다.</li>
        </ul>

        <h2>문항 분석</h2>
        <ul>
          <li>문항은 기본적으로 오름차순으로 정렬됩니다.</li>
          <li><b>정답률</b>은 전체 학생 중 정답자의 비율입니다.</li>
          <li><b>변별도</b>는 해당 문항을 맞힌 학생과 총점이 높은 학생 사이의 관련성을 봅니다.</li>
          <li>성취수준별 정답률이 낮은 칸은 음영으로 표시되어 재검토할 문항을 찾기 쉽습니다.</li>
        </ul>

        <h2>답지반응분포</h2>
        <p>각 성취수준 학생들이 어느 선택지를 골랐는지 봅니다. 오답 매력도가 높은 선택지나 특정 수준에서 흔한 오개념을 찾을 때 유용합니다.</p>

        <h2>성취기준 분석 결과</h2>
        <p>성취기준별 평균 정답률과 성취수준별 정답률을 함께 보여줍니다. 어느 성취기준에서 어려움이 컸는지 확인할 수 있습니다.</p>

        <h2>예상정답률 입력</h2>
        <p>분석된 실제 학생 자료를 근거로 새 문항의 예상정답률을 O/X 방식으로 조정합니다.
        A, B, C, D, E 수준 최소능력자가 맞힐지 판단하면 분할점수 산정용 예상정답률로 환산됩니다.</p>
        <p>이 앱에서 <b>A 수준 문항</b>이라는 말은 A 수준 학생 대표 3명 중 약 2명, 즉 2/3 정도가 맞힐 수 있는 문항이라는 뜻입니다.
        B, C, D, E 문항도 같은 방식으로 이해하면 됩니다.</p>
        <p>분석자료 없이도 기본 문항으로 바로 사용할 수 있습니다. 분석자료가 전달되면 지난 정기시험의
        난이도·성취수준별 응답 자료를 기준으로 기본 예상정답률을 맞춘 뒤, 교사가 새 시험 설계에 맞게 조정합니다.</p>
        <p><b>문항 추가</b>로 새 문항을 만들고, 표에서 체크한 문항은 <b>선택 제거</b>로 한 번에 삭제할 수 있습니다.
        행 끝의 휴지통 아이콘은 해당 문항 하나만 지울 때 사용합니다.</p>
        <p><b>시험지 자동 반영</b>은 HWPX, PDF, DOCX, 엑셀, 텍스트 자료를 문항 단위로 읽어 예상정답률 계산기에 바로 넣습니다.
        HWP 파일은 로컬에 <code>kordoc</code> 또는 <code>hwp5txt</code> 변환기가 설치된 경우 자동으로 시도하며, 안정성을 위해 HWPX 저장본을 권장합니다.</p>
        <p><b>문항 구성안 제안</b>은 사용자가 입력한 문항 수에 맞춰 성취수준 목표와 배점을 먼저 제시합니다.
        분석자료가 있으면 기존 문항의 정답률·난이도·배점 분포를 참고하고, 없으면 100점 기준 기본 구성안을 만듭니다.</p>
        <p>지필평가는 문항별 예상정답률을 합산해 분할점수를 만들고, 수행평가는 평가요소별 예상점수를 합산해 분할점수를 만듭니다.
        두 기능은 모두 “최소능력자가 어느 정도 수행할 수 있는가”를 숫자로 옮기는 같은 구조입니다.</p>

        <h2>AI 문항 검토 방향</h2>
        <p><b>AI 문항 검토</b> 탭은 시험 문제 자료를 먼저 문항 단위로 읽고, 별도로 넣은 성취기준·성취수준 자료와 대조해 검토 초안을 만드는 작업 공간입니다.</p>
        <ul>
          <li><b>문항 자료</b>: 시험 문제 HWPX/PDF/DOCX, 문항정보표, 수행평가 채점기준표를 문항 자료 칸으로 불러옵니다. PDF는 텍스트 추출 가능한 파일이어야 합니다.</li>
          <li><b>성취기준·수준 자료</b>: 성취기준, 성취수준 A~E 설명, 최소능력자 특성 자료를 참고자료 칸으로 불러옵니다.</li>
          <li><b>현재 문항정보표</b>: 이미 분석한 문항정보표를 검토 원문으로 가져옵니다.</li>
          <li><b>검토 초안 생성</b>: 문항 자료를 먼저 나누고 참고자료와 대조해 성취기준 후보, 평가유형, 목표수준 후보, 난이도 후보, 근거, 추가 확인 질문을 표로 정리합니다.</li>
          <li><b>AI로 보강</b>: AI 설정에 지정한 Ollama 로컬 모델 또는 Codex CLI OAuth가 문항 자료와 성취기준·수준 자료를 함께 읽고 검토표를 다시 제안합니다. 기본값은 외부 전송이 없는 로컬 초안입니다.</li>
          <li><b>A~E 예상</b>: 지필 문항은 각 수준 대표 3명 중 몇 명이 맞힐지 <code>2/3</code>처럼 표시하고, 수행평가는 <code>80%</code> 또는 <code>8점</code>처럼 예상점수를 표시합니다.</li>
          <li><b>AI 설정</b>: Ollama 로컬과 Codex CLI 클라우드 OAuth를 선택합니다. Codex 경로는 API Key를 쓰지 않고 터미널의 <code>codex login</code> 상태를 사용합니다.</li>
          <li><b>개인정보 제거</b>: 외부 서버로 보낼 때 학생 이름, 반/번호, 전화번호, 이메일을 가능한 한 제거합니다. 그래도 최종 전송 여부는 교사가 확인합니다.</li>
          <li><b>지필→예상정답률</b>: 선택형·서답형 문항 초안과 A~E 예상 O/X 값을 예상정답률 계산기로 보내 문항별 판단을 이어갑니다.</li>
          <li><b>수행→재산정</b>: 수행평가 평가요소 초안과 A~E 예상점수를 수행평가 분할점수 재산정 표로 보내 합산을 이어갑니다.</li>
          <li><b>AI 프롬프트</b>: 향후 AI 연결 또는 별도 AI 검토에 바로 사용할 수 있는 안전한 검토 요청문을 만듭니다.</li>
        </ul>
        <p>AI 판정은 자동 확정이 아니라, 근거 문장과 함께 제안하고 교사가 최종 수정하는 방식으로 사용합니다.</p>

        <h2>단축키와 조작</h2>
        <ul>
          <li><code>Ctrl+R</code>: 분석 실행</li>
          <li><code>Ctrl+E</code>: 결과 CSV 내보내기</li>
          <li><code>Ctrl+1/2/3</code>: 라이트 / 다크 / 자동 테마</li>
          <li><code>Ctrl + + / -</code>: 화면 배율 조절</li>
          <li><code>Ctrl+Z</code>, <code>Ctrl+Shift+Z</code>: 예상정답률 계산기에서 입력 되돌리기 / 다시 실행</li>
          <li>예상정답률 계산기: <b>문항 추가</b>로 새 문항 생성, <b>선택 제거</b>로 체크한 문항 삭제</li>
          <li>그래프 위 마우스 휠: 확대/축소, 더블클릭: 원상복귀</li>
        </ul>

        <h2>계산 기준</h2>
        <ul>
          <li>환산점수 = 지필점수 × 지필반영비율 + 수행환산점수 × 수행반영비율</li>
          <li>성취수준은 환산점수를 반올림한 정수 점수와 분할점수를 비교해 판정합니다.</li>
          <li>Cronbach α는 지필 문항의 내적 일관성을 확인하는 참고 지표입니다.</li>
        </ul>
        """)
        layout.addWidget(txt)

    # --------------------------------------------------------------- 메뉴
    def _build_menu(self):
        m_file = self.menuBar().addMenu("파일")
        a_run = QAction("분석 실행", self); a_run.setShortcut("Ctrl+R")
        a_run.triggered.connect(self.run_analysis); m_file.addAction(a_run)
        a_exp = QAction("결과 CSV 내보내기…", self); a_exp.setShortcut("Ctrl+E")
        a_exp.triggered.connect(self.export_csv); m_file.addAction(a_exp)
        a_spliter = QAction("예상정답률 근거 엑셀 저장", self)
        a_spliter.triggered.connect(self.export_spliter_evidence); m_file.addAction(a_spliter)
        m_file.addSeparator()
        a_quit = QAction("종료", self); a_quit.setShortcut("Ctrl+Q")
        a_quit.triggered.connect(self.close); m_file.addAction(a_quit)

        m_view = self.menuBar().addMenu("보기")
        self._theme_actions = {}
        if self.theme is not None:
            grp = QActionGroup(self); grp.setExclusive(True)
            for key, label, shortcut in (
                ("light", "라이트 모드", "Ctrl+1"),
                ("dark", "다크 모드", "Ctrl+2"),
                ("auto", "자동 (시간대 기반)", "Ctrl+3"),
            ):
                a = QAction(label, self, checkable=True)
                a.setShortcut(shortcut)
                a.triggered.connect(lambda _=False, k=key: self._set_theme(k))
                grp.addAction(a); m_view.addAction(a)
                self._theme_actions[key] = a
            self._theme_actions[self.theme.mode].setChecked(True)
            m_view.addSeparator()
            a_refresh = QAction("차트 다시 그리기", self); a_refresh.setShortcut("F5")
            a_refresh.triggered.connect(self._refresh_all_charts)
            m_view.addAction(a_refresh)

        m_help = self.menuBar().addMenu("도움말")
        a_about = QAction("프로그램 정보", self)
        a_about.triggered.connect(self._show_about); m_help.addAction(a_about)

    def _set_theme(self, key: str):
        if self.theme is None:
            return
        self.theme.set_mode(key)
        # _on_theme_changed가 차트 재렌더 처리

    def _build_statusbar(self):
        sb = self.statusBar()
        right = QLabel(f"{APP_COPYRIGHT}   v{APP_VERSION}")
        right.setProperty("role", "muted")
        sb.addPermanentWidget(right)

    def _on_theme_changed(self, eff: str):
        charts.set_theme(self.theme.colors)
        # 메뉴 체크 상태 + 헤더 버튼 동기화
        if hasattr(self, "_theme_actions") and self.theme.mode in self._theme_actions:
            self._theme_actions[self.theme.mode].setChecked(True)
        if hasattr(self, "btn_theme"):
            self._refresh_header_icons()
        self._send_spliter_theme()
        self._apply_zoom_to_surfaces()
        self._refresh_all_charts()

    def _refresh_all_charts(self):
        if self.exam is None:
            return
        self._render_data_tab()
        self._render_monitor_tab()
        self._render_overview()
        self._render_perform_tab()
        self._render_items()
        self._render_choice()
        self._render_standard()
        self._render_spliter_tab()

    def _show_about(self):
        QMessageBox.information(
            self, "프로그램 정보",
            f"<h3>Goedu-Split – 성취평가 결과 분석 데스크톱</h3>"
            f"<p>v{APP_VERSION} · 2026</p>"
            f"<p><b>제작자:</b> {APP_AUTHOR}</p>"
            f"<p>{APP_COPYRIGHT}</p>"
            "<p>본 데스크톱 버전은 PySide6/Qt와 matplotlib으로 재구성한 로컬 분석 도구이며, "
            "데이터는 본 PC를 떠나지 않습니다.</p>"
        )

    # --------------------------------------------------------- 분석 실행
    def run_analysis(self):
        if not self.fs_response.path() or not self.fs_iteminfo.path():
            QMessageBox.warning(self, "입력 부족",
                                "정오표와 문항정보표 두 파일을 먼저 지정해 주세요.")
            return
        try:
            exam = load_exam(self.fs_iteminfo.path(), self.fs_response.path())
        except Exception as e:
            QMessageBox.critical(self, "파싱 오류",
                                 f"파일을 읽는 중 오류:\n{e}\n\n{traceback.format_exc()}")
            return
        exam.cut_scores = {lv: float(self.spin_cuts[lv].value()) for lv in ["A","B","C","D","E"]}

        # 수행평가 결합
        perform_data = None
        if self.chk_perform.isChecked() and self.fs_perform.path():
            try:
                perform_data = load_perform(self.fs_perform.path())
            except Exception as e:
                QMessageBox.warning(self, "수행평가 파일 오류",
                                    f"수행평가 파일을 읽지 못했습니다. 지필만 분석합니다.\n{e}")
        apply_perform(
            exam, perform_data,
            float(self.spin_pencil_ratio.value()),
            float(self.spin_perform_ratio.value()),
        )

        try:
            score_mat, _, _ = build_score_matrix(exam)
            overall = analyze_overall(exam, score_mat)
            items = analyze_items(exam)[0]
        except Exception as e:
            QMessageBox.critical(self, "분석 오류", f"{e}\n\n{traceback.format_exc()}")
            return

        self.exam, self.overall, self.item_stats, self.perform_data = exam, overall, items, perform_data
        # 마지막 분석 시점 입력 스냅샷 저장 (실수 복원용)
        self._snapshot_inputs()
        if hasattr(self, "btn_revert"):
            self.btn_revert.setEnabled(True)
        self.exam_info_lbl.setText(
            f"<b>{exam.subject or '(과목 미상)'}</b><br>"
            f"{exam.semester} {exam.grade}<br>"
            f"학생 {len(exam.students)}명 · 문항 {len(exam.items)}문항"
        )
        if perform_data is not None and hasattr(self, "table_perform_recalc") and not self._perform_recalc_dirty:
            self._recommend_perform_recalc_rows()
        self._render_data_tab()
        self._render_monitor_tab()
        self._render_overview()
        self._render_perform_tab()
        self._render_items()
        self._render_choice_combo()
        self._render_choice()
        self._render_standard()
        self._render_spliter_tab()
        self.refresh_portfolio_tab()
        if self.spliter_view is not None:
            self._spliter_pending_payload = self._build_spliter_evidence_payload()
            self._flush_spliter_payload()
        self.statusBar().showMessage(
            f"분석 완료 · N={overall.n_students} · α={overall.cronbach_alpha:.3f} "
            f"({reliability_label(overall.cronbach_alpha)}) · "
            f"분할점수 A≥{int(exam.cut_scores['A'])} B≥{int(exam.cut_scores['B'])} "
            f"C≥{int(exam.cut_scores['C'])} D≥{int(exam.cut_scores['D'])} E≥{int(exam.cut_scores['E'])}"
        )

    # ------------------------------------------------------------- 렌더링
    def _render_data_tab(self):
        if hasattr(self, "data_empty_state"):
            self.data_empty_state.setVisible(False)
        ov = self.overall
        self.kpi_n.value_label.setText(f"{ov.n_students} 명")
        self.kpi_n_items.value_label.setText(f"{len(self.exam.items)} 문항")
        self.kpi_subject.value_label.setText(self.exam.subject or "-")

        totals = [s.final_score for s in self.exam.students]
        search_text = self.le_search.text() if hasattr(self, "le_search") else ""
        self._render_score_histogram_for_search(search_text)
        self.canvas_score_normal.set_figure(
            charts.fig_score_normal_monitoring(totals, self.exam.cut_scores, ov.level_dist_pct)
        )
        if totals:
            mean = sum(totals) / len(totals)
            std = float((sum((x - mean) ** 2 for x in totals) / max(1, len(totals) - 1)) ** 0.5)
            expected_a = (1 - _normal_cdf(float(self.exam.cut_scores["A"]), mean, std)) * 100 if std > 0 else 0.0
            self.lbl_normal_note.setText(
                f"평균 {mean:.2f}점 · 표준편차 {std:.2f}점 · "
                f"실제 A {ov.level_dist_pct.get('A', 0):.1f}% · "
                f"정규분포 기준 기대 A {expected_a:.1f}%"
            )

        # 학생별 응답표 — 학번 제거, 반/번호+이름을 앞쪽 두 컬럼에 고정 폭으로 둬
        # 가로 스크롤해도 항상 식별 가능. 점수 컬럼들도 앞쪽으로 옮겨 핵심을 우선 노출.
        items = sorted([it for it in self.exam.items if it.item_type == "선택형"],
                       key=lambda x: x.number)
        # KICE 원본 웹앱과 동일: 환산점수는 '반올림 정수 원점수'로 표시
        grade9_labels, grade5_labels = self._current_relative_grades()
        score_headers = ["성취도", "9등급", "5등급", "원점수"]
        if self.exam.use_perform:
            score_headers += ["수행환산", "지필총점"]
        else:
            score_headers += ["지필총점"]
        item_headers = [f"문{it.number}" for it in items]
        headers = ["반/번호", "이름"] + score_headers + item_headers
        self._data_level_col = 2
        self._data_grade9_col = 3
        self._data_grade5_col = 4
        self._data_score_col = 5
        self.table_data.setColumnCount(len(headers))
        self.table_data.setHorizontalHeaderLabels(headers)
        self.table_data.setRowCount(len(self.exam.students))

        # 컬럼 폭 — 반/번호와 이름이 잘리지 않게 충분히 넓힘 (좌측 고정 효과)
        hh = self.table_data.horizontalHeader()
        hh.setSectionResizeMode(QHeaderView.Fixed)
        self._set_scaled_column_width(self.table_data, 0, 95)   # 반/번호
        self._set_scaled_column_width(self.table_data, 1, 110)  # 이름
        score_n = len(score_headers)
        for i in range(2, 2 + score_n):
            width = 74 if i in (self._data_grade9_col, self._data_grade5_col) else 92
            self._set_scaled_column_width(self.table_data, i, width)
        for i in range(2 + score_n, len(headers)):
            self._set_scaled_column_width(self.table_data, i, 56)

        # 정렬을 위해 채우는 동안 sorting OFF
        self.table_data.setSortingEnabled(False)
        # 성취도 정렬을 위해 A~E,미도달을 숫자로 매핑
        level_order = {"A": 6, "B": 5, "C": 4, "D": 3, "E": 2, "미도달": 1}
        for r, s in enumerate(self.exam.students):
            level = ov.levels_arr[r]

            # 반/번호 — 표시는 '1/3' 그대로, 정렬은 (반,번호) 튜플 비교 (NaturalItem)
            cls_text = s.class_no or ""
            sort_value = _class_no_sort_value(cls_text)
            it_cls = NaturalItem(cls_text, sort_value)
            it_cls.setTextAlignment(Qt.AlignCenter)
            it_cls.setData(Qt.UserRole + 101, r)
            it_cls.setData(PRIVACY_REAL_TEXT_ROLE, cls_text)
            self.table_data.setItem(r, 0, it_cls)

            # 이름
            it_name = _set_item(self.table_data, r, 1, s.name or "-", align_left=True,
                                tooltip=s.name)
            it_name.setData(PRIVACY_REAL_TEXT_ROLE, s.name or "-")

            col = 2
            # 성취도 — 표시는 A/B/.../미도달, 정렬은 1~6 (NaturalItem)
            it_lv = NaturalItem(level, level_order.get(level, 0))
            it_lv.setTextAlignment(Qt.AlignCenter)
            f = it_lv.font(); f.setBold(True); it_lv.setFont(f)
            level_color = charts.COLOR_LEVELS.get(level, "#ffffff")
            it_lv.setBackground(QBrush(QColor(level_color)))
            it_lv.setForeground(QBrush(_contrast_text_for_fill(level_color)))
            self.table_data.setItem(r, col, it_lv); col += 1

            grade9 = grade9_labels[r] if r < len(grade9_labels) else "-"
            it_g9 = NaturalItem(grade9, int(grade9) if str(grade9).isdigit() else 99)
            it_g9.setTextAlignment(Qt.AlignCenter)
            it_g9.setToolTip("9등급제 상대 석차 기준")
            self.table_data.setItem(r, col, it_g9); col += 1

            grade5 = grade5_labels[r] if r < len(grade5_labels) else "-"
            it_g5 = NaturalItem(grade5, int(grade5) if str(grade5).isdigit() else 99)
            it_g5.setTextAlignment(Qt.AlignCenter)
            it_g5.setToolTip("5등급제 상대 석차 기준")
            self.table_data.setItem(r, col, it_g5); col += 1

            # 원점수 — 표시는 반올림 정수, 정렬은 float
            it_fs = NaturalItem(f"{round(s.final_score)}", float(s.final_score))
            it_fs.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            it_fs.setToolTip(f"환산점수 {s.final_score:.2f} → 반올림 {round(s.final_score)}")
            f = it_fs.font(); f.setBold(True); it_fs.setFont(f)
            self.table_data.setItem(r, col, it_fs); col += 1

            if self.exam.use_perform:
                it_ps = NaturalItem(f"{s.perform_score:.1f}", float(s.perform_score))
                it_ps.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.table_data.setItem(r, col, it_ps); col += 1

            it_ts = NaturalItem(f"{s.total:.1f}", float(s.total))
            it_ts.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.table_data.setItem(r, col, it_ts); col += 1

            for it in items:
                v = s.answers.get(it.number, "")
                _set_item(self.table_data, r, col, v, align_right=True); col += 1

        self.table_data.setSortingEnabled(True)
        self.table_data.sortByColumn(0, Qt.AscendingOrder)
        # 첫 두 컬럼(반/번호, 이름)을 좌측에 틀고정 — 가로 스크롤해도 항상 보임
        if not hasattr(self.table_data, "frozenView") or self.table_data.frozenView is None:
            install_frozen_columns(self.table_data, frozen_count=2)
        else:
            refresh_frozen_columns(self.table_data)
        # setSortingEnabled 직후 자동 정렬 상태가 남는 경우가 있어 최종 배치에서 한 번 더 고정한다.
        self.table_data.sortByColumn(0, Qt.AscendingOrder)
        self._filter_data_table(search_text)

    def _render_overview(self):
        ov = self.overall
        # 누적 막대 안 ABCDE 라벨이 겹치는 문제 → 도넛 차트로 교체
        self.canvas_level_stack.set_figure(charts.fig_level_donut(ov.level_dist_pct, ov.level_dist))
        self.canvas_level_means.set_figure(charts.fig_level_means_with_error(ov.level_mean, ov.level_std))
        # 표
        rows = []
        for lv in LEVELS:
            rows.append([lv, ov.level_dist[lv], f"{ov.level_dist_pct[lv]:.1f}",
                         f"{ov.level_mean[lv]:.1f}", f"{ov.level_std[lv]:.1f}"])
        rows.append(["전체", ov.n_students, "100.0", f"{ov.mean:.1f}", f"{ov.std:.1f}"])
        self.table_level.setRowCount(len(rows))
        for r, row in enumerate(rows):
            for c, v in enumerate(row):
                bold = (r == len(rows) - 1)
                _set_item(self.table_level, r, c, v, align_right=(c > 0), bold=bold)
        self.table_level.resizeColumnsToContents()
        self.canvas_class.set_figure(charts.fig_class_means(ov.by_class))

    def _perform_record_for_student(self, student):
        if self.perform_data is None:
            return None
        return (
            self.perform_data.records.get(student.sid)
            or self.perform_data.by_classno.get(student.class_no)
        )

    def _perform_matches(self):
        if self.exam is None or self.overall is None or self.perform_data is None:
            return []
        levels_arr = list(getattr(self.overall, "levels_arr", []) or [])
        matches = []
        for idx, student in enumerate(self.exam.students):
            rec = self._perform_record_for_student(student)
            if rec is None:
                continue
            level = levels_arr[idx] if idx < len(levels_arr) else grade_level(student.final_score, self.exam.cut_scores)
            matches.append((idx, student, level, rec))
        return matches

    @staticmethod
    def _mean_std(values: list[float]) -> tuple[float, float]:
        if not values:
            return 0.0, 0.0
        mean = sum(values) / len(values)
        if len(values) <= 1:
            return mean, 0.0
        var = sum((value - mean) ** 2 for value in values) / (len(values) - 1)
        return mean, math.sqrt(max(0.0, var))

    def _perform_area_rows(self):
        if self.perform_data is None:
            return []
        matches = self._perform_matches()
        rows = []
        for area in self.perform_data.areas:
            scores = []
            rates = []
            by_level = {lv: [] for lv in LEVELS}
            for _, _, level, rec in matches:
                score = float(rec.scores.get(area.name, 0.0))
                rate = (score / area.max_score * 100.0) if area.max_score > 0 else 0.0
                scores.append(score)
                rates.append(rate)
                by_level.setdefault(level, []).append(rate)
            mean_score, std_score = self._mean_std(scores)
            mean_rate, std_rate = self._mean_std(rates)
            n = len(rates) or 1
            rows.append({
                "name": area.name,
                "max_score": float(area.max_score),
                "ratio_pct": float(area.ratio_pct),
                "mean_score": mean_score,
                "std_score": std_score,
                "mean_rate": mean_rate,
                "std_rate": std_rate,
                "full_pct": sum(1 for score in scores if score >= area.max_score - 1e-9) / n * 100.0,
                "low_pct": sum(1 for rate in rates if rate < 50.0) / n * 100.0,
                "by_level": {
                    lv: (sum(vals) / len(vals) if vals else None)
                    for lv, vals in by_level.items()
                },
            })
        return rows

    def _perform_corr(self, matches) -> float | None:
        if len(matches) <= 1:
            return None
        x = [float(student.total) for _, student, _, _ in matches]
        y = [float(student.perform_score) for _, student, _, _ in matches]
        mx, sx = self._mean_std(x)
        my, sy = self._mean_std(y)
        if sx <= 0 or sy <= 0:
            return None
        cov = sum((a - mx) * (b - my) for a, b in zip(x, y)) / (len(x) - 1)
        return cov / (sx * sy)

    def _render_perform_tab(self):
        if not hasattr(self, "table_perform_areas"):
            return

        no_data = (
            self.exam is None
            or self.overall is None
            or self.perform_data is None
            or not self.exam.use_perform
        )
        if no_data:
            for card in getattr(self, "perform_cards", {}).values():
                card.value_label.setText("-")
            self.table_perform_areas.setRowCount(0)
            self.table_perform_students.setRowCount(0)
            self.canvas_perform_area.set_figure(charts.fig_perform_area_rates([]))
            self.canvas_perform_scatter.set_figure(charts.fig_perform_scatter([], [], []))
            if hasattr(self, "lbl_perform_tab_note"):
                self.lbl_perform_tab_note.setText(
                    "수행평가 파일을 지정하고 ‘수행평가도 포함하여 분석합니다’를 체크한 뒤 분석을 실행하면 "
                    "영역별 점수율, 지필-수행 관계, 학생별 차이가 표시됩니다."
                )
            self._render_perform_recalc_results()
            return

        matches = self._perform_matches()
        area_rows = self._perform_area_rows()
        corr = self._perform_corr(matches)
        self.perform_cards["ratio"].value_label.setText(f"{self.exam.weight_perform:.0f}%")
        self.perform_cards["areas"].value_label.setText(f"{len(self.perform_data.areas)}개")
        self.perform_cards["matched"].value_label.setText(f"{len(matches)}/{len(self.exam.students)}명")
        self.perform_cards["corr"].value_label.setText("-" if corr is None else f"{corr:.2f}")
        self.lbl_perform_tab_note.setText(
            f"{self.perform_data.subject or self.exam.subject or '(과목 미상)'} · "
            f"수행 원점수 만점 {self.perform_data.max_total:.1f}점 · "
            f"수행 자체 반영비율 합 {self.perform_data.ratio_total:.0f}% · "
            "2/3 기준선은 해당 영역이 목표 성취수준 학생에게 적절히 해결 가능한지 보는 참고선입니다."
        )

        self.canvas_perform_area.set_figure(charts.fig_perform_area_rates(area_rows))
        self.canvas_perform_scatter.set_figure(
            charts.fig_perform_scatter(
                [student.total for _, student, _, _ in matches],
                [student.perform_score for _, student, _, _ in matches],
                [level for _, _, level, _ in matches],
            )
        )

        c = self.theme.colors if self.theme else {}
        shade = QColor(c.get("shade", "#e1e6ee"))
        warn = QColor("#f8dfa3"); warn.setAlpha(120)
        self.table_perform_areas.setSortingEnabled(False)
        self.table_perform_areas.setRowCount(len(area_rows))
        for r, row in enumerate(area_rows):
            _set_item(self.table_perform_areas, r, 0, row["name"], align_left=True, tooltip=row["name"])
            _set_item(self.table_perform_areas, r, 1, f"{row['max_score']:.1f}", align_right=True)
            _set_item(self.table_perform_areas, r, 2, f"{row['ratio_pct']:.0f}%", align_right=True)
            _set_item(self.table_perform_areas, r, 3, f"{row['mean_score']:.2f}", align_right=True)
            _set_item(self.table_perform_areas, r, 4, f"{row['std_score']:.2f}", align_right=True)
            rate_item = NaturalItem(f"{row['mean_rate']:.1f}", row["mean_rate"])
            rate_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            ItemBarDelegate.set_bar(rate_item, row["mean_rate"], 100, "#2563eb", light=True)
            if row["mean_rate"] < 66.7:
                rate_item.setBackground(QBrush(shade))
            self.table_perform_areas.setItem(r, 5, rate_item)
            _set_item(self.table_perform_areas, r, 6, f"{row['full_pct']:.1f}%", align_right=True)
            _set_item(self.table_perform_areas, r, 7, f"{row['low_pct']:.1f}%", align_right=True,
                      bg=warn if row["low_pct"] >= 20 else None)
            for i, lv in enumerate(LEVELS):
                value = row["by_level"].get(lv)
                if value is None:
                    _set_item(self.table_perform_areas, r, 8 + i, "-", align_right=True)
                else:
                    _set_item(self.table_perform_areas, r, 8 + i, f"{value:.1f}", align_right=True,
                              bg=shade if value < 66.7 else None)
        self.table_perform_areas.setSortingEnabled(True)
        for col, width in enumerate([180, 70, 82, 74, 82, 92, 78, 82, 68, 68, 68, 68, 68, 78]):
            self._set_scaled_column_width(self.table_perform_areas, col, width)

        student_rows = []
        for _, student, level, _ in matches:
            diff = float(student.perform_score) - float(student.total)
            if diff >= 10:
                note = "수행 강점"
            elif diff <= -10:
                note = "수행 보완"
            elif abs(diff) >= 5:
                note = "차이 확인"
            else:
                note = "유사"
            student_rows.append((abs(diff), diff, student, level, note))
        student_rows.sort(key=lambda row: (-row[0], _class_no_sort_value(row[2].class_no)))

        self.table_perform_students.setSortingEnabled(False)
        self.table_perform_students.setRowCount(len(student_rows))
        level_order = {"A": 6, "B": 5, "C": 4, "D": 3, "E": 2, "미도달": 1}
        for r, (_, diff, student, level, note) in enumerate(student_rows):
            cls_item = NaturalItem(student.class_no or "", _class_no_sort_value(student.class_no or ""))
            cls_item.setTextAlignment(Qt.AlignCenter)
            self.table_perform_students.setItem(r, 0, cls_item)
            _set_item(self.table_perform_students, r, 1, student.name or "-", align_left=True)
            lv_item = NaturalItem(level, level_order.get(level, 0))
            lv_item.setTextAlignment(Qt.AlignCenter)
            level_color = charts.COLOR_LEVELS.get(level, "#ffffff")
            lv_item.setBackground(QBrush(QColor(level_color)))
            lv_item.setForeground(QBrush(_contrast_text_for_fill(level_color)))
            f = lv_item.font(); f.setBold(True); lv_item.setFont(f)
            self.table_perform_students.setItem(r, 2, lv_item)
            _set_item(self.table_perform_students, r, 3, f"{student.total:.1f}", align_right=True)
            _set_item(self.table_perform_students, r, 4, f"{student.perform_score:.1f}", align_right=True)
            diff_item = NaturalItem(f"{diff:+.1f}", abs(diff))
            diff_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            if abs(diff) >= 10:
                diff_item.setBackground(QBrush(warn))
            self.table_perform_students.setItem(r, 5, diff_item)
            _set_item(self.table_perform_students, r, 6, note, tooltip=note)
        self.table_perform_students.setSortingEnabled(True)
        self.table_perform_students.sortByColumn(5, Qt.DescendingOrder)
        for col, width in enumerate([90, 110, 72, 82, 90, 90, 110]):
            self._set_scaled_column_width(self.table_perform_students, col, width)
        self._render_perform_recalc_results()
        self._apply_zoom_to_surfaces()

    def _render_items(self):
        ov = self.overall
        self.lbl_alpha.setText(f"지필평가 신뢰도 (Cronbach's alpha): {ov.cronbach_alpha:.3f} "
                               f"({reliability_label(ov.cronbach_alpha)})")

        self.canvas_pvalue.set_figure(charts.fig_item_difficulty(self.item_stats))
        self.canvas_discr.set_figure(charts.fig_item_discrimination(self.item_stats))

        c = self.theme.colors if self.theme else {}
        col_p_bar = QColor(c.get("p_bar", "#cfe3fa")); col_p_bar.setAlpha(110)
        col_d_bar = QColor(c.get("d_bar", "#cfe9d6")); col_d_bar.setAlpha(110)
        col_highlight = QColor(c.get("highlight", "#fff5a4"))
        col_shade = QColor(c.get("shade", "#e1e6ee"))

        self.table_items.setRowCount(len(self.item_stats))
        self.table_items.setSortingEnabled(False)
        for r, s in enumerate(self.item_stats):
            # 문항번호 — 표시는 '문1', 정렬은 1 (NaturalItem). 기본은 문1→문18 순.
            it_no = NaturalItem(f"문{s.item.number}", int(s.item.number))
            it_no.setTextAlignment(Qt.AlignCenter)
            f = it_no.font(); f.setBold(True); it_no.setFont(f)
            it_no.setToolTip(f"문항 {s.item.number} · {s.item.content_area}")
            self.table_items.setItem(r, 0, it_no)
            _set_item(self.table_items, r, 1, s.item.difficulty or "-")
            # 정답률 — 표시는 '85.4', 정렬은 float + 막대
            it_p = NaturalItem(f"{s.p_value*100:.1f}", float(s.p_value * 100))
            it_p.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            it_p.setToolTip(f"정답률 {s.p_value*100:.2f}%")
            ItemBarDelegate.set_bar(it_p, s.p_value * 100, 100, "#2563eb", light=True)
            self.table_items.setItem(r, 2, it_p)
            # 변별도 — 표시는 '0.602', 정렬은 float + 막대
            it_d = NaturalItem(f"{s.discrimination:.3f}", float(s.discrimination))
            it_d.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            it_d.setToolTip(f"변별도 {s.discrimination:.3f}")
            ItemBarDelegate.set_bar(it_d, max(0.0, s.discrimination) * 100, 100, "#16a34a", light=True)
            self.table_items.setItem(r, 3, it_d)
            for i, ch in enumerate([1, 2, 3, 4, 5, 0]):
                v = s.choice_dist.get(ch, 0) * 100
                col = 4 + i
                bg = col_highlight if str(ch) == str(s.item.answer) else None
                _set_item(self.table_items, r, col, f"{v:.1f}", align_right=True, bg=bg)
            for i, lv in enumerate(LEVELS):
                col = 10 + i
                p = s.p_by_level.get(lv)
                if p is None:
                    _set_item(self.table_items, r, col, "-", align_right=True)
                else:
                    bg = col_shade if p < 2.0 / 3.0 else None
                    _set_item(self.table_items, r, col, f"{p*100:.1f}", align_right=True, bg=bg)
        self.table_items.setSortingEnabled(True)
        self.table_items.sortByColumn(0, Qt.AscendingOrder)  # 기본 정렬: 문1 → 문18
        # 문항번호 컬럼 틀고정
        if not hasattr(self.table_items, "frozenView") or self.table_items.frozenView is None:
            install_frozen_columns(self.table_items, frozen_count=1)
        else:
            refresh_frozen_columns(self.table_items)

        # ─── 서답형 분석 표 채우기 ─────────────────────────────────────
        sd = analyze_serdap(self.exam, ov.levels_arr)
        self.table_serdap.setRowCount(1 if sd.n_items > 0 else 0)
        if sd.n_items > 0:
            row = 0
            _set_item(self.table_serdap, row, 0, f"서답형 전체 ({sd.n_items}문항·{int(sd.max_total)}점)",
                      align_left=True)
            _set_item(self.table_serdap, row, 1, f"{sd.min_v:.1f}", align_right=True)
            _set_item(self.table_serdap, row, 2, f"{sd.max_v:.1f}", align_right=True)
            _set_item(self.table_serdap, row, 3, f"{sd.mean:.2f}", align_right=True)
            _set_item(self.table_serdap, row, 4, f"{sd.std:.2f}", align_right=True)
            # 정답률 + 막대
            it_p = NaturalItem(f"{sd.correct_rate*100:.1f}", float(sd.correct_rate*100))
            it_p.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            it_p.setToolTip(f"서답형 정답률 {sd.correct_rate*100:.2f}%")
            ItemBarDelegate.set_bar(it_p, sd.correct_rate*100, 100, "#2563eb", light=True)
            self.table_serdap.setItem(row, 5, it_p)
            # 변별도 + 막대
            it_d = NaturalItem(f"{sd.discrimination:.3f}", float(sd.discrimination))
            it_d.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            it_d.setToolTip(f"변별도 {sd.discrimination:.3f}")
            ItemBarDelegate.set_bar(it_d, max(0.0, sd.discrimination)*100, 100, "#16a34a", light=True)
            self.table_serdap.setItem(row, 6, it_d)
            # 성취수준별 정답률
            col_shade = QColor(c.get("shade", "#e1e6ee"))
            for i, lv in enumerate(LEVELS):
                p_lv = sd.by_level.get(lv)
                col = 7 + i
                if p_lv is None:
                    _set_item(self.table_serdap, row, col, "-", align_right=True)
                else:
                    pct = p_lv * 100
                    bg = col_shade if pct < 50 else None
                    _set_item(self.table_serdap, row, col, f"{pct:.1f}",
                              align_right=True, bg=bg)
            self.table_serdap.resizeColumnsToContents()

    def _render_choice_combo(self):
        self.combo_item.blockSignals(True)
        self.combo_item.clear()
        for s in self.item_stats:
            self.combo_item.addItem(f"문{s.item.number} – {s.item.content_area}", s)
        self.combo_item.blockSignals(False)
        if self.item_stats:
            self.combo_item.setCurrentIndex(0)

    def _render_choice(self):
        self._refresh_choice()

    def _refresh_choice(self):
        if self.combo_item.count() == 0:
            return
        s = self.combo_item.currentData()
        if s is None:
            return
        self.lbl_choice_info.setText(
            f"<b>문항 {s.item.number}</b> · {s.item.content_area} · "
            f"예상난이도(NEIS): {s.item.difficulty or '-'} · "
            f"배점 {s.item.score} · 정답: <b>{s.item.answer}</b><br>"
            f"전체 정답률 {s.p_value*100:.1f}% · 변별도 {s.discrimination:.3f} "
            f"({s.diff_label} / {s.discr_label})<br>"
            f"성취기준: {s.item.standard_code} {s.item.standard}"
        )
        # 매트릭스 표
        rows = []
        for lv in LEVELS:
            d = s.choice_dist_by_level.get(lv)
            if d is None:
                rows.append([lv, "-", "-", "-", "-", "-", "-", "-"])
                continue
            p = s.p_by_level.get(lv, 0) * 100
            rows.append([lv, f"{p:.1f}",
                         f"{d.get(1,0)*100:.1f}", f"{d.get(2,0)*100:.1f}",
                         f"{d.get(3,0)*100:.1f}", f"{d.get(4,0)*100:.1f}",
                         f"{d.get(5,0)*100:.1f}", f"{d.get(0,0)*100:.1f}"])
        # 전체 행
        d_all = s.choice_dist
        rows.append(["전체", f"{s.p_value*100:.1f}",
                     f"{d_all.get(1,0)*100:.1f}", f"{d_all.get(2,0)*100:.1f}",
                     f"{d_all.get(3,0)*100:.1f}", f"{d_all.get(4,0)*100:.1f}",
                     f"{d_all.get(5,0)*100:.1f}", f"{d_all.get(0,0)*100:.1f}"])
        self.table_choice.setRowCount(len(rows))
        try:
            correct_idx = int(s.item.answer)
        except (TypeError, ValueError):
            correct_idx = -1
        c_theme = self.theme.colors if self.theme else {}
        col_p_bar = QColor(c_theme.get("p_bar", "#cfe3fa")); col_p_bar.setAlpha(110)
        col_highlight = QColor(c_theme.get("highlight", "#fff5a4"))
        for r, row in enumerate(rows):
            bold = (r == len(rows) - 1)
            for cc, v in enumerate(row):
                bg = None
                if cc >= 2 and cc - 2 < 5 and (cc - 1) == correct_idx:
                    bg = col_highlight
                if cc == 1:  # 정답률 컬럼
                    bg = col_p_bar
                _set_item(self.table_choice, r, cc, v, align_right=(cc > 0), bold=bold, bg=bg)
        self.table_choice.resizeColumnsToContents()
        self.canvas_choice.set_figure(
            charts.fig_choice_dist_by_level(s, title=f"문항 {s.item.number} 답지반응분포 – 정답: {s.item.answer}")
        )

    def _render_standard(self):
        rows = self.overall.by_standard
        c = self.theme.colors if self.theme else {}
        col_shade = QColor(c.get("shade", "#e1e6ee"))
        self.table_std.setSortingEnabled(False)
        self.table_std.setRowCount(len(rows))
        for r, row in enumerate(rows):
            _set_item(self.table_std, r, 0, row["key"], align_left=True, tooltip=row["key"])
            it_n = NaturalItem(str(row["n_items"]), int(row["n_items"]))
            it_n.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.table_std.setItem(r, 1, it_n)
            p = row["mean_p"] * 100
            it_p = NaturalItem(f"{p:.1f}", float(p))
            it_p.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            it_p.setToolTip(f"성취율(전체) {p:.2f}%")
            ItemBarDelegate.set_bar(it_p, p, 100, "#2563eb", light=True)
            self.table_std.setItem(r, 2, it_p)
            for i, lv in enumerate(LEVELS):
                p_lv = row["by_level"].get(lv)
                col = 3 + i
                if p_lv is None:
                    _set_item(self.table_std, r, col, "-", align_right=True)
                else:
                    p_pct = p_lv * 100
                    bg = col_shade if p_pct < 50 else None
                    _set_item(self.table_std, r, col, f"{p_pct:.1f}", align_right=True, bg=bg)
        self.table_std.setSortingEnabled(True)
        self.table_std.sortByColumn(2, Qt.DescendingOrder)
        hh = self.table_std.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.Stretch)
        for i in range(1, self.table_std.columnCount()):
            hh.setSectionResizeMode(i, QHeaderView.Fixed)
            self._set_scaled_column_width(self.table_std, i, 72)
        self.table_std.verticalHeader().setDefaultSectionSize(self._px(30))
        self.canvas_std.setMinimumHeight(max(340, min(1400, 30 * len(rows) + 70)))
        self.canvas_std.set_figure(charts.fig_standard_attainment(rows))

    # ---------------------------------------------------------- 내보내기
    def _build_spliter_evidence_payload(self):
        select_items = sorted(self.exam.select_items, key=lambda it: it.number)
        serdap_items = sorted(self.exam.serdap_items, key=lambda it: it.number)
        serdap_max_total = sum(float(it.score) for it in serdap_items)
        serdap_stats = analyze_serdap(self.exam, levels_arr=getattr(self.overall, "levels_arr", None)) if serdap_items else None
        serdap_evidence_number = max([it.number for it in self.exam.items] or [0]) + 10000
        serdap_difficulties = {it.difficulty for it in serdap_items if it.difficulty}
        serdap_difficulty = next(iter(serdap_difficulties)) if len(serdap_difficulties) == 1 else ""
        stats_by_number = {s.item.number: s for s in self.item_stats}
        levels_arr = list(getattr(self.overall, "levels_arr", []) or [])

        source_files = dict(getattr(self.exam, "source_files", {}) or {})
        if self.fs_cuts.path():
            source_files["cuts"] = self.fs_cuts.path()

        def item_payload(it):
            stats = stats_by_number.get(it.number)
            return {
                "number": it.number,
                "type": it.item_type,
                "difficulty": it.difficulty,
                "score": round(float(it.score), 2),
                "answer": it.answer,
                "contentArea": it.content_area,
                "standardCode": it.standard_code,
                "standard": it.standard,
                "pValue": round(stats.p_value * 100, 2) if stats else None,
                "discrimination": round(stats.discrimination, 3) if stats else None,
                "pByLevel": {
                    lv: round(stats.p_by_level.get(lv, 0) * 100, 2)
                    for lv in LEVELS
                    if stats and lv in stats.p_by_level
                },
            }

        def serdap_payload():
            if not serdap_items or serdap_max_total <= 0:
                return None
            return {
                "number": serdap_evidence_number,
                "type": "서답형",
                "difficulty": serdap_difficulty,
                "score": round(float(serdap_max_total), 2),
                "answer": "부분점수",
                "contentArea": "서답형 전체",
                "standardCode": "",
                "standard": "서답형 전체 점수율",
                "pValue": round(serdap_stats.correct_rate * 100, 2) if serdap_stats else None,
                "discrimination": round(serdap_stats.discrimination, 3) if serdap_stats else None,
                "pByLevel": {
                    lv: round(serdap_stats.by_level.get(lv, 0) * 100, 2)
                    for lv in LEVELS
                    if serdap_stats and lv in serdap_stats.by_level
                },
                "isAggregate": True,
            }

        evidence_items = [item_payload(it) for it in select_items]
        serdap_evidence_item = serdap_payload()
        if serdap_evidence_item:
            evidence_items.append(serdap_evidence_item)

        students = []
        for idx, st in enumerate(self.exam.students):
            level = levels_arr[idx] if idx < len(levels_arr) else grade_level(st.final_score, self.exam.cut_scores)
            item_results = []
            for it in select_items:
                choice = str(st.answers.get(it.number, "")).strip()
                is_correct = choice == "."
                item_results.append({
                    "number": it.number,
                    "correct": is_correct,
                    "scoreRate": 100.0 if is_correct else 0.0,
                    "choice": choice,
                    "answer": it.answer,
                    "type": it.item_type,
                    "difficulty": it.difficulty,
                    "score": round(float(it.score), 2),
                    "contentArea": it.content_area,
                    "standardCode": it.standard_code,
                    "standard": it.standard,
                })
            if serdap_evidence_item and serdap_max_total > 0:
                serdap_rate = max(0.0, min(1.0, float(st.serdap_score) / serdap_max_total))
                item_results.append({
                    "number": serdap_evidence_number,
                    "correct": serdap_rate >= 0.5,
                    "scoreRate": round(serdap_rate * 100, 2),
                    "choice": f"{round(float(st.serdap_score), 2)}/{round(float(serdap_max_total), 2)}",
                    "answer": "부분점수",
                    "type": "서답형",
                    "difficulty": serdap_difficulty,
                    "score": round(float(serdap_max_total), 2),
                    "contentArea": "서답형 전체",
                    "standardCode": "",
                    "standard": "서답형 전체 점수율",
                    "isAggregate": True,
                })
            students.append({
                "id": st.sid,
                "classNo": st.class_no,
                "gradeClass": st.grade_class,
                "name": st.name,
                "level": level,
                "finalScore": round(float(st.final_score), 2),
                "total": round(float(st.total), 2),
                "pencilScore": round(float(st.total), 2),
                "performScore": round(float(st.perform_score), 2),
                "itemResults": item_results,
            })

        payload = {
            "kind": "goedusplit-spliter-evidence",
            "version": 1,
            "exportedAt": datetime.now().isoformat(timespec="seconds"),
            "subject": self.exam.subject,
            "grade": self.exam.grade,
            "semester": self.exam.semester,
            "cuts": {lv: round(float(self.exam.cut_scores.get(lv, 0)), 2) for lv in ["A", "B", "C", "D", "E"]},
            "weights": {
                "pencil": round(float(self.exam.weight_pencil), 2),
                "perform": round(float(self.exam.weight_perform), 2),
            },
            "sourceFiles": source_files,
            "levelSummary": {lv: int(self.overall.level_dist.get(lv, 0)) for lv in LEVELS},
            "items": evidence_items,
            "students": students,
        }
        return payload

    @staticmethod
    def _excel_cell_value(value):
        if value is None:
            return ""
        if isinstance(value, (str, int, float, bool)):
            return value
        return json.dumps(value, ensure_ascii=False)

    def _write_spliter_evidence_xlsx(self, path: str, payload: dict):
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "요약"
        header_fill = PatternFill("solid", fgColor="DDEDEA")
        title_font = Font(bold=True)

        def append(ws_, row):
            ws_.append([self._excel_cell_value(value) for value in row])

        append(ws, ["항목", "값"])
        append(ws, ["내보낸 시각", payload.get("exportedAt", "")])
        append(ws, ["과목", payload.get("subject", "")])
        append(ws, ["학년", payload.get("grade", "")])
        append(ws, ["학기", payload.get("semester", "")])
        append(ws, ["학생 수", len(payload.get("students", []))])
        append(ws, ["근거 문항 수", len(payload.get("items", []))])
        append(ws, ["지필 반영비율", payload.get("weights", {}).get("pencil", "")])
        append(ws, ["수행 반영비율", payload.get("weights", {}).get("perform", "")])
        for lv, label in [("A", "A/B"), ("B", "B/C"), ("C", "C/D"), ("D", "D/E"), ("E", "E/미도달")]:
            append(ws, [f"{label} 분할점수", payload.get("cuts", {}).get(lv, "")])
        ws.freeze_panes = "A2"

        ws_levels = wb.create_sheet("성취수준 요약")
        append(ws_levels, ["성취수준", "인원"])
        for lv in LEVELS:
            append(ws_levels, [lv, payload.get("levelSummary", {}).get(lv, 0)])
        ws_levels.freeze_panes = "A2"

        ws_items = wb.create_sheet("문항 근거")
        item_headers = [
            "문항", "유형", "난이도", "배점", "정답", "내용영역", "성취기준 코드", "성취기준",
            "전체 정답률(%)", "변별도", "A(%)", "B(%)", "C(%)", "D(%)", "E(%)", "미도달(%)", "서답형 집계",
        ]
        append(ws_items, item_headers)
        for item in payload.get("items", []):
            p_by_level = item.get("pByLevel") or {}
            append(ws_items, [
                item.get("number", ""),
                item.get("type", ""),
                item.get("difficulty", ""),
                item.get("score", ""),
                item.get("answer", ""),
                item.get("contentArea", ""),
                item.get("standardCode", ""),
                item.get("standard", ""),
                item.get("pValue", ""),
                item.get("discrimination", ""),
                p_by_level.get("A", ""),
                p_by_level.get("B", ""),
                p_by_level.get("C", ""),
                p_by_level.get("D", ""),
                p_by_level.get("E", ""),
                p_by_level.get("미도달", ""),
                "예" if item.get("isAggregate") else "",
            ])
        ws_items.freeze_panes = "A2"

        ws_students = wb.create_sheet("학생")
        student_headers = ["학생ID", "반/번호", "학년반", "이름", "성취도", "환산점수", "지필점수", "수행점수"]
        append(ws_students, student_headers)
        for st in payload.get("students", []):
            append(ws_students, [
                st.get("id", ""),
                st.get("classNo", ""),
                st.get("gradeClass", ""),
                st.get("name", ""),
                st.get("level", ""),
                st.get("finalScore", ""),
                st.get("pencilScore", ""),
                st.get("performScore", ""),
            ])
        ws_students.freeze_panes = "A2"

        ws_responses = wb.create_sheet("학생 응답")
        response_headers = [
            "학생ID", "반/번호", "이름", "성취도", "문항", "정오", "점수율(%)", "응답", "정답",
            "유형", "난이도", "배점", "성취기준 코드", "성취기준", "서답형 집계",
        ]
        append(ws_responses, response_headers)
        for st in payload.get("students", []):
            for result in st.get("itemResults", []):
                append(ws_responses, [
                    st.get("id", ""),
                    st.get("classNo", ""),
                    st.get("name", ""),
                    st.get("level", ""),
                    result.get("number", ""),
                    "정답" if result.get("correct") else "오답",
                    result.get("scoreRate", ""),
                    result.get("choice", ""),
                    result.get("answer", ""),
                    result.get("type", ""),
                    result.get("difficulty", ""),
                    result.get("score", ""),
                    result.get("standardCode", ""),
                    result.get("standard", ""),
                    "예" if result.get("isAggregate") else "",
                ])
        ws_responses.freeze_panes = "A2"

        ws_meta = wb.create_sheet("_앱내부데이터")
        append(ws_meta, ["이 시트는 앱 내부 복원용 원본 데이터입니다. 일반 사용자는 요약/문항/학생 시트를 보시면 됩니다."])
        append(ws_meta, [json.dumps(payload, ensure_ascii=False, indent=2)])
        ws_meta.sheet_state = "hidden"

        for sheet in wb.worksheets:
            for cell in sheet[1]:
                cell.font = title_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            for col in sheet.columns:
                letter = col[0].column_letter
                max_len = max(len(str(cell.value or "")) for cell in col[:80])
                sheet.column_dimensions[letter].width = min(max(max_len + 2, 10), 42)
        ws_items.column_dimensions["H"].width = 60
        ws_responses.column_dimensions["N"].width = 60
        wb.save(path)

    def _round_neis_rate(self, value: float | int | None) -> int:
        try:
            number = float(value)
        except (TypeError, ValueError):
            return 0
        if not math.isfinite(number):
            return 0
        return int(max(0, min(100, round(number / 5) * 5)))

    def _round_neis_level_rate(self, level: str, value: float | int | None, *, avoid_100: bool = True) -> int:
        try:
            number = float(value)
        except (TypeError, ValueError):
            return 0
        if not math.isfinite(number):
            return 0
        cap = 95 if avoid_100 else 100
        number = max(0, min(cap, number))
        if level == "E":
            return int(round(number))
        return int(max(0, min(cap, round(number / 5) * 5)))

    def _normalize_neis_item_type(self, value: str | None) -> str:
        text = str(value or "").strip()
        if re.search(r"서답|논술|주관|단답|서술", text):
            return "서답형"
        return "선택형"

    def _neis_item_key_from_parts(self, item_type: str | None, number) -> str:
        try:
            num = int(number)
        except (TypeError, ValueError):
            num = 0
        return f"{self._normalize_neis_item_type(item_type)}:{num}"

    def _neis_item_key(self, item) -> str:
        return self._neis_item_key_from_parts(getattr(item, "item_type", ""), getattr(item, "number", 0))

    def _neis_item_id(self, item) -> str:
        item_type = self._normalize_neis_item_type(getattr(item, "item_type", ""))
        prefix = "select" if item_type == "선택형" else "serdap"
        return f"{prefix}-{int(getattr(item, 'number', 0) or 0)}"

    def _enforce_neis_rate_order(self, rates: dict[str, float | int | None], *, avoid_100: bool = True) -> dict[str, int]:
        ordered = {
            level: self._round_neis_level_rate(level, rates.get(level, 0), avoid_100=avoid_100)
            for level in LEVELS_AE
        }
        for idx in range(len(LEVELS_AE) - 2, -1, -1):
            upper = LEVELS_AE[idx]
            lower = LEVELS_AE[idx + 1]
            ordered[upper] = max(ordered[upper], ordered[lower])
        return ordered

    def _minimum_target_rate(self, sample_size: int | None = None) -> int:
        if sample_size:
            sample = max(1, min(20, int(sample_size)))
            return int(math.ceil(math.ceil(sample * 2 / 3) / sample * 100))
        return TARGET_RATE_MIN_PERCENT

    def _enforce_target_rate_rules(
        self,
        rates: dict[str, float | int | None],
        target: str,
        sample_size: int | None = None,
    ) -> dict[str, int]:
        target = self._ai_review_level(target, "C")
        ordered = self._enforce_neis_rate_order(rates)
        target_index = LEVELS_AE.index(target)
        ordered[target] = max(ordered[target], self._minimum_target_rate(sample_size))
        for idx in range(target_index - 1, -1, -1):
            level = LEVELS_AE[idx]
            lower = LEVELS_AE[idx + 1]
            ordered[level] = max(ordered[level], ordered[lower])
        for idx in range(target_index + 1, len(LEVELS_AE)):
            level = LEVELS_AE[idx]
            upper = LEVELS_AE[idx - 1]
            ordered[level] = min(ordered[level], ordered[upper])
        return ordered

    def _item_stats_by_neis_key(self) -> dict[str, object]:
        return {self._neis_item_key(stats.item): stats for stats in self.item_stats}

    def _serdap_analysis_rates(self) -> dict[str, float] | None:
        if self.exam is None or not self.exam.serdap_items:
            return None
        try:
            stats = analyze_serdap(self.exam, levels_arr=getattr(self.overall, "levels_arr", None))
        except Exception:
            return None
        if not stats or not stats.by_level:
            return None
        return {level: float(stats.by_level.get(level, 0)) * 100 for level in LEVELS_AE}

    def _analysis_rates_for_item(self, item) -> dict[str, float] | None:
        item_type = self._normalize_neis_item_type(getattr(item, "item_type", ""))
        if item_type == "선택형":
            stats = self._item_stats_by_neis_key().get(self._neis_item_key(item))
            if not stats or not getattr(stats, "p_by_level", None):
                return None
            return {level: float(stats.p_by_level.get(level, 0)) * 100 for level in LEVELS_AE}
        return self._serdap_analysis_rates()

    def _item_default_target_level(self, item) -> str:
        item_type = self._normalize_neis_item_type(getattr(item, "item_type", ""))
        if item_type == "선택형":
            stats = self._item_stats_by_neis_key().get(self._neis_item_key(item))
            if stats:
                p_by_level = getattr(stats, "p_by_level", {}) or {}
                for level in ["E", "D", "C", "B", "A"]:
                    if p_by_level.get(level, -1) >= 0.5:
                        return level
        difficulty = (getattr(item, "difficulty", "") or "").strip()
        if difficulty == "쉬움":
            return "E"
        if difficulty == "어려움":
            return "B"
        return "C"

    def _normalize_target_rate_presets(self, value=None) -> dict[str, dict[str, int]]:
        source = value if isinstance(value, dict) else {}
        presets: dict[str, dict[str, int]] = {}
        for target in LEVELS_AE:
            base = DEFAULT_TARGET_RATE_PRESETS[target]
            raw_row = source.get(target) if isinstance(source.get(target), dict) else {}
            row = {}
            for level in LEVELS_AE:
                try:
                    row[level] = int(max(0, min(100, round(float(raw_row.get(level, base[level]))))))
                except Exception:
                    row[level] = int(base[level])
            presets[target] = self._enforce_target_rate_rules(row, target)
        return presets

    def _load_target_rate_presets(self) -> dict[str, dict[str, int]]:
        if not hasattr(self, "settings"):
            return self._normalize_target_rate_presets(DEFAULT_TARGET_RATE_PRESETS)
        raw = self.settings.value("spliter/target_rate_presets", "")
        if not raw:
            return self._normalize_target_rate_presets(DEFAULT_TARGET_RATE_PRESETS)
        try:
            return self._normalize_target_rate_presets(json.loads(str(raw)))
        except Exception:
            return self._normalize_target_rate_presets(DEFAULT_TARGET_RATE_PRESETS)

    def _save_target_rate_presets(self, presets: dict[str, dict[str, int]]):
        normalized = self._normalize_target_rate_presets(presets)
        self.settings.setValue("spliter/target_rate_presets", json.dumps(normalized, ensure_ascii=False))

    def _open_target_rate_presets_dialog(self, on_saved=None):
        dialog = QDialog(self)
        dialog.setWindowTitle(TARGET_RATE_PRESET_TITLE)
        dialog.resize(self._px(720), self._px(430))
        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(14, 14, 14, 14)
        layout.setSpacing(10)
        note = QLabel(
            "목표 성취수준을 바꿀 때 문항별 예상정답률 요약에 처음 적용할 기준입니다. "
            "각 행은 A≥B≥C≥D≥E 순서로 정리되고, 목표수준 학생은 최소 2/3 이상 맞히도록 보정됩니다."
        )
        note.setWordWrap(True)
        note.setProperty("role", "muted")
        layout.addWidget(note)

        table = QTableWidget(len(LEVELS_AE), len(LEVELS_AE) + 1)
        table.setHorizontalHeaderLabels(["목표수준", *[f"{level} 학생" for level in LEVELS_AE]])
        table.verticalHeader().setVisible(False)
        table.setEditTriggers(QTableWidget.DoubleClicked | QTableWidget.EditKeyPressed | QTableWidget.AnyKeyPressed)
        presets = self._load_target_rate_presets()
        for row, target in enumerate(LEVELS_AE):
            target_item = QTableWidgetItem(target)
            target_item.setTextAlignment(Qt.AlignCenter)
            target_item.setFlags(target_item.flags() & ~Qt.ItemIsEditable)
            table.setItem(row, 0, target_item)
            for col, level in enumerate(LEVELS_AE, start=1):
                item = QTableWidgetItem(str(presets[target][level]))
                item.setTextAlignment(Qt.AlignCenter)
                table.setItem(row, col, item)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(table, 1)

        buttons = QHBoxLayout()
        btn_reset = QPushButton("기본값으로")
        btn_cancel = QPushButton("취소")
        btn_save = QPushButton("저장하고 적용")
        btn_save.setProperty("role", "primary")
        buttons.addWidget(btn_reset)
        buttons.addStretch(1)
        buttons.addWidget(btn_cancel)
        buttons.addWidget(btn_save)
        layout.addLayout(buttons)

        def fill_defaults():
            defaults = self._normalize_target_rate_presets(DEFAULT_TARGET_RATE_PRESETS)
            for row, target in enumerate(LEVELS_AE):
                for col, level in enumerate(LEVELS_AE, start=1):
                    table.item(row, col).setText(str(defaults[target][level]))

        def collect() -> dict[str, dict[str, int]]:
            collected = {}
            for row, target in enumerate(LEVELS_AE):
                collected[target] = {}
                for col, level in enumerate(LEVELS_AE, start=1):
                    item = table.item(row, col)
                    try:
                        collected[target][level] = int(round(float((item.text() if item else "").strip())))
                    except Exception:
                        collected[target][level] = DEFAULT_TARGET_RATE_PRESETS[target][level]
            return self._normalize_target_rate_presets(collected)

        def save():
            choice = QMessageBox(dialog)
            choice.setWindowTitle(TARGET_RATE_PRESET_TITLE)
            choice.setIcon(QMessageBox.Question)
            choice.setText("설정을 저장한 뒤 현재 문항별 예상정답률 표도 바로 다시 계산할까요?")
            choice.setInformativeText("직접%로 입력한 값은 유지하고, 기준표로 만든 O/X 값만 다시 계산합니다.")
            save_only = choice.addButton("설정만 저장", QMessageBox.AcceptRole)
            save_apply = choice.addButton("저장 후 현재 표 재계산", QMessageBox.ActionRole)
            cancel = choice.addButton("취소", QMessageBox.RejectRole)
            choice.setDefaultButton(save_only)
            choice.exec()
            clicked = choice.clickedButton()
            if clicked is cancel:
                return
            apply_current = clicked is save_apply
            self._save_target_rate_presets(collect())
            if callable(on_saved):
                try:
                    on_saved(apply_current)
                except TypeError:
                    if apply_current:
                        on_saved()
            self._send_spliter_teacher_presets(apply_current=apply_current)
            message = (
                f"{TARGET_RATE_PRESET_TITLE}을 저장하고 현재 표 재계산을 요청했습니다."
                if apply_current
                else f"{TARGET_RATE_PRESET_TITLE}을 저장했습니다."
            )
            self.statusBar().showMessage(message, 6000)
            dialog.accept()

        btn_reset.clicked.connect(fill_defaults)
        btn_cancel.clicked.connect(dialog.reject)
        btn_save.clicked.connect(save)
        dialog.exec()

    def _target_level_rates(self, target: str, difficulty: str, sample_size: int = 20) -> dict[str, int]:
        target = self._ai_review_level(target, "C")
        difficulty = self._ai_review_difficulty(difficulty)
        base_by_target = self._load_target_rate_presets()
        difficulty_delta = {"쉬움": 10, "보통": 0, "어려움": -10}.get(difficulty, 0)
        rates = {level: base_by_target[target][level] + difficulty_delta for level in LEVELS_AE}
        return self._enforce_target_rate_rules(rates, target, sample_size)

    def _expected_rates_for_target(
        self,
        target: str,
        difficulty: str,
        sample_size: int = 20,
        *,
        item=None,
        rate_mode: str = "target",
    ) -> dict[str, int]:
        target_rates = self._target_level_rates(target, difficulty, sample_size)
        analysis_rates = self._analysis_rates_for_item(item) if item is not None else None
        if not analysis_rates or rate_mode == "target":
            return self._enforce_target_rate_rules(target_rates, target, sample_size)
        if rate_mode == "analysis":
            merged = {level: analysis_rates.get(level, target_rates[level]) for level in LEVELS_AE}
            return self._enforce_target_rate_rules(merged, target, sample_size)
        if rate_mode == "blend":
            merged = {
                level: target_rates[level] * 0.7 + analysis_rates.get(level, target_rates[level]) * 0.3
                for level in LEVELS_AE
            }
            return self._enforce_target_rate_rules(merged, target, sample_size)
        return self._enforce_target_rate_rules(target_rates, target, sample_size)

    def _judgments_from_rates(self, rates: dict[str, int], sample_size: int, judges: list[dict], target: str = "C") -> dict:
        sample_size = max(1, min(20, int(sample_size or 20)))
        target = self._ai_review_level(target, "C")
        min_target_count = math.ceil(sample_size * 2 / 3)
        judgments = {}
        for judge in judges:
            by_level = {}
            for level in LEVELS_AE:
                count = max(0, min(sample_size, round(rates[level] / 100 * sample_size)))
                if level == target:
                    count = max(count, min_target_count)
                by_level[level] = {
                    "correct": [idx < count for idx in range(sample_size)],
                    "targetRate": rates[level],
                    "overrideRate": rates[level],
                }
            judgments[judge["id"]] = by_level
        return judgments

    def _exam_item_by_neis_key(self) -> dict[str, object]:
        if self.exam is None:
            return {}
        return {self._neis_item_key(item): item for item in self.exam.items}

    def _default_neis_design_items(self) -> list[dict]:
        if self.exam is None:
            return []
        designs = []
        type_order = {"선택형": 0, "서답형": 1}
        for item in sorted(
            self.exam.items,
            key=lambda it: (
                type_order.get(self._normalize_neis_item_type(getattr(it, "item_type", "")), 9),
                getattr(it, "number", 0),
            ),
        ):
            item_type = self._normalize_neis_item_type(getattr(item, "item_type", ""))
            key = self._neis_item_key(item)
            designs.append({
                "id": self._neis_item_id(item),
                "key": key,
                "source_key": key,
                "number": int(getattr(item, "number", 0) or 0),
                "type": item_type,
                "difficulty": getattr(item, "difficulty", "") or "보통",
                "points": round(float(getattr(item, "score", 0) or 0), 2),
                "target": self._item_default_target_level(item),
                "standard": " ".join(
                    part for part in [getattr(item, "standard_code", ""), getattr(item, "standard", "")]
                    if part
                ).strip(),
                "source_item": item,
                "rates": None,
            })
        return designs

    def _parse_neis_int(self, value, default: int = 1) -> int:
        try:
            text = str(value).replace(",", "").strip()
            return max(1, int(round(float(text))))
        except Exception:
            return max(1, int(default))

    def _parse_neis_float(self, value, default: float = 0.0) -> float:
        try:
            text = str(value).replace(",", "").strip()
            return max(0.0, float(text))
        except Exception:
            return max(0.0, float(default))

    def _normalize_neis_difficulty(self, value) -> str:
        text = str(value or "").strip()
        return text if text in {"쉬움", "보통", "어려움"} else "보통"

    def _normalize_neis_rates(self, rates) -> dict[str, float] | None:
        if not isinstance(rates, dict):
            return None
        normalized = {}
        for level in LEVELS_AE:
            try:
                normalized[level] = max(0.0, min(100.0, float(rates.get(level, 0))))
            except Exception:
                return None
        return normalized

    def _neis_rate_text(self, rates: dict[str, float | int]) -> str:
        def fmt(value) -> str:
            try:
                number = float(value)
            except Exception:
                return "0"
            if abs(number - round(number)) < 0.01:
                return str(int(round(number)))
            return f"{number:.1f}".rstrip("0").rstrip(".")

        return " / ".join(f"{level}{fmt(rates.get(level, 0))}" for level in LEVELS_AE)

    def _neis_design_item_from_cells(self, table: QTableWidget, row: int) -> dict | None:
        number_item = table.item(row, 0)
        type_combo = table.cellWidget(row, 1)
        difficulty_combo = table.cellWidget(row, 2)
        points_item = table.item(row, 3)
        target_combo = table.cellWidget(row, 4)
        standard_item = table.item(row, 6)
        if number_item is None or not isinstance(type_combo, QComboBox) or not isinstance(target_combo, QComboBox):
            return None
        number = self._parse_neis_int(number_item.text(), row + 1)
        item_type = self._normalize_neis_item_type(type_combo.currentText())
        if isinstance(difficulty_combo, QComboBox):
            difficulty = self._normalize_neis_difficulty(difficulty_combo.currentText())
        else:
            difficulty = self._normalize_neis_difficulty(table.item(row, 2).text() if table.item(row, 2) else "")
        points = self._parse_neis_float(points_item.text() if points_item else 0.0, 0.0)
        target = self._ai_review_level(target_combo.currentText(), "C")
        key = self._neis_item_key_from_parts(item_type, number)
        source_key = number_item.data(NEIS_SOURCE_KEY_ROLE) or number_item.data(Qt.UserRole) or key
        item_map = self._exam_item_by_neis_key()
        source_item = item_map.get(key) or item_map.get(str(source_key))
        rates = self._normalize_neis_rates(number_item.data(NEIS_RATES_ROLE))
        return {
            "id": str(number_item.data(Qt.UserRole) or key),
            "key": key,
            "source_key": str(source_key),
            "number": number,
            "type": item_type,
            "difficulty": difficulty,
            "points": round(points, 2),
            "target": target,
            "standard": standard_item.text().strip() if standard_item else "",
            "source_item": source_item,
            "rates": rates,
            "sampleSize": number_item.data(Qt.UserRole + 1),
        }

    def _collect_neis_design_items(self, table: QTableWidget) -> list[dict]:
        designs = []
        for row in range(table.rowCount()):
            design = self._neis_design_item_from_cells(table, row)
            if design is not None:
                designs.append(design)
        return designs

    def _project_from_neis_targets(self, design_items: list[dict], sample_size: int = 20, rate_mode: str = "target") -> dict:
        sample_size = max(1, min(20, int(sample_size or 20)))
        judges = [{"id": "teacher-1", "name": "검토안 1"}, {"id": "teacher-2", "name": "검토안 2"}]
        items = []
        for idx, design in enumerate(design_items, start=1):
            item_type = self._normalize_neis_item_type(design.get("type", ""))
            difficulty = self._normalize_neis_difficulty(design.get("difficulty", ""))
            target = self._ai_review_level(design.get("target", ""), "C")
            source_item = design.get("source_item")
            item_sample = max(1, min(20, int(design.get("sampleSize") or sample_size)))
            rates = self._normalize_neis_rates(design.get("rates"))
            if rates is None:
                rates = self._expected_rates_for_target(
                    target,
                    difficulty,
                    item_sample,
                    item=source_item,
                    rate_mode=rate_mode,
                )
            rates = self._enforce_target_rate_rules(rates, target, item_sample)
            number = self._parse_neis_int(design.get("number", idx), idx)
            standard = str(design.get("standard", "")).strip()
            items.append({
                "id": str(design.get("id") or f"neis-{idx}-{item_type}-{number}"),
                "number": number,
                "title": f"{item_type} {number}번",
                "standard": standard,
                "points": round(float(design.get("points", 0) or 0), 2),
                "sampleSize": item_sample,
                "type": item_type,
                "difficulty": difficulty,
                "targetLevel": target,
                "judgmentsByJudge": self._judgments_from_rates(rates, item_sample, judges, target),
                "evidence": ["NEIS 설계표", "계산기 현재값" if design.get("rates") else rate_mode],
                "note": "NEIS 입력표에서 수정한 문항 수, 문항구분, 난이도, 배점, 목표 성취수준을 반영했습니다.",
            })
        project = {
            "version": 1,
            "judges": judges,
            "activeJudgeId": "teacher-1",
            "items": items,
            "evidenceMode": "difficultyAverage",
        }
        if self.exam is not None and self.overall is not None:
            project["evidenceData"] = self._build_spliter_evidence_payload()
        return project

    def _build_neis_expected_rows(self, design_items: list[dict], sample_size: int = 20, rate_mode: str = "target") -> list[dict]:
        groups: dict[tuple[str, str], dict] = {}
        for design in design_items:
            item_type = self._normalize_neis_item_type(design.get("type", ""))
            difficulty = self._normalize_neis_difficulty(design.get("difficulty", ""))
            target = self._ai_review_level(design.get("target", ""), "C")
            points = float(design.get("points", 0) or 0)
            rates = self._normalize_neis_rates(design.get("rates"))
            if rates is None:
                rates = self._expected_rates_for_target(
                    target,
                    difficulty,
                    sample_size,
                    item=design.get("source_item"),
                    rate_mode=rate_mode,
                )
            rates = self._enforce_target_rate_rules(rates, target, sample_size)
            number = self._parse_neis_int(design.get("number", 0), 0)
            group_key = (item_type, difficulty)
            group = groups.setdefault(group_key, {
                "문항구분": item_type,
                "난이도": difficulty,
                "문항번호": [],
                "목표수준": [],
                "문항수": 0,
                "배점합": 0.0,
                "weighted": {level: 0.0 for level in LEVELS_AE},
                "weight_sum": 0.0,
            })
            group["문항번호"].append(number)
            group["목표수준"].append(f"{number}:{target}")
            group["문항수"] += 1
            group["배점합"] += points
            weight = points if points > 0 else 1.0
            group["weight_sum"] += weight
            for level in LEVELS_AE:
                group["weighted"][level] += float(rates[level]) * weight
        order_type = {"선택형": 0, "서답형": 1}
        order_diff = {"쉬움": 0, "보통": 1, "어려움": 2}
        rows = []
        for (_type, _difficulty), group in sorted(
            groups.items(),
            key=lambda kv: (order_type.get(kv[0][0], 9), order_diff.get(kv[0][1], 9), kv[0][1]),
        ):
            denom = group["weight_sum"] if group["weight_sum"] > 0 else max(1.0, float(group["문항수"]))
            row = {
                "문항구분": group["문항구분"],
                "난이도": group["난이도"],
                "해당문항번호": ", ".join(str(n) for n in sorted(group["문항번호"])),
                "목표수준": ", ".join(group["목표수준"]),
                "문항수": group["문항수"],
                "배점합": round(group["배점합"], 2),
            }
            for level in LEVELS_AE:
                row[level] = group["weighted"][level] / denom
            ordered = self._enforce_neis_rate_order(row)
            for level in LEVELS_AE:
                row[level] = ordered[level]
            rows.append(row)
        return rows

    def _neis_row_headers(self) -> list[str]:
        return ["문항구분", "난이도", "해당문항번호", "문항수", "배점합", "A", "B", "C", "D", "E", "목표수준"]

    def _neis_rows_to_tsv(self, rows: list[dict]) -> str:
        headers = self._neis_row_headers()
        lines = ["\t".join(headers)]
        for row in rows:
            lines.append("\t".join(str(row.get(header, "")) for header in headers))
        return "\n".join(lines)

    def _neis_rows_to_markdown(self, rows: list[dict]) -> str:
        headers = self._neis_row_headers()

        def cell(value) -> str:
            return str(value if value is not None else "").replace("|", "\\|").replace("\n", "<br>")

        lines = ["| " + " | ".join(headers) + " |"]
        lines.append("| " + " | ".join("---" for _ in headers) + " |")
        for row in rows:
            lines.append("| " + " | ".join(cell(row.get(header, "")) for header in headers) + " |")
        return "\n".join(lines)

    def _send_neis_targets_to_spliter(self, table: QTableWidget, sample_combo: QComboBox, mode_combo: QComboBox, dialog: QDialog | None = None):
        if self.spliter_view is None:
            QMessageBox.information(self, "예상정답률 계산기", "이 환경에서는 내장 예상정답률 계산기를 열 수 없습니다.")
            return
        designs = self._collect_neis_design_items(table)
        if not designs:
            QMessageBox.warning(self, "NEIS 입력표", "계산기로 보낼 문항이 없습니다.")
            return
        sample_size = int(sample_combo.currentText())
        rate_mode = mode_combo.currentData() or "target"
        self._spliter_pending_project_payload = self._project_from_neis_targets(designs, sample_size, rate_mode)
        if self.exam is not None and self.overall is not None:
            self._spliter_pending_payload = self._build_spliter_evidence_payload()
        self._activate_spliter_tab_for_pending_payloads()
        if dialog is not None:
            dialog.accept()
        self.statusBar().showMessage(f"수정한 {len(designs)}문항 설계안을 예상정답률 계산기로 보냈습니다.", 7000)

    def _copy_neis_expected_rows(self, table: QTableWidget, sample_combo: QComboBox, mode_combo: QComboBox, *, markdown: bool = False):
        rows = self._build_neis_expected_rows(
            self._collect_neis_design_items(table),
            int(sample_combo.currentText()),
            mode_combo.currentData() or "target",
        )
        if not rows:
            QMessageBox.warning(self, "NEIS 입력표", "복사할 문항 묶음이 없습니다.")
            return
        text = self._neis_rows_to_markdown(rows) if markdown else self._neis_rows_to_tsv(rows)
        QApplication.clipboard().setText(text)
        label = "마크다운" if markdown else "엑셀용"
        self.statusBar().showMessage(f"NEIS 입력용 표 {len(rows)}행을 {label} 형식으로 복사했습니다.", 6000)

    def _save_neis_rows_xlsx(self, rows: list[dict]):
        if not rows:
            QMessageBox.warning(self, "NEIS 입력표", "저장할 문항 묶음이 없습니다.")
            return
        safe_subject = "".join(
            "_" if ch in "\\/:*?\"<>|" else ch
            for ch in ((self.exam.subject if self.exam else "") or "goedusplit")
        ).strip() or "goedusplit"
        path, _ = QFileDialog.getSaveFileName(
            self,
            "NEIS 예상정답률 입력표 저장",
            str(Path.home() / "Desktop" / f"{safe_subject}_NEIS_예상정답률_입력표.xlsx"),
            "Excel 통합문서 (*.xlsx)",
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "NEIS 입력표"
            headers = self._neis_row_headers()
            ws.append(headers)
            for row in rows:
                ws.append([row.get(header, "") for header in headers])
            fill = PatternFill("solid", fgColor="DDEDEA")
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.freeze_panes = "A2"
            for column in ws.columns:
                letter = column[0].column_letter
                width = max(len(str(cell.value or "")) for cell in column[:80]) + 2
                ws.column_dimensions[letter].width = min(max(width, 10), 46)
            wb.save(path)
        except Exception as exc:
            QMessageBox.critical(self, "NEIS 입력표", f"저장하지 못했습니다.\n{exc}")
            return
        QMessageBox.information(self, "NEIS 입력표", f"저장했습니다.\n{path}")

    def _save_neis_expected_rows_xlsx(self, table: QTableWidget, sample_combo: QComboBox, mode_combo: QComboBox):
        rows = self._build_neis_expected_rows(
            self._collect_neis_design_items(table),
            int(sample_combo.currentText()),
            mode_combo.currentData() or "target",
        )
        self._save_neis_rows_xlsx(rows)

    def _fetch_spliter_project(self, callback, *, title: str = "NEIS 입력표"):
        if self.spliter_view is None:
            QMessageBox.information(self, title, "이 환경에서는 내장 예상정답률 계산기를 열 수 없습니다.")
            return
        if not self._spliter_loaded:
            QMessageBox.information(self, title, "예상정답률 계산기 탭을 먼저 연 뒤 다시 시도해 주세요.")
            return
        script = "window.__GOEDUSPLIT_GET_PROJECT__ ? window.__GOEDUSPLIT_GET_PROJECT__() : null"
        self.spliter_view.page().runJavaScript(script, callback)

    def _rate_from_spliter_judgment(self, judgment) -> float | None:
        if not isinstance(judgment, dict):
            return None
        override = judgment.get("overrideRate")
        if override not in (None, ""):
            try:
                return max(0.0, min(100.0, float(override)))
            except Exception:
                pass
        correct = judgment.get("correct")
        if isinstance(correct, list) and correct:
            return sum(1 for value in correct if bool(value)) / len(correct) * 100
        target = judgment.get("targetRate")
        if target not in (None, ""):
            try:
                return max(0.0, min(100.0, float(target)))
            except Exception:
                pass
        return None

    def _rates_from_spliter_item(self, item: dict, judges: list[dict]) -> dict[str, int]:
        judgments = item.get("judgmentsByJudge") if isinstance(item, dict) else None
        if not isinstance(judgments, dict):
            judgments = {}
        judge_ids = [str(j.get("id")) for j in judges if isinstance(j, dict) and j.get("id")]
        if not judge_ids:
            judge_ids = [str(key) for key in judgments.keys()]
        fallback = self._target_level_rates(
            self._ai_review_level(item.get("targetLevel", ""), "C"),
            self._normalize_neis_difficulty(item.get("difficulty", "")),
        )
        rates = {}
        for level in LEVELS_AE:
            values = []
            for judge_id in judge_ids:
                by_level = judgments.get(judge_id)
                if isinstance(by_level, dict):
                    rate = self._rate_from_spliter_judgment(by_level.get(level))
                    if rate is not None:
                        values.append(rate)
            rates[level] = sum(values) / len(values) if values else fallback[level]
        return self._enforce_target_rate_rules(
            rates,
            self._ai_review_level(item.get("targetLevel", ""), "C"),
            self._parse_neis_int(item.get("sampleSize", 20), 20),
        )

    def _neis_design_items_from_spliter_project(self, project) -> list[dict]:
        if not isinstance(project, dict):
            return []
        raw_items = project.get("items")
        if not isinstance(raw_items, list):
            return []
        judges = project.get("judges") if isinstance(project.get("judges"), list) else []
        item_map = self._exam_item_by_neis_key()
        designs = []
        for idx, item in enumerate(raw_items, start=1):
            if not isinstance(item, dict):
                continue
            number = self._parse_neis_int(item.get("number", idx), idx)
            item_type = self._normalize_neis_item_type(item.get("type", ""))
            key = self._neis_item_key_from_parts(item_type, number)
            source_item = item_map.get(key)
            designs.append({
                "id": str(item.get("id") or f"calc-{idx}-{number}"),
                "key": key,
                "source_key": key,
                "number": number,
                "type": item_type,
                "difficulty": self._normalize_neis_difficulty(item.get("difficulty", "")),
                "points": round(self._parse_neis_float(item.get("points", 0.0), 0.0), 2),
                "target": self._ai_review_level(item.get("targetLevel", ""), "C"),
                "standard": str(item.get("standard", "") or "").strip(),
                "source_item": source_item,
                "rates": self._rates_from_spliter_item(item, judges),
                "sampleSize": self._parse_neis_int(item.get("sampleSize", 20), 20),
            })
        return designs

    def open_neis_expected_rate_dialog(self):
        if self.exam is None or self.overall is None:
            QMessageBox.warning(self, "NEIS 입력표", "먼저 분석을 실행해 주세요.")
            return
        dialog = QDialog(self)
        dialog.setWindowTitle("NEIS 예상정답률 입력표")
        dialog.resize(self._px(1220), self._px(700))
        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(14, 14, 14, 14)
        layout.setSpacing(10)
        note = QLabel(
            "이전 분석자료를 기준으로 다음 시험 문항 설계표를 만듭니다. 문항 수, 문항구분, 난이도, 배점, "
            "목표수준을 수정할 수 있고, 계산기 현재값을 불러오면 계산기에서 교사가 고친 정답률을 "
            "NEIS 표에 그대로 반영합니다. 논술형·주관식은 NEIS 기준에 맞춰 서답형으로 통일합니다."
        )
        note.setProperty("role", "muted")
        note.setWordWrap(True)
        layout.addWidget(note)

        top = QHBoxLayout()
        top.addWidget(QLabel("가상학생"))
        sample_combo = QComboBox()
        for value in [10, 20]:
            sample_combo.addItem(str(value))
        sample_combo.setCurrentText("20")
        top.addWidget(sample_combo)
        top.addWidget(QLabel("산출 기준"))
        mode_combo = QComboBox()
        mode_combo.addItem("목표수준 기준", "target")
        mode_combo.addItem("목표+분석 혼합", "blend")
        mode_combo.addItem("분석자료 우선", "analysis")
        mode_combo.setCurrentIndex(0)
        top.addWidget(mode_combo)
        btn_defaults = QPushButton(f"{TARGET_RATE_PRESET_TITLE}…")
        btn_defaults.setToolTip("목표 성취수준 A~E를 선택했을 때 처음 적용할 A~E 학생 예상정답률 기준표를 정합니다.")
        top.addWidget(btn_defaults)
        btn_add = QPushButton("행 추가")
        top.addWidget(btn_add)
        btn_delete = QPushButton("선택 행 삭제")
        top.addWidget(btn_delete)
        btn_reload_analysis = QPushButton("분석자료 기준 새로 채우기")
        top.addWidget(btn_reload_analysis)
        btn_load_calc = QPushButton("계산기 현재값 불러오기")
        btn_load_calc.setToolTip("예상정답률 계산기에서 문항 수, 목표수준, O/X, 직접%를 수정한 현재 상태를 NEIS 설계표로 가져옵니다.")
        top.addWidget(btn_load_calc)
        top.addStretch(1)
        layout.addLayout(top)

        table = QTableWidget(0, 8)
        table.setHorizontalHeaderLabels(["문항", "문항구분", "난이도", "배점", "목표수준", "A~E 제안", "성취기준", ""])
        table.setSelectionBehavior(QTableWidget.SelectRows)
        table.setEditTriggers(QTableWidget.DoubleClicked | QTableWidget.EditKeyPressed | QTableWidget.AnyKeyPressed)
        table.verticalHeader().setVisible(False)

        def row_for_widget(widget: QWidget) -> int:
            for r in range(table.rowCount()):
                for c in (1, 2, 4, 7):
                    if table.cellWidget(r, c) is widget:
                        return r
            return -1

        def update_row(row: int, *, clear_custom: bool = False):
            if row < 0 or row >= table.rowCount():
                return
            number_cell = table.item(row, 0)
            if number_cell is None:
                return
            if clear_custom:
                number_cell.setData(NEIS_RATES_ROLE, None)
            design = self._neis_design_item_from_cells(table, row)
            if design is None:
                return
            rates = self._normalize_neis_rates(design.get("rates"))
            from_calculator = rates is not None
            if rates is None:
                rates = self._expected_rates_for_target(
                    design["target"],
                    design["difficulty"],
                    int(sample_combo.currentText()),
                    item=design.get("source_item"),
                    rate_mode=mode_combo.currentData() or "target",
                )
            rates = self._enforce_target_rate_rules(rates, design["target"], int(sample_combo.currentText()))
            preview = QTableWidgetItem(self._neis_rate_text(rates))
            preview.setTextAlignment(Qt.AlignCenter)
            preview.setFlags(preview.flags() & ~Qt.ItemIsEditable)
            if from_calculator:
                preview.setToolTip("계산기에서 가져온 현재 정답률입니다. 목표수준·문항구분·난이도를 바꾸면 새 제안값으로 다시 계산됩니다.")
            table.setItem(row, 5, preview)

        def update_all_rows():
            for r in range(table.rowCount()):
                update_row(r)

        def combo_changed(combo: QComboBox):
            update_row(row_for_widget(combo), clear_custom=True)

        def delete_button_clicked(button: QPushButton):
            row = row_for_widget(button)
            if row >= 0:
                table.removeRow(row)

        def set_row(row: int, design: dict):
            table.blockSignals(True)
            try:
                while table.rowCount() <= row:
                    table.insertRow(table.rowCount())
                number = self._parse_neis_int(design.get("number", row + 1), row + 1)
                item_type = self._normalize_neis_item_type(design.get("type", ""))
                key = str(design.get("key") or self._neis_item_key_from_parts(item_type, number))
                number_cell = QTableWidgetItem(str(number))
                number_cell.setTextAlignment(Qt.AlignCenter)
                number_cell.setData(Qt.UserRole, str(design.get("id") or key))
                number_cell.setData(Qt.UserRole + 1, design.get("sampleSize"))
                number_cell.setData(NEIS_SOURCE_KEY_ROLE, str(design.get("source_key") or key))
                rates = self._normalize_neis_rates(design.get("rates"))
                if rates is not None:
                    number_cell.setData(NEIS_RATES_ROLE, rates)
                table.setItem(row, 0, number_cell)

                type_combo = QComboBox()
                type_combo.addItems(["선택형", "서답형"])
                type_combo.setCurrentText(item_type)
                table.setCellWidget(row, 1, type_combo)

                difficulty_combo = QComboBox()
                difficulty_combo.addItems(["쉬움", "보통", "어려움"])
                difficulty_combo.setCurrentText(self._normalize_neis_difficulty(design.get("difficulty", "")))
                table.setCellWidget(row, 2, difficulty_combo)

                score_cell = QTableWidgetItem(str(round(float(design.get("points", 0) or 0), 2)))
                score_cell.setTextAlignment(Qt.AlignCenter)
                table.setItem(row, 3, score_cell)

                target_combo = QComboBox()
                target_combo.addItems(LEVELS_AE)
                target_combo.setCurrentText(self._ai_review_level(design.get("target", ""), "C"))
                table.setCellWidget(row, 4, target_combo)

                standard_cell = QTableWidgetItem(str(design.get("standard", "") or ""))
                table.setItem(row, 6, standard_cell)

                delete_button = QPushButton("삭제")
                delete_button.setProperty("role", "danger")
                table.setCellWidget(row, 7, delete_button)
            finally:
                table.blockSignals(False)
            type_combo.currentTextChanged.connect(lambda _text, combo=type_combo: combo_changed(combo))
            difficulty_combo.currentTextChanged.connect(lambda _text, combo=difficulty_combo: combo_changed(combo))
            target_combo.currentTextChanged.connect(lambda _text, combo=target_combo: combo_changed(combo))
            delete_button.clicked.connect(lambda _checked=False, button=delete_button: delete_button_clicked(button))
            update_row(row)

        def populate_table(designs: list[dict]):
            table.blockSignals(True)
            try:
                table.setRowCount(0)
            finally:
                table.blockSignals(False)
            for row, design in enumerate(designs):
                set_row(row, design)
            update_all_rows()

        def add_blank_row():
            numbers = []
            for r in range(table.rowCount()):
                item = table.item(r, 0)
                if item is not None:
                    numbers.append(self._parse_neis_int(item.text(), r + 1))
            next_number = max(numbers or [0]) + 1
            set_row(table.rowCount(), {
                "id": f"manual-{datetime.now().timestamp()}-{next_number}",
                "number": next_number,
                "type": "선택형",
                "difficulty": "보통",
                "points": 5.0,
                "target": "C",
                "standard": "",
            })

        def delete_selected_rows():
            rows = sorted({index.row() for index in table.selectionModel().selectedRows()}, reverse=True)
            if not rows:
                return
            for row in rows:
                table.removeRow(row)

        def on_item_changed(item: QTableWidgetItem):
            if item.column() == 0:
                update_row(item.row(), clear_custom=True)
            elif item.column() in (3, 6):
                update_row(item.row())

        def load_current_calculator_project():
            def done(project):
                designs = self._neis_design_items_from_spliter_project(project)
                if not designs:
                    QMessageBox.information(self, "NEIS 입력표", "계산기에서 가져올 문항 현재값이 없습니다.")
                    return
                populate_table(designs)
                self.statusBar().showMessage(f"계산기 현재값 {len(designs)}문항을 NEIS 입력표로 불러왔습니다.", 6000)

            self._fetch_spliter_project(done, title="NEIS 입력표")

        table.itemChanged.connect(on_item_changed)
        btn_add.clicked.connect(add_blank_row)
        btn_delete.clicked.connect(delete_selected_rows)
        btn_reload_analysis.clicked.connect(lambda: populate_table(self._default_neis_design_items()))
        btn_load_calc.clicked.connect(load_current_calculator_project)
        btn_defaults.clicked.connect(
            lambda: self._open_target_rate_presets_dialog(
                lambda apply_current=False: update_all_rows() if apply_current else None
            )
        )
        populate_table(self._default_neis_design_items())
        sample_combo.currentTextChanged.connect(lambda _text: update_all_rows())
        mode_combo.currentIndexChanged.connect(lambda _idx: update_all_rows())
        table.horizontalHeader().setSectionResizeMode(6, QHeaderView.Stretch)
        for col, width in enumerate([62, 92, 78, 70, 92, 210, 260, 62]):
            table.setColumnWidth(col, self._px(width))
        layout.addWidget(table, 1)

        buttons = QHBoxLayout()
        btn_copy_excel = QPushButton("엑셀용 복사")
        btn_copy_excel.clicked.connect(lambda: self._copy_neis_expected_rows(table, sample_combo, mode_combo, markdown=False))
        buttons.addWidget(btn_copy_excel)
        btn_copy_md = QPushButton("마크다운 복사")
        btn_copy_md.clicked.connect(lambda: self._copy_neis_expected_rows(table, sample_combo, mode_combo, markdown=True))
        buttons.addWidget(btn_copy_md)
        btn_save = QPushButton("엑셀 저장…")
        btn_save.clicked.connect(lambda: self._save_neis_expected_rows_xlsx(table, sample_combo, mode_combo))
        buttons.addWidget(btn_save)
        buttons.addStretch(1)
        btn_send = QPushButton("수정 설계로 계산기에 보내기")
        btn_send.setProperty("role", "primary")
        btn_send.clicked.connect(lambda: self._send_neis_targets_to_spliter(table, sample_combo, mode_combo, dialog))
        buttons.addWidget(btn_send)
        btn_close = QPushButton("닫기")
        btn_close.clicked.connect(dialog.reject)
        buttons.addWidget(btn_close)
        layout.addLayout(buttons)
        dialog.exec()

    def export_spliter_evidence(self):
        if self.exam is None or self.overall is None:
            QMessageBox.warning(self, "예상정답률 계산기", "먼저 분석을 실행해 주세요.")
            return

        safe_subject = "".join(
            "_" if ch in "\\/:*?\"<>|" else ch for ch in (self.exam.subject or "goedusplit")
        ).strip() or "goedusplit"
        default_name = f"{safe_subject}_expected_rate_근거.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self,
            "예상정답률 근거 엑셀 저장",
            str(Path.home() / "Desktop" / default_name),
            "Excel 통합문서 (*.xlsx)",
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        payload = self._build_spliter_evidence_payload()
        students = payload["students"]
        evidence_items = payload["items"]
        select_count = len(self.exam.select_items)
        serdap_count = len(self.exam.serdap_items)
        serdap_text = f" · 서답형 {serdap_count}문항은 전체 점수율로 반영" if serdap_count else ""

        try:
            self._write_spliter_evidence_xlsx(path, payload)
        except Exception as e:
            QMessageBox.critical(self, "예상정답률 계산기", f"근거 엑셀 저장 중 오류가 발생했습니다.\n{e}")
            return

        self.statusBar().showMessage(f"예상정답률 근거 엑셀 저장 완료 · {path}", 8000)
        QMessageBox.information(
            self,
            "예상정답률 계산기",
            f"근거 엑셀을 저장했습니다.\n학생 {len(students)}명 · 근거 문항 {len(evidence_items)}개"
            f" (선택형 {select_count}개{serdap_text})\n\n{path}",
        )

    def export_csv(self):
        if self.exam is None:
            QMessageBox.warning(self, "내보내기", "먼저 분석을 실행해 주세요.")
            return
        directory = QFileDialog.getExistingDirectory(self, "CSV 저장 폴더 선택")
        if not directory:
            return

        # 그래프 PNG도 함께 저장할지 물어보기 (체크박스 있는 다이얼로그)
        msg = QMessageBox(self)
        msg.setIcon(QMessageBox.Question)
        msg.setWindowTitle("내보내기 옵션")
        msg.setText("결과 CSV 4개를 저장합니다.\n그래프(.png)도 함께 저장하시겠습니까?")
        msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel)
        msg.setDefaultButton(QMessageBox.Yes)
        msg.button(QMessageBox.Yes).setText("CSV + 그래프 PNG")
        msg.button(QMessageBox.No).setText("CSV만")
        msg.button(QMessageBox.Cancel).setText("취소")
        ans = msg.exec()
        if ans == QMessageBox.Cancel:
            return
        with_graphs = (ans == QMessageBox.Yes)

        out = Path(directory)
        ov = self.overall
        try:
            self._write_csv(out / "학생결과.csv",
                ["학번","반/번호","이름","학급","선택형","서답형","기타","지필총점","수행환산","환산점수","성취도"],
                [[s.sid, s.class_no, s.name, s.grade_class, s.multi_score, s.serdap_score,
                  s.etc_score, round(s.total,2), round(s.perform_score,2),
                  round(s.final_score,2), ov.levels_arr[i]]
                 for i, s in enumerate(self.exam.students)])
            self._write_csv(out / "문항분석.csv",
                ["문항","예상난이도","정답률(%)","변별도",
                 "응답1(%)","응답2(%)","응답3(%)","응답4(%)","응답5(%)","무응답(%)",
                 "A(%)","B(%)","C(%)","D(%)","E(%)","미도달(%)",
                 "내용영역","성취기준코드","성취기준"],
                [[s.item.number, s.item.difficulty,
                  round(s.p_value*100,2), round(s.discrimination,3),
                  *[round(s.choice_dist.get(c,0)*100,2) for c in (1,2,3,4,5,0)],
                  *[round(s.p_by_level.get(lv,0)*100,2) if lv in s.p_by_level else "" for lv in LEVELS],
                  s.item.content_area, s.item.standard_code, s.item.standard]
                 for s in self.item_stats])
            self._write_csv(out / "성취기준별.csv",
                ["성취기준","문항수","성취율(전체,%)","A(%)","B(%)","C(%)","D(%)","E(%)","미도달(%)"],
                [[r["key"], r["n_items"], round(r["mean_p"]*100,2)] +
                 [round(r["by_level"].get(lv,0)*100,2) if r["by_level"].get(lv) is not None else "" for lv in LEVELS]
                 for r in ov.by_standard])
            self._write_csv(out / "성취수준분포.csv",
                ["성취도","인원","비율(%)","평균","표준편차"],
                [[lv, ov.level_dist[lv], round(ov.level_dist_pct[lv],2),
                  round(ov.level_mean[lv],2), round(ov.level_std[lv],2)] for lv in LEVELS])

            saved_pngs = []
            if with_graphs:
                graph_dir = out / "그래프"
                graph_dir.mkdir(exist_ok=True)
                # 도넛
                f = charts.fig_level_donut(ov.level_dist_pct, ov.level_dist)
                f.savefig(graph_dir / "01_성취수준_도넛.png", dpi=150, bbox_inches="tight")
                saved_pngs.append("01_성취수준_도넛.png")
                # 환산점수 히스토그램
                f = charts.fig_score_histogram_colored(
                    [s.final_score for s in self.exam.students], ov.levels_arr)
                f.savefig(graph_dir / "02_환산점수_분포.png", dpi=150, bbox_inches="tight")
                saved_pngs.append("02_환산점수_분포.png")
                # 평균±표준편차
                f = charts.fig_level_means_with_error(ov.level_mean, ov.level_std)
                f.savefig(graph_dir / "03_성취수준별_평균.png", dpi=150, bbox_inches="tight")
                saved_pngs.append("03_성취수준별_평균.png")
                # 학급별 평균
                f = charts.fig_class_means(ov.by_class)
                f.savefig(graph_dir / "04_학급별_평균.png", dpi=150, bbox_inches="tight")
                saved_pngs.append("04_학급별_평균.png")
                # 문항별 정답률/변별도
                f = charts.fig_item_difficulty(self.item_stats)
                f.savefig(graph_dir / "05_문항_정답률.png", dpi=150, bbox_inches="tight")
                saved_pngs.append("05_문항_정답률.png")
                f = charts.fig_item_discrimination(self.item_stats)
                f.savefig(graph_dir / "06_문항_변별도.png", dpi=150, bbox_inches="tight")
                saved_pngs.append("06_문항_변별도.png")
                # 성취기준
                f = charts.fig_standard_attainment(ov.by_standard)
                f.savefig(graph_dir / "07_성취기준별_정답률.png", dpi=150, bbox_inches="tight")
                saved_pngs.append("07_성취기준별_정답률.png")

            extra = f"\n그래프 {len(saved_pngs)}장 → 그래프/ 폴더" if saved_pngs else ""
            QMessageBox.information(self, "완료",
                                    f"CSV 4개{extra}\n저장 위치: {out}")
        except Exception as e:
            QMessageBox.critical(self, "오류", f"{e}")

    @staticmethod
    def _write_csv(path: Path, headers, rows):
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(headers)
            w.writerows(rows)


def run():
    app = QApplication.instance() or QApplication(sys.argv)
    app.setApplicationName(APP_TITLE)
    app.setOrganizationName(APP_AUTHOR)
    icon_path = _app_icon_path()
    if icon_path is not None:
        app.setWindowIcon(QIcon(str(icon_path)))
    # 테마 매니저 초기화 (자동 모드)
    tm = ThemeManager(app)
    tm.set_mode("auto")
    win = MainWindow(theme_manager=tm)
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    run()
