from __future__ import annotations

import math
import re


GRADE5_CUMULATIVE = [("1", 10.0), ("2", 34.0), ("3", 66.0), ("4", 90.0), ("5", 100.0)]
NON_SCORE_STATUS_KEYWORDS = (
    "인정결", "질병결", "미인정결", "기타결", "결시", "결석",
    "자퇴", "전출", "전입", "위탁", "면제", "유예",
)


def format_score(value: float) -> str:
    text = f"{float(value):.2f}"
    return text.rstrip("0").rstrip(".")


def _score_or_none(value, max_score: float = 100.0) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        score = float(value)
        if 0 <= score <= max_score + max(1.0, max_score * 0.01):
            return score
    return None


def _detect_max_score(ws) -> float:
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12), values_only=True):
        text = " ".join(str(v) for v in row if v not in (None, ""))
        match = re.search(r"만점\s*[:：]\s*([0-9]+(?:\.[0-9]+)?)", text)
        if match:
            try:
                return max(1.0, float(match.group(1)))
            except Exception:
                pass
    return 100.0


def _detect_subject(ws, fallback: str) -> str:
    joined = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 8), values_only=True):
        joined.append(" ".join(str(v) for v in row if v not in (None, "")))
    text = " ".join(joined)
    match = re.search(r"교과목\s*[:：]\s*(.*?)\s+만점", text)
    if match and match.group(1).strip():
        return match.group(1).strip()
    for line in joined:
        if "교과목" in line:
            return line.strip()
    return fallback


def _extract_matrix_scores(ws, max_score: float) -> tuple[list[float], str, list[dict]]:
    header_row = None
    label_col = None
    for row_idx in range(1, min(ws.max_row, 30) + 1):
        for col_idx in range(1, min(ws.max_column, 12) + 1):
            first = str(ws.cell(row_idx, col_idx).value or "").replace(" ", "").replace("\n", "")
            filled_after = [
                next_col
                for next_col in range(col_idx + 1, ws.max_column + 1)
                if ws.cell(row_idx, next_col).value not in (None, "")
            ]
            if ("반번호" in first or first in {"번호", "반/번호"}) and len(filled_after) >= 2:
                header_row = row_idx
                label_col = col_idx
                break
        if header_row:
            break
    if not header_row:
        return [], "", []

    columns = [
        col_idx
        for col_idx in range((label_col or 1) + 1, ws.max_column + 1)
        if ws.cell(header_row, col_idx).value not in (None, "")
    ]
    scores: list[float] = []
    excluded_entries: list[dict] = []
    row_count = 0
    for row_idx in range(header_row + 1, ws.max_row + 1):
        first_text = str(ws.cell(row_idx, label_col or 1).value or "").replace(" ", "").replace("\n", "")
        if any(key in first_text for key in ("응시생수", "총점", "평균", "표준편차")):
            break
        row_scores = []
        for col_idx in columns:
            value = ws.cell(row_idx, col_idx).value
            score = _score_or_none(value, max_score)
            if score is not None:
                row_scores.append(score)
            else:
                text = str(value or "").strip()
                compact = text.replace(" ", "").replace("\n", "")
                if compact and any(keyword in compact for keyword in NON_SCORE_STATUS_KEYWORDS):
                    excluded_entries.append({
                        "row": row_idx,
                        "col": col_idx,
                        "class": str(ws.cell(header_row, col_idx).value or "").strip(),
                        "number": str(ws.cell(row_idx, label_col or 1).value or "").strip(),
                        "status": text,
                    })
        if row_scores:
            scores.extend(row_scores)
            row_count += 1
    source_note = f"반번호 점수표 · {len(columns)}개 반 열 × {row_count}개 번호 행"
    if excluded_entries:
        source_note += f" · 비점수 제외 {len(excluded_entries)}건"
    return scores, source_note, excluded_entries


def _report_priority(report: dict) -> int:
    sheet = str(report.get("sheet", ""))
    note = str(report.get("source_note", ""))
    score = 0
    if "일람표" in sheet:
        score += 30
    if "자료" == sheet:
        score += 10
    if "반번호 점수표" in note:
        score += 5
    if "등급" in sheet or "학생 정보" in sheet:
        score -= 10
    return score


def _extract_column_scores(ws, max_score: float) -> tuple[list[float], str]:
    preferred = [
        ("환산점수", 120), ("환산점", 115), ("과목총점", 105), ("총점", 100),
        ("합계", 90), ("원점수", 85), ("점수", 65),
    ]
    avoid = ("반번호", "번호", "학번", "성명", "이름", "석차", "등급", "성취", "평균", "표준", "응시", "문항", "배점")
    best: tuple[int, int, int, list[float], str] | None = None
    for header_row in range(1, min(ws.max_row, 35) + 1):
        headers = [
            str(ws.cell(header_row, col_idx).value or "").replace(" ", "")
            for col_idx in range(1, ws.max_column + 1)
        ]
        if sum(1 for header in headers if header) < 2:
            continue
        for col_idx, header in enumerate(headers, start=1):
            if not header or any(word in header for word in avoid):
                continue
            priority = 0
            for term, weight in preferred:
                if term in header:
                    priority = max(priority, weight)
            if priority <= 0:
                continue
            values = []
            for row_idx in range(header_row + 1, ws.max_row + 1):
                score = _score_or_none(ws.cell(row_idx, col_idx).value, max_score)
                if score is not None:
                    values.append(score)
            if len(values) < 3:
                continue
            candidate = (priority, len(values), -header_row, values, f"'{header}' 열 · {len(values)}명")
            if best is None or candidate[:3] > best[:3]:
                best = candidate
    if best:
        return best[3], best[4]
    return [], ""


def load_grade5_cut_reports(path: str) -> list[dict]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=False)
    reports = []
    signatures: dict[tuple[float, ...], int] = {}

    def add_report(report: dict):
        signature = tuple(sorted(round(float(score), 6) for score in report["scores"]))
        previous_idx = signatures.get(signature)
        if previous_idx is None:
            signatures[signature] = len(reports)
            reports.append(report)
        elif _report_priority(report) > _report_priority(reports[previous_idx]):
            reports[previous_idx] = report

    for sheet in workbook.worksheets:
        max_score = _detect_max_score(sheet)
        subject = _detect_subject(sheet, sheet.title)
        scores, source_note, excluded_entries = _extract_matrix_scores(sheet, max_score)
        if len(scores) < 3:
            scores, source_note = _extract_column_scores(sheet, max_score)
            excluded_entries = []
        if len(scores) >= 3:
            report = {
                "sheet": sheet.title,
                "subject": subject,
                "max_score": max_score,
                "scores": scores,
                "source_note": source_note,
                "excluded_entries": excluded_entries,
            }
            add_report(report)

    if not reports:
        raise ValueError("점수표를 찾지 못했습니다. '반번호' 점수표 또는 환산점수/총점/원점수 열이 있는 엑셀인지 확인해 주세요.")
    return reports


def _rounded_cumulative_limits(total: int, cumulative: list[tuple[str, float]]) -> list[tuple[str, int, float]]:
    limits = []
    for grade, boundary in cumulative:
        limit = max(1, min(total, int(math.floor(total * boundary / 100.0 + 0.5))))
        limits.append((grade, limit, boundary))
    if limits:
        last_grade, _, last_boundary = limits[-1]
        limits[-1] = (last_grade, total, last_boundary)
    return limits


def _grade_for_rank(rank: float, limits: list[tuple[str, int, float]]) -> str:
    for grade, limit, _ in limits:
        if rank <= limit + 1e-9:
            return grade
    return limits[-1][0]


def _grade_for_middle_percent(mid_rank: float, total: int, cumulative: list[tuple[str, float]]) -> str:
    pct = mid_rank / total * 100.0
    for grade, boundary in cumulative:
        if pct <= boundary + 1e-9:
            return grade
    return cumulative[-1][0]


def official_grade_groups(scores: list[float], cumulative: list[tuple[str, float]]) -> list[dict]:
    """Return score groups graded by school-record rules.

    Basis:
    - School Records Support Portal, School Record Creation and Management
      Guideline effective 2026-03-01, Ministry of Education Directive No. 555.
    - 2026 High School School Records Writing Guide.
      https://star.moe.go.kr/web/contents/m20103.do
      https://star.moe.go.kr/web/contents/m21100.do

    First, cumulative student counts are computed by rounding
    total_students * cumulative_percent. If a same-score group crosses one of
    those rounded boundaries, that group is graded by middle-rank percentage.
    """
    values = [float(score) for score in scores]
    total = len(values)
    if total == 0:
        return []
    counts = {value: values.count(value) for value in set(values)}
    limits = _rounded_cumulative_limits(total, cumulative)
    groups = []
    greater = 0
    for score in sorted(counts, reverse=True):
        count = counts[score]
        start = greater + 1
        end = greater + count
        mid = start + (count - 1) / 2.0
        crosses_boundary = any(start <= limit < end for _, limit, _ in limits[:-1])
        if crosses_boundary:
            grade = _grade_for_middle_percent(mid, total, cumulative)
        else:
            grade = _grade_for_rank(start, limits)
        groups.append({
            "score": score,
            "count": count,
            "start": start,
            "end": end,
            "middle": mid,
            "middle_pct": mid / total * 100.0,
            "grade": grade,
            "crosses_boundary": crosses_boundary,
        })
        greater += count
    return groups


def grade_boundary_notes(scores: list[float], cumulative: list[tuple[str, float]]) -> list[dict]:
    groups = official_grade_groups(scores, cumulative)
    total = len(scores)
    if not groups or total == 0:
        return []
    limits = _rounded_cumulative_limits(total, cumulative)
    cut_rows = cumulative_cut_rows(scores, cumulative)
    by_score = {float(group["score"]): group for group in groups}
    notes = []
    for idx, cut_row in enumerate(cut_rows):
        grade = str(cut_row["grade"])
        score = float(cut_row["score"])
        cut_group = by_score.get(score)
        official_limit = limits[idx][1] if idx < len(limits) else total
        official_group = next(
            (group for group in groups if group["start"] <= official_limit <= group["end"]),
            cut_group,
        )
        messages: list[str] = []
        if idx < len(cut_rows) - 1 and int(cut_row["rank"]) != official_limit:
            official_score = official_group["score"] if official_group else score
            if abs(float(official_score) - score) > 1e-9:
                messages.append(
                    f"생활기록부 누적인원 기준은 {official_limit}명, 하한 점수는 {format_score(official_score)}점입니다."
                )
            else:
                messages.append(f"생활기록부 누적인원 기준은 {official_limit}명입니다.")
        if idx < len(cut_rows) - 1 and cut_group and int(cut_group["count"]) > 1:
            messages.append(
                f"컷 점수 동점 {cut_group['count']}명({cut_group['start']}~{cut_group['end']}등)."
            )
        if official_group and official_group.get("crosses_boundary"):
            messages.append(
                f"경계 동점은 중간석차 {official_group['middle']:.1f}명"
                f"({official_group['middle_pct']:.2f}%) 기준으로 {official_group['grade']}등급입니다."
            )
        if not messages:
            messages.append("경계 동점 특이사항 없음.")
        notes.append({
            "grade": grade,
            "score": score,
            "rank": cut_row["rank"],
            "official_limit": official_limit,
            "messages": messages,
        })
    return notes


def relative_grade_labels(scores: list[float], cumulative: list[tuple[str, float]]) -> list[str]:
    cut_rows = cumulative_cut_rows(scores, cumulative)
    if not cut_rows:
        return []
    labels = []
    for score in scores:
        value = float(score)
        grade = cut_rows[-1]["grade"]
        for row in cut_rows:
            if value >= float(row["score"]) - 1e-9:
                grade = row["grade"]
                break
        labels.append(grade)
    return labels


def relative_grade_cut_points(scores: list[float], cumulative: list[tuple[str, float]], prefix: str) -> list[dict]:
    cut_points = []
    for row in cumulative_cut_rows(scores, cumulative)[:-1]:
        grade = str(row["grade"])
        cut_points.append({
            "score": float(row["score"]),
            "label": f"{prefix} {grade}/{int(grade) + 1 if str(grade).isdigit() else ''}".rstrip("/"),
            "kind": prefix,
        })
    return cut_points


def cumulative_cut_rows(scores: list[float], cumulative: list[tuple[str, float]]) -> list[dict]:
    sorted_scores = sorted([float(value) for value in scores], reverse=True)
    total = len(sorted_scores)
    if total == 0:
        return []
    cut_rows = []
    for grade, boundary in cumulative:
        raw_rank = total * boundary / 100.0
        # The reference workbook uses Excel INT(ratio * total) as the LARGE rank.
        rank = max(1, min(total, int(math.floor(raw_rank))))
        score = sorted_scores[rank - 1]
        cut_rows.append({
            "grade": grade,
            "boundary": boundary,
            "rank": rank,
            "raw_rank": raw_rank,
            "score": score,
        })
    return cut_rows


def grade5_cut_summary(scores: list[float]) -> dict:
    sorted_scores = sorted([float(value) for value in scores], reverse=True)
    total = len(sorted_scores)
    cut_rows = cumulative_cut_rows(sorted_scores, GRADE5_CUMULATIVE)
    labels = relative_grade_labels(sorted_scores, GRADE5_CUMULATIVE)
    counts = {grade: labels.count(grade) for grade, _ in GRADE5_CUMULATIVE}
    official_groups = official_grade_groups(sorted_scores, GRADE5_CUMULATIVE)
    official_counts = {
        grade: sum(int(group["count"]) for group in official_groups if str(group["grade"]) == grade)
        for grade, _ in GRADE5_CUMULATIVE
    }
    return {
        "n": total,
        "min": min(sorted_scores),
        "max": max(sorted_scores),
        "mean": sum(sorted_scores) / total,
        "cut_rows": cut_rows,
        "counts": counts,
        "official_counts": official_counts,
        "boundary_notes": grade_boundary_notes(sorted_scores, GRADE5_CUMULATIVE),
    }
