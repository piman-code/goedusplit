"""
NEIS 지필 정오표, 문항정보표, (선택) 수행평가 결과 엑셀 파일을 읽어
표준화된 파이썬 객체로 변환한다.

- 입력 파일은 셀 병합·머리글이 복잡하므로, 키워드 셀 위치를 탐색한 뒤
  상대 좌표로 데이터를 추출한다.
- 본 모듈은 GUI에 종속되지 않으므로, CLI/테스트에서도 동일하게 사용 가능.

주의(추정): 첨부된 샘플(2026 1학기 1차) 두 개를 기준으로 작성했으며,
  실제 학교/학년/과목별 양식이 일부 다를 수 있다. 그런 경우 _find_* 함수의
  키워드 인식 부분을 보강하면 된다.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl


# ---------------------------------------------------------------------------
# 데이터 구조
# ---------------------------------------------------------------------------

@dataclass
class ItemInfo:
    """문항 한 개의 메타 정보."""
    number: int                  # 문항번호 (선택형/서답형 각각 1부터)
    item_type: str               # '선택형' or '서답형'
    content_area: str = ""       # 내용영역 (e.g. '다항식의 연산')
    standard: str = ""           # 성취기준 (전체 문장)
    standard_code: str = ""      # 성취기준 코드 (e.g. '[10공수1-01-01]')
    difficulty: str = ""         # '쉬움' / '보통' / '어려움'
    score: float = 0.0           # 배점
    answer: str = ""             # 정답 (선택형은 1~5, 서답형은 임의 문자열)


@dataclass
class StudentResponse:
    """학생 한 명의 응답."""
    sid: str                     # 학번 (NEIS의 '반/번호'에 더해 학번까지)
    class_no: str                # '1/3' 형식
    grade_class: str = ""        # 학급 (e.g. '1')
    name: str = ""               # 학생 이름 (NEIS 정오표 '번호' 컬럼)
    answers: dict = field(default_factory=dict)  # item_number -> 응답 값. 선택형: '.' 또는 '1'~'5'
    serdap_score: float = 0.0    # 서답형 점수
    multi_score: float = 0.0     # 선택형 점수
    etc_score: float = 0.0       # 기타점수
    total: float = 0.0           # 과목 총점 (지필 100점 만점)
    perform_score: float = 0.0   # 수행평가 환산점수 (100점 만점)
    final_score: float = 0.0     # 반영비율 합산 환산점수 (100점 만점)


@dataclass
class ExamData:
    """한 시험(단일 시트) 분석에 필요한 모든 정보."""
    subject: str = ""            # 과목명
    grade: str = ""              # 학년
    semester: str = ""           # 학기/회차
    items: list[ItemInfo] = field(default_factory=list)  # 선택형 + 서답형
    students: list[StudentResponse] = field(default_factory=list)
    cut_scores: dict = field(default_factory=lambda: {  # 성취수준별 분할점수
        "A": 90.0, "B": 80.0, "C": 70.0, "D": 60.0, "E": 40.0
    })
    source_files: dict = field(default_factory=dict)
    # 반영비율 (지필 + 수행 = 100). 수행 0이면 지필만 분석
    weight_pencil: float = 100.0
    weight_perform: float = 0.0
    use_perform: bool = False

    @property
    def select_items(self) -> list[ItemInfo]:
        return [it for it in self.items if it.item_type == "선택형"]

    @property
    def serdap_items(self) -> list[ItemInfo]:
        return [it for it in self.items if it.item_type == "서답형"]


# ---------------------------------------------------------------------------
# 시트 헬퍼
# ---------------------------------------------------------------------------

def _ws_iter(ws):
    """모든 (row, col, value) 셀 값(2D list)으로 반환."""
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _find_cell(grid, keyword) -> tuple[int, int] | None:
    """grid에서 keyword가 들어간 첫 셀의 (row, col)을 반환. 1-indexed."""
    for r, row in enumerate(grid):
        for c, v in enumerate(row):
            if v is None:
                continue
            if isinstance(v, str) and keyword in v:
                return (r, c)
    return None


# ---------------------------------------------------------------------------
# 문항정보표 파서
# ---------------------------------------------------------------------------

_STD_CODE_RE = re.compile(r"^\s*(\[[^\]]+\])\s*(.*)$", re.DOTALL)


def _parse_standard(raw):
    """성취기준 셀에서 코드와 본문을 분리.

    예: '[10공수1-01-01]다항식의 사칙연산의 원리를 ...'
        -> ('[10공수1-01-01]', '다항식의 사칙연산의 원리를 ...')
    """
    if not raw:
        return "", ""
    s = str(raw)
    m = _STD_CODE_RE.match(s)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return "", s.strip()


def _difficulty_from_columns(row, cols_easy, cols_mid, cols_hard) -> str:
    """난이도 열에서 ○가 표시된 위치로 난이도 결정."""
    def has_mark(idxs):
        for i in idxs:
            v = row[i] if i < len(row) else None
            if v is not None and str(v).strip():
                return True
        return False

    if has_mark(cols_hard):
        return "어려움"
    if has_mark(cols_mid):
        return "보통"
    if has_mark(cols_easy):
        return "쉬움"
    return ""


def load_item_info(path) -> tuple[list[ItemInfo], dict]:
    """문항정보표 엑셀 -> ItemInfo 리스트 + 메타.

    동작 요약:
      1. '문항번호' 셀 위치(헤더)를 찾는다.
      2. 그 아래 행부터 숫자 문항번호가 있는 행을 누적한다.
      3. '서답형 문항' 헤더가 나오면 이후의 문항을 '서답형'으로 분류.
      4. '난이도' 옆 컬럼들(어려움/보통/쉬움) 위치는 헤더 두 줄 아래에서 추정.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    grid = _ws_iter(ws)

    # 헤더 위치
    header = _find_cell(grid, "문항번호")
    if header is None:
        raise ValueError("문항정보표에서 '문항번호' 헤더를 찾을 수 없습니다.")

    hr, hc = header
    # 난이도 헤더 컬럼
    diff_col = None
    score_col = None
    answer_col = None
    content_col = None
    standard_col = None
    for c, v in enumerate(grid[hr]):
        if v is None:
            continue
        s = str(v).strip()
        if s == "내용영역":
            content_col = c
        elif s == "성취기준":
            standard_col = c
        elif s == "난이도":
            diff_col = c
        elif s == "배점":
            score_col = c
        elif s == "정답":
            answer_col = c

    # 난이도는 다음 행에 어려움/보통/쉬움이 가로로 펼쳐짐
    sub_header = grid[hr + 1] if hr + 1 < len(grid) else []
    cols_easy, cols_mid, cols_hard = [], [], []
    if diff_col is not None:
        for c in range(max(0, diff_col - 1), min(len(sub_header), diff_col + 8)):
            v = sub_header[c]
            if v is None:
                continue
            s = str(v).strip()
            if s == "쉬움":
                cols_easy.append(c)
            elif s == "보통":
                cols_mid.append(c)
            elif s == "어려움":
                cols_hard.append(c)

    # 행 순회
    items: list[ItemInfo] = []
    item_type = "선택형"
    seen_select_header = True   # '선택형 문항' 헤더는 이미 hr 위에 있다고 가정

    for r in range(hr + 2, len(grid)):
        row = grid[r]
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue

        # '서답형 문항' 헤더 감지
        joined = " ".join(str(v) for v in row if v is not None)
        if "서답형 문항" in joined:
            item_type = "서답형"
            continue
        if "선택형 문항" in joined:
            item_type = "선택형"
            continue
        # 합계 행 등 스킵
        if "총합계" in joined or "비율(%)" in joined:
            continue
        if "/" in joined and len(joined) < 30:
            # 페이지 번호/푸터 스킵
            continue

        # 문항번호
        no_val = row[hc] if hc < len(row) else None
        if no_val is None:
            continue
        try:
            no = int(no_val)
        except (TypeError, ValueError):
            continue

        content = str(row[content_col]).strip() if content_col is not None and row[content_col] else ""
        std_raw = row[standard_col] if standard_col is not None else None
        std_code, std_body = _parse_standard(std_raw)
        diff = _difficulty_from_columns(row, cols_easy, cols_mid, cols_hard)
        score = row[score_col] if score_col is not None else None
        try:
            score = float(score) if score is not None and score != "" else 0.0
        except (TypeError, ValueError):
            score = 0.0
        ans = row[answer_col] if answer_col is not None else ""
        ans = str(ans).strip() if ans is not None else ""

        items.append(ItemInfo(
            number=no, item_type=item_type,
            content_area=content,
            standard=std_body, standard_code=std_code,
            difficulty=diff, score=score, answer=ans,
        ))

    # 메타 추출
    meta = {"path": str(path)}
    subj = _find_cell(grid, "과목")
    if subj:
        # 셀 자체가 'XXX 과목' 형태일 수도 있고, 별도 라벨일 수도 있음. 여기선 셀 통째로 보관.
        meta["subject_cell"] = str(grid[subj[0]][subj[1]])
    grade = _find_cell(grid, "학년")
    if grade:
        meta["grade_cell"] = str(grid[grade[0]][grade[1]])

    return items, meta


# ---------------------------------------------------------------------------
# 학생답 정오표 파서
# ---------------------------------------------------------------------------

def load_student_responses(path) -> tuple[list[StudentResponse], dict, dict]:
    """학생답 정오표 -> StudentResponse 리스트 + 정답표(item->answer) + 메타.

    형식 (샘플 기준):
      4행: '반/번호', None, '번호', 1, 2, ..., N, '선택형점수', '서답형점수', '기타점수', '과목총점'
      5행: 정답 행 (각 문항 정답)
      6행: 배점 행
      7행 이후: 학생 데이터. 응답값이 '.'이면 정답, 숫자면 학생이 고른 오답 번호.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    grid = _ws_iter(ws)

    header_pos = _find_cell(grid, "반/번호")
    if header_pos is None:
        raise ValueError("정오표에서 '반/번호' 헤더를 찾을 수 없습니다.")
    hr, _ = header_pos
    header = grid[hr]

    # 문항번호 컬럼 매핑
    item_cols: dict[int, int] = {}
    for c, v in enumerate(header):
        if isinstance(v, (int, float)) and float(v).is_integer():
            item_cols[int(v)] = c
    # 마지막 컬럼들
    sum_cols = {}
    for c, v in enumerate(header):
        if isinstance(v, str):
            s = v.strip()
            if s in ("선택형점수", "서답형점수", "기타점수", "과목총점"):
                sum_cols[s] = c

    answer_row = grid[hr + 1] if hr + 1 < len(grid) else []
    score_row = grid[hr + 2] if hr + 2 < len(grid) else []
    answers: dict[int, str] = {}
    item_scores: dict[int, float] = {}
    for no, c in item_cols.items():
        a = answer_row[c] if c < len(answer_row) else None
        s = score_row[c] if c < len(score_row) else None
        answers[no] = str(a).strip() if a is not None else ""
        try:
            item_scores[no] = float(s) if s not in (None, "") else 0.0
        except (TypeError, ValueError):
            item_scores[no] = 0.0

    students: list[StudentResponse] = []
    sid_col = 0
    name_col = None     # NEIS '번호' 헤더 컬럼 = 학생 이름이 들어있는 셀
    # '반/번호' 셀의 컬럼 위치
    for c, v in enumerate(header):
        if isinstance(v, str) and v.strip() == "반/번호":
            sid_col = c
        if isinstance(v, str) and v.strip() in ("번호", "성명", "이름", "학생명"):
            name_col = c

    for r in range(hr + 3, len(grid)):
        row = grid[r]
        if not row:
            continue
        sid = row[sid_col] if sid_col < len(row) else None
        if sid is None or (isinstance(sid, str) and not sid.strip()):
            continue
        # 학번이 정수/문자 모두 가능
        try:
            sid_str = str(int(sid))
        except (TypeError, ValueError):
            sid_str = str(sid).strip()

        # '반/번호' 컬럼 다음 셀이 '1/3'식 학급/번호 표기 (NEIS 양식 기준)
        cls_val = row[sid_col + 1] if sid_col + 1 < len(row) else ""
        cls_str = str(cls_val).strip() if cls_val else ""
        # '/'가 없거나 명백히 비정상이면 빈 문자열로
        if "/" not in cls_str:
            cls_str = ""
        grade_class = cls_str.split("/")[0] if "/" in cls_str else ""

        # 학생 이름 (있을 때)
        name_val = ""
        if name_col is not None and name_col < len(row):
            v = row[name_col]
            if v is not None and str(v).strip() not in ("", ".", "정답", "배점"):
                name_val = str(v).strip()

        ans_map = {}
        for no, c in item_cols.items():
            v = row[c] if c < len(row) else None
            ans_map[no] = "" if v is None else str(v).strip()

        def gnum(name):
            c = sum_cols.get(name)
            if c is None or c >= len(row):
                return 0.0
            v = row[c]
            try:
                return float(v) if v not in (None, "") else 0.0
            except (TypeError, ValueError):
                return 0.0

        total_v = gnum("과목총점")
        students.append(StudentResponse(
            sid=sid_str, class_no=cls_str, grade_class=grade_class,
            name=name_val,
            answers=ans_map,
            multi_score=gnum("선택형점수"),
            serdap_score=gnum("서답형점수"),
            etc_score=gnum("기타점수"),
            total=total_v,
            final_score=total_v,   # 수행평가 미적용 시 지필점수가 곧 환산점수
        ))

    meta = {"path": str(path)}
    # 시험 정보 (3행에 보통 '2026학년도 ...' 문자열)
    info_pos = _find_cell(grid, "학년도")
    if info_pos:
        meta["info_text"] = str(grid[info_pos[0]][info_pos[1]])
    return students, answers, meta


# ---------------------------------------------------------------------------
# 통합 로딩
# ---------------------------------------------------------------------------

def apply_perform(exam: "ExamData", perform_data, weight_pencil: float, weight_perform: float):
    """수행평가 결과를 학생 레코드에 합치고 환산점수(final_score)를 계산.

    환산점수 = 지필점수(100점) × 지필반영비율 + 수행평가(100점환산) × 수행반영비율
    반영비율은 합 100을 가정. 0이거나 합이 100이 아닐 때 자동 정규화.
    """
    total_w = float(weight_pencil) + float(weight_perform)
    if total_w <= 0:
        wp, ws = 1.0, 0.0
    else:
        wp = float(weight_pencil) / total_w
        ws = float(weight_perform) / total_w
    exam.weight_pencil = float(weight_pencil)
    exam.weight_perform = float(weight_perform)
    exam.use_perform = perform_data is not None and weight_perform > 0

    name_idx = {}
    if perform_data is not None:
        for sid, rec in perform_data.records.items():
            name_idx[(sid, rec.class_no, rec.name)] = rec
            name_idx[(sid,)] = rec

    for st in exam.students:
        rec = None
        if perform_data is not None:
            rec = perform_data.records.get(st.sid) or perform_data.by_classno.get(st.class_no)
        if rec is not None:
            st.perform_score = rec.pct100
            # 정오표에 이름이 비어있다면 수행평가 이름으로 보충
            if not st.name and rec.name:
                st.name = rec.name
        else:
            st.perform_score = 0.0
        if exam.use_perform:
            st.final_score = st.total * wp + st.perform_score * ws
        else:
            st.final_score = st.total


def load_exam(item_info_path, response_path) -> ExamData:
    items, item_meta = load_item_info(item_info_path)
    students, answers, resp_meta = load_student_responses(response_path)

    # 정답 일관성 보정: 정오표에 정답이 더 신뢰 가능 → 우선 적용
    for it in items:
        if it.item_type == "선택형" and it.number in answers:
            ans_from_resp = answers[it.number]
            if ans_from_resp:
                it.answer = ans_from_resp

    exam = ExamData(items=items, students=students)
    exam.source_files["item_info"] = str(item_info_path)
    exam.source_files["responses"] = str(response_path)

    info_text = resp_meta.get("info_text", "")
    if info_text:
        m = re.search(r"(\d{4})학년도\s+(\d+)학기", info_text)
        if m:
            exam.semester = f"{m.group(1)}학년도 {m.group(2)}학기"
        # '2026학년도' 다음의 '1학년' 식 표기를 우선 매치
        m2 = re.search(r"(?<!학년)(?<!\d)(\d{1,2})\s*학년(?!도)", info_text)
        if m2:
            exam.grade = f"{m2.group(1)}학년"
        m3 = re.search(r"수학[:\s]*([^\s,]+)", info_text)
        if m3:
            exam.subject = m3.group(1)

    if not exam.subject:
        sc = item_meta.get("subject_cell", "")
        m = re.search(r"\(\s*([^\)]+)\s*\)\s*과목", sc)
        if m:
            exam.subject = m.group(1).strip()

    return exam


if __name__ == "__main__":
    import sys, json
    if len(sys.argv) < 3:
        print("usage: data_loader.py <item_info.xlsx> <responses.xlsx>")
        sys.exit(1)
    e = load_exam(sys.argv[1], sys.argv[2])
    print("subject:", e.subject, "grade:", e.grade, "semester:", e.semester)
    print("items:", len(e.items), "students:", len(e.students))
    for it in e.items[:3]:
        print(" -", it)
    for st in e.students[:2]:
        print(" *", st)
